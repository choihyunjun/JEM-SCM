import csv
import io
from datetime import datetime
from decimal import Decimal
from typing import Iterable, Dict, Any, List, Tuple, Optional

from openpyxl import load_workbook

def _normalize(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "").replace("_", "")

def parse_erp_stock_file(file_name: str, content: bytes) -> List[Dict[str, Any]]:
    """
    Return list of dict rows with keys: warehouse_code, part_no, qty_onhand.
    Accepts CSV or XLSX. Uses flexible header matching.
    """
    if file_name.lower().endswith(".csv"):
        text = content.decode("utf-8-sig", errors="ignore")
        return _parse_csv(text)
    # default: try xlsx
    return _parse_xlsx(content)

def _map_headers(headers: List[str]) -> Dict[str, int]:
    norm = [_normalize(h) for h in headers]

    def find_any(cands: Iterable[str]) -> Optional[int]:
        for c in cands:
            c = _normalize(c)
            for i, h in enumerate(norm):
                if c in h or h in c:
                    return i
        return None

    wh = find_any(["warehouse", "wh", "창고", "저장위치", "store", "location", "보관장소"])
    part = find_any(["partno", "품번", "item", "itemcode", "품목코드", "품목", "자재코드"])
    qty = find_any(["qty", "수량", "재고", "onhand", "현재고", "재고수량", "잔량"])

    if wh is None or part is None or qty is None:
        raise ValueError("필수 컬럼(창고/품번/수량)을 찾지 못했습니다. 헤더명을 확인해주세요.")
    return {"warehouse_code": wh, "part_no": part, "qty_onhand": qty}

def _parse_csv(text: str) -> List[Dict[str, Any]]:
    f = io.StringIO(text)
    reader = csv.reader(f)
    rows = list(reader)
    if not rows:
        return []
    header = rows[0]
    idx = _map_headers(header)
    out = []
    for r in rows[1:]:
        if not r or len(r) <= max(idx.values()):
            continue
        wh = str(r[idx["warehouse_code"]]).strip()
        part = str(r[idx["part_no"]]).strip()
        qty_raw = str(r[idx["qty_onhand"]]).strip()
        if not (wh and part):
            continue
        try:
            qty = Decimal(qty_raw.replace(",", ""))
        except Exception:
            continue
        out.append({"warehouse_code": wh, "part_no": part, "qty_onhand": qty})
    return out

def _parse_xlsx(content: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = ["" if v is None else str(v) for v in rows[0]]
    idx = _map_headers(header)
    out = []
    for r in rows[1:]:
        if r is None:
            continue
        # pad
        r = list(r)
        if len(r) <= max(idx.values()):
            continue
        wh = "" if r[idx["warehouse_code"]] is None else str(r[idx["warehouse_code"]]).strip()
        part = "" if r[idx["part_no"]] is None else str(r[idx["part_no"]]).strip()
        qty_cell = r[idx["qty_onhand"]]
        if not (wh and part):
            continue
        try:
            if qty_cell is None:
                continue
            qty = Decimal(str(qty_cell).replace(",", ""))
        except Exception:
            continue
        out.append({"warehouse_code": wh, "part_no": part, "qty_onhand": qty})
    return out
