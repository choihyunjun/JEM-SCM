# orders/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User
from django.db.models import Q, Sum, Case, When, Value, IntegerField, Max, Count, F
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.db import transaction
from functools import wraps
import openpyxl
import datetime
from datetime import timedelta, date

# SCM 모델 임포트 (ReturnLog, VendorMonthlyPerformance, Notice, QnA, UserProfile, InventoryUploadLog 추가)
from .models import Order, Vendor, Part, Inventory, Incoming, LabelPrintLog, DeliveryOrder, DeliveryOrderItem, Demand, ReturnLog, VendorMonthlyPerformance, Notice, QnA, UserProfile, Organization, InventoryUploadLog

# [신규] 타 앱(WMS, QMS) 모델 임포트 (연동용)
try:
    from material.models import Warehouse, MaterialStock, MaterialTransaction
    from qms.models import ImportInspection
except ImportError:
    Warehouse = None
    MaterialStock = None
    MaterialTransaction = None
    ImportInspection = None

# ==========================================
# [0. 필수 공통 로직 및 권한 설정]
# ==========================================

def _get_profile(user):
    return getattr(user, 'profile', None)

def _get_role(user) -> str:
    profile = _get_profile(user)
    if not profile:
        return 'VENDOR'
    role = getattr(profile, 'role', None)
    if role == 'ADMIN':
        return 'ADMIN'
    if role == 'VENDOR':
        return 'VENDOR'
    return 'STAFF'

def _is_internal(user) -> bool:
    profile = _get_profile(user)
    if not profile:
        return False
    if hasattr(profile, 'account_type'):
        return profile.account_type == 'INTERNAL'
    return _get_role(user) != 'VENDOR'

def _get_user_vendor(user):
    """협력사 계정의 Vendor 객체 반환 - 기존 방식과 새 방식 모두 지원"""
    # 1. 기존 방식: Vendor.user 필드로 직접 연결
    vendor = Vendor.objects.filter(user=user).first()
    if vendor:
        return vendor

    # 2. 새 방식: UserProfile.org.linked_vendor로 연결
    try:
        if hasattr(user, 'profile') and user.profile.org and user.profile.org.linked_vendor:
            return user.profile.org.linked_vendor
    except Exception:
        pass

    return None

# [Legacy] 기존 role 기반 권한 - 폴백용으로 유지
ROLE_MENU_PERMS = {
    'ADMIN': {'can_view_orders', 'can_register_orders', 'can_view_inventory', 'can_manage_incoming', 'can_access_scm_admin', 'can_manage_parts', 'can_view_reports'},
    'STAFF': {'can_view_orders', 'can_register_orders', 'can_view_inventory', 'can_manage_incoming'},
    'VENDOR': {'can_view_orders', 'can_view_inventory'},
}

ROLE_ACTION_PERMS = {
    'ADMIN': {
        'order.upload', 'order.delete', 'order.close', 'order.approve', 'order.approve_all', 'order.export',
        'inv.upload', 'inv.adjust', 'inv.export',
        'incoming.scan', 'incoming.cancel', 'incoming.export',
        'demand.upload', 'demand.edit', 'demand.delete', 'demand.delete_all', 'demand.export',
        'label.print', 'delivery.print', 'delivery.register', 'delivery.delete',
    },
    'STAFF': {
        'order.upload', 'order.delete', 'order.approve', 'order.export',
        'inv.upload', 'inv.export',
        'incoming.scan', 'incoming.export',
        'demand.upload', 'demand.edit', 'demand.delete', 'demand.export',
        'label.print', 'delivery.print', 'delivery.register', 'delivery.delete',
    },
    'VENDOR': {
        'delivery.register', 'delivery.delete',  # 본인 납품서 삭제 가능
        'label.print', 'delivery.print',
    },
}

def role_has_menu_perm(user, permission_field: str) -> bool:
    """
    메뉴 권한 체크 - UserProfile의 boolean 필드 우선, 없으면 role 기반 폴백
    레거시 필드와 새 필드 모두 체크
    """
    if getattr(user, 'is_superuser', False):
        return True

    # 레거시 → 새 권한 필드 매핑
    LEGACY_TO_NEW = {
        'can_view_orders': 'can_scm_order_view',
        'can_register_orders': 'can_scm_order_edit',
        'can_view_inventory': 'can_scm_inventory_view',
        'can_manage_incoming': 'can_scm_incoming_view',
        'can_manage_parts': 'can_scm_admin',
        'can_view_reports': 'can_scm_report',
        'can_access_scm_admin': 'can_scm_admin',
        'can_view_order': 'can_scm_order_view',  # vendor_delivery_report에서 사용
    }

    profile = _get_profile(user)
    if not profile:
        return False

    # 1. 새 권한 필드 체크 (레거시 필드명이 들어온 경우 매핑)
    new_field = LEGACY_TO_NEW.get(permission_field, permission_field)
    if hasattr(profile, new_field) and getattr(profile, new_field, False):
        return True

    # 2. 레거시 필드도 체크 (호환성)
    if hasattr(profile, permission_field) and getattr(profile, permission_field, False):
        return True

    # 3. 폴백: 기존 role 기반 체크
    role = _get_role(user)
    allowed = ROLE_MENU_PERMS.get(role, set())
    return permission_field in allowed

def has_action_perm(user, action: str) -> bool:
    """
    액션 권한 체크 - role 기반 (기존 방식 유지)
    """
    if getattr(user, 'is_superuser', False):
        return True
    role = _get_role(user)
    return action in ROLE_ACTION_PERMS.get(role, set())

def require_action_perm(request, action: str):
    if has_action_perm(request.user, action):
        return
    messages.error(request, f"권한이 없습니다. (필요 권한: {action})")
    return redirect('order_list')

def scope_qs_for_user(user, qs):
    if _get_role(user) == 'VENDOR':
        v = _get_user_vendor(user)
        if not v:
            return qs.none()
        if hasattr(qs.model, 'vendor_id') or 'vendor' in [f.name for f in qs.model._meta.fields]:
            try:
                return qs.filter(vendor=v)
            except Exception:
                return qs.none()
    return qs

def menu_permission_required(permission_field):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
            if role_has_menu_perm(request.user, permission_field):
                return view_func(request, *args, **kwargs)

            if request.resolver_match.url_name == 'order_list':
                messages.error(request, f"귀하의 계정은 '{permission_field}' 권한이 활성화되지 않았습니다. 관리자에게 문의하세요.")
                return render(request, 'order_list.html', {'orders': [], 'vendor_name': '권한 없음'})

            messages.error(request, "해당 메뉴에 대한 접근 권한이 없습니다.")
            return redirect('order_list')
        return _wrapped_view
    return decorator

@login_required
def login_success(request):
    """로그인 후 권한에 따라 적절한 페이지로 redirect"""
    user = request.user

    # superuser는 SCM 대시보드로
    if user.is_superuser:
        return redirect('scm_alert_dashboard')

    # 프로필이 없으면 기본 대시보드로
    profile = getattr(user, 'profile', None)
    if not profile:
        return redirect('scm_alert_dashboard')

    # 권한에 따라 적절한 페이지로 redirect
    # SCM 권한이 있으면 SCM으로
    if profile.can_scm_order_view or profile.can_scm_label_view or profile.can_scm_incoming_view or profile.can_scm_admin:
        return redirect('scm_alert_dashboard')

    # WMS 권한이 있으면 WMS로
    if profile.can_wms_stock_view or profile.can_wms_inout_view or profile.can_wms_bom_view:
        return redirect('material:dashboard')

    # QMS 권한이 있으면 QMS로
    if profile.can_qms_4m_view or profile.can_qms_inspection_view:
        return redirect('qms:m4_list')

    # 레거시 권한 체크 (호환성)
    if profile.can_view_orders or profile.can_register_orders or profile.can_manage_incoming:
        return redirect('scm_alert_dashboard')

    if profile.can_access_wms or profile.can_wms_inout:
        return redirect('material:dashboard')

    if profile.can_access_qms or profile.can_qms_4m:
        return redirect('qms:m4_list')

    # 기본: SCM 대시보드
    return redirect('scm_alert_dashboard')

# ==========================================
# [1. 발주 조회 화면]
# ==========================================

@login_required
@menu_permission_required('can_view_orders')
def order_list(request):
    user = request.user

    # 협력업체 사용자 판별 (2가지 경로)
    user_vendor = Vendor.objects.filter(user=user).first()
    if not user_vendor and not user.is_superuser:
        try:
            if hasattr(user, 'profile') and user.profile.org and user.profile.org.linked_vendor:
                user_vendor = user.profile.org.linked_vendor
        except Exception:
            pass

    vendor_list = Vendor.objects.all().order_by('name') if user.is_superuser else []
    sort_by = request.GET.get('sort', 'due_date') or 'due_date'

    order_queryset = Order.objects.select_related('vendor').annotate(
        status_priority=Case(
            When(is_closed=True, then=Value(2)),
            When(approved_at__isnull=True, then=Value(0)),
            default=Value(1),
            output_field=IntegerField(),
        )
    )

    if user.is_superuser:
        orders = order_queryset.all().order_by('status_priority', sort_by, '-created_at')
        vendor_name = "전체 관리자"
    elif user_vendor:
        orders = order_queryset.filter(vendor=user_vendor).order_by('status_priority', sort_by, '-created_at')
        vendor_name = user_vendor.name
    else:
        orders = order_queryset.all().order_by('status_priority', sort_by, '-created_at')
        vendor_name = "시스템 운영자"

    selected_vendor = request.GET.get('vendor_id')
    if (user.is_superuser or not user_vendor) and selected_vendor:
        orders = orders.filter(vendor_id=selected_vendor)

    status_filter = request.GET.get('status')
    if status_filter:
        if status_filter == 'unapproved':
            orders = orders.filter(approved_at__isnull=True, is_closed=False)
        elif status_filter == 'approved':
            orders = orders.filter(approved_at__isnull=False, is_closed=False)
        elif status_filter == 'closed':
            orders = orders.filter(is_closed=True)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    q = request.GET.get('q', '')

    if start_date and end_date:
        orders = orders.filter(due_date__range=[start_date, end_date])
    if q:
        orders = orders.filter(Q(part_no__icontains=q) | Q(part_name__icontains=q))

    today = timezone.localtime().date()
    overdue_list = []
    active_overdue = Order.objects.filter(
        due_date__lt=today, is_closed=False, approved_at__isnull=False
    ).select_related('vendor')

    if not (user.is_superuser or (_is_internal(user))):
        active_overdue = active_overdue.filter(vendor=user_vendor) if user_vendor else active_overdue.none()
    elif selected_vendor:
        active_overdue = active_overdue.filter(vendor_id=selected_vendor)

    # N+1 최적화: 모든 미납 발주의 입고 수량을 한 번에 계산
    overdue_ids = list(active_overdue.values_list('id', flat=True))
    overdue_erp_nos = list(active_overdue.exclude(erp_order_no__isnull=True).exclude(erp_order_no='').values_list('erp_order_no', flat=True))

    # linked_order로 연결된 입고 수량 집계
    linked_incoming_by_order = {}
    if overdue_ids:
        linked_qs = DeliveryOrderItem.objects.filter(
            linked_order_id__in=overdue_ids,
            order__status__in=['RECEIVED', 'APPROVED']
        ).values('linked_order_id').annotate(total=Sum('total_qty'))
        for item in linked_qs:
            linked_incoming_by_order[item['linked_order_id']] = item['total'] or 0

    # ERP 발주번호로 연결된 입고 수량 집계
    linked_incoming_by_erp = {}
    if overdue_erp_nos:
        erp_qs = DeliveryOrderItem.objects.filter(
            erp_order_no__in=overdue_erp_nos,
            order__status__in=['RECEIVED', 'APPROVED']
        ).values('erp_order_no').annotate(total=Sum('total_qty'))
        for item in erp_qs:
            linked_incoming_by_erp[item['erp_order_no']] = item['total'] or 0

    for o in active_overdue.order_by('due_date'):
        # 미리 계산된 입고 수량 사용
        linked_incoming = linked_incoming_by_order.get(o.id, 0)

        # linked_order로 입고가 없으면 ERP 발주번호로 확인
        if linked_incoming == 0 and o.erp_order_no:
            linked_incoming = linked_incoming_by_erp.get(o.erp_order_no, 0)

        rem = o.quantity - linked_incoming

        if rem > 0:
            overdue_list.append({
                'due_date': o.due_date,
                'vendor_name': o.vendor.name if o.vendor else "미지정",
                'part_no': o.part_no,
                'order_qty': o.quantity,
                'incoming_qty': linked_incoming,
                'remain_qty': rem
            })

    # ========== 알림 대시보드 요약 데이터 (최적화) ==========
    from material.models import MaterialStock, Warehouse
    from collections import defaultdict

    target_warehouses = Warehouse.objects.filter(code__in=['2000', '4200'])
    target_wh_ids = list(target_warehouses.values_list('id', flat=True))

    # 협력사 필터링을 위한 parts 쿼리
    if user_vendor:
        alert_parts = Part.objects.filter(vendor=user_vendor)
    else:
        alert_parts = Part.objects.all()

    # 1) 재고 부족 예상 품목 수 (N+1 최적화: 배치 쿼리)
    part_ids = list(alert_parts.values_list('id', flat=True))

    # 1-1) 재고 일괄 조회
    stock_map = {}
    if part_ids:
        if target_wh_ids:
            stock_qs = MaterialStock.objects.filter(
                part_id__in=part_ids, warehouse_id__in=target_wh_ids
            ).values('part_id').annotate(total=Sum('quantity'))
        else:
            stock_qs = MaterialStock.objects.filter(
                part_id__in=part_ids
            ).values('part_id').annotate(total=Sum('quantity'))
        stock_map = {item['part_id']: item['total'] or 0 for item in stock_qs}

    # 1-2) 소요량 일괄 조회 (향후 7일)
    demand_map = {}
    if part_ids:
        demand_qs = Demand.objects.filter(
            part_id__in=part_ids,
            due_date__gte=today,
            due_date__lte=today + timedelta(days=7)
        ).values('part_id').annotate(total=Sum('quantity'))
        demand_map = {item['part_id']: item['total'] or 0 for item in demand_qs}

    # 1-3) 입고 예정 일괄 조회 (향후 7일, part_no 기준)
    part_no_map = {p.id: p.part_no for p in alert_parts.only('id', 'part_no')}
    part_nos = list(part_no_map.values())
    incoming_map = {}
    if part_nos:
        incoming_qs = Order.objects.filter(
            part_no__in=part_nos,
            is_closed=False,
            due_date__gte=today,
            due_date__lte=today + timedelta(days=7)
        ).values('part_no').annotate(total=Sum('quantity'))
        incoming_map = {item['part_no']: item['total'] or 0 for item in incoming_qs}

    # 1-4) 부족 품목 수 계산 (메모리에서)
    shortage_count = 0
    for part_id in part_ids:
        wms_stock = stock_map.get(part_id, 0)
        future_demand = demand_map.get(part_id, 0)
        part_no = part_no_map.get(part_id, '')
        pending_incoming = incoming_map.get(part_no, 0)

        expected_stock = wms_stock + pending_incoming - future_demand
        if expected_stock < 0:
            shortage_count += 1

    # 2) 납기 D-3 임박 품목 수
    due_soon_query = Order.objects.filter(
        is_closed=False,
        approved_at__isnull=False,
        due_date__gte=today,
        due_date__lte=today + timedelta(days=3)
    )
    if user_vendor:
        due_soon_query = due_soon_query.filter(vendor=user_vendor)
    due_soon_count = due_soon_query.count()

    # 3) 장기 미입고 (납기 경과 7일 이상)
    long_overdue_query = Order.objects.filter(
        is_closed=False,
        approved_at__isnull=False,
        due_date__lt=today - timedelta(days=7)
    )
    if user_vendor:
        long_overdue_query = long_overdue_query.filter(vendor=user_vendor)
    long_overdue_count = long_overdue_query.count()

    return render(request, 'order_list.html', {
        'orders': orders, 'user_name': user.username, 'vendor_name': vendor_name,
        'q': q, 'vendor_list': vendor_list, 'selected_vendor': selected_vendor,
        'status_filter': status_filter, 'start_date': start_date, 'end_date': end_date,
        'active_menu': 'list', 'current_sort': sort_by,
        'overdue_orders': overdue_list,
        # 알림 대시보드 요약
        'shortage_count': shortage_count,
        'due_soon_count': due_soon_count,
        'long_overdue_count': long_overdue_count,
        'today': today,
    })

# ==========================================
# [2. 발주 관련 액션]
# ==========================================

@login_required
@menu_permission_required('can_scm_order_edit')
def order_upload(request):
    resp = require_action_perm(request, 'order.upload')
    if resp:
        return resp
    if not request.user.is_superuser and _get_role(request.user) != 'STAFF':
        return redirect('order_list')
    return render(request, 'order_upload.html', {'active_menu': 'upload'})

@login_required
def order_upload_template(request):
    """발주 등록용 엑셀 양식 다운로드"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발주등록"

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더 (협력사명 제거 - 품목마스터에서 자동 조회)
    headers = ['품번', '수량', '납기일']
    col_widths = [25, 12, 15]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 예시 데이터 추가 (납기일은 실제 날짜 객체로)
    import datetime as dt
    example_data = [
        ['P9R-12323', 100, dt.date(2026, 2, 1)],
        ['ABC-12345', 200, dt.date(2026, 2, 5)],
    ]
    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 2:  # 수량
                cell.alignment = Alignment(horizontal='right')
            elif col_idx == 3:  # 납기일 - 날짜 서식 적용
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = 'YYYY-MM-DD'

    # C열 전체에 날짜 서식 적용 (사용자 입력용)
    for row_idx in range(4, 1000):  # 4행부터 999행까지 미리 서식 지정
        cell = ws.cell(row=row_idx, column=3)
        cell.number_format = 'YYYY-MM-DD'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="order_upload_template.xlsx"'
    wb.save(response)
    return response

@login_required
@require_POST
def order_upload_preview(request):
    resp = require_action_perm(request, 'order.upload')
    if resp:
        return resp

    if not request.FILES.get('excel_file'):
        messages.error(request, "파일을 선택해주세요.")
        return redirect('order_upload')

    preview_data = []

    try:
        wb = openpyxl.load_workbook(request.FILES['excel_file'], data_only=True)
        ws = wb.active

        # 새 양식: 품번, 수량, 납기일, ERP발주번호(선택), ERP순번(선택)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 필수 필드: 품번, 수량, 납기일
            if not row[0] or not row[1] or not row[2]:
                continue

            part_no = str(row[0]).strip()
            quantity = int(row[1]) if row[1] else 0

            raw_date = row[2]
            if isinstance(raw_date, datetime.datetime):
                fmt_date = raw_date.strftime("%Y-%m-%d")
            elif isinstance(raw_date, str):
                fmt_date = raw_date[:10]
            else:
                fmt_date = str(raw_date)

            # 품번으로 품목마스터 조회
            part_obj = Part.objects.filter(part_no=part_no).first()

            # 오류 상태 판단
            error_type = None
            if not part_obj:
                error_type = 'part_not_found'  # 품번 없음
            elif not part_obj.vendor:
                error_type = 'vendor_not_assigned'  # 협력사 미지정

            part_name = part_obj.part_name if part_obj else ""
            part_group = part_obj.part_group if part_obj else ""
            vendor_name = part_obj.vendor.name if part_obj and part_obj.vendor else ""
            is_valid = error_type is None

            item = {
                'vendor': vendor_name,
                'part_no': part_no,
                'part_name': part_name,
                'part_group': part_group,
                'is_valid': is_valid,
                'error_type': error_type,
                'quantity': quantity,
                'due_date': fmt_date,
            }
            preview_data.append(item)

        if not preview_data:
            messages.warning(request, "유효한 데이터가 없습니다. 엑셀 양식을 확인해주세요.")
            return redirect('order_upload')

        valid_count = sum(1 for item in preview_data if item['is_valid'])
        error_count = len(preview_data) - valid_count

        if valid_count == 0:
            messages.warning(request, "등록 가능한 정상 품목이 없습니다. 품번 또는 협력사 지정을 확인해주세요.")
        else:
            messages.info(request, f"총 {len(preview_data)}건 중 정상 {valid_count}건, 오류 {error_count}건이 확인되었습니다.")

    except Exception as e:
        messages.error(request, f"엑셀 처리 중 오류: {str(e)}")
        return redirect('order_upload')

    return render(request, 'order_upload.html', {
        'active_menu': 'upload',
        'preview_data': preview_data,
        'valid_count': valid_count,
        'error_count': error_count
    })

@login_required
@require_POST
def order_create_confirm(request):
    resp = require_action_perm(request, 'order.upload')
    if resp:
        return resp

    # 품번으로 품목마스터에서 협력사 자동 조회
    part_nos = request.POST.getlist('part_no_list[]')
    quantities = request.POST.getlist('quantity_list[]')
    due_dates = request.POST.getlist('due_date_list[]')

    success_count = 0
    skip_count = 0

    try:
        with transaction.atomic():
            for i in range(len(part_nos)):
                part_obj = Part.objects.filter(part_no=part_nos[i]).first()
                if not part_obj or not part_obj.vendor:
                    skip_count += 1
                    continue

                Order.objects.create(
                    vendor=part_obj.vendor,
                    part_group=part_obj.part_group or '',
                    part_no=part_obj.part_no,
                    part_name=part_obj.part_name or '',
                    quantity=int(quantities[i]),
                    due_date=due_dates[i],
                )
                success_count += 1

        if skip_count > 0:
            messages.warning(request, f"총 {success_count}건 등록, {skip_count}건 제외 (품번/협력사 미확인)")
        else:
            messages.success(request, f"총 {success_count}건의 발주가 정상적으로 등록되었습니다.")
        return redirect('order_list')

    except Exception as e:
        messages.error(request, f"저장 중 오류 발생: {str(e)}")
        return redirect('order_upload')

@login_required
@require_POST
def order_delete(request):
    if not request.user.is_superuser:
        return redirect('order_list')
    Order.objects.filter(id__in=request.POST.getlist('order_ids')).delete()
    return redirect('order_list')

@login_required
@require_POST
def order_close_action(request):
    if not request.user.is_superuser:
        return redirect('order_list')
    Order.objects.filter(id__in=request.POST.getlist('order_ids')).update(is_closed=True)
    return redirect('order_list')

@login_required
@menu_permission_required('can_scm_order_edit')
def order_approve_all(request):
    user = request.user
    q = Order.objects.filter(approved_at__isnull=True, is_closed=False)

    # 협력업체 사용자 판별 (2가지 경로)
    user_vendor = Vendor.objects.filter(user=user).first()
    if not user_vendor and not user.is_superuser:
        try:
            if hasattr(user, 'profile') and user.profile.org and user.profile.org.linked_vendor:
                user_vendor = user.profile.org.linked_vendor
        except Exception:
            pass

    # 협력업체 사용자는 자신의 발주만 승인 가능
    if not user.is_superuser and user_vendor:
        q = q.filter(vendor=user_vendor)

    q.update(approved_at=timezone.now())
    return redirect('order_list')

@login_required
@menu_permission_required('can_scm_order_edit')
def order_approve(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    if not order.approved_at and not order.is_closed:
        order.approved_at = timezone.now()
        order.save()
    return redirect('order_list')

@login_required
@menu_permission_required('can_view_orders')
def order_export(request):
    user = request.user

    # 협력업체 사용자 판별 (2가지 경로)
    user_vendor = Vendor.objects.filter(user=user).first()
    if not user_vendor and not user.is_superuser:
        try:
            if hasattr(user, 'profile') and user.profile.org and user.profile.org.linked_vendor:
                user_vendor = user.profile.org.linked_vendor
        except Exception:
            pass

    # 관리자는 전체, 협력업체는 자신의 발주만
    if user.is_superuser:
        orders = Order.objects.all().order_by('-created_at')
    elif user_vendor:
        orders = Order.objects.filter(vendor=user_vendor).order_by('-created_at')
    else:
        orders = Order.objects.all().order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['상태', '등록일', '승인일', '협력사', '품번', '품명', '수량', '납기일', 'ERP번호'])

    for o in orders:
        status = "마감" if o.is_closed else ("승인" if o.approved_at else "미확인")
        ws.append([status, o.created_at.date(), o.approved_at.date() if o.approved_at else "-", o.vendor.name, o.part_no, o.part_name, o.quantity, str(o.due_date), o.erp_order_no])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
    wb.save(response)
    return response

# ==========================================
# [3. 과부족/소요량 로직]
# ==========================================

@login_required
@menu_permission_required('can_view_inventory')
def inventory_list(request):
    """
    과부족 현황 조회 (최적화 버전)
    - N+1 쿼리 문제 해결: 모든 데이터를 미리 조회 후 메모리에서 처리
    """
    user = request.user
    today = timezone.localtime().date()

    # 협력업체 사용자 판별 (2가지 경로)
    # 1. Vendor.user 필드 (구 방식)
    user_vendor = Vendor.objects.filter(user=user).first()
    # 2. UserProfile.org → Organization.linked_vendor (신 방식)
    if not user_vendor and not user.is_superuser:
        try:
            if hasattr(user, 'profile') and user.profile.org and user.profile.org.linked_vendor:
                user_vendor = user.profile.org.linked_vendor
        except Exception:
            pass

    if MaterialStock is None:
        messages.error(request, "WMS(MaterialStock) 연동 모델을 불러올 수 없습니다. material 앱/모델 연결을 확인해주세요.")
        return redirect('order_list')

    if user.is_superuser or not user_vendor:
        max_due = Demand.objects.aggregate(Max('due_date'))['due_date__max']
        standard_end = today + datetime.timedelta(days=31)
        end_date = max_due if max_due and max_due > standard_end else standard_end
    else:
        end_date = today + datetime.timedelta(days=14)

    date_range = [today + datetime.timedelta(days=i) for i in range((end_date - today).days + 1)]

    show_all = request.GET.get('show_all') == 'true'
    selected_v = request.GET.get('vendor_id')
    q = request.GET.get('q', '')
    search_submitted = request.GET.get('search') == '1'  # 조회 버튼 클릭 여부

    part_qs = Part.objects.select_related('vendor').filter(vendor__isnull=False).order_by('vendor__name', 'part_name')

    # 협력업체 사용자인지 여부
    is_vendor_user = bool(user_vendor) and not user.is_superuser

    if user.is_superuser or not user_vendor:
        vendor_list = Vendor.objects.all().order_by('name')
        if selected_v:
            part_qs = part_qs.filter(vendor_id=selected_v)
    elif user_vendor:
        part_qs = part_qs.filter(vendor=user_vendor)
        vendor_list = []
        # 협력업체는 자동으로 조회 실행 (조회 버튼 필요 없음)
        search_submitted = True
    else:
        return redirect('order_list')

    if q:
        part_qs = part_qs.filter(Q(part_no__icontains=q) | Q(part_name__icontains=q))

    # ============================================
    # 조회 버튼을 눌렀을 때만 데이터 로드
    # ============================================
    inventory_data = []

    if search_submitted:
        if not show_all:
            # 소요 품목만: Demand(소요량)가 있는 품목만 표시
            demand_pnos = Demand.objects.filter(due_date__range=[today, end_date]).values_list('part__part_no', flat=True).distinct()
            part_qs = part_qs.filter(part_no__in=demand_pnos)

        # ============================================
        # 최적화: 모든 데이터를 미리 조회 (N+1 쿼리 방지)
        # ============================================
        from material.models import Warehouse
        from collections import defaultdict

        # 1. 품목 리스트 확정 (쿼리 실행)
        parts = list(part_qs)
        part_ids = [p.id for p in parts]

        # 2. 자재창고(2000, 4200) 조회
        target_warehouses = Warehouse.objects.filter(code__in=['2000', '4200'])
        target_wh_ids = list(target_warehouses.values_list('id', flat=True))

        # 3. MaterialStock 일괄 조회 (part별 재고 합계)
        if target_wh_ids:
            stock_qs = MaterialStock.objects.filter(
                part_id__in=part_ids,
                warehouse_id__in=target_wh_ids
            ).values('part_id').annotate(total_qty=Sum('quantity'))
        else:
            stock_qs = MaterialStock.objects.filter(
                part_id__in=part_ids
            ).values('part_id').annotate(total_qty=Sum('quantity'))

        stock_map = {item['part_id']: item['total_qty'] or 0 for item in stock_qs}

        # 4. Incoming 일괄 조회 (part별, 날짜별 입고량)
        incoming_qs = Incoming.objects.filter(
            part_id__in=part_ids,
            in_date__range=[today, end_date]
        ).values('part_id', 'in_date').annotate(total_qty=Sum('quantity'))

        incoming_map = defaultdict(lambda: defaultdict(int))
        for item in incoming_qs:
            incoming_map[item['part_id']][item['in_date']] = item['total_qty'] or 0

        # 5. Demand 일괄 조회 (part별, 날짜별 소요량)
        demand_qs = Demand.objects.filter(
            part_id__in=part_ids,
            due_date__range=[today, end_date]
        ).values('part_id', 'due_date').annotate(total_qty=Sum('quantity'))

        demand_map = defaultdict(lambda: defaultdict(int))
        for item in demand_qs:
            demand_map[item['part_id']][item['due_date']] = item['total_qty'] or 0

        # ============================================
        # 메모리에서 과부족 계산
        # ============================================
        for part in parts:
            daily_status = []

            # WMS 현재 재고
            current_wms_stock = stock_map.get(part.id, 0)

            # 오늘 입고량 (WMS에 이미 반영된 금일 입고)
            today_incoming = incoming_map[part.id].get(today, 0)

            # 시업재고 = WMS 현재 재고 - 오늘 입고량
            opening_stock = current_wms_stock - today_incoming
            temp_stock = opening_stock

            for dt in date_range:
                dq = demand_map[part.id].get(dt, 0)
                iq = incoming_map[part.id].get(dt, 0)

                # 입고/소요 반영
                temp_stock = temp_stock - dq + iq

                daily_status.append({
                    'date': dt,
                    'demand_qty': dq,
                    'in_qty': iq,
                    'stock': temp_stock,
                    'is_danger': temp_stock < 200  # 안전재고 200개 미만 시 위험 표시
                })

            inventory_data.append({
                'vendor_name': part.vendor.name if part.vendor else '(미연결)',
                'part_no': part.part_no,
                'part_name': part.part_name,
                'base_stock': opening_stock,
                'daily_status': daily_status
            })

    latest_inv_date = None
    last_inv_obj = Inventory.objects.exclude(last_inventory_date__isnull=True).order_by('-last_inventory_date').first()
    if last_inv_obj:
        latest_inv_date = last_inv_obj.last_inventory_date

    # 미확인 발주 목록 (approved_at이 null인 것) - 중복 발주 방지용
    pending_orders = Order.objects.filter(
        approved_at__isnull=True,
        is_closed=False
    ).values_list('part_no', 'due_date')
    pending_order_keys = [f"{po[0]}_{po[1]}" for po in pending_orders]
    # 품번만으로 관련 발주 있는지 확인용 (날짜 무관하게 경고 표시)
    pending_order_parts = list(set(po[0] for po in pending_orders))

    # 선택된 업체명 (검색 팝업에 표시용)
    selected_vendor_name = ''
    if selected_v:
        selected_vendor_obj = next((v for v in vendor_list if str(v.id) == selected_v), None)
        if selected_vendor_obj:
            selected_vendor_name = selected_vendor_obj.name

    return render(request, 'inventory_list.html', {
        'date_range': date_range,
        'inventory_data': inventory_data,
        'vendor_list': vendor_list,
        'active_menu': 'inventory',
        'show_all': show_all,
        'selected_vendor_id': selected_v,
        'selected_vendor_name': selected_vendor_name,
        'user_name': user.username,
        'vendor_name': user_vendor.name if user_vendor else "관리자",
        'q': q,
        'inventory_ref_date': latest_inv_date,
        'pending_order_keys': pending_order_keys,
        'pending_order_parts': pending_order_parts,
        'search_submitted': search_submitted,
        'is_vendor_user': is_vendor_user,  # 협력업체 사용자 여부
    })

@login_required
@menu_permission_required('can_view_inventory')
def inventory_export(request):
    user = request.user
    user_vendor = Vendor.objects.filter(user=user).first()

    # ✅ 직원/관리자(벤더가 아닌 계정)만 inv.export 권한 체크
    if (not user.is_superuser) and (not user_vendor):
        resp = require_action_perm(request, 'inv.export')
        if resp:
            return resp

    today = timezone.localtime().date()

    max_due = Demand.objects.aggregate(Max('due_date'))['due_date__max']
    end_date = max_due if max_due and max_due > (today + datetime.timedelta(days=31)) else (today + datetime.timedelta(days=31))
    dr = [today + datetime.timedelta(days=i) for i in range((end_date - today).days + 1)]

    items = Inventory.objects.select_related('part', 'part__vendor').all()
    if (not user.is_superuser) and user_vendor:
        items = items.filter(part__vendor=user_vendor)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['협력사', '품번', '품명', '구분'] + [d.strftime('%m/%d') for d in dr])

    for item in items:
        ref = item.last_inventory_date or date(2000, 1, 1)

        hist_dem = Demand.objects.filter(part=item.part, due_date__gt=ref, due_date__lt=today).aggregate(Sum('quantity'))['quantity__sum'] or 0
        hist_in = Incoming.objects.filter(part=item.part, in_date__gt=ref, in_date__lt=today).aggregate(Sum('quantity'))['quantity__sum'] or 0

        stock = item.base_stock - hist_dem + hist_in

        r1 = [item.part.vendor.name, item.part.part_no, item.part.part_name, '소요량']
        r2 = ['', '', '', '입고량']
        r3 = ['', '', '', '재고']

        for dt in dr:
            dq = Demand.objects.filter(part=item.part, due_date=dt).aggregate(Sum('quantity'))['quantity__sum'] or 0
            iq = Incoming.objects.filter(part=item.part, in_date=dt).aggregate(Sum('quantity'))['quantity__sum'] or 0

            stock = stock - dq + iq

            r1.append(dq)
            r2.append(iq)
            r3.append(stock)

        ws.append(r1)
        ws.append(r2)
        ws.append(r3)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Inventory_{today}.xlsx'
    wb.save(response)
    return response


@login_required
@require_POST
def quick_order_action(request):
    user = request.user
    if not (user.is_superuser or (_is_internal(user))):
        messages.error(request, "발주 등록 권한이 없습니다.")
        return redirect('inventory_list')

    v_name = request.POST.get('vendor_name')
    p_no = request.POST.get('part_no')
    qty = request.POST.get('quantity')
    due = request.POST.get('due_date')

    try:
        part = Part.objects.filter(part_no=p_no, vendor__name=v_name).first()
        if part:
            Order.objects.create(
                vendor=part.vendor,
                part_no=p_no,
                part_name=part.part_name,
                part_group=part.part_group,
                quantity=int(qty),
                due_date=due
            )
            messages.success(request, f"발주 완료: {p_no}")
    except Exception as e:
        messages.error(request, str(e))

    return redirect('inventory_list')


@login_required
@require_POST
def bulk_shortage_order(request):
    """부족품 일괄 발주 처리"""
    import json
    user = request.user
    if not (user.is_superuser or (_is_internal(user))):
        return JsonResponse({'success': False, 'error': '발주 등록 권한이 없습니다.'})

    try:
        orders_json = request.POST.get('orders', '[]')
        orders = json.loads(orders_json)

        created_count = 0
        for item in orders:
            vendor_name = item.get('vendor')
            part_no = item.get('part_no')
            due_date = item.get('due_date')
            quantity = int(item.get('quantity', 0))

            if not all([vendor_name, part_no, due_date, quantity > 0]):
                continue

            part = Part.objects.filter(part_no=part_no, vendor__name=vendor_name).first()
            if part:
                Order.objects.create(
                    vendor=part.vendor,
                    part_no=part_no,
                    part_name=part.part_name,
                    part_group=part.part_group,
                    quantity=quantity,
                    due_date=due_date
                )
                created_count += 1

        return JsonResponse({'success': True, 'created': created_count})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@menu_permission_required('can_view_inventory')
def demand_manage(request):
    if not request.user.is_superuser:
        return redirect('inventory_list')

    # 복수 업체 ID 지원 (콤마로 구분)
    vendor_ids_str = request.GET.get('vendor_ids', '')
    vendor_ids = [v.strip() for v in vendor_ids_str.split(',') if v.strip()]

    p_no = request.GET.get('part_no', '')
    sd = request.GET.get('start_date', '')
    ed = request.GET.get('end_date', '')

    demands = Demand.objects.select_related('part', 'part__vendor').all().order_by('-due_date')

    # 복수 업체 필터링
    if vendor_ids:
        demands = demands.filter(part__vendor_id__in=vendor_ids)
    if p_no:
        demands = demands.filter(part__part_no__icontains=p_no)
    if sd and ed:
        demands = demands.filter(due_date__range=[sd, ed])

    # 선택된 업체 이름 조회 (뱃지 표시용)
    selected_vendor_names = ''
    if vendor_ids:
        names = list(Vendor.objects.filter(id__in=vendor_ids).values_list('name', flat=True))
        selected_vendor_names = ', '.join(names)

    # 전체 업체 목록 조회
    vendor_list = Vendor.objects.all().order_by('name')

    # 소요량 있는 업체 ID 목록 (모달 내 필터용)
    vendors_with_demand_ids = list(Demand.objects.values_list('part__vendor_id', flat=True).distinct())

    return render(
        request,
        'demand_manage.html',
        {
            'demands': demands[:500],
            'vendor_list': vendor_list,
            'vendors_with_demand_ids': vendors_with_demand_ids,
            'active_menu': 'inventory',
            'selected_vendor_ids': vendor_ids_str,
            'selected_vendor_names': selected_vendor_names,
            'part_no': p_no,
            'start_date': sd,
            'end_date': ed,
        }
    )

@login_required
@require_POST
def demand_delete_action(request):
    resp = require_action_perm(request, 'demand.delete')
    if resp:
        return resp

    if not request.user.is_superuser:
        return redirect('inventory_list')
    Demand.objects.filter(id__in=request.POST.getlist('demand_ids')).delete()
    messages.success(request, "삭제 완료.")
    return redirect(request.META.get('HTTP_REFERER', 'demand_manage'))

@login_required
@require_POST
def delete_all_demands(request):
    resp = require_action_perm(request, 'demand.delete_all')
    if resp:
        return resp

    if not request.user.is_superuser:
        return redirect('inventory_list')
    Demand.objects.all().delete()
    messages.success(request, "전체 삭제 완료.")
    return redirect('inventory_list')

@login_required
@require_POST
def demand_upload_action(request):
    resp = require_action_perm(request, 'demand.upload')
    if resp:
        return resp

    if not request.user.is_superuser:
        return redirect('inventory_list')
    if request.FILES.get('demand_file'):
        try:
            wb = openpyxl.load_workbook(request.FILES['demand_file'], read_only=True, data_only=True)
            ws = wb.active
            c_count = 0
            all_parts = {p.part_no: p for p in Part.objects.select_related('vendor').all()}

            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    p_no = str(row[0]).strip() if row[0] else None
                    if not p_no or p_no not in all_parts:
                        continue
                    Demand.objects.update_or_create(part=all_parts[p_no], due_date=row[2], defaults={'quantity': row[1] or 0})
                    c_count += 1

            messages.success(request, f"소요량 {c_count}건 반영 완료")

        except Exception as e:
            messages.error(request, f"업로드 중 오류: {str(e)}")

    return redirect('inventory_list')

@login_required
@require_POST
def demand_update_ajax(request):
    resp = require_action_perm(request, 'demand.edit')
    if resp:
        return resp

    if not request.user.is_superuser:
        return JsonResponse({'status': 'error'}, status=403)

    p_no, d_date, qty = request.POST.get('part_no'), request.POST.get('due_date'), request.POST.get('quantity')

    try:
        part = Part.objects.get(part_no=p_no)

        if int(qty) <= 0:
            Demand.objects.filter(part=part, due_date=d_date).delete()
        else:
            Demand.objects.update_or_create(part=part, due_date=d_date, defaults={'quantity': int(qty)})

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

# ==========================================
# [4. 라벨/입고 관리]
# ==========================================

@login_required
@menu_permission_required('can_scm_label_view')
def label_list(request):
    user = request.user
    selected_v = request.GET.get('vendor_id')
    status_filter = request.GET.get('status')
    q = request.GET.get('q', '')

    # 협력업체 사용자 판별 (2가지 경로)
    # 1. Vendor.user 필드 (구 방식)
    user_vendor = Vendor.objects.filter(user=user).first()
    # 2. UserProfile.org → Organization.linked_vendor (신 방식)
    if not user_vendor and not user.is_superuser:
        try:
            if hasattr(user, 'profile') and user.profile.org and user.profile.org.linked_vendor:
                user_vendor = user.profile.org.linked_vendor
        except Exception:
            pass

    pnos_with_delivery = DeliveryOrderItem.objects.values_list('part_no', flat=True).distinct()
    vendor_ids = Part.objects.filter(part_no__in=pnos_with_delivery).values_list('vendor_id', flat=True).distinct()

    profile = getattr(user, 'profile', None)
    is_staff_or_admin = user.is_superuser or (profile and _is_internal(user))

    vendor_list = Vendor.objects.filter(id__in=vendor_ids).order_by('name') if is_staff_or_admin else []
    
    # [수정] 기본 쿼리셋: vendor 필드 없이 items__part__vendor 등을 통해 역추적해야 함
    # 하지만 items는 ManyToMany or ReverseFK 이므로 .distinct() 주의
    # 만약 DeliveryOrder 모델에 'vendor' 필드가 없다면 select_related('vendor')는 제거해야 함.
    # 안전하게 제거하고 진행합니다.
    do_qs = DeliveryOrder.objects.prefetch_related('items').order_by('-created_at')

    if not is_staff_or_admin:
        if user_vendor:
            # Part(vendor=user_vendor) → part_no 목록
            vendor_part_nos = Part.objects.filter(
                vendor=user_vendor
            ).values_list('part_no', flat=True)

            do_ids = DeliveryOrderItem.objects.filter(
                part_no__in=vendor_part_nos
            ).values_list('order_id', flat=True)

            do_qs = do_qs.filter(id__in=do_ids).distinct()
        else:
            do_qs = do_qs.none()

    elif selected_v:
        selected_part_nos = Part.objects.filter(
            vendor_id=selected_v
        ).values_list('part_no', flat=True)

        do_ids = DeliveryOrderItem.objects.filter(
            part_no__in=selected_part_nos
        ).values_list('order_id', flat=True)

        do_qs = do_qs.filter(id__in=do_ids).distinct()
    # 2. 상태별 리스트 분리
    recent_orders = do_qs.exclude(status='REJECTED')

    if status_filter == 'registered':
        recent_orders = recent_orders.filter(status='PENDING')
    elif status_filter == 'received':
        recent_orders = recent_orders.filter(status__in=['RECEIVED', 'APPROVED'])

    recent_orders = recent_orders[:20]

    # [수정] 부적합/반출 관리 탭 데이터 (ReturnLog 기준)
    if user_vendor:
        return_logs_qs = ReturnLog.objects.filter(
            part__vendor=user_vendor
        ).select_related(
            'delivery_order',
            'part'
        ).order_by(
            'is_confirmed',
            '-created_at'
        )
    else:
        return_logs_qs = ReturnLog.objects.all().select_related(
            'delivery_order',
            'part'
        ).order_by(
            'is_confirmed',
            '-created_at'
        )

    # 🔴 미확인 건수 (뱃지용)
    return_pending_count = return_logs_qs.filter(
        is_confirmed=False
    ).count()

    return_logs = return_logs_qs
        
    # 3. 라벨 발행 데이터 (잔량 계산 로직)
    label_data = []
    
    order_q = Order.objects.filter(is_closed=False, approved_at__isnull=False)
    if not is_staff_or_admin and user_vendor:
        order_q = order_q.filter(vendor=user_vendor)
    elif selected_v:
        order_q = order_q.filter(vendor_id=selected_v)

    if q:
        order_q = order_q.filter(Q(part_no__icontains=q) | Q(part_name__icontains=q))

    # [A] ERP 발주 건
    erp_orders = order_q.exclude(erp_order_no__isnull=True).exclude(erp_order_no='')
    for o in erp_orders:
        printed = LabelPrintLog.objects.filter(order=o).aggregate(Sum('printed_qty'))['printed_qty__sum'] or 0
        
        # 반출 확인 수량 (ERP 번호 매칭)
        returned = ReturnLog.objects.filter(
            part__part_no=o.part_no,
            is_confirmed=True,
            delivery_order__items__erp_order_no=o.erp_order_no
        ).distinct().aggregate(Sum('quantity'))['quantity__sum'] or 0

        valid_printed = printed - returned
        remain = o.quantity - valid_printed

        if remain > 0:
            label_data.append({
                'is_erp': True,
                'order_id': o.id,
                'erp_no': o.erp_order_no,
                'erp_seq': o.erp_order_seq,
                'part_no': o.part_no,
                'part_name': o.part_name,
                'total_order': o.quantity,
                'remain': remain,
                'due_date': o.due_date
            })

    # [B] 수기 발주 건
    manual_orders = order_q.filter(Q(erp_order_no__isnull=True) | Q(erp_order_no=''))
    manual_pnos = manual_orders.values_list('part_no', flat=True).distinct()

    for p_no in manual_pnos:
        sub_orders = manual_orders.filter(part_no=p_no)
        total_qty = sub_orders.aggregate(Sum('quantity'))['quantity__sum'] or 0
        part_first = sub_orders.first()
        part_name = part_first.part_name
        due_date = sub_orders.order_by('due_date').first().due_date
        
        printed = LabelPrintLog.objects.filter(part_no=p_no, order__isnull=True).aggregate(Sum('printed_qty'))['printed_qty__sum'] or 0
        
        returned = ReturnLog.objects.filter(
            part__part_no=p_no,
            is_confirmed=True,
            delivery_order__items__erp_order_no=''
        ).distinct().aggregate(Sum('quantity'))['quantity__sum'] or 0

        remain = total_qty - (printed - returned)

        if remain > 0:
            label_data.append({
                'is_erp': False,
                'order_id': None,
                'erp_no': '-',
                'erp_seq': '-',
                'part_no': p_no,
                'part_name': part_name,
                'total_order': total_qty,
                'remain': remain,
                'due_date': due_date
            })

    label_data.sort(key=lambda x: x['due_date'])

    # ✅✅✅ [요청 반영 1] 템플릿 경로만 orders/로 변경 ✅✅✅
    return render(request, 'label_list.html', {
        'label_data': label_data,
        'orders': recent_orders,
        'return_logs': return_logs,
        'return_pending_count': return_pending_count,
        'vendor_list': vendor_list,
        'selected_vendor_id': selected_v,
        'status_filter': status_filter,
        'active_menu': 'label',
        'q': q
    })

@login_required
@menu_permission_required('can_scm_label_edit')
@require_POST
def delete_delivery_order(request, order_id):
    resp = require_action_perm(request, 'delivery.delete')
    if resp:
        return resp

    order = get_object_or_404(DeliveryOrder, pk=order_id)

    # 권한 체크 (items를 통해 vendor 확인)
    first_item = order.items.first()
    part = Part.objects.filter(part_no=first_item.part_no).first() if first_item else None
    item_vendor = part.vendor if part else None

    # 사용자의 협력사 정보 가져오기
    user_vendor = _get_user_vendor(request.user)

    # 관리자가 아니고, 협력사 계정인 경우 본인 납품서만 삭제 가능
    if not request.user.is_superuser and user_vendor and user_vendor != item_vendor:
        messages.error(request, "삭제 권한이 없습니다.")
        return redirect('label_list')

    if order.status != 'PENDING' and order.status != 'REJECTED':
        messages.error(request, "이미 처리된 납품서는 삭제할 수 없습니다.")
        return redirect('label_list')

    with transaction.atomic():
        for item in order.items.all():
            LabelPrintLog.objects.filter(
                part_no=item.part_no,
                printed_qty=item.total_qty,
                printed_at__date=order.created_at.date()
            ).delete()
        order.delete()
        messages.success(request, "납품서가 삭제되었습니다.")

    return redirect('label_list')

@login_required
@menu_permission_required('can_scm_label_view')
def label_print_action(request):
    return redirect('label_list')

@login_required
@menu_permission_required('can_scm_label_edit')
@require_POST
def create_delivery_order(request):
    resp = require_action_perm(request, 'delivery.register')
    if resp:
        return resp

    p_nos = request.POST.getlist('part_nos[]')
    snps = request.POST.getlist('snps[]')
    b_counts = request.POST.getlist('box_counts[]')
    order_ids = request.POST.getlist('order_ids[]')
    lot_nos = request.POST.getlist('lot_nos[]')

    if _get_role(request.user) == 'VENDOR' and not request.user.is_superuser:
        user_vendor = _get_user_vendor(request.user)
        if not user_vendor:
            messages.error(request, "협력사 정보가 연결되어 있지 않습니다.")
            return redirect('label_list')

        allowed_pnos = set(Part.objects.filter(vendor=user_vendor).values_list('part_no', flat=True))
        bad = [p for p in p_nos if p not in allowed_pnos]
        if bad:
            messages.error(request, f"권한이 없는 품번이 포함되어 있습니다.")
            return redirect('label_list')

    with transaction.atomic():
        # [수정] vendor 필드 제거 (DeliveryOrder에 vendor가 없다면)
        do = DeliveryOrder.objects.create(order_no="DO-"+timezone.now().strftime("%Y%m%d-%H%M%S"))

        for i in range(len(p_nos)):
            part = Part.objects.filter(part_no=p_nos[i]).first()
            if not part:
                continue

            qty = int(snps[i]) * int(b_counts[i])
            if qty <= 0:
                continue

            linked_order = None
            erp_no = ''
            erp_seq = ''

            if len(order_ids) > i and order_ids[i] and order_ids[i] != 'None':
                try:
                    linked_order = Order.objects.get(id=order_ids[i])
                    erp_no = linked_order.erp_order_no
                    erp_seq = linked_order.erp_order_seq
                except Order.DoesNotExist:
                    linked_order = None

            # LOT 정보 추출 (날짜 형식으로 통일)
            lot_no = lot_nos[i] if len(lot_nos) > i else None

            DeliveryOrderItem.objects.create(
                order=do,
                part_no=p_nos[i],
                part_name=part.part_name,
                snp=int(snps[i]),
                box_count=int(b_counts[i]),
                total_qty=qty,
                linked_order=linked_order,
                erp_order_no=erp_no,
                erp_order_seq=erp_seq,
                lot_no=lot_no
            )

            LabelPrintLog.objects.create(
                vendor=part.vendor,
                part=part,
                part_no=p_nos[i],
                printed_qty=qty,
                snp=int(snps[i]),
                order=linked_order
            )

    return redirect('label_list')

@login_required
@menu_permission_required('can_scm_label_view')
def label_print(request, order_id):
    resp = require_action_perm(request, 'label.print')
    if resp:
        return resp

    order = get_object_or_404(DeliveryOrder, pk=order_id)
    queue = []
    first_item = order.items.first()
    part = Part.objects.filter(part_no=first_item.part_no).first() if first_item else None
    v_name = part.vendor.name if part else "알수없음"

    for item in order.items.all():
        for box_seq in range(1, item.box_count + 1):
            # TAG_ID 생성: DLV-{order_id:05d}-{item_id:05d}-{box_seq:03d}
            tag_id = f"DLV-{order.pk:05d}-{item.pk:05d}-{box_seq:03d}"
            lot_str = item.lot_no.strftime('%Y-%m-%d') if item.lot_no else ''
            queue.append({
                'tag_id': tag_id,
                'vendor_name': v_name,
                'part_name': item.part_name,
                'part_no': item.part_no,
                'snp': item.snp,
                'lot_no': item.lot_no,
                'lot_str': lot_str,
                'print_date': timezone.now()
            })

    return render(request, 'label_print_popup.html', {'box_count': queue, 'vendor_name': v_name})

@login_required
@menu_permission_required('can_scm_label_view')
def delivery_note_print(request, order_id):
    resp = require_action_perm(request, 'delivery.print')
    if resp:
        return resp

    do = get_object_or_404(DeliveryOrder, pk=order_id)
    items = do.items.all()
    first_item = items.first()
    part = Part.objects.filter(part_no=first_item.part_no).first() if first_item else None
    vendor = part.vendor if part else None

    return render(request, 'print_delivery_note.html', {
        'order': do,
        'items': items,
        'total_qty': items.aggregate(Sum('total_qty'))['total_qty__sum'] or 0,
        'total_box': items.aggregate(Sum('box_count'))['box_count__sum'] or 0,
        'print_date': timezone.localtime().date(),
        'vendor': vendor
    })

# ==========================================
# [5. 입고 및 반출 관리]
# ==========================================

@login_required
@require_POST
@menu_permission_required('can_scm_incoming_edit')
def receive_delivery_order_scan(request):
    qr_code = request.POST.get('qr_code', '').strip()
    do = DeliveryOrder.objects.filter(order_no=qr_code).first()

    if not do:
        messages.error(request, f"납품서 번호 [{qr_code}]를 찾을 수 없습니다.")
        return redirect('incoming_list')

    if do.is_received:
        messages.warning(request, f"이미 입고 처리된 납품서입니다. ({do.order_no})")
        return redirect('incoming_list')

    if Warehouse is None:
        warehouses = []
    else:
        warehouses = Warehouse.objects.exclude(code__in=['8100', '8200']).order_by('code')

    # =====================================================================
    # [FIFO 경고 체크] 입고 확인 화면에 표시할 FIFO 경고 생성
    # =====================================================================
    fifo_warnings = []

    if MaterialStock is not None:
        mat_warehouse = Warehouse.objects.filter(code='2000').first()  # 자재창고

        if mat_warehouse:
            for item in do.items.all():
                part = Part.objects.filter(part_no=item.part_no).first()
                if not part or not item.lot_no:
                    continue

                # FIFO 경고 체크: 자재창고(2000)에 입고 LOT보다 최근 생산품이 있는지 확인
                # (이미 더 최근 LOT가 있는데 과거 LOT를 입고하면 FIFO 위반!)
                newer_lots = MaterialStock.objects.filter(
                    warehouse=mat_warehouse,
                    part=part,
                    lot_no__gt=item.lot_no,  # 입고 LOT보다 생산일이 나중
                    quantity__gt=0
                ).order_by('lot_no')

                if newer_lots.exists():
                    newest_lot = newer_lots.first()
                    days_diff = (newest_lot.lot_no - item.lot_no).days

                    # FIFO 위반 - 무조건 경고 표시
                    fifo_warnings.append({
                        'level': 'danger',
                        'icon': '🚨',
                        'label': 'FIFO 위반 경고',
                        'part_no': item.part_no,
                        'incoming_lot': item.lot_no.strftime('%Y-%m-%d'),
                        'existing_lot': newest_lot.lot_no.strftime('%Y-%m-%d'),
                        'days_diff': days_diff
                    })

    return render(request, 'incoming_check.html', {
        'order': do,
        'warehouses': warehouses,
        'fifo_warnings': fifo_warnings
    })

@login_required
@menu_permission_required('can_scm_incoming_edit')
@require_POST
def incoming_cancel(request):
    resp = require_action_perm(request, 'incoming.cancel')
    if resp:
        return resp

    inc_id = request.POST.get('incoming_id')
    mode = request.POST.get('cancel_mode')
    target_inc = get_object_or_404(Incoming, id=inc_id)
    do_no = target_inc.delivery_order_no
    do = DeliveryOrder.objects.filter(order_no=do_no).first()

    with transaction.atomic():
        if mode == 'item':
            if do:
                LabelPrintLog.objects.filter(part_no=target_inc.part.part_no, printed_qty=target_inc.quantity).delete()
                DeliveryOrderItem.objects.filter(order=do, part_no=target_inc.part.part_no, total_qty=target_inc.quantity).delete()

            target_inc.delete()
            messages.success(request, f"품목 {target_inc.part.part_no} 입고 취소 및 잔량이 복구되었습니다.")

        elif mode == 'all':
            Incoming.objects.filter(delivery_order_no=do_no).delete()
            if do:
                do.is_received = False
                do.save()

            messages.success(request, f"납품서 {do_no} 입고 취소 완료. (품목 데이터는 보존됩니다)")

    return redirect('incoming_list')

@login_required
@menu_permission_required('can_manage_incoming')
def incoming_list(request):
    user = request.user
    selected_v = request.GET.get('vendor_id')
    sd, ed, q = request.GET.get('start_date'), request.GET.get('end_date'), request.GET.get('q', '')

    user_vendor = Vendor.objects.filter(user=user).first()

    incomings = Incoming.objects.select_related('part', 'part__vendor').all().order_by('-in_date', '-created_at')

    profile = getattr(user, 'profile', None)
    vendor_ids = Incoming.objects.values_list('part__vendor_id', flat=True).distinct()
    vendor_list = Vendor.objects.filter(id__in=vendor_ids).order_by('name') if (user.is_superuser or (profile and _is_internal(user))) else []

    if not user.is_superuser and (profile and profile.role != 'STAFF'):
        incomings = incomings.filter(part__vendor=user_vendor) if user_vendor else incomings.none()
    elif selected_v:
        incomings = incomings.filter(part__vendor_id=selected_v)

    if sd and ed:
        incomings = incomings.filter(in_date__range=[sd, ed])
    if q:
        incomings = incomings.filter(Q(part__part_no__icontains=q) | Q(part__part_name__icontains=q))

    return render(request, 'incoming_list.html', {
        'incomings': incomings,
        'active_menu': 'incoming',
        'start_date': sd,
        'end_date': ed,
        'q': q,
        'vendor_list': vendor_list,
        'selected_vendor_id': selected_v
    })

@login_required
@menu_permission_required('can_manage_incoming')
def incoming_export(request):
    # 1. 권한 및 사용자 확인
    user_vendor = Vendor.objects.filter(user=request.user).first()

    # 슈퍼유저가 아니고, 협력사도 아닌 경우에만 권한 체크
    if (not request.user.is_superuser) and (not user_vendor):
        resp = require_action_perm(request, 'incoming.export')
        if resp:
            return resp

    # 2. 기본 쿼리셋 생성
    incomings = Incoming.objects.select_related('part', 'part__vendor').all().order_by('-in_date', '-created_at')

    # 3. 필터링 적용 (화면 조회와 동일한 로직)
    # 3-1. 협력사 계정인 경우 본인 데이터만 필터링
    if (not request.user.is_superuser) and user_vendor:
        incomings = incomings.filter(part__vendor=user_vendor)
    # (관리자 페이지 등에서 특정 업체만 선택해서 조회했을 경우 대응이 필요하다면 아래 주석 해제)
    # elif request.GET.get('vendor_id'):
    #     incomings = incomings.filter(part__vendor_id=request.GET.get('vendor_id'))

    # 3-2. 날짜 필터 (시작일~종료일)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date and end_date:
        incomings = incomings.filter(in_date__range=[start_date, end_date])

    # 3-3. 검색어 필터 (품번/품명)
    q = request.GET.get('q', '')
    if q:
        incomings = incomings.filter(Q(part__part_no__icontains=q) | Q(part__part_name__icontains=q))

    # 4. 엑셀 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    # 헤더 작성
    ws.append(['입고일자', '협력사', '품번', '품명', '입고수량(확정)', '처리일시'])

    # 데이터 작성
    for i in incomings:
        ws.append([
            i.in_date,
            i.part.vendor.name,
            i.part.part_no,
            i.part.part_name,
            i.confirmed_qty,  # [수정] 납품수량(quantity) 대신 확정수량(confirmed_qty) 사용
            i.created_at.strftime("%Y-%m-%d %H:%M")
        ])

    # 5. 응답 반환
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Incomings.xlsx'
    wb.save(response)
    return response


@staff_member_required
@menu_permission_required('can_access_scm_admin')
def scm_admin_main(request):
    today = timezone.localtime().date()
    overdue_list = []
    active_overdue_orders = Order.objects.filter(due_date__lt=today, is_closed=False, approved_at__isnull=False).order_by('due_date')

    for order in active_overdue_orders:
        total_printed = LabelPrintLog.objects.filter(part_no=order.part_no).aggregate(Sum('printed_qty'))['printed_qty__sum'] or 0
        closed_qty = Order.objects.filter(part_no=order.part_no, is_closed=True).aggregate(Sum('quantity'))['quantity__sum'] or 0
        current_printed = max(0, total_printed - closed_qty)
        remain = order.quantity - current_printed

        if remain > 0:
            overdue_list.append({
                'due_date': order.due_date,
                'vendor_name': order.vendor.name,
                'part_no': order.part_no,
                'part_name': order.part_name,
                'remain_qty': remain,
                'days_diff': (today - order.due_date).days
            })

    summary = {
        'total_vendors': Vendor.objects.count(),
        'total_parts': Part.objects.count(),
        'unapproved_orders': Order.objects.filter(approved_at__isnull=True, is_closed=False).count(),
        'today_incoming': Incoming.objects.filter(in_date=today).aggregate(Sum('quantity'))['quantity__sum'] or 0,
        'overdue_count': len(overdue_list)
    }

    recent_incomings = Incoming.objects.select_related('part', 'part__vendor').all().order_by('-created_at')[:10]

    return render(request, 'scm_admin_main.html', {
        'summary': summary,
        'recent_incomings': recent_incomings,
        'overdue_orders': overdue_list[:10],
        'active_menu': 'admin_main',
        'user_name': request.user.username,
        'vendor_name': "시스템 관리자"
    })

# ==========================================
# [맨 마지막] 납품서 입고 확정 (교체 반영)
# ==========================================

@login_required
@menu_permission_required('can_scm_incoming_edit')
@require_POST
def receive_delivery_order_confirm(request):
    from django.db import transaction as db_transaction

    order_id = request.POST.get('order_id')
    inspection_needed = request.POST.get('inspection_needed')
    direct_warehouse_code = request.POST.get('direct_warehouse_code')
    target_warehouse_code = request.POST.get('target_warehouse_code', '2000')  # 수입검사 후 입고될 창고

    do = get_object_or_404(DeliveryOrder, pk=order_id)
    if do.is_received:
        return redirect('incoming_list')

    if Warehouse is None or MaterialStock is None or MaterialTransaction is None:
        messages.error(request, "WMS 연동 모델을 불러올 수 없습니다.")
        return redirect('incoming_list')

    if inspection_needed == 'yes' and ImportInspection is None:
        messages.error(request, "QMS 연동 모델을 불러올 수 없습니다.")
        return redirect('incoming_list')

    try:
        # =====================================================================
        # [FIFO 경고 체크] 실제 입고 처리 전에 FIFO 위반 여부 확인
        # =====================================================================
        mat_warehouse = Warehouse.objects.filter(code='2000').first()  # 자재창고

        if mat_warehouse:  # 수입검사/무검사 모두 체크
            for item in do.items.all():
                part = Part.objects.filter(part_no=item.part_no).first()
                if not part or not item.lot_no:
                    continue

                # FIFO 경고 체크: 자재창고(2000)에 입고 LOT보다 최근 생산품이 있는지 확인
                # (이미 더 최근 LOT가 있는데 과거 LOT를 입고하면 FIFO 위반!)
                newer_lots = MaterialStock.objects.filter(
                    warehouse=mat_warehouse,
                    part=part,
                    lot_no__gt=item.lot_no,  # 입고 LOT보다 생산일이 나중
                    quantity__gt=0
                ).order_by('lot_no')

                if newer_lots.exists():
                    newest_lot = newer_lots.first()
                    days_diff = (newest_lot.lot_no - item.lot_no).days

                    # FIFO 위반 - 무조건 경고 메시지
                    messages.error(
                        request,
                        f"🚨 FIFO 위반: [{item.part_no}] 입고 LOT({item.lot_no.strftime('%Y-%m-%d')})보다 "
                        f"{days_diff}일 최근 생산품이 이미 있습니다! (기존 LOT: {newest_lot.lot_no.strftime('%Y-%m-%d')})"
                    )

        # =====================================================================
        # 실제 입고 처리 시작
        # =====================================================================
        with transaction.atomic():
            do.is_received = True

            if inspection_needed == 'yes':
                do.status = 'RECEIVED'
                target_wh = Warehouse.objects.filter(code='8100').first()
                if not target_wh:
                    target_wh = Warehouse.objects.filter(name__contains='검사').first()
                remark_msg = "[SCM연동] 수입검사 대기 입고 (8100)"
            else:
                do.status = 'APPROVED'
                if direct_warehouse_code:
                    target_wh = Warehouse.objects.filter(code=direct_warehouse_code).first()
                else:
                    target_wh = Warehouse.objects.filter(code='4200').first()

                remark_msg = f"[SCM연동] 무검사 직납 입고 ({target_wh.name if target_wh else '미지정'})"

            if not target_wh:
                raise Exception("입고할 창고 정보를 찾을 수 없습니다.")

            do.save()

            for item in do.items.all():
                part = Part.objects.filter(part_no=item.part_no).first()
                if not part:
                    continue

                # LOT 정보 포함하여 재고 저장
                # select_for_update로 동시성 문제 방지
                with db_transaction.atomic():
                    # 중복 레코드가 있으면 첫 번째만 사용
                    existing_stocks = MaterialStock.objects.filter(
                        warehouse=target_wh,
                        part=part,
                        lot_no=item.lot_no
                    ).select_for_update()

                    if existing_stocks.exists():
                        # 중복이 있으면 첫 번째만 남기고 나머지는 수량 합산 후 삭제
                        stock = existing_stocks.first()
                        if existing_stocks.count() > 1:
                            total_qty = sum(s.quantity for s in existing_stocks)
                            existing_stocks.exclude(id=stock.id).delete()
                            stock.quantity = total_qty
                            stock.save()

                        # 입고 수량 추가
                        stock.quantity = F('quantity') + item.total_qty
                        stock.save()
                        stock.refresh_from_db()
                    else:
                        # 신규 생성
                        stock = MaterialStock.objects.create(
                            warehouse=target_wh,
                            part=part,
                            lot_no=item.lot_no,
                            quantity=item.total_qty
                        )

                # ERP 발주 연결 여부로 발주입고/예외입고 구분
                is_erp_order = bool(item.erp_order_no and item.erp_order_no.strip())
                trx_type = 'IN_SCM' if is_erp_order else 'IN_MANUAL'
                trx_prefix = 'IN-SCM' if is_erp_order else 'IN-MAN'

                trx_no = f"{trx_prefix}-{timezone.now().strftime('%y%m%d%H%M%S')}-{item.id}"
                trx = MaterialTransaction.objects.create(
                    transaction_no=trx_no,
                    transaction_type=trx_type,
                    part=part,
                    lot_no=item.lot_no,
                    quantity=item.total_qty,
                    warehouse_to=target_wh,
                    result_stock=stock.quantity,  # 입고 후 재고량
                    vendor=part.vendor,
                    actor=request.user,
                    ref_delivery_order=do.order_no,
                    remark=remark_msg
                )

                if inspection_needed == 'yes':
                    ImportInspection.objects.create(
                        inbound_transaction=trx,
                        lot_no=item.lot_no,
                        target_warehouse_code=target_warehouse_code,
                        status='PENDING'
                    )
                else:
                    Incoming.objects.create(
                        part=part,
                        quantity=item.total_qty,
                        in_date=timezone.localtime().date(),
                        delivery_order_no=do.order_no,
                        erp_order_no=item.erp_order_no,
                        erp_order_seq=item.erp_order_seq
                    )

                    # ERP 입고등록 (무검사 직납 - 전체수량)
                    try:
                        from material.erp_api import register_erp_incoming
                        erp_ok, erp_no, erp_err = register_erp_incoming(
                            trx, item.total_qty, target_wh.code,
                            erp_order_no=item.erp_order_no or '',
                            erp_order_seq=item.erp_order_seq or ''
                        )
                        if erp_ok:
                            messages.info(request, f'ERP 입고등록 완료: {erp_no} ({item.part_no})')
                        elif erp_err:
                            messages.warning(request, f'ERP 연동 실패: {erp_err} ({item.part_no})')
                    except Exception as e:
                        import logging
                        logging.getLogger(__name__).error(f'ERP 입고등록 예외(SCM): {e}')

            msg = f"{'수입검사 요청' if inspection_needed == 'yes' else '직납 입고'} 완료 (입고창고: {target_wh.name})"
            messages.success(request, f"납품서 처리 완료: {msg}")

    except Exception as e:
        messages.error(request, f"처리 중 오류 발생: {str(e)}")

    return redirect('incoming_list')


# orders/views.py 의 confirm_return 함수 교체

@login_required
@require_POST
def confirm_return(request, pk):
    """
    [협력사 액션] 부적합 반출 확인 (단순 확인용)
    - WMS 재고 차감은 관리자가 이미 수행했다고 가정함.
    - 여기서는 협력사가 '확인' 버튼을 누르면 납품 가능 수량(Remain)만 복구해줌.
    """
    return_log = get_object_or_404(ReturnLog, pk=pk)
    
    # 1. 권한 체크 (본인 회사 물건인지)
    if not request.user.is_superuser:
        user_vendor = _get_user_vendor(request.user)
        if (not user_vendor) or (user_vendor != return_log.part.vendor):
            messages.error(request, "권한이 없습니다.")
            return redirect('label_list')

    # 2. 중복 체크
    if return_log.is_confirmed:
        messages.warning(request, "이미 확인 처리된 건입니다.")
        return redirect('label_list')

    try:
        # 3. 상태 업데이트 (단순 마킹)
        # 재고 로직(WMS)은 일절 개입하지 않음
        return_log.is_confirmed = True
        return_log.confirmed_at = timezone.now()
        return_log.save()

        messages.success(request, f"반출 확인 완료. ({return_log.quantity}ea 만큼 납품 가능 수량이 복구되었습니다.)")

    except Exception as e:
        messages.error(request, f"처리 중 오류 발생: {str(e)}")

    return redirect('label_list')
# ==========================================
# [LOT 관리] LOT별 재고 상세 조회 API
# ==========================================
@login_required
@menu_permission_required('can_view_inventory')
def get_lot_details(request, part_no):
    """
    특정 품목의 LOT별 재고 상세 정보를 JSON으로 반환
    """
    try:
        part = Part.objects.filter(part_no=part_no).first()
        if not part:
            return JsonResponse({'error': '품목을 찾을 수 없습니다.'}, status=404)

        # MaterialStock에서 해당 품목의 LOT별 재고 조회
        if MaterialStock is None:
            return JsonResponse({'error': 'WMS 모듈이 연결되지 않았습니다.'}, status=500)

        lot_stocks = MaterialStock.objects.filter(part=part, quantity__gt=0).select_related('warehouse').order_by('lot_no')

        lot_data = []
        total_qty = 0
        oldest_lot = None

        for stock in lot_stocks:
            lot_info = {
                'warehouse': stock.warehouse.name,
                'warehouse_code': stock.warehouse.code,
                'lot_no': stock.lot_no.strftime('%Y-%m-%d') if stock.lot_no else '-',
                'quantity': stock.quantity,
                'days_old': (timezone.now().date() - stock.lot_no).days if stock.lot_no else 0
            }
            lot_data.append(lot_info)
            total_qty += stock.quantity

            # 가장 오래된 LOT 추적 (FIFO 경고용)
            if stock.lot_no and (oldest_lot is None or stock.lot_no < oldest_lot):
                oldest_lot = stock.lot_no

        # FIFO 경고 판정 (60일 이상 된 LOT가 있으면 경고)
        fifo_warning = False
        if oldest_lot:
            days_old = (timezone.now().date() - oldest_lot).days
            if days_old >= 60:
                fifo_warning = True

        return JsonResponse({
            'part_no': part.part_no,
            'part_name': part.part_name,
            'vendor_name': part.vendor.name,
            'total_quantity': total_qty,
            'lot_details': lot_data,
            'fifo_warning': fifo_warning,
            'oldest_lot': oldest_lot.strftime('%Y-%m-%d') if oldest_lot else None,
            'oldest_days': (timezone.now().date() - oldest_lot).days if oldest_lot else 0
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ==========================================
# [협력사 납기준수율 리포트]
# ==========================================
def _calculate_vendor_monthly_stats(vendor, year_month):
    """협력사 월별 실적 계산 (공통 함수)"""
    from datetime import datetime
    from calendar import monthrange

    # 해당 월의 시작일/종료일 계산
    year, month = map(int, year_month.split('-'))
    start_dt = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    end_dt = date(year, month, last_day)
    today = timezone.localtime().date()

    # 해당 협력사의 발주 (해당 월 납기)
    orders = Order.objects.filter(
        vendor=vendor,
        due_date__gte=start_dt,
        due_date__lte=end_dt,
        approved_at__isnull=False
    )

    order_qty = orders.aggregate(Sum('quantity'))['quantity__sum'] or 0

    if order_qty == 0:
        return None

    # 해당 협력사 품목 목록
    vendor_part_nos = Part.objects.filter(vendor=vendor).values_list('part_no', flat=True)

    # 해당 협력사의 납품서 아이템들 (해당 월 입고)
    delivery_items = DeliveryOrderItem.objects.filter(
        order__created_at__date__gte=start_dt,
        order__created_at__date__lte=end_dt,
        order__status__in=['RECEIVED', 'APPROVED'],
        part_no__in=vendor_part_nos
    ).select_related('linked_order', 'order')

    incoming_qty = delivery_items.aggregate(Sum('total_qty'))['total_qty__sum'] or 0

    # 납기 준수/지연 수량 계산
    on_time_qty = 0
    delayed_qty = 0
    total_lead_time = 0
    lead_time_count = 0

    for item in delivery_items:
        delivery_date = item.order.created_at.date()
        qty = item.total_qty

        if item.linked_order and item.linked_order.due_date:
            due_date = item.linked_order.due_date

            if item.linked_order.created_at:
                lead_time = (delivery_date - item.linked_order.created_at.date()).days
                total_lead_time += max(0, lead_time)
                lead_time_count += 1

            if delivery_date <= due_date:
                on_time_qty += qty
            else:
                delayed_qty += qty
        else:
            on_time_qty += qty

    # 납기가 지난 미입고 수량을 "지연"으로 처리
    for o in orders:
        if o.due_date < today:
            order_incoming = DeliveryOrderItem.objects.filter(
                linked_order=o,
                order__status__in=['RECEIVED', 'APPROVED']
            ).aggregate(Sum('total_qty'))['total_qty__sum'] or 0

            if order_incoming == 0 and o.erp_order_no:
                order_incoming = DeliveryOrderItem.objects.filter(
                    erp_order_no=o.erp_order_no,
                    order__status__in=['RECEIVED', 'APPROVED']
                ).aggregate(Sum('total_qty'))['total_qty__sum'] or 0

            undelivered = max(0, o.quantity - order_incoming)
            if undelivered > 0:
                delayed_qty += undelivered

    # 준수율 계산
    total_qty_for_rate = on_time_qty + delayed_qty
    compliance_rate = (on_time_qty / total_qty_for_rate * 100) if total_qty_for_rate > 0 else 100
    avg_lead_time = (total_lead_time / lead_time_count) if lead_time_count > 0 else 0
    incoming_rate = (incoming_qty / order_qty * 100) if order_qty > 0 else 0

    # 등급 계산
    if compliance_rate >= 95:
        grade = 'A'
    elif compliance_rate >= 85:
        grade = 'B'
    else:
        grade = 'C'

    return {
        'vendor': vendor,
        'order_qty': order_qty,
        'incoming_qty': incoming_qty,
        'incoming_rate': round(incoming_rate, 1),
        'on_time_qty': on_time_qty,
        'delayed_qty': delayed_qty,
        'compliance_rate': round(compliance_rate, 1),
        'avg_lead_time': round(avg_lead_time, 1),
        'grade': grade,
    }


@login_required
@menu_permission_required('can_view_order')
def vendor_delivery_report(request):
    """협력사별 납기준수율 리포트 (기간별 조회 + 마감 기능)"""
    from datetime import datetime
    from calendar import monthrange

    def add_months(d, months):
        """날짜에 월을 더하는 헬퍼 함수 (dateutil 없이)"""
        month = d.month - 1 + months
        year = d.year + month // 12
        month = month % 12 + 1
        day = min(d.day, monthrange(year, month)[1])
        return d.replace(year=year, month=month, day=day)

    today = timezone.localtime().date()

    # 기간 선택 (기본: 이번 달)
    default_month = today.strftime('%Y-%m')
    start_month = request.GET.get('start_month', default_month)
    end_month = request.GET.get('end_month', default_month)
    selected_vendor_id = request.GET.get('vendor_id', '')

    # 시작월이 종료월보다 큰 경우 스왑
    if start_month > end_month:
        start_month, end_month = end_month, start_month

    # 범위 모드 여부 (시작월 != 종료월)
    is_range_mode = start_month != end_month

    # 선택 가능한 월 목록 생성 (최근 12개월)
    month_list = []
    for i in range(12):
        m = today - timedelta(days=30 * i)
        ym = m.strftime('%Y-%m')
        if ym not in [x['value'] for x in month_list]:
            month_list.append({
                'value': ym,
                'label': m.strftime('%Y년 %m월')
            })

    # 협력사 검색 (드롭다운 대신 검색)
    vendor_search = request.GET.get('vendor_search', '').strip()

    # 조회 대상 월 목록 생성
    target_months = []
    current = datetime.strptime(start_month + '-01', '%Y-%m-%d').date()
    end_date = datetime.strptime(end_month + '-01', '%Y-%m-%d').date()
    while current <= end_date:
        target_months.append(current.strftime('%Y-%m'))
        current = add_months(current, 1)

    # 단일 월인 경우 마감 여부 확인 (범위 조회 시에는 마감 버튼 비활성화)
    selected_month = start_month  # 단일월 선택 시 사용
    is_month_closed = False
    if not is_range_mode:
        closed_records = VendorMonthlyPerformance.objects.filter(
            year_month=selected_month, is_closed=True
        )
        is_month_closed = closed_records.exists()

    vendor_stats = []
    total_order_qty = 0
    total_incoming_qty = 0
    total_on_time_qty = 0
    total_delayed_qty = 0

    # 협력사 필터링: 납품 이력이 있는 업체만 (Incoming 테이블 기준)
    if selected_vendor_id:
        target_vendors = Vendor.objects.filter(id=selected_vendor_id)
    elif vendor_search:
        # 검색어가 있으면 검색 결과 중 납품 이력 있는 업체만
        target_vendors = Vendor.objects.filter(
            Q(name__icontains=vendor_search) | Q(code__icontains=vendor_search),
            part__incoming__isnull=False
        ).distinct().order_by('name')
    else:
        # 납품 이력이 있는 업체만 (Incoming 통해서)
        target_vendors = Vendor.objects.filter(
            part__incoming__isnull=False
        ).distinct().order_by('name')

    # 협력사별 통계 계산 (범위 누적)
    for vendor in target_vendors:
        vendor_total = {
            'vendor': vendor,
            'order_qty': 0,
            'incoming_qty': 0,
            'on_time_qty': 0,
            'delayed_qty': 0,
            'lead_time_sum': 0,
            'lead_time_count': 0,
        }

        for ym in target_months:
            # 해당 월이 마감된 경우 DB에서 조회
            closed_rec = VendorMonthlyPerformance.objects.filter(
                vendor=vendor, year_month=ym, is_closed=True
            ).first()

            if closed_rec:
                vendor_total['order_qty'] += closed_rec.order_qty
                vendor_total['incoming_qty'] += closed_rec.incoming_qty
                vendor_total['on_time_qty'] += closed_rec.on_time_qty
                vendor_total['delayed_qty'] += closed_rec.delayed_qty
                if closed_rec.avg_lead_time > 0:
                    vendor_total['lead_time_sum'] += float(closed_rec.avg_lead_time) * closed_rec.incoming_qty
                    vendor_total['lead_time_count'] += closed_rec.incoming_qty
            else:
                # 미마감 월: 실시간 계산
                stats = _calculate_vendor_monthly_stats(vendor, ym)
                if stats:
                    vendor_total['order_qty'] += stats['order_qty']
                    vendor_total['incoming_qty'] += stats['incoming_qty']
                    vendor_total['on_time_qty'] += stats['on_time_qty']
                    vendor_total['delayed_qty'] += stats['delayed_qty']
                    if stats['avg_lead_time'] > 0:
                        vendor_total['lead_time_sum'] += stats['avg_lead_time'] * stats['incoming_qty']
                        vendor_total['lead_time_count'] += stats['incoming_qty']

        # 데이터가 있는 경우만 추가
        if vendor_total['order_qty'] > 0 or vendor_total['incoming_qty'] > 0:
            # 준수율 계산
            total_for_rate = vendor_total['on_time_qty'] + vendor_total['delayed_qty']
            compliance_rate = (vendor_total['on_time_qty'] / total_for_rate * 100) if total_for_rate > 0 else 0
            incoming_rate = (vendor_total['incoming_qty'] / vendor_total['order_qty'] * 100) if vendor_total['order_qty'] > 0 else 0
            avg_lead_time = (vendor_total['lead_time_sum'] / vendor_total['lead_time_count']) if vendor_total['lead_time_count'] > 0 else 0

            # 등급 결정
            if compliance_rate >= 95:
                grade = 'A'
            elif compliance_rate >= 85:
                grade = 'B'
            else:
                grade = 'C'

            vendor_stats.append({
                'vendor': vendor,
                'order_qty': vendor_total['order_qty'],
                'incoming_qty': vendor_total['incoming_qty'],
                'incoming_rate': round(incoming_rate, 1),
                'on_time_qty': vendor_total['on_time_qty'],
                'delayed_qty': vendor_total['delayed_qty'],
                'compliance_rate': round(compliance_rate, 1),
                'avg_lead_time': round(avg_lead_time, 1),
                'grade': grade,
            })

            total_order_qty += vendor_total['order_qty']
            total_incoming_qty += vendor_total['incoming_qty']
            total_on_time_qty += vendor_total['on_time_qty']
            total_delayed_qty += vendor_total['delayed_qty']

    # 준수율 기준 정렬
    vendor_stats.sort(key=lambda x: x['compliance_rate'], reverse=True)

    # 전체 준수율
    total_qty_for_rate = total_on_time_qty + total_delayed_qty
    total_compliance_rate = (total_on_time_qty / total_qty_for_rate * 100) if total_qty_for_rate > 0 else 0
    total_incoming_rate = (total_incoming_qty / total_order_qty * 100) if total_order_qty > 0 else 0

    # 등급별 분류
    grade_a = len([v for v in vendor_stats if v['compliance_rate'] >= 95])
    grade_b = len([v for v in vendor_stats if 85 <= v['compliance_rate'] < 95])
    grade_c = len([v for v in vendor_stats if v['compliance_rate'] < 85])

    # 마감 가능 여부 (단일 월만 마감 가능, 이번 달은 불가)
    can_close = not is_range_mode and selected_month < today.strftime('%Y-%m') and not is_month_closed

    context = {
        'start_month': start_month,
        'end_month': end_month,
        'selected_month': selected_month,  # 마감용 (단일 월)
        'is_range_mode': is_range_mode,
        'month_list': month_list,
        'vendor_search': vendor_search,
        'selected_vendor_id': int(selected_vendor_id) if selected_vendor_id else '',
        'vendor_stats': vendor_stats,
        'total_order_qty': total_order_qty,
        'total_incoming_qty': total_incoming_qty,
        'total_incoming_rate': round(total_incoming_rate, 1),
        'total_on_time_qty': total_on_time_qty,
        'total_delayed_qty': total_delayed_qty,
        'total_compliance_rate': round(total_compliance_rate, 1),
        'grade_a': grade_a,
        'grade_b': grade_b,
        'grade_c': grade_c,
        'is_month_closed': is_month_closed,
        'can_close': can_close,
        'active_menu': 'report',
    }

    return render(request, 'vendor_delivery_report.html', context)


@login_required
@staff_member_required
@require_POST
def vendor_delivery_close_month(request):
    """월별 납기준수율 마감 처리"""
    year_month = request.POST.get('year_month')
    today = timezone.localtime().date()

    # 유효성 검사
    if not year_month:
        messages.error(request, '마감할 월을 선택해주세요.')
        return redirect('vendor_delivery_report')

    if year_month >= today.strftime('%Y-%m'):
        messages.error(request, '현재 월은 마감할 수 없습니다. 지난 달부터 마감 가능합니다.')
        return redirect('vendor_delivery_report')

    # 이미 마감 여부 확인
    if VendorMonthlyPerformance.objects.filter(year_month=year_month, is_closed=True).exists():
        messages.warning(request, f'{year_month}월은 이미 마감되었습니다.')
        return redirect('vendor_delivery_report')

    # 모든 협력사에 대해 실적 계산 및 저장
    vendors = Vendor.objects.all()
    saved_count = 0

    for vendor in vendors:
        stats = _calculate_vendor_monthly_stats(vendor, year_month)
        if stats:
            VendorMonthlyPerformance.objects.update_or_create(
                vendor=vendor,
                year_month=year_month,
                defaults={
                    'order_qty': stats['order_qty'],
                    'incoming_qty': stats['incoming_qty'],
                    'on_time_qty': stats['on_time_qty'],
                    'delayed_qty': stats['delayed_qty'],
                    'compliance_rate': stats['compliance_rate'],
                    'incoming_rate': stats['incoming_rate'],
                    'avg_lead_time': stats['avg_lead_time'],
                    'grade': stats['grade'],
                    'is_closed': True,
                    'closed_at': timezone.now(),
                    'closed_by': request.user,
                }
            )
            saved_count += 1

    messages.success(request, f'{year_month}월 납기준수율이 마감되었습니다. (총 {saved_count}개 협력사)')
    return redirect(f'/report/vendor-delivery/?month={year_month}')


# ==========================================
# [리포트] 알림/모니터링 대시보드
# ==========================================

@login_required
def scm_alert_dashboard(request):
    """SCM 종합 대시보드 - 발주/입고 현황 + 알림"""
    user = request.user
    # 협력사 감지 - 기존 방식(Vendor.user)과 새 방식(UserProfile.org.linked_vendor) 모두 지원
    user_vendor = Vendor.objects.filter(user=user).first()
    if not user_vendor and not user.is_superuser:
        try:
            if hasattr(user, 'profile') and user.profile.org and user.profile.org.linked_vendor:
                user_vendor = user.profile.org.linked_vendor
        except Exception:
            pass
    today = timezone.localtime().date()
    this_month_start = today.replace(day=1)

    # MaterialStock 사용 (WMS 실시간 재고 - 과부족 조회와 동일 기준)
    from material.models import MaterialStock, Warehouse
    target_warehouses = Warehouse.objects.filter(code__in=['2000', '4200'])

    # ========== 1. 발주 통계 ==========
    order_qs = Order.objects.all()
    if user_vendor:
        order_qs = order_qs.filter(vendor=user_vendor)

    # 금일 발주
    today_orders = order_qs.filter(created_at__date=today)
    today_order_count = today_orders.count()
    today_order_qty = today_orders.aggregate(total=Sum('quantity'))['total'] or 0

    # 이번달 발주
    month_orders = order_qs.filter(created_at__date__gte=this_month_start)
    month_order_count = month_orders.count()
    month_order_qty = month_orders.aggregate(total=Sum('quantity'))['total'] or 0

    # 승인 대기 발주
    pending_approval = order_qs.filter(approved_at__isnull=True, is_closed=False).count()

    # 미완료 발주 (승인됨, 미마감)
    open_orders = order_qs.filter(approved_at__isnull=False, is_closed=False).count()

    # ========== 2. 입고 통계 ==========
    incoming_qs = Incoming.objects.all()
    if user_vendor:
        incoming_qs = incoming_qs.filter(part__vendor=user_vendor)

    # 금일 입고
    today_incoming = incoming_qs.filter(in_date=today)
    today_incoming_count = today_incoming.count()
    today_incoming_qty = today_incoming.aggregate(total=Sum('confirmed_qty'))['total'] or 0

    # 이번달 입고
    month_incoming = incoming_qs.filter(in_date__gte=this_month_start)
    month_incoming_count = month_incoming.count()
    month_incoming_qty = month_incoming.aggregate(total=Sum('confirmed_qty'))['total'] or 0

    # ========== 3. 납품서 통계 ==========
    delivery_qs = DeliveryOrder.objects.all()
    if user_vendor:
        # DeliveryOrder는 vendor 필드가 없음 - items__linked_order__vendor로 필터링
        delivery_qs = delivery_qs.filter(items__linked_order__vendor=user_vendor).distinct()

    pending_delivery = delivery_qs.filter(status='PENDING').count()
    today_delivery = delivery_qs.filter(created_at__date=today).count()

    # ========== 4. 협력사별 미입고 현황 (상위 10개) ==========
    if not user_vendor:
        vendor_order_stats = Order.objects.filter(
            approved_at__isnull=False,
            is_closed=False
        ).values(
            'vendor__id', 'vendor__name'
        ).annotate(
            order_count=Count('id'),
            order_qty=Sum('quantity')
        ).order_by('-order_count')[:10]
    else:
        vendor_order_stats = []

    # ========== 5. 최근 입고 이력 ==========
    recent_incoming = incoming_qs.select_related('part', 'part__vendor').order_by('-in_date', '-created_at')[:10]

    # ========== 6. 품목/협력사 현황 ==========
    if user_vendor:
        total_parts = Part.objects.filter(vendor=user_vendor).count()
    else:
        total_parts = Part.objects.count()
    total_vendors = Vendor.objects.count()

    # 협력사 필터링 - 소요량이 있는 품목만 조회 (성능 최적화)
    parts_with_demand = Demand.objects.filter(
        due_date__gte=today,
        due_date__lte=today + timedelta(days=7)
    ).values_list('part_id', flat=True).distinct()

    if user_vendor:
        parts = Part.objects.select_related('vendor').filter(vendor=user_vendor, id__in=parts_with_demand)
    else:
        parts = Part.objects.select_related('vendor').filter(id__in=parts_with_demand)

    # 1. 재고 부족 품목 (과부족 D+7 기준 부족 예상 품목)
    shortage_items = []

    for part in parts:
        # MaterialStock 기준 현재고 (창고 2000, 4200 합산)
        if target_warehouses.exists():
            wms_stock = MaterialStock.objects.filter(part=part, warehouse__in=target_warehouses).aggregate(total=Sum('quantity'))['total'] or 0
        else:
            wms_stock = MaterialStock.objects.filter(part=part).aggregate(total=Sum('quantity'))['total'] or 0
        current_stock = wms_stock

        # D+7까지의 소요량 합산
        future_demand = Demand.objects.filter(
            part=part,
            due_date__gte=today,
            due_date__lte=today + timedelta(days=7)
        ).aggregate(total=Sum('quantity'))['total'] or 0

        # D+7까지의 입고 예정 (PENDING 상태의 납품서 - 등록되었으나 아직 스캔 안된 것)
        pending_incoming = DeliveryOrderItem.objects.filter(
            order__status='PENDING',
            part_no=part.part_no
        ).aggregate(total=Sum('total_qty'))['total'] or 0

        # 예상 재고 = 현재고 + 입고예정 - 소요량
        expected_stock = current_stock + pending_incoming - future_demand

        if expected_stock < 0:
            shortage_items.append({
                'part': part,
                'current_stock': current_stock,
                'future_demand': future_demand,
                'pending_incoming': pending_incoming,
                'expected_shortage': abs(expected_stock),
            })

    # 부족량 순 정렬
    shortage_items.sort(key=lambda x: x['expected_shortage'], reverse=True)

    # 2. 납기 D-3 임박 품목 (3일 이내 납기 도래하는 미입고 발주)
    due_soon_orders = Order.objects.filter(
        is_closed=False,
        approved_at__isnull=False,  # 승인된 발주만
        due_date__gte=today,
        due_date__lte=today + timedelta(days=3)
    ).select_related('vendor')
    if user_vendor:
        due_soon_orders = due_soon_orders.filter(vendor=user_vendor)
    due_soon_orders = due_soon_orders.order_by('due_date')

    # 입고 완료된 수량 계산
    due_soon_list = []
    for order in due_soon_orders:
        part = Part.objects.filter(vendor=order.vendor, part_no=order.part_no).first()
        if part:
            # ERP 발주번호 기반으로 입고 수량 조회
            incoming_qty = Incoming.objects.filter(
                part=part,
                erp_order_no=order.erp_order_no,
                erp_order_seq=order.erp_order_seq
            ).aggregate(total=Sum('confirmed_qty'))['total'] or 0

            remain_qty = order.quantity - incoming_qty
            if remain_qty > 0:
                days_left = (order.due_date - today).days
                due_soon_list.append({
                    'order': order,
                    'part': part,
                    'incoming_qty': incoming_qty,
                    'remain_qty': remain_qty,
                    'days_left': days_left,
                })

    # 3. 미납 발주 (납기일 경과)
    overdue_orders = Order.objects.filter(
        is_closed=False,
        approved_at__isnull=False,
        due_date__lt=today  # 납기일이 지난 모든 발주
    ).select_related('vendor')
    if user_vendor:
        overdue_orders = overdue_orders.filter(vendor=user_vendor)
    overdue_orders = overdue_orders.order_by('due_date')

    overdue_list = []
    for order in overdue_orders:
        part = Part.objects.filter(vendor=order.vendor, part_no=order.part_no).first()
        if part:
            # ERP 발주번호 기반으로 입고 수량 조회
            incoming_qty = Incoming.objects.filter(
                part=part,
                erp_order_no=order.erp_order_no,
                erp_order_seq=order.erp_order_seq
            ).aggregate(total=Sum('confirmed_qty'))['total'] or 0

            remain_qty = order.quantity - incoming_qty
            if remain_qty > 0:
                overdue_days = (today - order.due_date).days
                overdue_list.append({
                    'order': order,
                    'part': part,
                    'incoming_qty': incoming_qty,
                    'remain_qty': remain_qty,
                    'overdue_days': overdue_days,
                })

    # 경과일 순 정렬
    overdue_list.sort(key=lambda x: x['overdue_days'], reverse=True)

    # 공지사항 (최근 5개, 활성화된 것만)
    notices = Notice.objects.filter(is_active=True)[:5]

    # QnA (협력사는 본인 글만, 관리자는 전체)
    if user_vendor:
        qna_list = QnA.objects.filter(vendor=user_vendor)[:10]
    else:
        qna_list = QnA.objects.all()[:10]

    context = {
        # 알림 현황
        'shortage_items': shortage_items[:20],
        'shortage_count': len(shortage_items),
        'due_soon_list': due_soon_list,
        'due_soon_count': len(due_soon_list),
        'overdue_list': overdue_list[:30],
        'overdue_count': len(overdue_list),

        # 발주 통계
        'today_order_count': today_order_count,
        'today_order_qty': today_order_qty,
        'month_order_count': month_order_count,
        'month_order_qty': month_order_qty,
        'pending_approval': pending_approval,
        'open_orders': open_orders,

        # 입고 통계
        'today_incoming_count': today_incoming_count,
        'today_incoming_qty': today_incoming_qty,
        'month_incoming_count': month_incoming_count,
        'month_incoming_qty': month_incoming_qty,

        # 납품서 통계
        'pending_delivery': pending_delivery,
        'today_delivery': today_delivery,

        # 협력사별 현황
        'vendor_order_stats': vendor_order_stats,

        # 최근 입고
        'recent_incoming': recent_incoming,

        # 품목/협력사 현황
        'total_parts': total_parts,
        'total_vendors': total_vendors,

        # 기타
        'today': today,
        'user_vendor': user_vendor,
        'notices': notices,
        'qna_list': qna_list,
    }

    return render(request, 'scm_alert_dashboard.html', context)


@login_required
@require_POST
def notice_create(request):
    """공지사항 등록 (관리자/직원 전용)"""
    # 권한 체크
    user = request.user
    if not user.is_superuser:
        profile = getattr(user, 'profile', None)
        if not profile or profile.role not in ['STAFF', 'ADMIN']:
            messages.error(request, '공지사항 등록 권한이 없습니다.')
            return redirect('scm_alert_dashboard')

    title = request.POST.get('title', '').strip()
    content = request.POST.get('content', '').strip()
    is_important = request.POST.get('is_important') == 'on'

    if not title or not content:
        messages.error(request, '제목과 내용을 모두 입력해주세요.')
        return redirect('scm_alert_dashboard')

    Notice.objects.create(
        title=title,
        content=content,
        is_important=is_important,
        created_by=request.user,
    )

    messages.success(request, '공지사항이 등록되었습니다.')
    return redirect('scm_alert_dashboard')


@login_required
@require_POST
def qna_create(request):
    """QnA 질문 등록"""
    title = request.POST.get('title', '').strip()
    content = request.POST.get('content', '').strip()

    if not title or not content:
        messages.error(request, '제목과 내용을 모두 입력해주세요.')
        return redirect('scm_alert_dashboard')

    user_vendor = Vendor.objects.filter(user=request.user).first()

    QnA.objects.create(
        title=title,
        content=content,
        author=request.user,
        vendor=user_vendor,
    )

    messages.success(request, '질문이 등록되었습니다. 답변을 기다려 주세요.')
    return redirect('scm_alert_dashboard')


@login_required
@require_POST
def qna_answer(request, qna_id):
    """QnA 답변 등록/수정 (관리자/직원 전용)"""
    # 권한 체크
    user = request.user
    if not user.is_superuser:
        profile = getattr(user, 'profile', None)
        if not profile or profile.role not in ['STAFF', 'ADMIN']:
            messages.error(request, '답변 권한이 없습니다.')
            return redirect('scm_alert_dashboard')

    qna = get_object_or_404(QnA, id=qna_id)
    answer = request.POST.get('answer', '').strip()

    if not answer:
        messages.error(request, '답변 내용을 입력해주세요.')
        return redirect('scm_alert_dashboard')

    qna.answer = answer
    qna.answered_by = request.user
    qna.answered_at = timezone.now()
    qna.save()

    messages.success(request, '답변이 등록되었습니다.')
    return redirect('scm_alert_dashboard')


# ==========================================
# [품목 마스터 관리]
# ==========================================

@login_required
@menu_permission_required('can_access_scm_admin')
def part_list(request):
    """품목 마스터 조회 및 업체 연결 관리"""
    user = request.user

    # 권한 체크: 관리자 또는 직원만 접근 가능
    if not user.is_superuser:
        profile = getattr(user, 'profile', None)
        if not profile or profile.role == 'VENDOR':
            messages.error(request, '품목 관리 권한이 없습니다.')
            return redirect('home')

    # 검색 및 필터
    search_q = request.GET.get('q', '').strip()
    vendor_filter = request.GET.get('vendor', '')
    group_filter = request.GET.get('group', '')
    wms_only = request.GET.get('wms_only', '')  # WMS 전용(업체 미연결) 필터

    parts = Part.objects.select_related('vendor').all()

    if search_q:
        parts = parts.filter(
            Q(part_no__icontains=search_q) | Q(part_name__icontains=search_q)
        )

    if vendor_filter:
        parts = parts.filter(vendor_id=vendor_filter)

    if group_filter:
        parts = parts.filter(part_group=group_filter)

    if wms_only == '1':
        parts = parts.filter(vendor__isnull=True)  # 업체 미연결 품목만
    elif not search_q:
        parts = parts.filter(vendor__isnull=False)  # 검색어 없을 때만 업체 연결된 품목 기본 표시

    parts = parts.order_by('-id')[:200]

    # 업체 목록 (필터용)
    vendors = Vendor.objects.all().order_by('name')

    # 품목군 목록 (필터용)
    part_groups = Part.objects.values_list('part_group', flat=True).distinct().order_by('part_group')

    # POST 요청 처리 (업체 연결)
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'connect_vendor':
            part_id = request.POST.get('part_id')
            vendor_id = request.POST.get('vendor_id')

            try:
                part = Part.objects.get(id=part_id)
                if vendor_id:
                    vendor = Vendor.objects.get(id=vendor_id)
                    part.vendor = vendor
                    part.save()
                    messages.success(request, f'품목 [{part.part_no}]에 업체 [{vendor.name}]이(가) 연결되었습니다.')
                else:
                    part.vendor = None
                    part.save()
                    messages.success(request, f'품목 [{part.part_no}]의 업체 연결이 해제되었습니다.')
            except Part.DoesNotExist:
                messages.error(request, '품목을 찾을 수 없습니다.')
            except Vendor.DoesNotExist:
                messages.error(request, '업체를 찾을 수 없습니다.')

            return redirect(request.get_full_path())

        elif action == 'update_part':
            part_id = request.POST.get('part_id')
            part_name = request.POST.get('part_name', '').strip()
            part_group = request.POST.get('part_group', '').strip()
            vendor_id = request.POST.get('vendor_id')

            try:
                part = Part.objects.get(id=part_id)
                if part_name:
                    part.part_name = part_name
                if part_group:
                    part.part_group = part_group
                if vendor_id:
                    part.vendor = Vendor.objects.get(id=vendor_id)
                else:
                    part.vendor = None
                part.save()
                messages.success(request, f'품목 [{part.part_no}] 정보가 수정되었습니다.')
            except Part.DoesNotExist:
                messages.error(request, '품목을 찾을 수 없습니다.')
            except Vendor.DoesNotExist:
                messages.error(request, '업체를 찾을 수 없습니다.')

            return redirect(request.get_full_path())

        elif action == 'upload_vendor_excel':
            # 엑셀로 품번-품목군-업체 일괄 연결
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, '엑셀 파일을 선택해주세요.')
                return redirect(request.get_full_path())

            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active

                updated_count = 0
                not_found_parts = []
                not_found_vendors = []

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not row[0]:
                        continue

                    part_no = str(row[0]).strip()
                    part_group = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    vendor_code = str(row[2]).strip() if len(row) > 2 and row[2] else ''

                    # 품번으로 Part 찾기
                    try:
                        part = Part.objects.get(part_no=part_no)
                    except Part.DoesNotExist:
                        not_found_parts.append(part_no)
                        continue

                    changed = False

                    # 품목군 업데이트
                    if part_group and part.part_group != part_group:
                        part.part_group = part_group
                        changed = True

                    # 업체 코드로 Vendor 찾기
                    if vendor_code:
                        try:
                            vendor = Vendor.objects.get(code=vendor_code)
                            if part.vendor != vendor:
                                part.vendor = vendor
                                changed = True
                        except Vendor.DoesNotExist:
                            not_found_vendors.append(vendor_code)
                    else:
                        # 업체 코드가 비어있으면 연결 해제
                        if part.vendor:
                            part.vendor = None
                            changed = True

                    if changed:
                        part.save()
                        updated_count += 1

                # 결과 메시지
                if updated_count > 0:
                    messages.success(request, f'{updated_count}건의 품목이 업데이트되었습니다.')

                if not_found_parts:
                    messages.warning(request, f'품번을 찾을 수 없음: {", ".join(not_found_parts[:5])}{"..." if len(not_found_parts) > 5 else ""}')

                if not_found_vendors:
                    unique_vendors = list(set(not_found_vendors))
                    messages.warning(request, f'업체코드를 찾을 수 없음: {", ".join(unique_vendors[:5])}{"..." if len(unique_vendors) > 5 else ""}')

            except Exception as e:
                messages.error(request, f'엑셀 처리 중 오류 발생: {str(e)}')

            return redirect(request.get_full_path())

    context = {
        'parts': parts,
        'vendors': vendors,
        'part_groups': part_groups,
        'search_q': search_q,
        'vendor_filter': vendor_filter,
        'group_filter': group_filter,
        'wms_only': wms_only,
    }
    return render(request, 'part_list.html', context)


@login_required
@menu_permission_required('can_access_scm_admin')
def part_vendor_template(request):
    """품번-품목군-업체 연결용 엑셀 템플릿 다운로드"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "품번-품목군-업체"

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더
    headers = ['품번', '품목군', '업체코드']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 데이터 포함 여부
    include_data = request.GET.get('include_data', '')

    if include_data:
        # 업체 미연결 품목 우선, 나머지는 품번 순
        parts = Part.objects.select_related('vendor').all().order_by('vendor', 'part_no')

        for row_idx, part in enumerate(parts, start=2):
            cell_a = ws.cell(row=row_idx, column=1, value=part.part_no)
            cell_a.border = thin_border
            cell_a.number_format = '@'  # 텍스트 형식

            cell_b = ws.cell(row=row_idx, column=2, value=part.part_group or '')
            cell_b.border = thin_border

            cell_c = ws.cell(row=row_idx, column=3, value=part.vendor.code if part.vendor else '')
            cell_c.border = thin_border
            cell_c.number_format = '@'  # 텍스트 형식 (00104 → "00104")

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15

    # 업체 목록 시트 추가
    ws_vendors = wb.create_sheet(title="업체목록(참고)")
    ws_vendors.cell(row=1, column=1, value="업체코드").fill = header_fill
    ws_vendors.cell(row=1, column=1).font = header_font
    ws_vendors.cell(row=1, column=2, value="업체명").fill = header_fill
    ws_vendors.cell(row=1, column=2).font = header_font

    vendors = Vendor.objects.all().order_by('name')
    for row_idx, vendor in enumerate(vendors, start=2):
        cell_code = ws_vendors.cell(row=row_idx, column=1, value=vendor.code)
        cell_code.number_format = '@'  # 텍스트 형식
        ws_vendors.cell(row=row_idx, column=2, value=vendor.name)

    ws_vendors.column_dimensions['A'].width = 15
    ws_vendors.column_dimensions['B'].width = 30

    # 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="part_vendor_template.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# 품목마스터 일괄 업로드
# =============================================================================

@login_required
@menu_permission_required('can_access_scm_admin')
def part_upload(request):
    """품목마스터 일괄 업로드 페이지"""
    # 최근 실패 로그 조회
    error_logs = InventoryUploadLog.objects.filter(
        upload_type='PART_MASTER'
    ).order_by('-uploaded_at')[:20]

    return render(request, 'part_upload.html', {
        'error_logs': error_logs,
    })


@login_required
@menu_permission_required('can_access_scm_admin')
def part_upload_preview(request):
    """품목마스터 업로드 미리보기"""
    if request.method != 'POST':
        return redirect('part_upload')

    upload_file = request.FILES.get('upload_file')
    if not upload_file:
        messages.error(request, '파일을 선택해주세요.')
        return redirect('part_upload')

    try:
        import openpyxl
        import csv
        import io

        # 파일 확장자 확인
        filename = upload_file.name.lower()
        data_rows = []

        if filename.endswith('.csv'):
            # CSV 파일 처리
            content = upload_file.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content))
            for row in reader:
                data_rows.append(row)
        else:
            # 엑셀 파일 처리
            wb = openpyxl.load_workbook(upload_file)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                data_rows.append(row)

        # 첫 행은 헤더로 가정하고 스킵
        preview_data = []
        new_count = 0
        update_count = 0
        error_count = 0

        # 업체 목록 미리 로드 (코드 -> 업체 매핑)
        vendors_by_code = {v.code: v for v in Vendor.objects.all()}

        # 계정구분 매핑
        account_type_map = {
            '원재료': 'RAW', 'RAW': 'RAW', '원자재': 'RAW',
            '상품': 'PRODUCT', 'PRODUCT': 'PRODUCT',
            '제품': 'FINISHED', 'FINISHED': 'FINISHED',
        }
        account_type_display = {
            'RAW': '원재료', 'PRODUCT': '상품', 'FINISHED': '제품'
        }

        for row_idx, row in enumerate(data_rows[1:], start=2):  # 헤더 스킵
            if not row or not row[0]:
                continue

            part_no = str(row[0]).strip()
            part_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            part_group = str(row[2]).strip() if len(row) > 2 and row[2] else '일반'
            account_type_raw = str(row[3]).strip() if len(row) > 3 and row[3] else '원재료'
            vendor_code = str(row[4]).strip() if len(row) > 4 and row[4] else ''

            # 계정구분 변환
            account_type = account_type_map.get(account_type_raw.upper(), 'RAW')
            account_type_disp = account_type_display.get(account_type, '원재료')

            # 업체 조회
            vendor = vendors_by_code.get(vendor_code) if vendor_code else None
            vendor_name = vendor.name if vendor else ''

            # 품번 검증
            if not part_no:
                status = 'error'
                note = '품번 누락'
                error_count += 1
            elif not part_name:
                status = 'error'
                note = '품명 누락'
                error_count += 1
            elif vendor_code and not vendor:
                status = 'error'
                note = f'업체코드 [{vendor_code}] 없음'
                error_count += 1
            else:
                # 기존 품번 존재 여부 확인
                existing_part = Part.objects.filter(part_no=part_no).first()
                if existing_part:
                    status = 'update'
                    note = '기존 품목 업데이트'
                    update_count += 1
                else:
                    status = 'new'
                    note = ''
                    new_count += 1

            preview_data.append({
                'part_no': part_no,
                'part_name': part_name,
                'part_group': part_group,
                'account_type': account_type,
                'account_type_display': account_type_disp,
                'vendor_code': vendor_code,
                'vendor_name': vendor_name,
                'status': status,
                'note': note,
            })

        return render(request, 'part_upload.html', {
            'preview_data': preview_data,
            'new_count': new_count,
            'update_count': update_count,
            'error_count': error_count,
        })

    except Exception as e:
        messages.error(request, f'파일 처리 중 오류: {str(e)}')
        return redirect('part_upload')


@login_required
@menu_permission_required('can_access_scm_admin')
def part_upload_confirm(request):
    """품목마스터 업로드 확정"""
    if request.method != 'POST':
        return redirect('part_upload')

    part_no_list = request.POST.getlist('part_no_list[]')
    part_name_list = request.POST.getlist('part_name_list[]')
    part_group_list = request.POST.getlist('part_group_list[]')
    account_type_list = request.POST.getlist('account_type_list[]')
    vendor_code_list = request.POST.getlist('vendor_code_list[]')

    # 업체 목록 미리 로드
    vendors_by_code = {v.code: v for v in Vendor.objects.all()}

    created_count = 0
    updated_count = 0

    for i in range(len(part_no_list)):
        part_no = part_no_list[i]
        part_name = part_name_list[i] if i < len(part_name_list) else ''
        part_group = part_group_list[i] if i < len(part_group_list) else '일반'
        account_type = account_type_list[i] if i < len(account_type_list) else 'RAW'
        vendor_code = vendor_code_list[i] if i < len(vendor_code_list) else ''

        vendor = vendors_by_code.get(vendor_code) if vendor_code else None

        # 기존 품목 확인
        existing_part = Part.objects.filter(part_no=part_no).first()

        if existing_part:
            # 업데이트
            existing_part.part_name = part_name
            existing_part.part_group = part_group
            existing_part.account_type = account_type
            if vendor:
                existing_part.vendor = vendor
            existing_part.save()
            updated_count += 1
        else:
            # 신규 생성
            Part.objects.create(
                part_no=part_no,
                part_name=part_name,
                part_group=part_group,
                account_type=account_type,
                vendor=vendor,
            )
            created_count += 1

    messages.success(request, f'품목마스터 등록 완료: 신규 {created_count}건, 업데이트 {updated_count}건')

    return render(request, 'part_upload.html', {
        'result': {
            'created': created_count,
            'updated': updated_count,
        }
    })


@login_required
@menu_permission_required('can_access_scm_admin')
def part_upload_template(request):
    """품목마스터 업로드용 엑셀 템플릿 다운로드"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "품목마스터"

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더
    headers = ['품번', '품명', '품목군', '계정구분', '업체코드']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 데이터 포함 여부
    include_data = request.GET.get('include_data', '')

    account_type_display = {'RAW': '원재료', 'PRODUCT': '상품', 'FINISHED': '제품'}

    if include_data:
        parts = Part.objects.select_related('vendor').all().order_by('part_no')

        for row_idx, part in enumerate(parts, start=2):
            ws.cell(row=row_idx, column=1, value=part.part_no).border = thin_border
            ws.cell(row=row_idx, column=2, value=part.part_name).border = thin_border
            ws.cell(row=row_idx, column=3, value=part.part_group or '일반').border = thin_border
            ws.cell(row=row_idx, column=4, value=account_type_display.get(part.account_type, '원재료')).border = thin_border
            cell_e = ws.cell(row=row_idx, column=5, value=part.vendor.code if part.vendor else '')
            cell_e.border = thin_border
            cell_e.number_format = '@'

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15

    # 업체 목록 시트 추가
    ws_vendors = wb.create_sheet(title="업체목록(참고)")
    ws_vendors.cell(row=1, column=1, value="업체코드").fill = header_fill
    ws_vendors.cell(row=1, column=1).font = header_font
    ws_vendors.cell(row=1, column=2, value="업체명").fill = header_fill
    ws_vendors.cell(row=1, column=2).font = header_font

    vendors = Vendor.objects.all().order_by('name')
    for row_idx, vendor in enumerate(vendors, start=2):
        cell_code = ws_vendors.cell(row=row_idx, column=1, value=vendor.code)
        cell_code.number_format = '@'
        ws_vendors.cell(row=row_idx, column=2, value=vendor.name)

    ws_vendors.column_dimensions['A'].width = 15
    ws_vendors.column_dimensions['B'].width = 30

    # 계정구분 안내 시트
    ws_help = wb.create_sheet(title="계정구분안내")
    ws_help.cell(row=1, column=1, value="계정구분").fill = header_fill
    ws_help.cell(row=1, column=1).font = header_font
    ws_help.cell(row=1, column=2, value="설명").fill = header_fill
    ws_help.cell(row=1, column=2).font = header_font

    help_data = [
        ('원재료', '제조에 투입되는 원재료'),
        ('상품', '외부에서 구매하여 그대로 판매하는 상품'),
        ('제품', '자사에서 생산한 완제품'),
    ]
    for row_idx, (acct, desc) in enumerate(help_data, start=2):
        ws_help.cell(row=row_idx, column=1, value=acct)
        ws_help.cell(row=row_idx, column=2, value=desc)

    ws_help.column_dimensions['A'].width = 15
    ws_help.column_dimensions['B'].width = 40

    # 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="part_master_template.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# 사용자 권한 관리
# =============================================================================

@login_required
@menu_permission_required('can_access_scm_admin')
def user_permission_manage(request):
    """사용자 권한 관리 페이지 (SCM 스타일)"""
    # 관리자만 접근 가능
    if not request.user.is_superuser:
        profile = getattr(request.user, 'profile', None)
        if not profile or profile.role != 'ADMIN':
            messages.error(request, '권한 관리 메뉴에 접근할 수 없습니다.')
            return redirect('home')

    # 권한 필드 정의 (카테고리 > 기능 > View/Edit 분리)
    PERMISSION_FIELDS = {
        'SCM (발주관리)': [
            ('can_scm_order_view', '📋 발주 조회/승인'),
            ('can_scm_order_edit', '✏️ 발주 등록'),
            ('can_scm_label_view', '📋 납품서 조회'),
            ('can_scm_label_edit', '✏️ 납품서 등록'),
            ('can_scm_incoming_view', '📋 입고 현황 조회'),
            ('can_scm_incoming_edit', '✏️ 입고 처리'),
            ('can_scm_inventory_view', '📋 재고/소요 조회'),
            ('can_scm_inventory_edit', '✏️ 소요량 수정'),
            ('can_scm_report', '📊 납기준수율 리포트'),
            ('can_scm_admin', '🔧 관리자'),
        ],
        'WMS (자재관리)': [
            ('can_wms_stock_view', '📋 재고/수불 조회'),
            ('can_wms_stock_edit', '✏️ 재고 조정/이동'),
            ('can_wms_inout_view', '📋 입출고 내역 조회'),
            ('can_wms_inout_edit', '✏️ 입출고 처리'),
            ('can_wms_bom_view', '📋 BOM 조회'),
            ('can_wms_bom_edit', '✏️ BOM 등록/수정'),
        ],
        'QMS (품질관리)': [
            ('can_qms_4m_view', '📋 4M 변경 조회'),
            ('can_qms_4m_edit', '✏️ 4M 등록/수정'),
            ('can_qms_inspection_view', '📋 검사 조회'),
            ('can_qms_inspection_edit', '✏️ 검사 등록/판정'),
            ('can_qms_nc_view', '📋 부적합/CAPA 조회'),
            ('can_qms_nc_edit', '✏️ 부적합/CAPA 등록'),
            ('can_qms_claim_view', '📋 클레임 조회'),
            ('can_qms_claim_edit', '✏️ 클레임 등록/처리'),
            ('can_qms_isir_view', '📋 ISIR 조회'),
            ('can_qms_isir_edit', '✏️ ISIR 등록/승인'),
            ('can_qms_rating_view', '📋 협력사평가 조회'),
            ('can_qms_rating_edit', '✏️ 협력사평가 등록'),
        ],
    }

    # 필터
    role_filter = request.GET.get('role', '')
    search_q = request.GET.get('q', '').strip()

    # 사용자 목록 (superuser 제외, profile 있는 사용자만)
    users = User.objects.filter(is_superuser=False).select_related('profile').order_by('username')

    if role_filter:
        users = users.filter(profile__role=role_filter)

    if search_q:
        users = users.filter(
            Q(username__icontains=search_q) |
            Q(profile__display_name__icontains=search_q)
        )

    # 선택된 사용자
    selected_user_id = request.GET.get('user_id') or request.POST.get('user_id')
    selected_user = None
    selected_profile = None

    if selected_user_id:
        try:
            selected_user = User.objects.get(id=selected_user_id)
            selected_profile, _ = UserProfile.objects.get_or_create(user=selected_user)
        except User.DoesNotExist:
            pass

    # POST: 권한 저장
    if request.method == 'POST' and selected_profile:
        action = request.POST.get('action')

        if action == 'save_permissions':
            # role 변경
            new_role = request.POST.get('role')
            if new_role in ['ADMIN', 'STAFF', 'VENDOR']:
                selected_profile.role = new_role

            # 기본 정보 저장 (표시이름, 부서)
            selected_profile.display_name = request.POST.get('display_name', '').strip() or None
            selected_profile.department = request.POST.get('department', '').strip() or None

            # 개별 권한 업데이트
            for category, fields in PERMISSION_FIELDS.items():
                for field_name, _ in fields:
                    value = request.POST.get(field_name) == 'on'
                    setattr(selected_profile, field_name, value)

            selected_profile.save()
            messages.success(request, f'{selected_user.username} 사용자의 권한이 저장되었습니다.')
            return redirect(f"{request.path}?user_id={selected_user_id}")

        elif action == 'grant_all':
            # 전체 권한 부여
            for category, fields in PERMISSION_FIELDS.items():
                for field_name, _ in fields:
                    setattr(selected_profile, field_name, True)
            selected_profile.save()
            messages.success(request, f'{selected_user.username} 사용자에게 전체 권한이 부여되었습니다.')
            return redirect(f"{request.path}?user_id={selected_user_id}")

        elif action == 'revoke_all':
            # 전체 권한 해제
            for category, fields in PERMISSION_FIELDS.items():
                for field_name, _ in fields:
                    setattr(selected_profile, field_name, False)
            selected_profile.save()
            messages.success(request, f'{selected_user.username} 사용자의 모든 권한이 해제되었습니다.')
            return redirect(f"{request.path}?user_id={selected_user_id}")

    context = {
        'users': users,
        'selected_user': selected_user,
        'selected_profile': selected_profile,
        'permission_fields': PERMISSION_FIELDS,
        'role_filter': role_filter,
        'search_q': search_q,
        'role_choices': UserProfile.ROLE_CHOICES,
    }
    return render(request, 'user_permission_manage.html', context)


# ============================================
# 사용자 관리 (등록/수정/삭제)
# ============================================

@login_required
def user_manage(request):
    """사용자 관리 페이지"""
    # 관리자만 접근 가능
    if not request.user.is_superuser:
        profile = getattr(request.user, 'profile', None)
        if not profile or profile.role != 'ADMIN':
            messages.error(request, '사용자 관리 메뉴에 접근할 수 없습니다.')
            return redirect('home')

    # 필터
    search_q = request.GET.get('q', '').strip()
    role_filter = request.GET.get('role', '')
    account_type_filter = request.GET.get('account_type', '')

    # 사용자 목록
    users = User.objects.select_related('profile', 'profile__org').order_by('-date_joined')

    if search_q:
        users = users.filter(
            Q(username__icontains=search_q) |
            Q(profile__display_name__icontains=search_q) |
            Q(profile__department__icontains=search_q)
        )

    if role_filter:
        users = users.filter(profile__role=role_filter)

    if account_type_filter:
        users = users.filter(profile__account_type=account_type_filter)

    # 조직(협력사) 목록 (드롭다운용)
    organizations = Organization.objects.filter(org_type='VENDOR').order_by('name')

    context = {
        'users': users,
        'search_q': search_q,
        'role_filter': role_filter,
        'account_type_filter': account_type_filter,
        'role_choices': UserProfile.ROLE_CHOICES,
        'organizations': organizations,
    }
    return render(request, 'user_manage.html', context)


@login_required
def user_create(request):
    """신규 사용자 등록"""
    if request.method != 'POST':
        return redirect('user_manage')

    # 관리자 권한 확인
    if not request.user.is_superuser:
        profile = getattr(request.user, 'profile', None)
        if not profile or profile.role != 'ADMIN':
            messages.error(request, '권한이 없습니다.')
            return redirect('user_manage')

    username = request.POST.get('username', '').strip()
    password = request.POST.get('password', '')
    display_name = request.POST.get('display_name', '').strip() or None
    department = request.POST.get('department', '').strip() or None
    role = request.POST.get('role', 'VENDOR')
    account_type = request.POST.get('account_type', 'VENDOR')
    org_id = request.POST.get('org_id', '') or None

    if not username or not password:
        messages.error(request, '사용자명과 비밀번호는 필수입니다.')
        return redirect('user_manage')

    if User.objects.filter(username=username).exists():
        messages.error(request, f'이미 존재하는 사용자명입니다: {username}')
        return redirect('user_manage')

    # 사용자 생성
    user = User.objects.create_user(username=username, password=password)

    # 프로필 생성/업데이트
    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.display_name = display_name
    profile.department = department
    profile.role = role
    profile.account_type = account_type
    if org_id:
        profile.org_id = org_id
    profile.save()

    messages.success(request, f'사용자 "{username}"이(가) 등록되었습니다.')
    return redirect('user_manage')


@login_required
def user_update(request):
    """사용자 정보 수정"""
    if request.method != 'POST':
        return redirect('user_manage')

    # 관리자 권한 확인
    if not request.user.is_superuser:
        profile = getattr(request.user, 'profile', None)
        if not profile or profile.role != 'ADMIN':
            messages.error(request, '권한이 없습니다.')
            return redirect('user_manage')

    user_id = request.POST.get('user_id')
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, '사용자를 찾을 수 없습니다.')
        return redirect('user_manage')

    # 비밀번호 변경 (입력된 경우에만)
    new_password = request.POST.get('new_password', '').strip()
    if new_password:
        user.set_password(new_password)
        user.save()

    # 프로필 업데이트
    profile, _ = UserProfile.objects.get_or_create(user=user)
    profile.display_name = request.POST.get('display_name', '').strip() or None
    profile.department = request.POST.get('department', '').strip() or None
    profile.role = request.POST.get('role', 'VENDOR')
    profile.account_type = request.POST.get('account_type', 'VENDOR')
    org_id = request.POST.get('org_id', '') or None
    profile.org_id = org_id if org_id else None
    profile.save()

    messages.success(request, f'사용자 "{user.username}" 정보가 수정되었습니다.')
    return redirect('user_manage')


@login_required
def user_delete(request):
    """사용자 삭제"""
    if request.method != 'POST':
        return redirect('user_manage')

    # 관리자 권한 확인
    if not request.user.is_superuser:
        profile = getattr(request.user, 'profile', None)
        if not profile or profile.role != 'ADMIN':
            messages.error(request, '권한이 없습니다.')
            return redirect('user_manage')

    user_id = request.POST.get('user_id')
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, '사용자를 찾을 수 없습니다.')
        return redirect('user_manage')

    # 슈퍼유저는 삭제 불가
    if user.is_superuser:
        messages.error(request, '슈퍼유저는 삭제할 수 없습니다.')
        return redirect('user_manage')

    username = user.username
    user.delete()
    messages.success(request, f'사용자 "{username}"이(가) 삭제되었습니다.')
    return redirect('user_manage')


# ============================================
# 협력사 관리
# ============================================

@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_manage(request):
    """협력사 관리 메인 페이지"""
    from django.core.paginator import Paginator
    from django.contrib.auth.models import User

    query = request.GET.get('q', '')
    has_user_filter = request.GET.get('has_user', '')

    vendors = Vendor.objects.select_related('user').all()

    if query:
        vendors = vendors.filter(
            Q(name__icontains=query) | Q(code__icontains=query)
        )

    if has_user_filter == 'yes':
        vendors = vendors.filter(user__isnull=False)
    elif has_user_filter == 'no':
        vendors = vendors.filter(user__isnull=True)

    vendors = vendors.order_by('code')

    paginator = Paginator(vendors, 50)
    page = request.GET.get('page', 1)
    vendors = paginator.get_page(page)

    # 연결 가능한 사용자 (다른 협력사에 연결되지 않은 모든 사용자)
    linked_user_ids = Vendor.objects.filter(user__isnull=False).values_list('user_id', flat=True)
    available_users = User.objects.exclude(id__in=linked_user_ids).order_by('username')

    context = {
        'vendors': vendors,
        'query': query,
        'has_user_filter': has_user_filter,
        'available_users': available_users,
    }
    return render(request, 'vendor_manage.html', context)


@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_detail(request, vendor_id):
    """협력사 상세 정보 (JSON)"""
    from django.http import JsonResponse

    try:
        vendor = Vendor.objects.get(id=vendor_id)
        return JsonResponse({
            'id': vendor.id,
            'code': vendor.code,
            'name': vendor.name,
            'biz_registration_number': vendor.biz_registration_number,
            'representative': vendor.representative,
            'address': vendor.address,
            'biz_type': vendor.biz_type,
            'biz_item': vendor.biz_item,
        })
    except Vendor.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_create(request):
    """협력사 신규 등록"""
    if request.method != 'POST':
        return redirect('vendor_manage')

    code = request.POST.get('code', '').strip()
    name = request.POST.get('name', '').strip()

    if not code or not name:
        messages.error(request, '업체코드와 업체명은 필수입니다.')
        return redirect('vendor_manage')

    if Vendor.objects.filter(code=code).exists():
        messages.error(request, f'업체코드 "{code}"가 이미 존재합니다.')
        return redirect('vendor_manage')

    try:
        Vendor.objects.create(
            code=code,
            name=name,
            erp_code=code,
            biz_registration_number=request.POST.get('biz_registration_number') or None,
            representative=request.POST.get('representative') or None,
            address=request.POST.get('address') or None,
            biz_type=request.POST.get('biz_type') or None,
            biz_item=request.POST.get('biz_item') or None,
        )
        messages.success(request, f'협력사 "{name}"이(가) 등록되었습니다.')
    except Exception as e:
        messages.error(request, f'등록 실패: {e}')

    return redirect('vendor_manage')


@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_update(request):
    """협력사 수정"""
    if request.method != 'POST':
        return redirect('vendor_manage')

    vendor_id = request.POST.get('vendor_id')
    try:
        vendor = Vendor.objects.get(id=vendor_id)
        vendor.name = request.POST.get('name', '').strip()
        vendor.biz_registration_number = request.POST.get('biz_registration_number') or None
        vendor.representative = request.POST.get('representative') or None
        vendor.address = request.POST.get('address') or None
        vendor.biz_type = request.POST.get('biz_type') or None
        vendor.biz_item = request.POST.get('biz_item') or None
        vendor.save()

        # Organization 이름도 동기화
        if hasattr(vendor, 'organization') and vendor.organization:
            vendor.organization.name = vendor.name
            vendor.organization.save()

        messages.success(request, f'협력사 "{vendor.name}"이(가) 수정되었습니다.')
    except Vendor.DoesNotExist:
        messages.error(request, '협력사를 찾을 수 없습니다.')
    except Exception as e:
        messages.error(request, f'수정 실패: {e}')

    return redirect('vendor_manage')


@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_delete(request):
    """협력사 삭제"""
    if request.method != 'POST':
        return redirect('vendor_manage')

    vendor_id = request.POST.get('vendor_id')
    try:
        vendor = Vendor.objects.get(id=vendor_id)

        # 연결된 데이터 체크
        if vendor.order_set.exists():
            messages.error(request, f'"{vendor.name}"에 연결된 발주 데이터가 있어 삭제할 수 없습니다.')
            return redirect('vendor_manage')

        if vendor.part_set.exists():
            messages.error(request, f'"{vendor.name}"에 연결된 품목 데이터가 있어 삭제할 수 없습니다.')
            return redirect('vendor_manage')

        name = vendor.name
        vendor.delete()
        messages.success(request, f'협력사 "{name}"이(가) 삭제되었습니다.')
    except Vendor.DoesNotExist:
        messages.error(request, '협력사를 찾을 수 없습니다.')
    except Exception as e:
        messages.error(request, f'삭제 실패: {e}')

    return redirect('vendor_manage')


@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_export(request):
    """협력사 전체 엑셀 다운로드"""
    vendors = Vendor.objects.all().order_by('code')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "협력사"

    # 헤더
    headers = ['코드', '업체명', '사업자번호', 'ERP코드', '대표자', '주소', '업태', '종목', '연결사용자']
    ws.append(headers)

    # 스타일
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=1, column=col).fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # 데이터
    for v in vendors:
        ws.append([
            v.code,
            v.name,
            v.biz_registration_number or '',
            v.erp_code or '',
            v.representative or '',
            v.address or '',
            v.biz_type or '',
            v.biz_item or '',
            v.user.username if v.user else '',
        ])

    # 열 너비 조정
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15

    today = timezone.localtime().strftime('%Y%m%d')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=vendors_{today}.xlsx'
    wb.save(response)
    return response


@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_link_user(request):
    """협력사에 사용자 연결 (Vendor.user OneToOneField 사용)"""
    if request.method != 'POST':
        return redirect('vendor_manage')

    from django.contrib.auth.models import User

    vendor_id = request.POST.get('vendor_id')
    user_id = request.POST.get('user_id')

    try:
        vendor = Vendor.objects.get(id=vendor_id)
        user = User.objects.get(id=user_id)

        # Vendor.user에 직접 연결
        vendor.user = user
        vendor.save()

        messages.success(request, f'"{vendor.name}"에 사용자 "{user.username}"이(가) 연결되었습니다.')
    except Vendor.DoesNotExist:
        messages.error(request, '협력사를 찾을 수 없습니다.')
    except User.DoesNotExist:
        messages.error(request, '사용자를 찾을 수 없습니다.')
    except Exception as e:
        messages.error(request, f'연결 실패: {e}')

    return redirect('vendor_manage')


@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_unlink_user(request):
    """협력사 사용자 연결 해제"""
    if request.method != 'POST':
        return redirect('vendor_manage')

    vendor_id = request.POST.get('vendor_id')

    try:
        vendor = Vendor.objects.get(id=vendor_id)
        username = vendor.user.username if vendor.user else ''
        vendor.user = None
        vendor.save()

        messages.success(request, f'"{vendor.name}" 협력사의 사용자 연결이 해제되었습니다.')
    except Vendor.DoesNotExist:
        messages.error(request, '협력사를 찾을 수 없습니다.')
    except Exception as e:
        messages.error(request, f'연결 해제 실패: {e}')

    return redirect('vendor_manage')


@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_search_users(request):
    """협력사에 연결 가능한 사용자 검색 API"""
    query = request.GET.get('q', '').strip()

    # 이미 다른 협력사에 연결된 사용자 ID
    linked_user_ids = Vendor.objects.filter(user__isnull=False).values_list('user_id', flat=True)

    # 검색 + 연결 안된 사용자만
    users = User.objects.exclude(id__in=linked_user_ids)

    if query:
        users = users.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )

    users = users.order_by('username')[:20]  # 최대 20개

    result = []
    for user in users:
        name = f"{user.last_name}{user.first_name}".strip()
        result.append({
            'id': user.id,
            'username': user.username,
            'name': name if name else None
        })

    return JsonResponse({'users': result})


# ============================================
# 협력사 일괄 업로드
# ============================================

@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_upload(request):
    """협력사 일괄 업로드 페이지"""
    return render(request, 'vendor_upload.html')


@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_upload_preview(request):
    """협력사 업로드 미리보기 (심플/ERP 양식 지원)"""
    if request.method != 'POST':
        return redirect('vendor_upload')

    upload_file = request.FILES.get('upload_file')
    format_type = request.POST.get('format_type', 'simple')  # simple or erp

    if not upload_file:
        messages.error(request, '파일을 선택해주세요.')
        return redirect('vendor_upload')

    import csv
    import io

    preview_data = []
    new_count = 0
    update_count = 0
    error_count = 0

    try:
        # CSV 파일 읽기
        if upload_file.name.endswith('.csv'):
            content = upload_file.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
        else:
            # Excel 파일
            import openpyxl
            wb = openpyxl.load_workbook(upload_file, data_only=True)
            ws = wb.active
            rows = [[cell.value or '' for cell in row] for row in ws.iter_rows()]

        if format_type == 'simple':
            # 심플 양식: 1행 헤더 + 1행/업체
            # 컬럼: 코드, 업체명, 사업자번호, ERP코드, 대표자, 주소, 업태, 종목, (연결사용자-무시)
            for i, row in enumerate(rows):
                if i == 0:  # 헤더 스킵
                    continue
                if len(row) < 2 or not row[0] or not str(row[0]).strip():
                    continue

                code = str(row[0]).strip()
                name = str(row[1]).strip() if len(row) > 1 else ''
                biz_reg = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                erp_code = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                representative = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                address = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                biz_type = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                biz_item = str(row[7]).strip() if len(row) > 7 and row[7] else ''

                if not name:
                    continue

                existing = Vendor.objects.filter(code=code).first()
                if existing:
                    status = 'update'
                    update_count += 1
                else:
                    status = 'new'
                    new_count += 1

                preview_data.append({
                    'code': code,
                    'name': name,
                    'biz_registration_number': biz_reg,
                    'erp_code': erp_code,
                    'representative': representative,
                    'address': address,
                    'biz_type': biz_type,
                    'biz_item': biz_item,
                    'status': status,
                })
        else:
            # ERP 양식: 2행이 1업체 (헤더 5행 스킵)
            i = 5
            while i < len(rows) - 1:
                row1 = rows[i]
                row2 = rows[i + 1] if i + 1 < len(rows) else [''] * 9

                # 코드가 숫자가 아니면 스킵
                if len(row1) < 2 or not row1[0] or not str(row1[0]).strip():
                    i += 1
                    continue

                code_val = str(row1[0]).strip()
                if not code_val.isdigit():
                    i += 1
                    continue

                code = code_val
                name = str(row1[1]).strip() if len(row1) > 1 else ''
                biz_reg = str(row1[3]).strip() if len(row1) > 3 else ''
                biz_type = str(row1[4]).strip() if len(row1) > 4 else ''

                address = str(row2[0]).strip() if len(row2) > 0 else ''
                representative = str(row2[3]).strip() if len(row2) > 3 else ''
                biz_item = str(row2[4]).strip() if len(row2) > 4 else ''

                if not name:
                    i += 2
                    continue

                # 기존 Vendor 존재 여부 확인
                existing = Vendor.objects.filter(code=code).first()
                if existing:
                    status = 'update'
                    update_count += 1
                else:
                    status = 'new'
                    new_count += 1

                preview_data.append({
                    'code': code,
                    'name': name,
                    'biz_registration_number': biz_reg,
                    'representative': representative,
                    'address': address,
                    'biz_type': biz_type,
                    'biz_item': biz_item,
                    'status': status,
                })

                i += 2

    except Exception as e:
        messages.error(request, f'파일 처리 중 오류: {e}')
        return redirect('vendor_upload')

    context = {
        'preview_data': preview_data,
        'new_count': new_count,
        'update_count': update_count,
        'error_count': error_count,
    }
    return render(request, 'vendor_upload.html', context)


@login_required
@menu_permission_required('can_access_scm_admin')
def vendor_upload_confirm(request):
    """협력사 업로드 최종 확정"""
    if request.method != 'POST':
        return redirect('vendor_upload')

    code_list = request.POST.getlist('code_list[]')
    name_list = request.POST.getlist('name_list[]')
    biz_reg_list = request.POST.getlist('biz_reg_list[]')
    erp_code_list = request.POST.getlist('erp_code_list[]')
    representative_list = request.POST.getlist('representative_list[]')
    address_list = request.POST.getlist('address_list[]')
    biz_type_list = request.POST.getlist('biz_type_list[]')
    biz_item_list = request.POST.getlist('biz_item_list[]')

    created = 0
    updated = 0

    for i in range(len(code_list)):
        code = code_list[i]
        name = name_list[i] if i < len(name_list) else ''
        biz_reg = biz_reg_list[i] if i < len(biz_reg_list) else ''
        erp_code = erp_code_list[i] if i < len(erp_code_list) else code
        rep = representative_list[i] if i < len(representative_list) else ''
        addr = address_list[i] if i < len(address_list) else ''
        biz_type = biz_type_list[i] if i < len(biz_type_list) else ''
        biz_item = biz_item_list[i] if i < len(biz_item_list) else ''

        try:
            vendor, was_created = Vendor.objects.update_or_create(
                code=code,
                defaults={
                    'name': name,
                    'erp_code': erp_code or code,
                    'biz_registration_number': biz_reg or None,
                    'representative': rep or None,
                    'address': addr or None,
                    'biz_type': biz_type or None,
                    'biz_item': biz_item or None,
                }
            )
            if was_created:
                created += 1
            else:
                updated += 1
        except Exception:
            pass

    messages.success(request, f'협력사 등록 완료: 신규 {created}개, 업데이트 {updated}개')
    return render(request, 'vendor_upload.html', {
        'result': {'created': created, 'updated': updated}
    })


# ============================================
# API 엔드포인트 (품번/협력사 검색)
# ============================================
@login_required
def api_part_search(request):
    """품번 검색 API - 품목마스터에서 검색"""
    q = request.GET.get('q', '').strip()
    exact = request.GET.get('exact', '0') == '1'

    if not exact and len(q) < 2:
        return JsonResponse({'results': []})

    # exact=1이면 품번 정확 일치, 아니면 부분 검색
    if exact:
        parts = Part.objects.filter(part_no=q).select_related('vendor')[:1]
    else:
        parts = Part.objects.filter(
            Q(part_no__icontains=q) | Q(part_name__icontains=q)
        ).select_related('vendor')[:50]

    results = []
    for p in parts:
        # Organization ID 찾기 (vendor와 연결된 organization)
        org_id = None
        org_name = ''
        if p.vendor:
            org = Organization.objects.filter(linked_vendor=p.vendor).first()
            if org:
                org_id = org.id
                org_name = org.name
            else:
                # Organization이 없으면 Vendor 이름 사용
                org_name = p.vendor.name

        results.append({
            'part_no': p.part_no,
            'part_name': p.part_name,
            'part_group': p.part_group,
            'vendor_id': org_id,  # Organization ID 반환
            'vendor_name': org_name,
        })

    return JsonResponse({'results': results})


@login_required
def api_vendor_search(request):
    """협력사 검색 API (Vendor 모델)"""
    q = request.GET.get('q', '').strip()
    if len(q) < 1:
        return JsonResponse({'results': []})

    vendors = Vendor.objects.filter(
        Q(name__icontains=q) | Q(code__icontains=q)
    )[:30]

    results = []
    for v in vendors:
        results.append({
            'id': v.id,
            'code': v.code,
            'name': v.name,
        })

    return JsonResponse({'results': results})


@login_required
def api_organization_search(request):
    """협력사(Organization) 검색 API - QMS용"""
    q = request.GET.get('q', '').strip()
    if len(q) < 1:
        return JsonResponse({'results': []})

    orgs = Organization.objects.filter(
        org_type='VENDOR',
        name__icontains=q
    )[:30]

    results = []
    for org in orgs:
        results.append({
            'id': org.id,
            'name': org.name,
        })

    return JsonResponse({'results': results})


@login_required
def api_employee_search(request):
    """직원 검색 API - 결재선 지정용"""
    q = request.GET.get('q', '').strip()
    if len(q) < 1:
        return JsonResponse({'results': []})

    # 내부 사용자 검색 (협력사 제외)
    from django.contrib.auth.models import User
    users = User.objects.filter(
        Q(is_superuser=True) |
        Q(is_staff=True) |
        Q(profile__role__in=['ADMIN', 'STAFF']) |
        Q(profile__is_jinyoung_staff=True) |
        Q(profile__account_type='INTERNAL')
    ).filter(
        Q(username__icontains=q) |
        Q(first_name__icontains=q) |
        Q(last_name__icontains=q) |
        Q(profile__display_name__icontains=q) |
        Q(profile__department__icontains=q)
    ).distinct().select_related('profile')[:30]

    results = []
    for u in users:
        profile = getattr(u, 'profile', None)
        dept = ''
        display_name = ''
        if profile:
            dept = getattr(profile, 'department', '') or ''
            display_name = getattr(profile, 'display_name', '') or ''
        if not display_name:
            display_name = u.get_full_name() or u.username

        label = f"{dept} {display_name}".strip() if dept else display_name

        results.append({
            'id': u.id,
            'username': u.username,
            'display_name': display_name,
            'department': dept,
            'label': label,
        })

    return JsonResponse({'results': results})


# =============================================================================
# ERP 발주 연동 → SCM 발주 등록
# =============================================================================

@login_required
@menu_permission_required('can_scm_order_edit')
def erp_po_sync(request):
    """
    ERP 발주 조회 → 선택 → SCM Order 등록
    - GET: 조회 폼
    - POST action=search: ERP API로 발주 조회
    - POST action=apply: 선택 항목을 SCM Order로 등록
    """
    from material.erp_api import fetch_erp_po_headers, fetch_erp_po_details
    from django.conf import settings as conf_settings

    erp_enabled = getattr(conf_settings, 'ERP_ENABLED', False)
    vendors_qs = Vendor.objects.exclude(
        erp_code__isnull=True
    ).exclude(erp_code='').order_by('name')
    today = timezone.now().date()

    # 거래처 표시명 조회
    sel_vendor_code = request.POST.get('vendor_erp_code', '')
    sel_vendor_display = ''
    if sel_vendor_code:
        v_obj = vendors_qs.filter(erp_code=sel_vendor_code).first()
        if v_obj:
            sel_vendor_display = f'[{v_obj.erp_code}] {v_obj.name}'

    context = {
        'erp_enabled': erp_enabled,
        'vendors': vendors_qs,
        'date_from': request.POST.get('date_from', (today - timedelta(days=30)).strftime('%Y-%m-%d')),
        'date_to': request.POST.get('date_to', today.strftime('%Y-%m-%d')),
        'selected_vendor': sel_vendor_code,
        'selected_vendor_display': sel_vendor_display,
        'item_search': request.POST.get('item_search', ''),
        'po_items': [],
        'summary': None,
    }

    if request.method != 'POST':
        return render(request, 'erp_po_sync.html', context)

    action = request.POST.get('action', '')

    # ── 조회 ──
    if action == 'search':
        date_from = request.POST.get('date_from', '')
        date_to = request.POST.get('date_to', '')
        vendor_erp_code = request.POST.get('vendor_erp_code', '')
        item_search = (request.POST.get('item_search', '') or '').strip().upper()

        if not date_from or not date_to:
            messages.warning(request, '발주일 시작/종료를 입력해주세요.')
            return render(request, 'erp_po_sync.html', context)

        # 날짜 포맷 변환 (YYYY-MM-DD → YYYYMMDD)
        dt_from = date_from.replace('-', '')
        dt_to = date_to.replace('-', '')

        context['date_from'] = date_from
        context['date_to'] = date_to
        context['selected_vendor'] = vendor_erp_code
        context['item_search'] = item_search

        # 1) 헤더 조회
        headers = fetch_erp_po_headers(dt_from, dt_to, tr_cd=vendor_erp_code or None)
        if not headers:
            messages.info(request, '해당 기간에 ERP 발주 내역이 없습니다.')
            return render(request, 'erp_po_sync.html', context)

        # 2) 이미 SCM에 등록된 건 조회
        existing = set(
            Order.objects.filter(erp_order_no__isnull=False)
            .exclude(erp_order_no='')
            .values_list('erp_order_no', 'erp_order_seq')
        )

        # 3) SCM Part 목록 (매칭용)
        part_map = {p.part_no: p for p in Part.objects.select_related('vendor').all()}

        # 4) 각 헤더의 디테일 조회 → 품목 리스트 구성
        po_items = []
        for header in headers:
            po_nb = header.get('poNb', '')
            po_dt = header.get('poDt', '')
            tr_cd_val = header.get('trCd', '')
            vendor_name = header.get('attrNm', '')

            details = fetch_erp_po_details(po_nb)
            for detail in details:
                item_cd = detail.get('itemCd', '')
                po_sq = str(detail.get('poSq', ''))
                po_qt = int(detail.get('poQt', 0) or 0)
                rcv_qt = int(detail.get('rcvQt', 0) or 0)
                due_dt = detail.get('dueDt', '')
                item_nm = detail.get('itemNm', '')

                # 품번/품명 필터
                if item_search:
                    if item_search not in item_cd.upper() and item_search not in (item_nm or '').upper():
                        continue

                # 이미 적용 여부
                is_applied = (po_nb, po_sq) in existing

                # Part 매칭
                part_obj = part_map.get(item_cd)
                part_matched = part_obj is not None

                # 날짜 포맷
                po_display = ''
                if po_dt and len(po_dt) == 8:
                    po_display = f'{po_dt[:4]}-{po_dt[4:6]}-{po_dt[6:8]}'
                due_display = ''
                if due_dt and len(due_dt) == 8:
                    due_display = f'{due_dt[:4]}-{due_dt[4:6]}-{due_dt[6:8]}'

                po_items.append({
                    'po_nb': po_nb,
                    'po_dt': po_display,
                    'po_sq': po_sq,
                    'tr_cd': tr_cd_val,
                    'vendor_name': vendor_name,
                    'item_cd': item_cd,
                    'item_nm': item_nm,
                    'po_qt': po_qt,
                    'rcv_qt': rcv_qt,
                    'remain_qt': po_qt - rcv_qt,
                    'due_dt': due_display,
                    'due_dt_raw': due_dt,
                    'is_applied': is_applied,
                    'part_matched': part_matched,
                })

        context['po_items'] = po_items
        context['summary'] = {
            'total': len(po_items),
            'applied': sum(1 for i in po_items if i['is_applied']),
            'new': sum(1 for i in po_items if not i['is_applied']),
            'unmatched': sum(1 for i in po_items if not i['part_matched']),
        }
        return render(request, 'erp_po_sync.html', context)

    # ── SCM 적용 ──
    elif action == 'apply':
        selected = request.POST.getlist('selected_items')
        if not selected:
            messages.warning(request, '적용할 항목을 선택해주세요.')
            return redirect('erp_po_sync')

        # 이미 등록된 건
        existing = set(
            Order.objects.filter(erp_order_no__isnull=False)
            .exclude(erp_order_no='')
            .values_list('erp_order_no', 'erp_order_seq')
        )

        part_map = {p.part_no: p for p in Part.objects.select_related('vendor').all()}
        vendor_map = {v.erp_code: v for v in Vendor.objects.exclude(erp_code__isnull=True).exclude(erp_code='')}

        created_count = 0
        skipped_count = 0
        error_list = []

        for item_data in selected:
            # 포맷: "poNb|poSq|itemCd|poQt|dueDt|trCd"
            try:
                fields = item_data.split('|')
                po_nb = fields[0]
                po_sq = fields[1]
                item_cd = fields[2]
                po_qt = int(fields[3])
                due_dt_raw = fields[4]
                tr_cd_val = fields[5] if len(fields) > 5 else ''
            except (IndexError, ValueError):
                error_list.append(f'데이터 파싱 오류: {item_data}')
                continue

            # 중복 체크
            if (po_nb, po_sq) in existing:
                skipped_count += 1
                continue

            # Part 매칭
            part_obj = part_map.get(item_cd)
            if not part_obj:
                error_list.append(f'품번 미매칭: {item_cd}')
                continue

            # Vendor: Part에 할당된 vendor 우선, 없으면 ERP 거래처코드로 매칭
            vendor = part_obj.vendor
            if not vendor and tr_cd_val:
                vendor = vendor_map.get(tr_cd_val)
            if not vendor:
                error_list.append(f'거래처 미매칭: {item_cd} (trCd={tr_cd_val})')
                continue

            # 납기일 파싱
            due_date_val = None
            if due_dt_raw and len(due_dt_raw) == 8:
                try:
                    due_date_val = date(int(due_dt_raw[:4]), int(due_dt_raw[4:6]), int(due_dt_raw[6:8]))
                except ValueError:
                    due_date_val = today
            if not due_date_val:
                due_date_val = today

            # Order 생성 (협력사 승인 대기)
            Order.objects.create(
                vendor=vendor,
                part_group=part_obj.part_group or '',
                part_no=part_obj.part_no,
                part_name=part_obj.part_name or '',
                quantity=po_qt,
                due_date=due_date_val,
                erp_order_no=po_nb,
                erp_order_seq=po_sq,
            )
            existing.add((po_nb, po_sq))
            created_count += 1

        # 결과 메시지
        msg = f'SCM 발주 {created_count}건 등록 완료 (협력사 승인 대기)'
        if skipped_count:
            msg += f', 중복 건너뜀 {skipped_count}건'
        if error_list:
            msg += f', 오류 {len(error_list)}건'
            for err in error_list[:5]:
                messages.warning(request, err)

        messages.success(request, msg)
        return redirect('erp_po_sync')

    return render(request, 'erp_po_sync.html', context)
