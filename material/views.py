# material/views.py
import logging
import re

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator  # 페이징 처리
from django.db import transaction
from django.db.models import F, Sum, Q, Value
from django.db.models.functions import Greatest
from django.utils import timezone
from functools import wraps

logger = logging.getLogger(__name__)

# SCM(Orders) 앱 모델
from orders.models import Part, Vendor, Inventory as OldInventory, Demand, InventoryUploadLog


# =============================================================================
# WMS 권한 체크 데코레이터
# =============================================================================

def _get_profile(user):
    """UserProfile을 안전하게 가져온다"""
    try:
        return getattr(user, 'profile', None)
    except Exception:
        return None

def wms_permission_required(permission_field):
    """
    WMS 메뉴 권한 체크 데코레이터
    - superuser는 모든 권한
    - 일반 사용자는 UserProfile의 해당 필드가 True여야 접근 가능
    """
    def decorator(view_func):
        @wraps(view_func)
        @login_required
        def _wrapped_view(request, *args, **kwargs):
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)

            profile = _get_profile(request.user)
            if profile:
                # 새 권한 필드 체크
                if getattr(profile, permission_field, False):
                    return view_func(request, *args, **kwargs)
                # 레거시 필드 폴백 (하위 호환)
                legacy_map = {
                    'can_wms_stock_view': ['can_wms_inout', 'can_wms_adjustment'],
                    'can_wms_stock_edit': ['can_wms_adjustment'],
                    'can_wms_inout_view': ['can_wms_inout'],
                    'can_wms_inout_edit': ['can_wms_inout'],
                    'can_wms_bom_view': ['can_wms_bom'],
                    'can_wms_bom_edit': ['can_wms_bom'],
                    'can_wms_label_view': ['can_wms_inout_view', 'can_wms_inout'],
                    'can_wms_label_edit': ['can_wms_inout_edit', 'can_wms_inout'],
                    'can_wms_field_view': ['can_wms_stock_view', 'can_wms_inout'],
                    'can_wms_field_edit': ['can_wms_stock_edit', 'can_wms_adjustment'],
                }
                for legacy_field in legacy_map.get(permission_field, []):
                    if getattr(profile, legacy_field, False):
                        return view_func(request, *args, **kwargs)

            messages.error(request, "해당 메뉴에 대한 접근 권한이 없습니다.")
            return redirect('material:dashboard')
        return _wrapped_view
    return decorator

# WMS(Material) 앱 모델
from .models import Warehouse, MaterialStock, MaterialTransaction

from django.http import JsonResponse, HttpResponse
import openpyxl

# [신규] QMS 앱 모델 (수입검사 연동용)
try:
    from qms.models import ImportInspection
except ImportError:
    ImportInspection = None

# =============================================================================
# 마감일 검증 헬퍼 함수
# =============================================================================

def check_closing_date(target_date):
    """
    수불 날짜가 마감 기간에 속하는지 확인
    Returns: (is_closed, warning_message, closing_month)
    - is_closed: True면 마감된 기간
    - warning_message: 경고 메시지
    - closing_month: 마감월 (표시용)
    """
    from .models import InventoryClosing
    from datetime import date
    from calendar import monthrange

    latest = InventoryClosing.get_latest_closing()
    if not latest:
        return False, None, None

    # 마감월의 마지막 날 계산
    closing_year = latest.closing_month.year
    closing_month = latest.closing_month.month
    _, last_day = monthrange(closing_year, closing_month)
    closing_end_date = date(closing_year, closing_month, last_day)

    # target_date를 date 객체로 변환
    if hasattr(target_date, 'date'):
        check_date = target_date.date()
    else:
        check_date = target_date

    if check_date <= closing_end_date:
        month_str = latest.closing_month.strftime('%Y년 %m월')
        warning_msg = f"[마감 경고] {month_str}은 이미 마감되었습니다. 마감 기간({closing_end_date.strftime('%Y-%m-%d')} 이전) 날짜로 수불을 입력하면 재고 정합성에 영향을 줄 수 있습니다."
        return True, warning_msg, month_str

    return False, None, None


# =============================================================================
# 1. 대시보드 및 재고 조회
# =============================================================================

@wms_permission_required('can_wms_stock_view')
def dashboard(request):
    """자재 관리 대시보드 - 종합 현황판"""
    from datetime import timedelta
    from django.db.models import Count, Min
    from .models import Product, BOMItem, ProcessTag, InventoryClosing

    today = timezone.now().date()
    this_month_start = today.replace(day=1)

    # ========== 1. 재고 현황 통계 ==========
    # 총 재고 품목 수 (재고가 있는 품목)
    total_parts = MaterialStock.objects.filter(quantity__gt=0).values('part').distinct().count()

    # 총 재고 수량
    total_qty = MaterialStock.objects.filter(quantity__gt=0).aggregate(total=Sum('quantity'))['total'] or 0

    # 창고 수
    warehouse_count = Warehouse.objects.filter(is_active=True).count()

    # ========== 2. 입출고 통계 ==========
    # 금일 입고
    today_in = MaterialTransaction.objects.filter(
        date__date=today,
        transaction_type__in=['IN_SCM', 'IN_MANUAL', 'IN_ERP', 'RCV_ERP']
    ).aggregate(
        count=Count('id'),
        qty=Sum('quantity')
    )

    # 금일 출고
    today_out = MaterialTransaction.objects.filter(
        date__date=today,
        transaction_type__in=['OUT_PROD', 'OUT_RETURN', 'OUT_MANUAL', 'OUT_ERP', 'ISU_ERP']
    ).aggregate(
        count=Count('id'),
        qty=Sum('quantity')
    )

    # 이번달 입고
    month_in = MaterialTransaction.objects.filter(
        date__date__gte=this_month_start,
        transaction_type__in=['IN_SCM', 'IN_MANUAL', 'IN_ERP', 'RCV_ERP']
    ).aggregate(
        count=Count('id'),
        qty=Sum('quantity')
    )

    # 이번달 출고
    month_out = MaterialTransaction.objects.filter(
        date__date__gte=this_month_start,
        transaction_type__in=['OUT_PROD', 'OUT_RETURN', 'OUT_MANUAL', 'OUT_ERP', 'ISU_ERP']
    ).aggregate(
        count=Count('id'),
        qty=Sum('quantity')
    )

    # ========== 3. (삭제됨 - 창고별 재고 현황) ==========

    # ========== 4. (삭제됨 - 품목군별 재고 현황) ==========

    # ========== 5. FIFO 경고 (30일 이상 된 LOT) ==========
    fifo_warning_date = today - timedelta(days=30)
    fifo_warnings = MaterialStock.objects.filter(
        quantity__gt=0,
        lot_no__isnull=False,
        lot_no__lt=fifo_warning_date
    ).select_related('warehouse', 'part').order_by('lot_no')[:10]

    fifo_warning_count = MaterialStock.objects.filter(
        quantity__gt=0,
        lot_no__isnull=False,
        lot_no__lt=fifo_warning_date
    ).count()

    # ========== 6. 금일 입고/출고 이력 (이동·보정 제외, 하루치 누적) ==========
    recent_inbound = MaterialTransaction.objects.select_related(
        'part', 'warehouse_from', 'warehouse_to', 'vendor', 'actor'
    ).filter(
        date__date=today,
        transaction_type__in=['IN_SCM', 'IN_MANUAL', 'IN_ERP', 'RCV_ERP']
    ).order_by('-date')

    recent_outbound = list(MaterialTransaction.objects.select_related(
        'part', 'part__vendor', 'warehouse_from', 'warehouse_to', 'vendor', 'actor'
    ).filter(
        date__date=today,
        transaction_type__in=['OUT_PROD', 'OUT_RETURN', 'OUT_MANUAL', 'OUT_ERP', 'ISU_ERP']
    ).order_by('-date'))

    # 거래처 표시: vendor → part.vendor → remark에서 추출
    import re as _re
    for tx in recent_outbound:
        if tx.vendor:
            tx.display_vendor = tx.vendor.name
        elif tx.part and tx.part.vendor:
            tx.display_vendor = tx.part.vendor.name
        elif tx.remark and '\u2192' in tx.remark:
            # remark 형식: "ERP출고(창고→거래처명) ..." → '→' 뒤에서 마지막 ')' 직전까지
            after_arrow = tx.remark.split('\u2192', 1)[1]
            # 마지막 ')' 이전까지 추출하고 뒤의 숫자/공백 제거
            vendor_str = _re.sub(r'\)\s*[\d\s]*$', '', after_arrow).strip()
            tx.display_vendor = vendor_str
        else:
            tx.display_vendor = ''

    # ========== 7. BOM 현황 ==========
    bom_stats = {
        'product_count': Product.objects.filter(is_active=True).count(),
        'bom_item_count': BOMItem.objects.filter(is_active=True).count(),
    }

    # ========== 8. 공정 현품표 현황 ==========
    tag_stats = {
        'today_printed': ProcessTag.objects.filter(printed_at__date=today).count(),
        'pending': ProcessTag.objects.filter(status='PRINTED').count(),
        'used_today': ProcessTag.objects.filter(used_at__date=today).count(),
    }

    # ========== 9. 마감 현황 ==========
    latest_closing = InventoryClosing.get_latest_closing()
    closing_info = None
    if latest_closing:
        closing_info = {
            'month': latest_closing.closing_month.strftime('%Y년 %m월'),
            'closed_at': latest_closing.closed_at,
            'closed_by': latest_closing.closed_by.username if latest_closing.closed_by else '-',
        }

    # ========== 10. 재고 부족 경고 (재고 0인 품목 중 최근 출고 이력 있는 것) ==========
    # 최근 7일 내 출고된 품목 중 현재 재고가 0인 것
    recent_out_parts = MaterialTransaction.objects.filter(
        date__date__gte=today - timedelta(days=7),
        transaction_type__in=['OUT_PROD', 'OUT_RETURN']
    ).values_list('part_id', flat=True).distinct()

    zero_stock_parts = []
    for part_id in recent_out_parts:
        total = MaterialStock.objects.filter(part_id=part_id).aggregate(total=Sum('quantity'))['total'] or 0
        if total <= 0:
            part = Part.objects.filter(id=part_id).first()
            if part:
                zero_stock_parts.append(part)
    zero_stock_count = len(zero_stock_parts)

    context = {
        # 재고 현황
        'total_parts': total_parts,
        'total_qty': total_qty,
        'warehouse_count': warehouse_count,
        'zero_stock_count': zero_stock_count,

        # 입출고 통계
        'today_in_count': today_in['count'] or 0,
        'today_in_qty': abs(today_in['qty'] or 0),
        'today_out_count': today_out['count'] or 0,
        'today_out_qty': abs(today_out['qty'] or 0),
        'month_in_count': month_in['count'] or 0,
        'month_in_qty': abs(month_in['qty'] or 0),
        'month_out_count': month_out['count'] or 0,
        'month_out_qty': abs(month_out['qty'] or 0),

        # FIFO 경고
        'fifo_warning_count': fifo_warning_count,
        'fifo_warnings': fifo_warnings,

        # 최근 입고/출고 이력
        'recent_inbound': recent_inbound,
        'recent_outbound': recent_outbound,

        # BOM 현황
        'bom_stats': bom_stats,

        # 현품표 현황
        'tag_stats': tag_stats,

        # 마감 현황
        'closing_info': closing_info,

        # 기타
        'today': today,
    }

    return render(request, 'material/dashboard.html', context)


@wms_permission_required('can_wms_stock_view')
def dashboard_api(request):
    """대시보드 AJAX 폴링용 JSON API"""
    from django.http import JsonResponse
    from datetime import timedelta
    from django.db.models import Count

    today = timezone.now().date()
    this_month_start = today.replace(day=1)

    # KPI
    total_parts = MaterialStock.objects.filter(quantity__gt=0).values('part').distinct().count()
    total_qty = MaterialStock.objects.filter(quantity__gt=0).aggregate(total=Sum('quantity'))['total'] or 0

    today_in = MaterialTransaction.objects.filter(
        date__date=today, transaction_type__in=['IN_SCM', 'IN_MANUAL', 'IN_ERP', 'RCV_ERP']
    ).aggregate(count=Count('id'), qty=Sum('quantity'))
    today_out = MaterialTransaction.objects.filter(
        date__date=today, transaction_type__in=['OUT_PROD', 'OUT_RETURN', 'OUT_MANUAL', 'OUT_ERP', 'ISU_ERP']
    ).aggregate(count=Count('id'), qty=Sum('quantity'))
    month_in = MaterialTransaction.objects.filter(
        date__date__gte=this_month_start, transaction_type__in=['IN_SCM', 'IN_MANUAL', 'IN_ERP', 'RCV_ERP']
    ).aggregate(count=Count('id'), qty=Sum('quantity'))
    month_out = MaterialTransaction.objects.filter(
        date__date__gte=this_month_start, transaction_type__in=['OUT_PROD', 'OUT_RETURN', 'OUT_MANUAL', 'OUT_ERP', 'ISU_ERP']
    ).aggregate(count=Count('id'), qty=Sum('quantity'))

    fifo_warning_date = today - timedelta(days=30)
    fifo_warning_count = MaterialStock.objects.filter(
        quantity__gt=0, lot_no__isnull=False, lot_no__lt=fifo_warning_date
    ).count()

    # 입고 이력 (금일 전체)
    inbound = list(MaterialTransaction.objects.filter(
        date__date=today,
        transaction_type__in=['IN_SCM', 'IN_MANUAL', 'IN_ERP', 'RCV_ERP']
    ).order_by('-date').values(
        'date', 'transaction_type', 'part__part_no', 'quantity',
        'warehouse_from__name', 'warehouse_to__name', 'vendor__name'
    ))
    # 출고 이력 (금일 전체)
    outbound = list(MaterialTransaction.objects.filter(
        date__date=today,
        transaction_type__in=['OUT_PROD', 'OUT_RETURN', 'OUT_MANUAL', 'OUT_ERP', 'ISU_ERP']
    ).order_by('-date').values(
        'date', 'transaction_type', 'part__part_no', 'quantity',
        'warehouse_from__name', 'warehouse_to__name', 'vendor__name',
        'part__vendor__name', 'remark'
    ))

    TX_DISPLAY = dict(MaterialTransaction.TYPE_CHOICES)
    import re as _re_api

    def _extract_vendor_from_remark(remark):
        if remark and '\u2192' in remark:
            after = remark.split('\u2192', 1)[1]
            return _re_api.sub(r'\)\s*[\d\s]*$', '', after).strip()
        return ''

    def fmt_tx(rows):
        result = []
        for r in rows:
            vendor_name = r.get('vendor__name') or r.get('part__vendor__name') or _extract_vendor_from_remark(r.get('remark', ''))
            result.append({
                'date': r['date'].strftime('%m/%d %H:%M') if r['date'] else '',
                'type': TX_DISPLAY.get(r['transaction_type'], r['transaction_type']),
                'part_no': r['part__part_no'] or '',
                'wh_from': r['warehouse_from__name'] or '',
                'wh_to': r['warehouse_to__name'] or '',
                'vendor': vendor_name,
                'qty': r['quantity'],
            })
        return result

    return JsonResponse({
        'total_parts': total_parts,
        'total_qty': total_qty,
        'today_in_count': today_in['count'] or 0,
        'today_in_qty': abs(today_in['qty'] or 0),
        'today_out_count': today_out['count'] or 0,
        'today_out_qty': abs(today_out['qty'] or 0),
        'month_in_count': month_in['count'] or 0,
        'month_in_qty': abs(month_in['qty'] or 0),
        'month_out_count': month_out['count'] or 0,
        'month_out_qty': abs(month_out['qty'] or 0),
        'fifo_warning_count': fifo_warning_count,
        'inbound': fmt_tx(inbound),
        'outbound': fmt_tx(outbound),
    })


@wms_permission_required('can_wms_stock_view')
def stock_list(request):
    """
    [WMS] 자재 재고 현황 조회
    """
    # 1. 파라미터 수신
    search_triggered = request.GET.get('search_triggered', '')
    q = request.GET.get('q', '')
    part_group = request.GET.get('part_group', '')
    warehouse_id = request.GET.get('warehouse_id', '')
    view_mode = request.GET.get('view_mode', 'detail')

    stock_data = []

    # 2. 검색 버튼이 눌렸을 때만 DB 조회 실행
    if search_triggered == 'yes':
        # 수량이 0이 아닌 것만 가져옴 (마이너스 재고 포함)
        stocks = MaterialStock.objects.select_related('warehouse', 'part', 'part__vendor').exclude(quantity=0)

        # 검색 및 필터 적용
        if q:
            stocks = stocks.filter(Q(part__part_no__icontains=q) | Q(part__part_name__icontains=q))

        if part_group:
            stocks = stocks.filter(part__part_group=part_group)

        if warehouse_id:
            stocks = stocks.filter(warehouse_id=warehouse_id)

        # 3. 조회 모드에 따른 데이터 가공
        if view_mode == 'summary':
            # 전체 합계 모드: 품목별 전체 재고 합계
            stock_data = stocks.values(
                'part__part_no',
                'part__part_name',
                'part__part_group',
                'part__vendor__name'
            ).annotate(total_qty=Sum('quantity')).order_by('part__part_no')
        else:
            # 창고별 상세 모드: 창고별 품목별 합계 (LOT 합산)
            stock_data = stocks.values(
                'warehouse__id',
                'warehouse__code',
                'warehouse__name',
                'part__id',
                'part__part_no',
                'part__part_name',
                'part__part_group',
                'part__vendor__name'
            ).annotate(total_qty=Sum('quantity')).order_by('warehouse__code', 'part__part_no')

    # 4. 필터용 데이터
    part_groups = Part.objects.values_list('part_group', flat=True).distinct().order_by('part_group')
    warehouses = Warehouse.objects.all().order_by('code')

    context = {
        'stocks': stock_data,
        'warehouses': warehouses,
        'part_groups': part_groups,
        'q': q,
        'selected_group': part_group,
        'selected_wh': warehouse_id,
        'view_mode': view_mode,
        'search_triggered': search_triggered,
    }
    return render(request, 'material/stock_list.html', context)


# =============================================================================
# 2. 입고 관리 (Inbound)
# =============================================================================

@wms_permission_required('can_wms_inout_edit')
def manual_incoming(request):
    """
    [WMS] 자재 수기 입고 처리 (LOT 포함)
    """
    if request.method == 'POST':
        try:
            date_str = request.POST.get('date', timezone.now().date())
            warehouse_id = request.POST.get('warehouse_id')
            vendor_id = request.POST.get('vendor_id')
            needs_inspection = request.POST.get('needs_inspection')

            target_warehouse_id = request.POST.get('target_warehouse_id')

            part_ids = request.POST.getlist('part_ids[]')
            lot_nos = request.POST.getlist('lot_nos[]')
            quantities = request.POST.getlist('quantities[]')
            remarks = request.POST.getlist('remarks[]')
            erp_order_nos = request.POST.getlist('erp_order_nos[]')  # ERP 발주번호
            erp_order_seqs = request.POST.getlist('erp_order_seqs[]')  # ERP 발주순번

            if not part_ids:
                messages.error(request, "입고할 품목이 리스트에 없습니다.")
                return redirect('material:manual_incoming')

            # 마감 기간 검증 (경고만 표시, 진행은 허용)
            from datetime import datetime
            if isinstance(date_str, str):
                check_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                check_date = date_str
            is_closed, warning_msg, _ = check_closing_date(check_date)
            if is_closed:
                messages.warning(request, warning_msg)

            # 입고 일자를 DateTimeField에 맞게 datetime 객체로 변환
            if isinstance(date_str, str):
                date_value = timezone.make_aware(datetime.strptime(date_str, '%Y-%m-%d'))
            else:
                date_value = timezone.now()

            success_count = 0

            with transaction.atomic():
                warehouse = Warehouse.objects.get(id=warehouse_id)
                vendor = Vendor.objects.get(id=vendor_id) if vendor_id else None

                for i in range(len(part_ids)):
                    p_id = part_ids[i]
                    qty = int(quantities[i])
                    rmk = remarks[i] if i < len(remarks) else ''
                    lot_no_str = lot_nos[i] if i < len(lot_nos) and lot_nos[i] else None

                    if qty <= 0:
                        continue

                    part = Part.objects.get(id=p_id)
                    system_remark = "[수입검사 대상] " if needs_inspection else ""
                    final_remark = f"{system_remark}{rmk}".strip()

                    # LOT 번호 처리 (날짜 형식)
                    from datetime import datetime
                    lot_date = None
                    if lot_no_str:
                        try:
                            lot_date = datetime.strptime(lot_no_str, '%Y-%m-%d').date()
                        except (ValueError, TypeError):
                            pass

                    # (1) 재고 증가 (LOT별로 분리)
                    stock, _ = MaterialStock.objects.get_or_create(
                        warehouse=warehouse,
                        part=part,
                        lot_no=lot_date,
                        defaults={'quantity': 0}
                    )
                    MaterialStock.objects.filter(pk=stock.pk).update(
                        quantity=F('quantity') + qty
                    )
                    stock.refresh_from_db()

                    # (2) 발주 연결 확인 → IN_SCM / IN_MANUAL 분기
                    erp_no = erp_order_nos[i] if i < len(erp_order_nos) else ''
                    erp_seq = erp_order_seqs[i] if i < len(erp_order_seqs) else ''

                    if erp_no:
                        trx_type = 'IN_SCM'
                        trx_prefix = 'IN-SCM'
                        order_info = f"[발주입고] ERP:{erp_no}-{erp_seq}"
                        final_remark = f"{order_info} {final_remark}".strip()
                    else:
                        trx_type = 'IN_MANUAL'
                        trx_prefix = 'IN-MAN'

                    trx_no = f"{trx_prefix}-{timezone.now().strftime('%y%m%d%H%M%S')}-{request.user.id}-{i}"
                    trx_obj = MaterialTransaction.objects.create(
                        transaction_no=trx_no,
                        transaction_type=trx_type,
                        date=date_value,
                        part=part,
                        quantity=qty,
                        lot_no=lot_date,
                        warehouse_to=warehouse,
                        result_stock=stock.quantity,
                        vendor=vendor,
                        actor=request.user,
                        remark=final_remark
                    )

                    # (3) 수입검사 요청
                    if needs_inspection and ImportInspection:
                        # 합격 후 입고 창고 코드 조회
                        target_wh_code = '2000'
                        if target_warehouse_id:
                            try:
                                target_wh = Warehouse.objects.get(id=target_warehouse_id)
                                target_wh_code = target_wh.code
                            except Warehouse.DoesNotExist:
                                pass
                        ImportInspection.objects.create(
                            inbound_transaction=trx_obj,
                            lot_no=lot_date,
                            target_warehouse_code=target_wh_code,
                            status='PENDING'
                        )
                    elif ImportInspection:
                        # 무검사 입고도 라벨 발행 대기 목록에 표시되도록
                        ImportInspection.objects.create(
                            inbound_transaction=trx_obj,
                            lot_no=lot_date,
                            target_warehouse_code=warehouse.code,
                            status='APPROVED',
                            inspected_at=timezone.now(),
                            qty_good=qty,
                            remark='무검사 수기입고 (수입검사 생략)',
                        )

                        # ERP 입고등록 (무검사 - 전체수량)
                        try:
                            from material.erp_api import register_erp_incoming
                            erp_ok, erp_rcv_no, erp_err = register_erp_incoming(
                                trx_obj, qty, warehouse.code,
                                erp_order_no=erp_no, erp_order_seq=erp_seq
                            )
                            if erp_ok:
                                messages.info(request, f'ERP 입고등록 완료: {erp_rcv_no}')
                            elif erp_err:
                                messages.warning(request, f'ERP 연동 실패: {erp_err}')
                            else:
                                messages.info(request, 'ERP 연동 건너뜀 (거래처 ERP코드 없음)')
                        except Exception as e:
                            logger.error(f'ERP 입고등록 예외: {e}')
                            messages.warning(request, f'ERP 연동 오류: {e}')

                    success_count += 1

            if success_count > 0:
                messages.success(request, f"총 {success_count}건 입고 처리가 완료되었습니다.")
            else:
                messages.warning(request, "저장된 항목이 없습니다.")

            return redirect('material:manual_incoming')

        except Exception as e:
            messages.error(request, f"오류 발생: {str(e)}")
            return redirect('material:manual_incoming')

    # === 입고처리 내역 (History) ===
    history_qs = MaterialTransaction.objects.filter(
        transaction_type__in=['IN_MANUAL', 'IN_SCM', 'IN_ERP']
    ).select_related(
        'part', 'warehouse_to', 'vendor', 'actor'
    ).order_by('-date', '-id')

    history_q = (request.GET.get('hq') or '').strip()
    if history_q:
        history_qs = history_qs.filter(
            Q(part__part_no__icontains=history_q) |
            Q(part__part_name__icontains=history_q) |
            Q(transaction_no__icontains=history_q)
        )

    history_wh = (request.GET.get('hwh') or '').strip()
    if history_wh:
        history_qs = history_qs.filter(warehouse_to_id=history_wh)

    history_start = (request.GET.get('hstart') or '').strip()
    history_end = (request.GET.get('hend') or '').strip()
    if history_start and history_start not in ('None', 'null'):
        history_qs = history_qs.filter(date__date__gte=history_start)
    if history_end and history_end not in ('None', 'null'):
        history_qs = history_qs.filter(date__date__lte=history_end)

    history_paginator = Paginator(history_qs, 15)
    history_page = history_paginator.get_page(request.GET.get('hpage'))

    from .models import RawMaterialLabel
    for item in history_page:
        item.label_count = RawMaterialLabel.objects.filter(
            incoming_transaction=item, label_type='PACKAGE'
        ).exclude(status='CANCELLED').count()
        try:
            item.inspection_status = item.inspection.status
        except Exception:
            item.inspection_status = None
        item.can_cancel = (item.label_count == 0)
        item.can_edit = (
            item.transaction_type in ('IN_MANUAL', 'IN_SCM')
            and item.label_count == 0
            and item.inspection_status in (None, 'PENDING')
        )

    warehouses_qs = Warehouse.objects.filter(is_active=True).order_by('code')

    # ERP 자동동기화 결과
    from django.core.cache import cache
    from django.conf import settings as django_settings
    erp_sync_result = cache.get('erp_incoming_sync_result')
    erp_auto_sync_enabled = getattr(django_settings, 'ERP_AUTO_SYNC_ENABLED', False)
    erp_auto_sync_interval = getattr(django_settings, 'ERP_AUTO_SYNC_INTERVAL_MINUTES', 10)

    context = {
        'warehouses': warehouses_qs,
        'today': timezone.now().date(),
        'history_page': history_page,
        'history_q': history_q,
        'history_wh': history_wh,
        'history_start': history_start,
        'history_end': history_end,
        'erp_sync_result': erp_sync_result,
        'erp_auto_sync_enabled': erp_auto_sync_enabled,
        'erp_auto_sync_interval': erp_auto_sync_interval,
    }
    return render(request, 'material/manual_incoming.html', context)


@wms_permission_required('can_wms_inout_edit')
def cancel_manual_incoming(request, trx_id):
    """[WMS] 수기 입고 삭제 / 입고만 취소 (수입검사 판정 초기화)"""
    # 리다이렉트 대상 결정 (입고 내역 조회에서 호출 시)
    redirect_to = request.POST.get('redirect_to', 'material:manual_incoming')
    if redirect_to not in ('material:manual_incoming', 'material:incoming_history'):
        redirect_to = 'material:manual_incoming'

    if request.method != 'POST':
        return redirect(redirect_to)

    trx = get_object_or_404(MaterialTransaction, pk=trx_id, transaction_type__in=['IN_MANUAL', 'IN_SCM', 'IN_ERP'])

    if trx.transaction_type == 'IN_ERP':
        messages.error(request, "ERP에서 동기화된 입고 건은 WMS에서 삭제할 수 없습니다. ERP(아마란스)에서 삭제해주세요.")
        return redirect(redirect_to)

    from .models import RawMaterialLabel
    label_count = RawMaterialLabel.objects.filter(incoming_transaction=trx).exclude(status='CANCELLED').count()
    if label_count > 0:
        messages.error(request, f"라벨이 {label_count}장 발행된 입고 건은 삭제할 수 없습니다. 라벨을 먼저 취소하세요.")
        return redirect(redirect_to)

    is_closed, warning_msg, _ = check_closing_date(
        trx.date.date() if hasattr(trx.date, 'date') and callable(trx.date.date) else trx.date
    )
    if is_closed:
        messages.error(request, f"마감된 기간의 입고 건은 삭제할 수 없습니다. ({warning_msg})")
        return redirect(redirect_to)

    cancel_action = request.POST.get('cancel_action', 'delete_all')
    trx_no = trx.transaction_no
    trx_qty = trx.quantity

    # ── 입고만 취소 (수입검사 판정 초기화) ──
    if cancel_action == 'cancel_incoming_only':
        inspection = None
        if ImportInspection:
            try:
                inspection = trx.inspection
            except Exception:
                pass

        if not inspection:
            messages.error(request, "수입검사 데이터가 없습니다. '전체 삭제'를 사용하세요.")
            return redirect(redirect_to)

        try:
            with transaction.atomic():
                part = trx.part
                lot_no = trx.lot_no

                if inspection.status == 'APPROVED' or inspection.status == 'REJECTED':
                    # 판정 완료 상태 → 목표 창고/부적합 창고 재고 원복 + 검사대기 복원

                    # ERP 입고 삭제 (이동 트랜잭션 삭제 전에 처리)
                    from material.erp_api import delete_erp_incoming as del_erp_cancel
                    for erp_trx in MaterialTransaction.objects.filter(
                        transaction_type='TRANSFER',
                        part=part, lot_no=lot_no,
                        warehouse_from=trx.warehouse_to,
                        remark__startswith='[수입검사]',
                        erp_incoming_no__isnull=False,
                    ).exclude(erp_incoming_no=''):
                        erp_ok, erp_err = del_erp_cancel(erp_trx.erp_incoming_no)
                        if erp_ok:
                            messages.info(request, f'ERP 입고 삭제 완료: {erp_trx.erp_incoming_no}')
                        else:
                            raise Exception(f'ERP 입고 삭제 실패: {erp_err} (ERP번호: {erp_trx.erp_incoming_no})')

                    if inspection.qty_good > 0:
                        target_code = inspection.target_warehouse_code or '2000'
                        wh_good = Warehouse.objects.filter(code=target_code).first()
                        if wh_good:
                            good_stock = MaterialStock.objects.filter(
                                warehouse=wh_good, part=part, lot_no=lot_no
                            ).first()
                            if good_stock and good_stock.quantity >= inspection.qty_good:
                                MaterialStock.objects.filter(pk=good_stock.pk).update(
                                    quantity=F('quantity') - inspection.qty_good
                                )
                            else:
                                raise Exception(f"목표 창고({target_code}) 양품 재고가 부족하여 취소할 수 없습니다.")

                    if inspection.qty_bad > 0:
                        wh_bad = Warehouse.objects.filter(code='8200').first()
                        if wh_bad:
                            bad_stock = MaterialStock.objects.filter(
                                warehouse=wh_bad, part=part, lot_no=lot_no
                            ).first()
                            if bad_stock and bad_stock.quantity >= inspection.qty_bad:
                                MaterialStock.objects.filter(pk=bad_stock.pk).update(
                                    quantity=F('quantity') - inspection.qty_bad
                                )

                    # 양품/불량 이동 트랜잭션 삭제
                    MaterialTransaction.objects.filter(
                        transaction_type='TRANSFER',
                        part=part, lot_no=lot_no,
                        warehouse_from=trx.warehouse_to,
                        remark__startswith='[수입검사]',
                    ).delete()

                    # 검사대기 창고에 원래 수량 복원
                    inspect_stock, _ = MaterialStock.objects.get_or_create(
                        warehouse=trx.warehouse_to, part=part, lot_no=lot_no
                    )
                    MaterialStock.objects.filter(pk=inspect_stock.pk).update(
                        quantity=F('quantity') + trx.quantity
                    )

                # PENDING이든 APPROVED이든 → 검사 판정 초기화
                inspection.status = 'PENDING'
                inspection.qty_good = 0
                inspection.qty_bad = 0
                inspection.inspector = None
                inspection.inspected_at = None
                inspection.check_report = False
                inspection.check_visual = False
                inspection.check_dimension = False
                inspection.remark = ''
                inspection.save()

            messages.success(request, f"[{trx_no}] 입고만 취소 완료 — 수입검사가 대기 상태로 초기화되었습니다. 품질팀에서 다시 판정해주세요.")

        except Exception as e:
            messages.error(request, f"입고 취소 중 오류 발생: {str(e)}")

        return redirect(redirect_to)

    # ── 전체 삭제 ──
    try:
        with transaction.atomic():
            part = trx.part
            lot_no = trx.lot_no

            # 수입검사 판정 완료 상태면 → 목표/부적합 창고 재고도 원복
            inspection = None
            if ImportInspection:
                try:
                    inspection = trx.inspection
                except Exception:
                    pass

            if inspection and inspection.status in ('APPROVED', 'REJECTED'):
                if inspection.qty_good > 0:
                    target_code = inspection.target_warehouse_code or '2000'
                    wh_good = Warehouse.objects.filter(code=target_code).first()
                    if wh_good:
                        good_stock = MaterialStock.objects.filter(
                            warehouse=wh_good, part=part, lot_no=lot_no
                        ).first()
                        if good_stock and good_stock.quantity >= inspection.qty_good:
                            MaterialStock.objects.filter(pk=good_stock.pk).update(
                                quantity=F('quantity') - inspection.qty_good
                            )
                        else:
                            raise Exception(f"목표 창고({target_code}) 양품 재고가 부족하여 삭제할 수 없습니다.")

                if inspection.qty_bad > 0:
                    wh_bad = Warehouse.objects.filter(code='8200').first()
                    if wh_bad:
                        bad_stock = MaterialStock.objects.filter(
                            warehouse=wh_bad, part=part, lot_no=lot_no
                        ).first()
                        if bad_stock and bad_stock.quantity >= inspection.qty_bad:
                            MaterialStock.objects.filter(pk=bad_stock.pk).update(
                                quantity=F('quantity') - inspection.qty_bad
                            )

                # 판정 시 생성된 ERP 입고 삭제
                from material.erp_api import delete_erp_incoming as del_erp_insp
                for erp_trx in MaterialTransaction.objects.filter(
                    transaction_type='TRANSFER',
                    part=part, lot_no=lot_no,
                    warehouse_from=trx.warehouse_to,
                    remark__startswith='[수입검사]',
                    erp_incoming_no__isnull=False,
                ).exclude(erp_incoming_no=''):
                    ok, err = del_erp_insp(erp_trx.erp_incoming_no)
                    if ok:
                        messages.info(request, f'ERP 입고 삭제 완료: {erp_trx.erp_incoming_no}')

                # 양품/불량 이동 트랜잭션 삭제
                MaterialTransaction.objects.filter(
                    transaction_type='TRANSFER',
                    part=part, lot_no=lot_no,
                    warehouse_from=trx.warehouse_to,
                    remark__startswith='[수입검사]',
                ).delete()
            else:
                # PENDING 또는 검사 없음 → 검사대기(또는 입고) 창고에서 차감
                stock = MaterialStock.objects.filter(
                    warehouse=trx.warehouse_to,
                    part=part,
                    lot_no=lot_no
                ).first()

                if not stock:
                    messages.error(request, "해당 재고를 찾을 수 없습니다.")
                    return redirect(redirect_to)

                if stock.quantity < trx.quantity:
                    messages.error(request, f"현재 재고({stock.quantity})가 입고 수량({trx.quantity})보다 적어 삭제할 수 없습니다.")
                    return redirect(redirect_to)

                MaterialStock.objects.filter(pk=stock.pk).update(
                    quantity=F('quantity') - trx.quantity
                )

            # ERP 입고 삭제 (실패 시 전체 롤백)
            erp_no = getattr(trx, 'erp_incoming_no', None)
            if erp_no:
                from material.erp_api import delete_erp_incoming
                erp_ok, erp_err = delete_erp_incoming(erp_no)
                if erp_ok:
                    messages.info(request, f'ERP 입고 삭제 완료: {erp_no}')
                else:
                    raise Exception(f'ERP 입고 삭제 실패: {erp_err} (ERP번호: {erp_no})')

            # ImportInspection 삭제
            if ImportInspection:
                try:
                    trx.inspection.delete()
                except Exception:
                    pass

            # 트랜잭션 삭제
            trx.delete()

        messages.success(request, f"입고 건 [{trx_no}] 삭제 완료 (재고 {trx_qty}개 차감)")

    except Exception as e:
        messages.error(request, f"취소 처리 중 오류 발생: {str(e)}")

    return redirect(redirect_to)


@wms_permission_required('can_wms_inout_edit')
def edit_manual_incoming(request, trx_id):
    """[WMS] 수기 입고 수정 - 수량/일자/LOT/비고 변경"""
    from django.http import JsonResponse
    import json
    from datetime import datetime
    from .models import RawMaterialLabel

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': '잘못된 요청 형식입니다.'}, status=400)

    trx = get_object_or_404(MaterialTransaction, pk=trx_id)

    # === 검증 ===

    # (1) 트랜잭션 타입
    if trx.transaction_type not in ('IN_MANUAL', 'IN_SCM'):
        return JsonResponse({'success': False, 'error': 'ERP 입고 건은 수정할 수 없습니다.'})

    # (2) 라벨 발행 여부
    label_count = RawMaterialLabel.objects.filter(
        incoming_transaction=trx
    ).exclude(status='CANCELLED').count()
    if label_count > 0:
        return JsonResponse({'success': False, 'error': f'라벨이 {label_count}장 발행된 입고 건은 수정할 수 없습니다.'})

    # (3) 검사 상태
    inspection = None
    if ImportInspection:
        try:
            inspection = trx.inspection
        except Exception:
            inspection = None

    if inspection and inspection.status in ('APPROVED', 'REJECTED'):
        return JsonResponse({'success': False, 'error': '검사 판정이 완료된 입고 건은 수정할 수 없습니다.'})

    # (4) 마감 기간 (기존 일자)
    is_closed, warning_msg, _ = check_closing_date(
        trx.date.date() if hasattr(trx.date, 'date') and callable(trx.date.date) else trx.date
    )
    if is_closed:
        return JsonResponse({'success': False, 'error': f'마감된 기간의 입고 건은 수정할 수 없습니다. ({warning_msg})'})

    # === 입력값 파싱 ===

    # 수량
    try:
        new_qty = int(data.get('quantity', 0))
        if new_qty <= 0:
            raise ValueError
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': '수량은 1 이상의 정수여야 합니다.'})

    # 일자
    try:
        new_date_str = data.get('date', '')
        new_date = timezone.make_aware(datetime.strptime(new_date_str, '%Y-%m-%d'))
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': '올바른 날짜 형식이 아닙니다. (YYYY-MM-DD)'})

    # (5) 마감 기간 (변경 일자)
    is_closed_new, warning_msg_new, _ = check_closing_date(new_date.date())
    if is_closed_new:
        return JsonResponse({'success': False, 'error': f'변경하려는 날짜가 마감 기간에 속합니다. ({warning_msg_new})'})

    # LOT
    new_lot_date = None
    new_lot_str = (data.get('lot_no') or '').strip()
    if new_lot_str and new_lot_str != '-':
        try:
            new_lot_date = datetime.strptime(new_lot_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'LOT 번호 형식이 올바르지 않습니다. (YYYY-MM-DD)'})

    new_remark = data.get('remark', '')

    # === 변경 감지 ===
    old_qty = trx.quantity
    old_lot = trx.lot_no
    warehouse = trx.warehouse_to
    qty_changed = (new_qty != old_qty)
    lot_changed = (new_lot_date != old_lot)

    # === 재고 조정 + 저장 (atomic) ===
    try:
        with transaction.atomic():
            if qty_changed or lot_changed:
                # (A) 기존 재고 차감
                old_stock = MaterialStock.objects.filter(
                    warehouse=warehouse, part=trx.part, lot_no=old_lot
                ).first()

                if not old_stock or old_stock.quantity < old_qty:
                    current = old_stock.quantity if old_stock else 0
                    return JsonResponse({
                        'success': False,
                        'error': f'현재 재고({current})가 기존 입고 수량({old_qty})보다 적어 수정할 수 없습니다.'
                    })

                MaterialStock.objects.filter(pk=old_stock.pk).update(
                    quantity=F('quantity') - old_qty
                )

                # (B) 새 재고 증가
                new_stock, _ = MaterialStock.objects.get_or_create(
                    warehouse=warehouse, part=trx.part, lot_no=new_lot_date,
                    defaults={'quantity': 0}
                )
                MaterialStock.objects.filter(pk=new_stock.pk).update(
                    quantity=F('quantity') + new_qty
                )
                new_stock.refresh_from_db()
                trx.result_stock = new_stock.quantity

            # (C) 트랜잭션 업데이트
            trx.quantity = new_qty
            trx.date = new_date
            trx.lot_no = new_lot_date
            trx.remark = new_remark
            trx.save(update_fields=['quantity', 'date', 'lot_no', 'remark', 'result_stock'])

            # (D) 검사 LOT 업데이트 (PENDING)
            if inspection and lot_changed:
                inspection.lot_no = new_lot_date
                inspection.save(update_fields=['lot_no'])

            # (E) ERP 재등록 (무검사 + ERP 연동 건)
            erp_no = trx.erp_incoming_no
            if erp_no and (qty_changed or lot_changed):
                from material.erp_api import delete_erp_incoming, register_erp_incoming
                del_ok, del_err = delete_erp_incoming(erp_no)
                if not del_ok:
                    raise Exception(f'ERP 입고 삭제 실패: {del_err}')

                trx.erp_incoming_no = None
                trx.erp_sync_status = 'PENDING'
                trx.save(update_fields=['erp_incoming_no', 'erp_sync_status'])

                reg_ok, reg_no, reg_err = register_erp_incoming(trx, new_qty, warehouse.code)
                if not reg_ok and reg_err:
                    logger.warning(f'ERP 재등록 실패: {reg_err}')

        return JsonResponse({
            'success': True,
            'message': f'입고 건 [{trx.transaction_no}] 수정 완료'
        })

    except Exception as e:
        logger.error(f'입고 수정 오류: {e}', exc_info=True)
        return JsonResponse({'success': False, 'error': f'수정 처리 중 오류: {str(e)}'})


@login_required
@wms_permission_required('can_wms_inout_edit')
def reregister_erp_price(request, trx_id):
    """[입고 내역] ERP 단가 재반영 - 기존 ERP 입고 삭제 후 최신 단가로 재등록"""
    from django.http import JsonResponse

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    trx = get_object_or_404(MaterialTransaction, pk=trx_id)

    # ERP입고(아마란스에서 온 건)는 제외
    if trx.transaction_type in ('IN_ERP', 'RCV_ERP'):
        return JsonResponse({'success': False, 'error': 'ERP입고 건은 아마란스에서 직접 수정하세요.'})

    try:
        from material.erp_api import delete_erp_incoming, register_erp_incoming

        erp_no = trx.erp_incoming_no

        # 창고 코드 결정 — 수입검사 대기장(1000)이면 TRANSFER의 최종 창고 사용
        wh = trx.warehouse_to
        warehouse_code = wh.code if wh else '2000'
        if warehouse_code == '1000':
            transfer = MaterialTransaction.objects.filter(
                part=trx.part, lot_no=trx.lot_no,
                transaction_type='TRANSFER',
                warehouse_from__code='1000',
            ).order_by('-date').first()
            if transfer and transfer.warehouse_to:
                warehouse_code = transfer.warehouse_to.code
            else:
                warehouse_code = '2000'

        # 발주번호 추출 (remark에서)
        erp_order_no, erp_order_seq = '', ''
        origin_trx = trx
        if trx.transaction_type == 'TRANSFER' and trx.remark and '수입검사' in trx.remark:
            # 검사입고인 경우 원본 입고 트랜잭션 찾기
            from qms.models import ImportInspection as II
            insp = II.objects.filter(
                inbound_transaction__part=trx.part,
                inbound_transaction__lot_no=trx.lot_no,
                inbound_transaction__vendor=trx.vendor,
                status='APPROVED'
            ).select_related('inbound_transaction').order_by('-inspected_at').first()
            if insp:
                origin_trx = insp.inbound_transaction

        import re as _re
        if origin_trx.remark:
            m = _re.search(r'ERP:(\S+)-(\d+)', origin_trx.remark or '')
            if m:
                erp_order_no = m.group(1)
                erp_order_seq = m.group(2)

        # 1) 기존 ERP 입고 삭제 (있는 경우만)
        if erp_no:
            del_ok, del_err = delete_erp_incoming(erp_no)
            if not del_ok:
                return JsonResponse({'success': False, 'error': f'ERP 입고 삭제 실패: {del_err}'})

            trx.erp_incoming_no = None
            trx.erp_sync_status = 'PENDING'
            trx.save(update_fields=['erp_incoming_no', 'erp_sync_status'])

        # 2) 최신 단가로 (재)등록
        print(f'[단가재반영] trx={trx.id}, part={trx.part.part_no}, qty={trx.quantity}, wh={warehouse_code}, vendor={trx.vendor.erp_code if trx.vendor else None}', flush=True)
        reg_ok, reg_no, reg_err = register_erp_incoming(
            trx, trx.quantity, warehouse_code,
            erp_order_no=erp_order_no, erp_order_seq=erp_order_seq
        )
        print(f'[단가재반영] 결과: ok={reg_ok}, no={reg_no}, err={reg_err}', flush=True)

        if reg_ok:
            action = '재반영' if erp_no else '등록'
            return JsonResponse({
                'success': True,
                'message': f'ERP 단가 {action} 완료 (ERP번호: {reg_no})',
                'new_erp_no': reg_no
            })
        else:
            err_detail = f' 기존 입고({erp_no})는 이미 삭제되었습니다.' if erp_no else ''
            return JsonResponse({
                'success': False,
                'error': f'ERP 등록 실패: {reg_err}.{err_detail}'
            })

    except Exception as e:
        logger.error(f'단가 재반영 오류: {e}', exc_info=True)
        return JsonResponse({'success': False, 'error': f'처리 중 오류: {str(e)}'})


@login_required
@wms_permission_required('can_wms_inout_edit')
def erp_incoming_sync(request):
    """[WMS] ERP 입고 내역 역방향 동기화 - 백그라운드 스레드"""
    if request.method != 'POST':
        return redirect('material:manual_incoming')

    from django.core.cache import cache

    if cache.get('erp_incoming_sync_running'):
        messages.warning(request, '이미 ERP 입고 동기화가 진행 중입니다.')
        return redirect('material:manual_incoming')

    import threading

    def _run_sync():
        try:
            from material.erp_api import sync_erp_incoming
            cache.set('erp_incoming_sync_running', True, timeout=600)
            synced, skipped, errors, error_list = sync_erp_incoming()
            cache.set('erp_incoming_sync_result', {
                'synced': synced, 'skipped': skipped,
                'errors': errors, 'error_list': error_list[:5],
                'finished_at': timezone.now().strftime('%Y-%m-%d %H:%M'),
            }, timeout=86400)
        except Exception as e:
            logger.error(f'ERP 입고 동기화 예외: {e}')
            cache.set('erp_incoming_sync_result', {
                'synced': 0, 'skipped': 0, 'errors': 1,
                'error_list': [str(e)],
                'finished_at': timezone.now().strftime('%Y-%m-%d %H:%M'),
            }, timeout=86400)
        finally:
            cache.delete('erp_incoming_sync_running')

    thread = threading.Thread(target=_run_sync, daemon=True)
    thread.start()

    messages.info(request, 'ERP 입고 동기화를 시작했습니다. 잠시 후 새로고침하여 결과를 확인하세요.')
    return redirect('material:manual_incoming')


@wms_permission_required('can_wms_inout_view')
def incoming_history(request):
    """
    [입고 내역 조회]
    - 구분은 'SCM입고' / '수기입고'만 표시
    - 수입검사 대기(8100) 입고는 제외
    - 부적합(8200)으로 간 건 제외
    - 수기입고 비고: 입력 비고
    - SCM입고 비고: 납품서번호(ref_delivery_order) (※ TRANSFER에는 ref가 없을 수 있음)
    """
    waiting_wh = Warehouse.objects.filter(code='8100').first()
    if not waiting_wh:
        waiting_wh = Warehouse.objects.filter(name__contains='수입검사').first()
    # 수입검사 대기장(1000)도 중간 단계 창고
    inspection_wh = Warehouse.objects.filter(code='1000').first()

    ng_wh = Warehouse.objects.filter(code='8200').first()
    if not ng_wh:
        ng_wh = Warehouse.objects.filter(name__contains='부적합').first()

    # ✅ "입고 확정" 기준
    # - 수기입고(IN_MANUAL/IN_SCM): 검사대기장(8100) 직행 건은 제외 (TRANSFER로 확정됨)
    # - 수입검사 경유: 검사 완료 후 TRANSFER(8100→정상창고)만 최종 입고로 표시
    # - ERP입고/생산입고: 그대로
    qs = MaterialTransaction.objects.filter(
        Q(transaction_type='IN_MANUAL')
        | Q(transaction_type='IN_ERP')
        | Q(transaction_type='RCV_ERP')
        | Q(transaction_type='IN_SCM')
        |
        (
            Q(transaction_type='TRANSFER')
            & Q(quantity__gt=0)
            & Q(warehouse_from__isnull=False)
            & Q(warehouse_to__isnull=False)
        )
    ).select_related('part', 'warehouse_to', 'warehouse_from', 'actor', 'vendor').order_by('-date', '-id')

    if waiting_wh:
        # 검사대기장으로 입고된 건 제외 (IN_MANUAL, IN_SCM 모두)
        qs = qs.exclude(warehouse_to=waiting_wh, transaction_type__in=['IN_MANUAL', 'IN_SCM'])

        # 수입검사 대기장(1000)으로 입고된 건도 제외 (TRANSFER에서 최종 입고로 표시)
        if inspection_wh:
            qs = qs.exclude(warehouse_to=inspection_wh, transaction_type__in=['IN_MANUAL', 'IN_SCM'])

        # TRANSFER는 "검사대기장/수입검사대기장 → 다른창고"만 입고확정으로 취급
        transfer_q = Q(transaction_type='TRANSFER') & Q(warehouse_from=waiting_wh)
        if inspection_wh:
            transfer_q = transfer_q | (Q(transaction_type='TRANSFER') & Q(warehouse_from=inspection_wh))
        qs = qs.filter(
            Q(transaction_type='IN_MANUAL') |
            Q(transaction_type='IN_ERP') |
            Q(transaction_type='RCV_ERP') |
            Q(transaction_type='IN_SCM') |
            transfer_q
        )
    else:
        qs = qs.exclude(transaction_type='TRANSFER')

    if ng_wh:
        qs = qs.exclude(warehouse_to=ng_wh)

    # 2) 검색 필터
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(part__part_no__icontains=q) |
            Q(part__part_name__icontains=q)
        )

    q_group = (request.GET.get('q_group') or '').strip()
    if q_group:
        qs = qs.filter(part__part_group__icontains=q_group)

    q_vendor = (request.GET.get('q_vendor') or '').strip()
    if q_vendor:
        qs = qs.filter(vendor__name__icontains=q_vendor)

    # 3) 날짜 필터 ('None' 방어)
    start_date = (request.GET.get('start_date') or '').strip()
    end_date = (request.GET.get('end_date') or '').strip()

    if start_date in ('None', 'null', 'NULL'):
        start_date = ''
    if end_date in ('None', 'null', 'NULL'):
        end_date = ''

    if start_date:
        qs = qs.filter(date__date__gte=start_date)
    if end_date:
        qs = qs.filter(date__date__lte=end_date)

    # 4) 페이징
    paginator = Paginator(qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ✅ 구분/비고 표시값 세팅 + 관리 가능 여부
    from .models import RawMaterialLabel
    for item in page_obj:
        remark = item.remark or ""

        if item.transaction_type == 'IN_MANUAL':
            item.display_type = "수기입고"
            item.badge_color = "success"
            remark = re.sub(r'\[수입검사 대상\]\s*', '', remark)
            remark = re.sub(r'\[발주입고\]\s*ERP:\S*\s*', '', remark)
            item.display_remark = remark.strip()
        elif item.transaction_type == 'IN_SCM':
            item.display_type = "발주입고"
            item.badge_color = "primary"
            remark = re.sub(r'\[수입검사 대상\]\s*', '', remark)
            remark = re.sub(r'\[발주입고\]\s*ERP:\S*\s*', '', remark)
            item.display_remark = remark.strip()
        elif item.transaction_type == 'IN_ERP':
            item.display_type = "ERP입고"
            item.badge_color = "info"
            item.display_remark = ""
        elif item.transaction_type == 'RCV_ERP':
            item.display_type = "ERP생산입고"
            item.badge_color = "info"
            item.display_remark = ""
        elif item.transaction_type == 'TRANSFER':
            item.display_type = "검사입고"
            item.badge_color = "warning"
            # 원본 입고 건의 비고를 가져옴
            origin_remark = ""
            origin_trx = MaterialTransaction.objects.filter(
                part=item.part, lot_no=item.lot_no,
                transaction_type__in=('IN_MANUAL', 'IN_SCM'),
            ).order_by('-date').first()
            if origin_trx and origin_trx.remark:
                origin_remark = re.sub(r'\[수입검사 대상\]\s*', '', origin_trx.remark)
                origin_remark = re.sub(r'\[발주입고\]\s*ERP:\S*\s*', '', origin_remark).strip()
            item.display_remark = origin_remark or item.remark or ""
        else:
            item.display_type = item.transaction_type
            item.badge_color = "secondary"
            item.display_remark = remark

        # 관리 가능 여부 판단
        item.can_edit = item.transaction_type in ('IN_MANUAL', 'IN_SCM')
        item.can_delete = item.transaction_type in ('IN_MANUAL', 'IN_SCM')
        # ERP입고(IN_ERP/RCV_ERP)는 아마란스에서 직접 수정 → 제외
        item.can_reregister = item.transaction_type not in ('IN_ERP', 'RCV_ERP')
        item.has_label = RawMaterialLabel.objects.filter(
            incoming_transaction=item
        ).exclude(status='CANCELLED').exists() if item.can_delete else False
        item.label_issued = RawMaterialLabel.objects.filter(
            incoming_transaction=item
        ).exclude(status='CANCELLED').exists()
        # 수입검사 첨부파일 조회
        item.inspection_attachment_url = ''
        item.inspection_attachment_name = ''
        try:
            insp = item.inspection
            if insp.attachment:
                item.inspection_attachment_url = insp.attachment.url
                item.inspection_attachment_name = insp.attachment.name.split('/')[-1]
        except Exception:
            # TRANSFER(검사입고)는 원본 입고 건에서 조회
            if item.transaction_type == 'TRANSFER':
                from qms.models import ImportInspection
                origin = ImportInspection.objects.filter(
                    inbound_transaction__part=item.part,
                    inbound_transaction__lot_no=item.lot_no,
                    attachment__isnull=False,
                ).exclude(attachment='').first()
                if origin and origin.attachment:
                    item.inspection_attachment_url = origin.attachment.url
                    item.inspection_attachment_name = origin.attachment.name.split('/')[-1]

        # 수입검사 대기장이면 최종 입고 창고 표시
        if item.warehouse_to and item.warehouse_to.code == '1000':
            final_transfer = MaterialTransaction.objects.filter(
                part=item.part, lot_no=item.lot_no,
                transaction_type='TRANSFER',
                warehouse_from__code='1000',
            ).select_related('warehouse_to').order_by('-date').first()
            if final_transfer and final_transfer.warehouse_to:
                item.display_warehouse = final_transfer.warehouse_to
            else:
                item.display_warehouse = item.warehouse_to
        else:
            item.display_warehouse = item.warehouse_to

    # 모달 선택용 목록
    part_groups = list(
        Part.objects.values_list('part_group', flat=True)
        .distinct().order_by('part_group')
    )
    vendors = list(
        Vendor.objects.values_list('name', flat=True)
        .order_by('name')
    )

    # 품번 자동완성용 목록
    parts = Part.objects.values_list('part_no', 'part_name').order_by('part_no')

    context = {
        'page_obj': page_obj,
        'q': q,
        'q_group': q_group,
        'q_vendor': q_vendor,
        'start_date': start_date,
        'end_date': end_date,
        'part_groups': part_groups,
        'vendors': vendors,
        'parts': parts,
    }
    return render(request, 'material/incoming_history.html', context)


# =============================================================================
# 3. 현장 지원 (태그 발행)
# =============================================================================

@wms_permission_required('can_wms_stock_view')
def process_tag_form(request):
    """현품표 발행 입력 폼"""
    parts = Part.objects.all().order_by('part_no')
    context = {'parts': parts}
    return render(request, 'material/process_tag_form.html', context)


@wms_permission_required('can_wms_stock_view')
def process_tag_print(request):
    """현품표 실제 출력 뷰"""
    from .models import ProcessTag
    from datetime import datetime

    if request.method == 'POST':
        part_no = request.POST.get('part_no')
        part_name = request.POST.get('part_name', '')
        quantity = request.POST.get('quantity')
        lot_no = request.POST.get('lot_no')
        print_count = int(request.POST.get('print_count', 1))
        print_mode = request.POST.get('print_mode', 'roll')
        size_type = request.POST.get('size_type', 'medium')
        custom_width = request.POST.get('custom_width', 100)
        custom_height = request.POST.get('custom_height', 60)
        use_shift = request.POST.get('use_shift') == '1'
        shift_type = request.POST.get('shift_type', '주간') if use_shift else ''
        use_serial = request.POST.get('use_serial') == '1'
        serial_start = int(request.POST.get('serial_start', 1)) if use_serial else 0

        # A4 모아찍기 모드용 레이아웃 계산
        if print_mode == 'sheet':
            # 라벨 사이즈에 따른 mm 단위 설정
            extra_row = use_shift or use_serial
            if size_type == 'small':
                label_w_mm, label_h_mm = 60, 70 if not extra_row else 80
            elif size_type == 'medium':
                label_w_mm = 95
                label_h_mm = 53 if extra_row else 45  # 12개→10개 (2×5)
            elif size_type == 'large':
                label_w_mm, label_h_mm = 210, 148
            elif size_type == 'custom':
                label_w_mm = int(custom_width)
                label_h_mm = int(custom_height)
            else:
                label_w_mm = 95
                label_h_mm = 53 if extra_row else 45

            # A4 용지 크기 (210mm x 297mm)
            # 여백 고려: 좌우 5mm, 상하 5mm
            usable_width = 210 - 10  # 200mm
            usable_height = 297 - 10  # 287mm

            # 한 줄에 들어갈 라벨 개수 (gap 3mm 고려)
            sheet_cols = max(1, int((usable_width + 3) / (label_w_mm + 3)))
            sheet_rows = max(1, int((usable_height + 3) / (label_h_mm + 3)))
            per_page = sheet_cols * sheet_rows
        else:
            label_w_mm, label_h_mm = 0, 0
            sheet_cols, per_page = 0, 0

        # ========================================
        # ProcessTag 레코드 생성 (중복 스캔 방지용)
        # ========================================
        # Part 조회 (없으면 None)
        part = Part.objects.filter(part_no=part_no).first()

        # LOT 날짜 파싱
        lot_date = None
        if lot_no:
            try:
                lot_date = datetime.strptime(lot_no, '%Y-%m-%d').date()
            except ValueError:
                pass  # LOT 형식이 날짜가 아닌 경우 None 유지

        # 태그 생성
        tag_list = []
        for _ in range(print_count):
            tag_id = ProcessTag.generate_tag_id()
            tag = ProcessTag.objects.create(
                tag_id=tag_id,
                part=part,
                part_no=part_no,
                part_name=part_name,
                quantity=int(quantity) if quantity else 0,
                lot_no=lot_date,
                status='PRINTED',
                printed_by=request.user if request.user.is_authenticated else None
            )
            serial_no = serial_start + len(tag_list) if use_serial else 0
            tag_list.append({
                'index': len(tag_list),
                'tag_id': tag_id,
                'serial_no': serial_no,
            })

        # 중량 정보
        weight_qty = float(part.weight_qty) if part and part.weight_qty else 0
        weight_unit = part.weight_unit if part else ''

        context = {
            'part_no': part_no,
            'part_name': part_name,
            'part_group': part.part_group if part else '',  # 품목군 추가
            'quantity': quantity,
            'lot_no': lot_no,
            'weight_qty': weight_qty,
            'weight_unit': weight_unit,
            'print_mode': print_mode,
            'size_type': size_type,
            'print_range': range(print_count),
            'tag_list': tag_list,  # 태그 ID 목록 전달
            'custom_width': custom_width,
            'custom_height': custom_height,
            'print_date': timezone.now().strftime('%Y-%m-%d'),
            'worker': request.user.username,
            # A4 모아찍기용 변수
            'label_w_mm': label_w_mm,
            'label_h_mm': label_h_mm,
            'sheet_cols': sheet_cols,
            'per_page': per_page,
            'use_shift': use_shift,
            'shift_type': shift_type,
            'use_serial': use_serial,
            'serial_start': serial_start,
        }
        return render(request, 'material/process_tag_print.html', context)

    return redirect('material:process_tag_form')


@login_required
def lot_allocation_print(request):
    """LOT 배분 후 현품표 일괄 출력 (복수 품번×LOT)"""
    if request.method != 'POST':
        return redirect('material:lot_allocation')

    import json
    from .models import ProcessTag
    from datetime import datetime

    items = json.loads(request.POST.get('items', '[]'))
    print_mode = request.POST.get('print_mode', 'roll')
    size_type = request.POST.get('size_type', 'medium')

    if not items:
        return redirect('material:lot_allocation')

    # A4 모아찍기 레이아웃 계산
    if print_mode == 'sheet':
        if size_type == 'small':
            label_w_mm, label_h_mm = 60, 70
        elif size_type == 'medium':
            label_w_mm, label_h_mm = 95, 45
        elif size_type == 'large':
            label_w_mm, label_h_mm = 210, 148
        else:
            label_w_mm, label_h_mm = 95, 45

        usable_width = 200  # 210 - 10
        usable_height = 287  # 297 - 10
        sheet_cols = max(1, int((usable_width + 3) / (label_w_mm + 3)))
        sheet_rows = max(1, int((usable_height + 3) / (label_h_mm + 3)))
        per_page = sheet_cols * sheet_rows
    else:
        label_w_mm = label_h_mm = 0
        sheet_cols = per_page = 0

    # 각 아이템별 ProcessTag 생성
    all_tags = []
    for item in items:
        part = Part.objects.filter(part_no=item['part_no']).first()
        lot_date = None
        if item.get('lot_no'):
            try:
                lot_date = datetime.strptime(item['lot_no'], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                pass

        tag_id = ProcessTag.generate_tag_id()
        ProcessTag.objects.create(
            tag_id=tag_id,
            part=part,
            part_no=item['part_no'],
            part_name=item.get('part_name', ''),
            quantity=int(item.get('quantity', 0)),
            lot_no=lot_date,
            status='PRINTED',
            printed_by=request.user if request.user.is_authenticated else None,
        )
        all_tags.append({
            'index': len(all_tags),
            'tag_id': tag_id,
            'part_no': item['part_no'],
            'part_name': item.get('part_name', ''),
            'part_group': item.get('part_group', ''),
            'quantity': int(item.get('quantity', 0)),
            'lot_no': item.get('lot_no', ''),
            'weight_qty': float(part.weight_qty) if part and part.weight_qty else 0,
            'weight_unit': part.weight_unit if part else '',
        })

    context = {
        'multi_mode': True,
        'tag_list': all_tags,
        'print_mode': print_mode,
        'size_type': size_type,
        'print_date': timezone.now().strftime('%Y-%m-%d'),
        'worker': request.user.username,
        'label_w_mm': label_w_mm,
        'label_h_mm': label_h_mm,
        'sheet_cols': sheet_cols,
        'per_page': per_page,
        # 단일모드 호환 변수
        'part_no': '',
        'part_name': '',
        'part_group': '',
        'quantity': '',
        'lot_no': '',
        'print_range': range(len(all_tags)),
    }
    return render(request, 'material/process_tag_print.html', context)


@wms_permission_required('can_wms_stock_edit')
def lot_allocation_rm_print(request):
    """LOT 배분 후 원재료 라벨(RM) 일괄 출력 - 3200 창고 기준"""
    if request.method != 'POST':
        return redirect('material:lot_allocation')

    import json
    from .models import RawMaterialLabel
    from datetime import datetime
    from decimal import Decimal

    items = json.loads(request.POST.get('items', '[]'))
    if not items:
        return redirect('material:lot_allocation')

    all_labels = []
    for item in items:
        part = Part.objects.filter(part_no=item['part_no']).first()
        if not part:
            continue

        lot_date = None
        if item.get('lot_no'):
            try:
                lot_date = datetime.strptime(item['lot_no'], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                pass

        qty = Decimal(str(item.get('quantity', 0)))
        unit = (part.weight_unit or 'KG').strip() or 'KG'

        label = RawMaterialLabel.objects.create(
            label_id=RawMaterialLabel.generate_label_id(),
            part=part,
            part_no=part.part_no,
            part_name=part.part_name or '',
            lot_no=lot_date,
            quantity=qty,
            unit=unit,
            status='INSTOCK',
            printed_by=request.user if request.user.is_authenticated else None,
        )
        all_labels.append(label)

    if not all_labels:
        messages.error(request, '출력할 라벨이 없습니다.')
        return redirect('material:lot_allocation')

    # 라벨 출력 페이지로 리다이렉트
    label_ids = ','.join([str(l.id) for l in all_labels])
    return redirect(f'/wms/raw-material/label-print/?ids={label_ids}')


@wms_permission_required('can_wms_stock_edit')
def lot_allocation_plt_print(request):
    """LOT 배분 후 파렛트 라벨(PLT) 일괄 출력 - 사용자 지정 단위로 분할"""
    if request.method != 'POST':
        return redirect('material:lot_allocation')

    import json
    from .models import RawMaterialLabel
    from datetime import datetime
    from decimal import Decimal, InvalidOperation

    items = json.loads(request.POST.get('items', '[]'))
    pallet_unit_input = request.POST.get('pallet_unit_qty', '').strip()

    if not items:
        return redirect('material:lot_allocation')

    pallet_unit_qty = None
    if pallet_unit_input:
        try:
            pallet_unit_qty = Decimal(pallet_unit_input)
            if pallet_unit_qty <= 0:
                pallet_unit_qty = None
        except (InvalidOperation, ValueError):
            pallet_unit_qty = None

    all_labels = []
    for item in items:
        part = Part.objects.filter(part_no=item['part_no']).first()
        if not part:
            continue

        lot_date = None
        if item.get('lot_no'):
            try:
                lot_date = datetime.strptime(item['lot_no'], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                pass

        total_qty = Decimal(str(item.get('quantity', 0)))
        unit = (part.weight_unit or 'KG').strip() or 'KG'

        # 분할 수량 계산
        split_qtys = []
        if pallet_unit_qty and total_qty > pallet_unit_qty:
            full_count = int(total_qty // pallet_unit_qty)
            remainder = total_qty - (pallet_unit_qty * full_count)
            for _ in range(full_count):
                split_qtys.append(pallet_unit_qty)
            if remainder > 0:
                split_qtys.append(remainder)
        else:
            split_qtys.append(total_qty)

        for split_qty in split_qtys:
            label = RawMaterialLabel.objects.create(
                label_id=RawMaterialLabel.generate_pallet_label_id(),
                label_type='PALLET',
                part=part,
                part_no=part.part_no,
                part_name=part.part_name or '',
                lot_no=lot_date,
                quantity=split_qty,
                unit=unit,
                status='INSTOCK',
                printed_by=request.user if request.user.is_authenticated else None,
            )
            all_labels.append(label)

    if not all_labels:
        messages.error(request, '출력할 라벨이 없습니다.')
        return redirect('material:lot_allocation')

    # 파렛트 라벨 출력 페이지 (ids로 복수 지원)
    label_ids = ','.join([str(l.id) for l in all_labels])
    return redirect(f'/wms/raw-material/pallet-label/print/?ids={label_ids}')


# =============================================================================
# 3-2. 현품표 스캔 API (중복 스캔 확인)
# =============================================================================

from django.http import JsonResponse
import json

@wms_permission_required('can_wms_stock_view')
def api_process_tag_scan(request):
    """
    현품표 QR 스캔 시 호출되는 API
    - 태그 ID로 ProcessTag 조회
    - 스캔 기록 저장
    - 중복 스캔 시 차단 (success: false)

    요청: POST { "tag_id": "TAG-20260124-0001", "warehouse_id": 1 (optional) }
    응답: {
        "success": true/false,
        "is_first_scan": true/false,
        "error": "중복 스캔 시 에러 메시지",
        "tag_info": { ... }
    }
    """
    from .models import ProcessTag, ProcessTagScanLog, RawMaterialLabel

    try:
        data = json.loads(request.body)
        raw_tag = data.get('tag_id', '').strip()
        # QR 데이터 형식: TAG_ID|품번|수량|LOT → 첫 번째 필드만 추출
        tag_id = raw_tag.split('|')[0].strip() if '|' in raw_tag else raw_tag
        warehouse_id = data.get('warehouse_id')
        # 부족 시 사용자가 입력한 조정 수량 (옵션)
        override_qty = data.get('override_qty')
        # 도착 창고 코드 (기본: 3000 현장)
        target_warehouse_code = (data.get('target_warehouse_code') or '3000').strip()

        if not tag_id:
            return JsonResponse({
                'success': False,
                'error': '태그 ID가 누락되었습니다.'
            }, status=400)

        # RM/PLT 라벨 스캔 처리 (원재료 라벨)
        if tag_id.startswith('RM-') or tag_id.startswith('PLT-'):
            rm_label = RawMaterialLabel.objects.filter(label_id=tag_id).first()
            if not rm_label:
                return JsonResponse({
                    'success': False,
                    'error': f'등록되지 않은 라벨입니다: {tag_id}',
                    'is_registered': False
                })

            label_kind = '파렛트' if tag_id.startswith('PLT-') else '원재료'

            if rm_label.status == 'USED':
                # 이미 투입된 라벨 → 투입 취소 처리 (3000 차감, 3200 복구, INSTOCK 복구)
                wh_3200 = Warehouse.objects.filter(code='3200').first()
                wh_3000 = Warehouse.objects.filter(code='3000').first()
                if not wh_3200 or not wh_3000:
                    return JsonResponse({
                        'success': False,
                        'error': '창고 마스터(3200/3000)가 등록되어 있지 않습니다.',
                    })

                # 가장 최근 투입 이력에서 실제 이동 수량 조회
                last_trx = MaterialTransaction.objects.filter(
                    part=rm_label.part,
                    transaction_type='TRANSFER',
                    warehouse_from=wh_3200,
                    warehouse_to=wh_3000,
                    remark__icontains=rm_label.label_id,
                ).order_by('-date').first()

                cancel_qty = float(last_trx.quantity) if last_trx else float(rm_label.quantity)
                lot_no_cancel = rm_label.lot_no

                from django.db.models import F as _F
                with transaction.atomic():
                    # 3000 재고 차감 (LOT 유지)
                    src_3000 = MaterialStock.objects.filter(
                        warehouse=wh_3000, part=rm_label.part, lot_no=lot_no_cancel
                    ).first()
                    if src_3000:
                        MaterialStock.objects.filter(pk=src_3000.pk).update(
                            quantity=_F('quantity') - cancel_qty
                        )
                    else:
                        MaterialStock.objects.create(
                            warehouse=wh_3000, part=rm_label.part, lot_no=lot_no_cancel,
                            quantity=-cancel_qty
                        )

                    # 3200 재고 복구 (LOT 유지)
                    tgt_3200, _c = MaterialStock.objects.get_or_create(
                        warehouse=wh_3200, part=rm_label.part, lot_no=lot_no_cancel,
                        defaults={'quantity': 0}
                    )
                    MaterialStock.objects.filter(pk=tgt_3200.pk).update(
                        quantity=_F('quantity') + cancel_qty
                    )
                    tgt_3200.refresh_from_db()

                    # 라벨 USED → INSTOCK 복구
                    rm_label.status = 'INSTOCK'
                    rm_label.used_at = None
                    rm_label.used_by = None
                    rm_label.save(update_fields=['status', 'used_at', 'used_by'])

                    # 취소 이력 (역방향 TRANSFER)
                    trx_no = f"RM-CANCEL-{timezone.now().strftime('%y%m%d%H%M%S%f')}-{request.user.id}"
                    lot_disp = lot_no_cancel.strftime('%Y-%m-%d') if lot_no_cancel else 'NO LOT'
                    label_kind_c = '파렛트' if rm_label.label_id.startswith('PLT-') else '원재료'
                    cancel_trx = MaterialTransaction.objects.create(
                        transaction_no=trx_no,
                        transaction_type='TRANSFER',
                        date=timezone.now(),
                        part=rm_label.part,
                        quantity=cancel_qty,
                        lot_no=lot_no_cancel,
                        warehouse_from=wh_3000,
                        warehouse_to=wh_3200,
                        result_stock=tgt_3200.quantity,
                        actor=request.user,
                        remark=f"[{label_kind_c} 라벨 투입취소] {rm_label.label_id} (LOT: {lot_disp})"
                    )

                # ERP 역방향 재고이동 등록 (atomic 밖)
                erp_msg_c = ''
                try:
                    from .erp_api import register_erp_stock_move
                    ok_erp, erp_no, err_erp = register_erp_stock_move(
                        cancel_trx, cancel_qty, wh_3000.code, wh_3200.code
                    )
                    if not ok_erp:
                        erp_msg_c = f' (ERP 연동 실패: {err_erp})'
                        logger.warning(f'RM 취소 ERP 이동 실패: {err_erp}')
                except Exception as _e:
                    erp_msg_c = f' (ERP 연동 예외)'
                    logger.warning(f'RM 취소 ERP 이동 예외: {_e}')

                return JsonResponse({
                    'success': True,
                    'is_first_scan': True,
                    'message': f'투입 취소 완료 ({cancel_qty}) - 라벨 재사용 가능' + erp_msg_c,
                    'tag_info': {
                        'tag_id': rm_label.label_id,
                        'part_no': rm_label.part_no,
                        'part_name': rm_label.part_name,
                        'quantity': cancel_qty,
                        'lot_no': str(lot_no_cancel) if lot_no_cancel else '',
                        'status': '재고',
                        'cancelled': True,
                    }
                })
            elif rm_label.status == 'CANCELLED':
                return JsonResponse({
                    'success': False,
                    'error': f'취소된 라벨입니다. ({rm_label.label_id})',
                })

            # 3200 원재료창고 (출발) / 도착 창고 조회
            wh_from = Warehouse.objects.filter(code='3200').first()
            wh_to = Warehouse.objects.filter(code=target_warehouse_code).first()
            if not wh_from:
                return JsonResponse({
                    'success': False,
                    'error': '3200 원재료창고가 등록되어 있지 않습니다.',
                })
            if not wh_to:
                return JsonResponse({
                    'success': False,
                    'error': f'도착 창고({target_warehouse_code})가 등록되어 있지 않습니다.',
                })

            # 현장(3000)으로 이동할 때만 라벨 USED 처리
            mark_used = (target_warehouse_code == '3000')

            # FIFO 검증 (3200 창고에서 같은 품번 더 오래된 LOT가 남아있으면 차단)
            if rm_label.lot_no:
                older = MaterialStock.objects.filter(
                    warehouse=wh_from, part=rm_label.part,
                    lot_no__lt=rm_label.lot_no, quantity__gt=0
                ).order_by('lot_no').first()
                if older:
                    return JsonResponse({
                        'success': False,
                        'error': f'[FIFO 위반] 더 오래된 LOT {older.lot_no} 재고({older.quantity})가 남아있습니다. 먼저 소진해주세요.',
                        'tag_info': {
                            'tag_id': rm_label.label_id,
                            'part_no': rm_label.part_no,
                            'lot_no': str(rm_label.lot_no),
                            'quantity': float(rm_label.quantity),
                        }
                    })

            # 3200 창고의 해당 LOT 재고 조회 (3200은 LOT 필수 창고)
            from django.db.models import F as _F
            lot_no = rm_label.lot_no

            # 라벨에 LOT가 없으면 차단
            if not lot_no:
                return JsonResponse({
                    'success': False,
                    'error': f'라벨에 LOT 정보가 없습니다. ({rm_label.label_id})',
                })

            src_stock = MaterialStock.objects.filter(
                warehouse=wh_from, part=rm_label.part, lot_no=lot_no
            ).first()

            available_lot = src_stock.quantity if src_stock else 0

            # LOT 일치 재고가 없으면 라벨 사용 불가
            if available_lot <= 0:
                # 같은 품번의 다른 LOT 재고 조회
                other_lots = MaterialStock.objects.filter(
                    warehouse=wh_from, part=rm_label.part, lot_no__isnull=False, quantity__gt=0
                ).exclude(lot_no=lot_no).values_list('lot_no', flat=True)[:3]

                if src_stock is not None:
                    # 해당 LOT 레코드는 있지만 수량 0 → 재고 없음
                    err_msg = f'[재고 없음] 라벨 LOT({lot_no})의 3200 창고 재고가 0입니다.'
                else:
                    # 해당 LOT 레코드 자체가 없음 → LOT 불일치
                    err_msg = f'[LOT 불일치] 라벨 LOT({lot_no}) 재고가 3200 창고에 없습니다.'
                if other_lots:
                    err_msg += f' (창고 내 다른 LOT: {", ".join(str(l) for l in other_lots)})'
                return JsonResponse({
                    'success': False,
                    'error': err_msg,
                    'tag_info': {
                        'tag_id': rm_label.label_id,
                        'part_no': rm_label.part_no,
                        'lot_no': str(lot_no),
                        'quantity': float(rm_label.quantity),
                    }
                })

            # NULL LOT은 더 이상 사용 안 함 (3200은 LOT 필수)
            null_stock = None
            available_null = 0
            available_total = max(available_lot, 0)

            label_qty = float(rm_label.quantity)

            # 사용자가 조정 수량을 입력하지 않았는데 재고 부족 → 조정 입력 요청
            if override_qty is None and available_total < label_qty:
                return JsonResponse({
                    'success': False,
                    'need_adjust': True,
                    'available_qty': float(available_total),
                    'label_qty': label_qty,
                    'error': f'재고 부족: 라벨 수량 {label_qty} > 재고 {available_total}',
                    'tag_info': {
                        'tag_id': rm_label.label_id,
                        'part_no': rm_label.part_no,
                        'part_name': rm_label.part_name,
                        'quantity': label_qty,
                        'lot_no': str(lot_no) if lot_no else '',
                    }
                })

            # 실제 이동할 수량 결정
            if override_qty is not None:
                try:
                    move_qty = float(override_qty)
                except (ValueError, TypeError):
                    return JsonResponse({'success': False, 'error': '이동 수량이 올바르지 않습니다.'})
                if move_qty <= 0:
                    return JsonResponse({'success': False, 'error': '이동 수량은 0보다 커야 합니다.'})
                if move_qty > available_total:
                    return JsonResponse({'success': False, 'error': f'이동 수량({move_qty})이 재고({available_total})보다 큽니다.'})
            else:
                move_qty = label_qty

            # 트랜잭션: 3200 차감 → 3000 추가 → 라벨 USED → 이력 생성
            with transaction.atomic():
                remaining = move_qty
                # LOT 재고 우선 차감
                if lot_no and src_stock and src_stock.quantity > 0:
                    deduct_lot = min(src_stock.quantity, remaining)
                    MaterialStock.objects.filter(pk=src_stock.pk).update(
                        quantity=_F('quantity') - deduct_lot
                    )
                    remaining -= deduct_lot
                # NULL LOT 차감
                if remaining > 0 and null_stock and null_stock.quantity > 0:
                    deduct_null = min(null_stock.quantity, remaining)
                    MaterialStock.objects.filter(pk=null_stock.pk).update(
                        quantity=_F('quantity') - deduct_null
                    )
                    remaining -= deduct_null
                # NULL LOT 부족분도 강제 차감 (마이너스 가능)
                if remaining > 0:
                    if null_stock:
                        MaterialStock.objects.filter(pk=null_stock.pk).update(
                            quantity=_F('quantity') - remaining
                        )
                    else:
                        MaterialStock.objects.create(
                            warehouse=wh_from, part=rm_label.part, lot_no=None, quantity=-remaining
                        )

                # 3000 창고에 추가 (LOT 유지)
                tgt_stock, _created = MaterialStock.objects.get_or_create(
                    warehouse=wh_to, part=rm_label.part, lot_no=lot_no,
                    defaults={'quantity': 0}
                )
                MaterialStock.objects.filter(pk=tgt_stock.pk).update(
                    quantity=_F('quantity') + move_qty
                )
                tgt_stock.refresh_from_db()

                # 현장(3000)으로 이동할 때만 라벨 USED 처리
                if mark_used:
                    rm_label.status = 'USED'
                    rm_label.used_at = timezone.now()
                    rm_label.used_by = request.user
                    rm_label.save(update_fields=['status', 'used_at', 'used_by'])

                # 이력 생성
                trx_no = f"RM-SCAN-{timezone.now().strftime('%y%m%d%H%M%S%f')}-{request.user.id}"
                lot_display = lot_no.strftime('%Y-%m-%d') if lot_no else 'NO LOT'
                action_label = '스캔 투입' if mark_used else '재고 이동'
                trx_for_erp = MaterialTransaction.objects.create(
                    transaction_no=trx_no,
                    transaction_type='TRANSFER',
                    date=timezone.now(),
                    part=rm_label.part,
                    quantity=move_qty,
                    lot_no=lot_no,
                    warehouse_from=wh_from,
                    warehouse_to=wh_to,
                    result_stock=tgt_stock.quantity,
                    actor=request.user,
                    remark=f"[{label_kind} 라벨 {action_label}] {rm_label.label_id} (LOT: {lot_display})"
                )

            # ERP 재고이동 등록 (atomic 밖 - ERP 실패가 SCM 롤백하지 않도록)
            erp_msg = ''
            try:
                from .erp_api import register_erp_stock_move
                ok_erp, erp_no, err_erp = register_erp_stock_move(
                    trx_for_erp, move_qty, wh_from.code, wh_to.code
                )
                if not ok_erp:
                    erp_msg = f' (ERP 연동 실패: {err_erp})'
                    logger.warning(f'RM 스캔 ERP 이동 실패: {err_erp}')
            except Exception as _e:
                erp_msg = f' (ERP 연동 예외)'
                logger.warning(f'RM 스캔 ERP 이동 예외: {_e}')

            msg = f'{wh_to.name}(으)로 이동 완료 ({move_qty})' + erp_msg
            return JsonResponse({
                'success': True,
                'is_first_scan': True,
                'message': msg,
                'tag_info': {
                    'tag_id': rm_label.label_id,
                    'part_no': rm_label.part_no,
                    'part_name': rm_label.part_name,
                    'quantity': move_qty,
                    'lot_no': str(lot_no) if lot_no else '',
                    'status': rm_label.get_status_display(),
                    'scan_count': 1,
                    'printed_at': timezone.localtime(rm_label.printed_at).strftime('%Y-%m-%d %H:%M') if rm_label.printed_at else '',
                    'used_at': timezone.now().strftime('%Y-%m-%d %H:%M'),
                    'target_warehouse': wh_to.code,
                }
            })

        # 태그 조회 (ProcessTag) - RM/PLT와 동일 로직으로 처리
        tag = ProcessTag.objects.filter(tag_id=tag_id).first()

        if not tag:
            return JsonResponse({
                'success': False,
                'error': f'등록되지 않은 태그입니다: {tag_id}',
                'is_registered': False
            })

        if tag.status == 'CANCELLED':
            return JsonResponse({
                'success': False,
                'error': f'취소된 태그입니다. ({tag.tag_id})',
            })

        # 3200 원재료창고 / 도착 창고 조회
        wh_from_tag = Warehouse.objects.filter(code='3200').first()
        wh_to_tag = Warehouse.objects.filter(code=target_warehouse_code).first()
        if not wh_from_tag:
            return JsonResponse({
                'success': False,
                'error': '3200 원재료창고가 등록되어 있지 않습니다.',
            })
        if not wh_to_tag:
            return JsonResponse({
                'success': False,
                'error': f'도착 창고({target_warehouse_code})가 등록되어 있지 않습니다.',
            })

        # USED 상태 → 투입 취소 (3000→3200 복구)
        if tag.status == 'USED':
            # 원래 투입처 조회 (이전 이동 이력에서 도착 창고 확인)
            last_trx_t = MaterialTransaction.objects.filter(
                part=tag.part,
                transaction_type='TRANSFER',
                warehouse_from__code='3200',
                remark__icontains=tag.tag_id,
            ).order_by('-date').first()

            if not last_trx_t:
                return JsonResponse({
                    'success': False,
                    'error': f'이동 이력을 찾을 수 없습니다. ({tag.tag_id})',
                })

            cancel_qty_t = float(last_trx_t.quantity)
            lot_no_tag = tag.lot_no
            wh_to_cancel = last_trx_t.warehouse_to  # 원래 투입처

            from django.db.models import F as _F
            with transaction.atomic():
                # 원래 투입처 재고 차감
                src_t = MaterialStock.objects.filter(
                    warehouse=wh_to_cancel, part=tag.part, lot_no=lot_no_tag
                ).first()
                if src_t:
                    MaterialStock.objects.filter(pk=src_t.pk).update(
                        quantity=_F('quantity') - cancel_qty_t
                    )
                else:
                    MaterialStock.objects.create(
                        warehouse=wh_to_cancel, part=tag.part, lot_no=lot_no_tag,
                        quantity=-cancel_qty_t
                    )

                # 3200 재고 복구
                tgt_t, _c = MaterialStock.objects.get_or_create(
                    warehouse=wh_from_tag, part=tag.part, lot_no=lot_no_tag,
                    defaults={'quantity': 0}
                )
                MaterialStock.objects.filter(pk=tgt_t.pk).update(
                    quantity=_F('quantity') + cancel_qty_t
                )
                tgt_t.refresh_from_db()

                # 태그 USED → PRINTED 복구
                tag.status = 'PRINTED'
                tag.used_at = None
                tag.used_by = None
                tag.used_warehouse = None
                tag.scan_count = 0
                tag.save(update_fields=['status', 'used_at', 'used_by', 'used_warehouse', 'scan_count'])

                # 취소 이력
                trx_no_t = f"TAG-CANCEL-{timezone.now().strftime('%y%m%d%H%M%S%f')}-{request.user.id}"
                lot_disp_t = lot_no_tag.strftime('%Y-%m-%d') if lot_no_tag else 'NO LOT'
                cancel_trx_t = MaterialTransaction.objects.create(
                    transaction_no=trx_no_t,
                    transaction_type='TRANSFER',
                    date=timezone.now(),
                    part=tag.part,
                    quantity=cancel_qty_t,
                    lot_no=lot_no_tag,
                    warehouse_from=wh_to_cancel,
                    warehouse_to=wh_from_tag,
                    result_stock=tgt_t.quantity,
                    actor=request.user,
                    remark=f"[현품표 투입취소] {tag.tag_id} (LOT: {lot_disp_t})"
                )

            # ERP 역방향
            erp_msg_tc = ''
            try:
                from .erp_api import register_erp_stock_move
                ok_erp, erp_no, err_erp = register_erp_stock_move(
                    cancel_trx_t, cancel_qty_t, wh_to_cancel.code, wh_from_tag.code
                )
                if not ok_erp:
                    erp_msg_tc = f' (ERP 연동 실패: {err_erp})'
            except Exception as _e:
                erp_msg_tc = f' (ERP 연동 예외)'
                logger.warning(f'TAG 취소 ERP 예외: {_e}')

            return JsonResponse({
                'success': True,
                'is_first_scan': True,
                'message': f'투입 취소 완료 ({cancel_qty_t}) - 라벨 재사용 가능' + erp_msg_tc,
                'tag_info': {
                    'tag_id': tag.tag_id,
                    'part_no': tag.part_no,
                    'part_name': tag.part_name,
                    'quantity': cancel_qty_t,
                    'lot_no': str(lot_no_tag) if lot_no_tag else '',
                    'status': 'PRINTED',
                    'cancelled': True,
                }
            })

        # PRINTED 상태 → 3200→도착 이동
        # FIFO 검증 (3200 창고 기준)
        if tag.lot_no:
            older = MaterialStock.objects.filter(
                warehouse=wh_from_tag, part=tag.part,
                lot_no__lt=tag.lot_no, quantity__gt=0
            ).order_by('lot_no').first()
            if older:
                return JsonResponse({
                    'success': False,
                    'error': f'[FIFO 위반] 더 오래된 LOT {older.lot_no} 재고({older.quantity})가 남아있습니다. 먼저 소진해주세요.',
                    'tag_info': {
                        'tag_id': tag.tag_id,
                        'part_no': tag.part_no,
                        'lot_no': str(tag.lot_no),
                        'quantity': int(tag.quantity),
                    }
                })

        # 재고 조회
        lot_no_t = tag.lot_no
        from django.db.models import F as _F

        # 태그에 LOT가 없으면 차단
        if not lot_no_t:
            return JsonResponse({
                'success': False,
                'error': f'태그에 LOT 정보가 없습니다. ({tag.tag_id})',
            })

        src_stock_t = MaterialStock.objects.filter(
            warehouse=wh_from_tag, part=tag.part, lot_no=lot_no_t
        ).first()

        avail_lot = src_stock_t.quantity if src_stock_t else 0

        # LOT 일치 재고가 없으면 태그 사용 불가
        if avail_lot <= 0:
            other_lots = MaterialStock.objects.filter(
                warehouse=wh_from_tag, part=tag.part, lot_no__isnull=False, quantity__gt=0
            ).exclude(lot_no=lot_no_t).values_list('lot_no', flat=True)[:3]

            if src_stock_t is not None:
                err_msg_t = f'[재고 없음] 태그 LOT({lot_no_t})의 3200 창고 재고가 0입니다.'
            else:
                err_msg_t = f'[LOT 불일치] 태그 LOT({lot_no_t}) 재고가 3200 창고에 없습니다.'
            if other_lots:
                err_msg_t += f' (창고 내 다른 LOT: {", ".join(str(l) for l in other_lots)})'
            return JsonResponse({
                'success': False,
                'error': err_msg_t,
                'tag_info': {
                    'tag_id': tag.tag_id,
                    'part_no': tag.part_no,
                    'lot_no': str(lot_no_t),
                    'quantity': int(tag.quantity),
                }
            })

        # NULL LOT은 더 이상 사용 안 함 (3200은 LOT 필수)
        null_stock_t = None
        available_total_t = max(avail_lot, 0)

        tag_qty_t = float(tag.quantity)

        # 부족 시 조정 수량 요청
        if override_qty is None and available_total_t < tag_qty_t:
            return JsonResponse({
                'success': False,
                'need_adjust': True,
                'available_qty': float(available_total_t),
                'label_qty': tag_qty_t,
                'error': f'재고 부족: 태그 수량 {tag_qty_t} > 재고 {available_total_t}',
                'tag_info': {
                    'tag_id': tag.tag_id,
                    'part_no': tag.part_no,
                    'part_name': tag.part_name,
                    'quantity': tag_qty_t,
                    'lot_no': str(lot_no_t) if lot_no_t else '',
                }
            })

        if override_qty is not None:
            try:
                move_qty_t = float(override_qty)
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'error': '이동 수량이 올바르지 않습니다.'})
            if move_qty_t <= 0:
                return JsonResponse({'success': False, 'error': '이동 수량은 0보다 커야 합니다.'})
            if move_qty_t > available_total_t:
                return JsonResponse({'success': False, 'error': f'이동 수량({move_qty_t})이 재고({available_total_t})보다 큽니다.'})
        else:
            move_qty_t = tag_qty_t

        # 현장(3000)으로 이동할 때만 USED 처리
        mark_used_t = (target_warehouse_code == '3000')

        with transaction.atomic():
            remaining = move_qty_t
            if lot_no_t and src_stock_t and src_stock_t.quantity > 0:
                deduct_lot = min(src_stock_t.quantity, remaining)
                MaterialStock.objects.filter(pk=src_stock_t.pk).update(
                    quantity=_F('quantity') - deduct_lot
                )
                remaining -= deduct_lot
            if remaining > 0 and null_stock_t and null_stock_t.quantity > 0:
                deduct_null = min(null_stock_t.quantity, remaining)
                MaterialStock.objects.filter(pk=null_stock_t.pk).update(
                    quantity=_F('quantity') - deduct_null
                )
                remaining -= deduct_null
            if remaining > 0:
                if null_stock_t:
                    MaterialStock.objects.filter(pk=null_stock_t.pk).update(
                        quantity=_F('quantity') - remaining
                    )
                else:
                    MaterialStock.objects.create(
                        warehouse=wh_from_tag, part=tag.part, lot_no=None, quantity=-remaining
                    )

            # 도착창고 재고 추가
            tgt_stock_t, _c = MaterialStock.objects.get_or_create(
                warehouse=wh_to_tag, part=tag.part, lot_no=lot_no_t,
                defaults={'quantity': 0}
            )
            MaterialStock.objects.filter(pk=tgt_stock_t.pk).update(
                quantity=_F('quantity') + move_qty_t
            )
            tgt_stock_t.refresh_from_db()

            # 현장(3000) 이동 시에만 USED 처리
            if mark_used_t:
                tag.status = 'USED'
                tag.used_at = timezone.now()
                tag.used_by = request.user
                tag.used_warehouse = wh_to_tag
                tag.scan_count = (tag.scan_count or 0) + 1
                tag.save(update_fields=['status', 'used_at', 'used_by', 'used_warehouse', 'scan_count'])

            # 이력 생성
            trx_no_t = f"TAG-SCAN-{timezone.now().strftime('%y%m%d%H%M%S%f')}-{request.user.id}"
            lot_display_t = lot_no_t.strftime('%Y-%m-%d') if lot_no_t else 'NO LOT'
            action_label_t = '스캔 투입' if mark_used_t else '재고 이동'
            trx_for_erp_t = MaterialTransaction.objects.create(
                transaction_no=trx_no_t,
                transaction_type='TRANSFER',
                date=timezone.now(),
                part=tag.part,
                quantity=move_qty_t,
                lot_no=lot_no_t,
                warehouse_from=wh_from_tag,
                warehouse_to=wh_to_tag,
                result_stock=tgt_stock_t.quantity,
                actor=request.user,
                remark=f"[현품표 {action_label_t}] {tag.tag_id} (LOT: {lot_display_t})"
            )

            # 스캔 로그
            ProcessTagScanLog.objects.create(
                tag=tag,
                scanned_by=request.user,
                warehouse=wh_to_tag,
                is_first_scan=mark_used_t,
                remark='' if mark_used_t else f'{wh_to_tag.code} 이동'
            )

        # ERP 연동
        erp_msg_t = ''
        try:
            from .erp_api import register_erp_stock_move
            ok_erp, erp_no, err_erp = register_erp_stock_move(
                trx_for_erp_t, move_qty_t, wh_from_tag.code, wh_to_tag.code
            )
            if not ok_erp:
                erp_msg_t = f' (ERP 연동 실패: {err_erp})'
        except Exception as _e:
            erp_msg_t = f' (ERP 연동 예외)'
            logger.warning(f'TAG 스캔 ERP 예외: {_e}')

        msg_t = f'{wh_to_tag.name}(으)로 이동 완료 ({move_qty_t})' + erp_msg_t
        return JsonResponse({
            'success': True,
            'is_first_scan': mark_used_t,
            'message': msg_t,
            'tag_info': {
                'tag_id': tag.tag_id,
                'part_no': tag.part_no,
                'part_name': tag.part_name,
                'quantity': move_qty_t,
                'lot_no': str(lot_no_t) if lot_no_t else '',
                'status': tag.get_status_display(),
                'scan_count': tag.scan_count,
                'printed_at': timezone.localtime(tag.printed_at).strftime('%Y-%m-%d %H:%M') if tag.printed_at else '',
                'used_at': timezone.now().strftime('%Y-%m-%d %H:%M'),
                'target_warehouse': wh_to_tag.code,
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': '잘못된 요청 형식입니다.'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'처리 중 오류 발생: {str(e)}'
        }, status=500)


@wms_permission_required('can_wms_stock_view')
def api_process_tag_info(request, tag_id):
    """
    현품표 정보 조회 API (GET)
    - 태그 상태 및 스캔 이력 조회
    """
    from .models import ProcessTag, ProcessTagScanLog

    # QR 데이터 형식: TAG_ID|품번|수량|LOT → 첫 번째 필드만 추출
    tag_id = tag_id.split('|')[0].strip() if '|' in tag_id else tag_id
    tag = ProcessTag.objects.filter(tag_id=tag_id).first()

    if not tag:
        return JsonResponse({
            'success': False,
            'error': f'등록되지 않은 태그입니다: {tag_id}'
        }, status=404)

    # 스캔 이력 조회 (최근 10건)
    scan_logs = tag.scan_logs.select_related('scanned_by', 'warehouse')[:10]
    logs_data = [{
        'scanned_at': timezone.localtime(log.scanned_at).strftime('%Y-%m-%d %H:%M:%S'),
        'scanned_by': log.scanned_by.username if log.scanned_by else '-',
        'warehouse': log.warehouse.name if log.warehouse else '-',
        'is_first_scan': log.is_first_scan,
    } for log in scan_logs]

    return JsonResponse({
        'success': True,
        'tag_info': {
            'tag_id': tag.tag_id,
            'part_no': tag.part_no,
            'part_name': tag.part_name,
            'quantity': tag.quantity,
            'lot_no': str(tag.lot_no) if tag.lot_no else '',
            'status': tag.get_status_display(),
            'scan_count': tag.scan_count,
            'printed_at': timezone.localtime(tag.printed_at).strftime('%Y-%m-%d %H:%M'),
            'printed_by': tag.printed_by.username if tag.printed_by else '-',
            'used_at': timezone.localtime(tag.used_at).strftime('%Y-%m-%d %H:%M') if tag.used_at else None,
            'used_by': tag.used_by.username if tag.used_by else '-',
        },
        'scan_logs': logs_data
    })


@wms_permission_required('can_wms_stock_view')
def api_scan_history_by_part(request):
    """
    품번별 투입(스캔) 이력 조회 API - 현장(3000) 투입 건만 표시
    - 원재료 레이아웃에서 랙 셀 클릭 시 사용
    GET ?part_no=11630-06360
    """
    from .models import ProcessTag, RawMaterialLabel

    part_no = request.GET.get('part_no', '').strip()
    # part_no 비어있으면 전체 품목 조회
    if not part_no:
        # 전체 모드: 모든 USED 라벨/태그 (최근 50건)
        tags_all = ProcessTag.objects.filter(
            status='USED', used_warehouse__code='3000'
        ).select_related('used_by').order_by('-used_at')[:50]
        items_all = []
        for t in tags_all:
            items_all.append({
                'tag_id': t.tag_id,
                'part_no': t.part_no,
                'part_name': t.part_name,
                'lot_no': str(t.lot_no) if t.lot_no else '-',
                'quantity': t.quantity,
                'used_at': timezone.localtime(t.used_at).strftime('%Y-%m-%d %H:%M') if t.used_at else '-',
                'used_by': t.used_by.username if t.used_by else '-',
                'stock_reflected': t.stock_reflected,
            })
        used_labels_all = RawMaterialLabel.objects.filter(
            status='USED'
        ).select_related('used_by').order_by('-used_at')[:50]
        for lbl in used_labels_all:
            items_all.append({
                'tag_id': lbl.label_id,
                'part_no': lbl.part_no,
                'part_name': lbl.part_name,
                'lot_no': lbl.lot_no.strftime('%Y-%m-%d') if lbl.lot_no else '-',
                'quantity': float(lbl.quantity),
                'used_at': timezone.localtime(lbl.used_at).strftime('%Y-%m-%d %H:%M') if lbl.used_at else '-',
                'used_by': lbl.used_by.username if lbl.used_by else '-',
                'stock_reflected': False,
            })
        items_all.sort(key=lambda x: x['used_at'], reverse=True)
        return JsonResponse({'success': True, 'part_no': '', 'items': items_all[:50]})

    items = []

    # 1) ProcessTag - 현재 USED 상태만 표시 (취소된 건 자동 제외)
    tags = ProcessTag.objects.filter(
        part_no=part_no, status='USED', used_warehouse__code='3000'
    ).select_related('used_by').order_by('-used_at')[:30]
    for t in tags:
        items.append({
            'tag_id': t.tag_id,
            'lot_no': str(t.lot_no) if t.lot_no else '-',
            'quantity': t.quantity,
            'used_at': timezone.localtime(t.used_at).strftime('%Y-%m-%d %H:%M') if t.used_at else '-',
            'used_by': t.used_by.username if t.used_by else '-',
            'stock_reflected': t.stock_reflected,
        })

    # 2) RM/PLT 라벨 - 현재 USED 상태만 표시 (취소되어 INSTOCK 복구된 건 자동 제외)
    used_labels = RawMaterialLabel.objects.filter(
        part_no=part_no, status='USED'
    ).select_related('used_by', 'part').order_by('-used_at')[:30]
    for lbl in used_labels:
        items.append({
            'tag_id': lbl.label_id,
            'lot_no': lbl.lot_no.strftime('%Y-%m-%d') if lbl.lot_no else '-',
            'quantity': float(lbl.quantity),
            'used_at': timezone.localtime(lbl.used_at).strftime('%Y-%m-%d %H:%M') if lbl.used_at else '-',
            'used_by': lbl.used_by.username if lbl.used_by else '-',
            'stock_reflected': False,
        })

    # 시간 역순 정렬 후 30건 제한
    items.sort(key=lambda x: x['used_at'], reverse=True)
    items = items[:30]

    return JsonResponse({'success': True, 'part_no': part_no, 'items': items})


@wms_permission_required('can_wms_stock_view')
def api_process_tag_cancel_scan(request):
    """
    투입(스캔) 취소 API - 태그를 PRINTED 상태로 되돌려 재스캔 가능하게 함
    - stock_reflected=True (이미 출고 반영됨)인 경우 취소 불가
    """
    from .models import ProcessTag, ProcessTagScanLog

    try:
        data = json.loads(request.body)
        tag_id = data.get('tag_id', '').strip()

        if not tag_id:
            return JsonResponse({'success': False, 'error': '태그 ID가 누락되었습니다.'}, status=400)

        tag = ProcessTag.objects.filter(tag_id=tag_id).first()
        if not tag:
            return JsonResponse({'success': False, 'error': f'등록되지 않은 태그: {tag_id}'}, status=404)

        if tag.status != 'USED':
            return JsonResponse({'success': False, 'error': f'투입 상태가 아닙니다. (현재: {tag.get_status_display()})'})

        if tag.stock_reflected:
            return JsonResponse({'success': False, 'error': '이미 출고 처리된 태그는 취소할 수 없습니다.'})

        # 투입 취소: PRINTED로 되돌리기
        tag.status = 'PRINTED'
        tag.used_at = None
        tag.used_by = None
        tag.used_warehouse = None
        tag.scan_count = 0
        tag.save()

        # 취소 로그 기록
        ProcessTagScanLog.objects.create(
            tag=tag,
            scanned_by=request.user,
            is_first_scan=False,
            remark=f'투입 취소 ({request.user.username})'
        )

        return JsonResponse({'success': True, 'message': f'{tag_id} 투입이 취소되었습니다.'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# 4. 기타 메뉴
# =============================================================================

@wms_permission_required('can_wms_stock_view')
def transaction_history(request):
    """
    품번별 수불 이력 - 품번 선택 시 해당 품번의 모든 이동 이력을 시간순으로 표시
    입고 -> 창고이동 -> 출고까지 전체 흐름 추적
    """
    from django.core.paginator import Paginator
    from orders.models import Part
    from datetime import datetime

    # ===== 필터 파라미터 =====
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    part_no = request.GET.get('part_no', '').strip()
    part_group = request.GET.get('part_group', '')
    lot_no = request.GET.get('lot_no', '').strip()
    q = request.GET.get('q', '').strip()

    # 기본값: 이번 달
    if not start_date:
        today = timezone.localtime().date()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = timezone.localtime().date().strftime('%Y-%m-%d')

    start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()

    # ===== 품목군 목록 =====
    part_groups = Part.objects.values_list('part_group', flat=True).distinct().order_by('part_group')

    # ===== 품번 목록 (트랜잭션이 있는 품번만) =====
    part_ids_with_trx = MaterialTransaction.objects.values_list('part_id', flat=True).distinct()
    parts_list = Part.objects.filter(id__in=part_ids_with_trx).order_by('part_no')

    if part_group:
        parts_list = parts_list.filter(part_group=part_group)
    if q:
        parts_list = parts_list.filter(Q(part_no__icontains=q) | Q(part_name__icontains=q))

    # ===== 선택된 품번의 수불 이력 조회 =====
    transactions = []
    selected_part = None
    lot_list = []  # 해당 품번의 LOT 목록

    if part_no:
        selected_part = Part.objects.filter(part_no=part_no).first()
        if selected_part:
            # 해당 품번의 LOT 목록 조회
            lot_list = MaterialTransaction.objects.filter(
                part=selected_part,
                lot_no__isnull=False
            ).values_list('lot_no', flat=True).distinct().order_by('-lot_no')

            # 트랜잭션 조회 (내부 재고 보정 제외)
            trx_qs = MaterialTransaction.objects.filter(
                part=selected_part,
                date__date__gte=start_dt,
                date__date__lte=end_dt
            ).exclude(
                transaction_type__in=['ADJ_ERP_IN', 'ADJ_ERP_OUT']
            )

            # LOT 필터 적용
            if lot_no:
                try:
                    lot_date = datetime.strptime(lot_no, '%Y-%m-%d').date()
                    trx_qs = trx_qs.filter(lot_no=lot_date)
                except ValueError:
                    pass

            transactions = trx_qs.select_related(
                'warehouse_from', 'warehouse_to', 'actor'
            ).order_by('date')

    # ===== 페이징 =====
    paginator = Paginator(transactions, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # ===== 합계 계산 =====
    total_in = sum(t.quantity for t in transactions if t.warehouse_to and not t.warehouse_from)
    total_out = sum(t.quantity for t in transactions if t.warehouse_from and not t.warehouse_to)
    total_move = sum(t.quantity for t in transactions if t.warehouse_from and t.warehouse_to)

    return render(request, 'material/transaction_history.html', {
        'page_obj': page_obj,
        'parts_list': parts_list,
        'part_groups': part_groups,
        'lot_list': lot_list,
        'start_date': start_date,
        'end_date': end_date,
        'part_no': part_no,
        'selected_part': selected_part,
        'part_group': part_group,
        'lot_no': lot_no,
        'q': q,
        'total_in': total_in,
        'total_out': total_out,
        'total_move': total_move,
        'result_count': len(transactions),
    })


@wms_permission_required('can_wms_stock_view')
def transaction_history_excel(request):
    """
    품번별 수불 이력 - Excel 다운로드
    """
    from datetime import datetime
    from orders.models import Part

    # ===== 필터 파라미터 =====
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    part_no = request.GET.get('part_no', '').strip()
    lot_no = request.GET.get('lot_no', '').strip()

    # 기본값: 이번 달
    if not start_date:
        today = timezone.localtime().date()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
    if not end_date:
        end_date = timezone.localtime().date().strftime('%Y-%m-%d')

    start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()

    # ===== 선택된 품번의 수불 이력 조회 =====
    selected_part = Part.objects.filter(part_no=part_no).first()
    transactions = []

    if selected_part:
        trx_qs = MaterialTransaction.objects.filter(
            part=selected_part,
            date__date__gte=start_dt,
            date__date__lte=end_dt
        )

        # LOT 필터 적용
        if lot_no:
            try:
                lot_date = datetime.strptime(lot_no, '%Y-%m-%d').date()
                trx_qs = trx_qs.filter(lot_no=lot_date)
            except ValueError:
                pass

        transactions = list(trx_qs.select_related(
            'warehouse_from', 'warehouse_to', 'actor'
        ).order_by('date'))

    # ===== Excel 파일 생성 =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "수불이력"

    # 헤더 스타일
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 제목 행
    ws.append([f"품번별 수불 이력 ({start_date} ~ {end_date})"])
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 품번 정보
    if selected_part:
        ws.append([f"품번: {selected_part.part_no} / 품명: {selected_part.part_name}"])
    else:
        ws.append(["품번 미선택"])
    ws.merge_cells('A2:H2')
    ws.append([])  # 빈 행

    # 헤더
    headers = ['No', '일시', '유형', 'LOT', '출발 창고', '도착 창고', '수량', '처리자']
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    for idx, trx in enumerate(transactions, 1):
        # 유형 결정
        if trx.warehouse_to and not trx.warehouse_from:
            trx_type = "입고"
        elif trx.warehouse_from and not trx.warehouse_to:
            trx_type = "출고"
        elif trx.warehouse_from and trx.warehouse_to:
            trx_type = "이동"
        else:
            trx_type = "기타"

        row = [
            idx,
            trx.date.strftime("%Y-%m-%d %H:%M"),
            trx_type,
            trx.lot_no.strftime("%Y-%m-%d") if trx.lot_no else "-",
            f"({trx.warehouse_from.code}) {trx.warehouse_from.name}" if trx.warehouse_from else "- 외부 입고 -",
            f"({trx.warehouse_to.code}) {trx.warehouse_to.name}" if trx.warehouse_to else "- 외부 출고 -",
            trx.quantity,
            trx.actor.username if trx.actor else "-"
        ]
        ws.append(row)
        for col_num in range(1, 9):
            cell = ws.cell(row=4 + idx, column=col_num)
            cell.border = thin_border
            if col_num == 7:
                cell.alignment = Alignment(horizontal='right')

    # 합계 행
    total_in = sum(t.quantity for t in transactions if t.warehouse_to and not t.warehouse_from)
    total_out = sum(t.quantity for t in transactions if t.warehouse_from and not t.warehouse_to)
    total_move = sum(t.quantity for t in transactions if t.warehouse_from and t.warehouse_to)

    total_row = 5 + len(transactions)
    ws.append([])  # 빈 행
    ws.append(['', '', '', '', '', '입고 합계:', total_in, ''])
    ws.append(['', '', '', '', '', '출고 합계:', total_out, ''])
    ws.append(['', '', '', '', '', '이동 합계:', total_move, ''])

    # 열 너비 조정
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    # 응답 반환
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"수불이력_{part_no}_{start_date}_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@wms_permission_required('can_wms_inout_edit')
def outbound_create(request):
    """생산 자재 불출"""
    return render(request, 'material/outbound_create.html')




@wms_permission_required('can_wms_stock_edit')
def stock_transfer(request):
    """
    [WMS] 재고 이동 처리
    - 기존 메뉴/URL(material:stock_transfer)을 유지하면서
      실제 처리는 stock_move로 위임
    """
    return stock_move(request)



@wms_permission_required('can_wms_stock_view')
def transfer_history(request):
    """재고 이동 현황 - 창고 간 이동 이력 조회"""
    from django.core.paginator import Paginator
    from django.db.models import Q
    from orders.models import Part

    # 이동(TRANSFER, TRF_ERP) 타입 조회
    qs = MaterialTransaction.objects.filter(
        transaction_type__in=['TRANSFER', 'TRF_ERP']
    ).select_related(
        'part', 'warehouse_from', 'warehouse_to', 'actor'
    ).order_by('-date', '-id')

    # ===== 필터 파라미터 =====
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    from_wh = request.GET.get('from_wh', '')
    to_wh = request.GET.get('to_wh', '')
    part_group = request.GET.get('part_group', '')
    q = request.GET.get('q', '').strip()

    # ===== 기간 필터 =====
    if start_date:
        qs = qs.filter(date__date__gte=start_date)
    if end_date:
        qs = qs.filter(date__date__lte=end_date)

    # ===== 창고 필터 =====
    if from_wh:
        qs = qs.filter(warehouse_from__code=from_wh)
    if to_wh:
        qs = qs.filter(warehouse_to__code=to_wh)

    # ===== 품목군 필터 =====
    if part_group:
        qs = qs.filter(part__part_group=part_group)

    # ===== 검색 (품번/품명) =====
    if q:
        qs = qs.filter(
            Q(part__part_no__icontains=q) |
            Q(part__part_name__icontains=q)
        )

    # ===== 페이징 =====
    paginator = Paginator(qs, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # ===== 창고 목록 (필터용) =====
    warehouses = Warehouse.objects.all().order_by('code')

    # ===== 품목군 목록 (필터용) =====
    part_groups = Part.objects.values_list('part_group', flat=True).distinct().order_by('part_group')

    return render(request, 'material/transfer_history.html', {
        'page_obj': page_obj,
        'warehouses': warehouses,
        'part_groups': part_groups,
        'start_date': start_date,
        'end_date': end_date,
        'from_wh': from_wh,
        'to_wh': to_wh,
        'part_group': part_group,
        'q': q,
    })


@wms_permission_required('can_wms_stock_view')
def stock_check(request):
    """[현장] 재고 실사 (수동)"""
    warehouses = Warehouse.objects.filter(is_active=True)
    return render(request, 'material/stock_check.html', {
        'warehouses': warehouses,
    })


@wms_permission_required('can_wms_stock_view')
def stock_check_result(request):
    """[관리] 재고 실사 결과"""
    warehouse_id = request.GET.get('warehouse')
    part_no = request.GET.get('part_no', '').strip()
    part_name = request.GET.get('part_name', '').strip()

    stocks = MaterialStock.objects.select_related('warehouse', 'part').filter(quantity__gt=0)

    if warehouse_id:
        stocks = stocks.filter(warehouse_id=warehouse_id)
    if part_no:
        stocks = stocks.filter(part__part_no__icontains=part_no)
    if part_name:
        stocks = stocks.filter(part__part_name__icontains=part_name)

    stocks = stocks.order_by('warehouse__code', 'part__part_no', 'lot_no')[:200]

    return render(request, 'material/stock_check_result.html', {
        'stocks': stocks,
    })


# =============================================================================
# 5. 불량 반품 및 반출증
# =============================================================================

@wms_permission_required('can_wms_stock_edit')
def stock_return(request):
    """
    [WMS] 불량 반품 처리
    """
    ng_wh = Warehouse.objects.filter(code='8200').first()
    if not ng_wh:
        ng_wh = Warehouse.objects.filter(name__contains='부적합').first()

    if request.method == 'POST':
        try:
            stock_id = request.POST.get('stock_id')
            return_qty = int(request.POST.get('quantity', 0))
            remark = request.POST.get('remark', '')

            with transaction.atomic():
                stock = get_object_or_404(MaterialStock, id=stock_id)

                if return_qty <= 0:
                    messages.error(request, "반품 수량은 1개 이상이어야 합니다.")
                    return redirect('material:stock_return')
                if stock.quantity < return_qty:
                    messages.error(request, "재고 수량보다 많이 반품할 수 없습니다.")
                    return redirect('material:stock_return')

                # (1) 재고 차감
                stock.quantity = F('quantity') - return_qty
                stock.save()

                # (2) 반품 이력 남기기
                trx_no = f"RET-{timezone.now().strftime('%y%m%d%H%M%S')}-{request.user.id}"
                trx = MaterialTransaction.objects.create(
                    transaction_no=trx_no,
                    transaction_type='OUT_RETURN',
                    date=timezone.now(),
                    part=stock.part,
                    quantity=return_qty,
                    warehouse_from=stock.warehouse,
                    vendor=stock.part.vendor,
                    actor=request.user,
                    remark=f"[반품] {remark}"
                )

            messages.success(request, f"[{stock.part.part_name}] {return_qty}개 반품 처리 완료.")
            return redirect('material:print_return_note', trx_id=trx.id)

        except Exception as e:
            messages.error(request, f"오류 발생: {str(e)}")
            return redirect('material:stock_return')

    ng_stocks = []
    if ng_wh:
        ng_stocks = MaterialStock.objects.filter(
            warehouse=ng_wh, quantity__gt=0
        ).select_related('part', 'part__vendor').order_by('part__part_no')

    return render(request, 'material/stock_return.html', {
        'ng_stocks': ng_stocks
    })


@wms_permission_required('can_wms_stock_view')
def print_return_note(request, trx_id):
    """
    [WMS] 반출증(Gate Pass) 인쇄 화면
    """
    trx = get_object_or_404(MaterialTransaction, pk=trx_id)
    return render(request, 'material/print_return_note.html', {
        'trx': trx,
        'today': timezone.now()
    })

# material/views.py 에 아래 함수를 추가/덮어쓰기 하세요.

# 기존 import 아래에 추가하거나, 필요한 import가 없다면 함께 추가하세요.
from django.db import transaction
from django.db.models import F
from .models import MaterialTransaction  # MaterialStock, Warehouse 등은 이미 import 되어 있을 것으로 예상됨

# material/views.py 맨 아래에 작성

@wms_permission_required('can_wms_stock_edit')
def stock_move(request):
    """
    [WMS] 재고 이동 처리 (다중 품목 일괄 처리 + LOT 기반)
    """

    # 1. 화면 보여주기 (GET)
    if request.method == 'GET':
        from django.core.paginator import Paginator
        from django.db.models import Q, Count

        # 전체 창고 조회
        warehouses = Warehouse.objects.all().order_by('code')

        # 이동 내역 조회
        hstart = request.GET.get('hstart', '')
        hend = request.GET.get('hend', '')
        hwh = request.GET.get('hwh', '')
        hq = request.GET.get('hq', '')

        history_qs = MaterialTransaction.objects.filter(
            transaction_type__in=['TRANSFER', 'TRF_ERP']
        ).select_related('part', 'warehouse_from', 'warehouse_to', 'actor').annotate(
            rm_label_count=Count('used_labels', distinct=True),
            tag_count=Count('used_tags', distinct=True),
        ).annotate(
            label_count=F('rm_label_count') + F('tag_count')
        ).order_by('-date')

        if hstart:
            history_qs = history_qs.filter(date__date__gte=hstart)
        if hend:
            history_qs = history_qs.filter(date__date__lte=hend)
        if hwh:
            history_qs = history_qs.filter(
                Q(warehouse_from__code=hwh) | Q(warehouse_to__code=hwh)
            )
        if hq:
            history_qs = history_qs.filter(
                Q(part__part_no__icontains=hq) | Q(part__part_name__icontains=hq) | Q(transaction_no__icontains=hq)
            )

        paginator = Paginator(history_qs, 20)
        page_num = request.GET.get('page', 1)
        history_page = paginator.get_page(page_num)

        production_codes = list(Warehouse.objects.filter(is_production=True).values_list('code', flat=True))

        context = {
            'warehouses': warehouses,
            'production_codes': production_codes,
            'today': timezone.now().strftime('%Y-%m-%d'),
            'history': history_page,
            'history_count': paginator.count,
            'hstart': hstart,
            'hend': hend,
            'hwh': hwh,
            'hq': hq,
        }
        return render(request, 'material/stock_move.html', context)

    # 2. 이동 처리 (POST) - LOT 기반 처리
    elif request.method == 'POST':
        part_nos = request.POST.getlist('part_no[]')
        stock_ids = request.POST.getlist('stock_id[]')  # LOT별 재고 ID
        from_locs = request.POST.getlist('from_loc[]')
        to_locs = request.POST.getlist('to_loc[]')
        move_qtys = request.POST.getlist('move_qty[]')
        label_ids_list = request.POST.getlist('label_ids[]')  # 라벨 선택 (JSON 문자열 리스트)
        fifo_reasons = request.POST.getlist('fifo_reason[]')  # FIFO 위반 사유

        # 이동 처리일 (사용자 선택, 미선택 시 오늘)
        from datetime import datetime as dt_cls
        transfer_date_str = request.POST.get('transfer_date', '').strip()
        if transfer_date_str:
            try:
                transfer_date = dt_cls.strptime(transfer_date_str, '%Y-%m-%d').date()
                transfer_datetime = timezone.make_aware(dt_cls.combine(transfer_date, timezone.localtime().time()))
            except ValueError:
                transfer_datetime = timezone.now()
        else:
            transfer_datetime = timezone.now()

        success_count = 0
        erp_pending = []  # ERP 연동 대기 목록 (atomic 밖에서 처리)

        try:
            with transaction.atomic():
                for i in range(len(part_nos)):
                    p_no = (part_nos[i] or "").strip()
                    stock_id = (stock_ids[i] or "").strip()
                    f_loc_val = (from_locs[i] or "").strip()
                    t_loc_val = (to_locs[i] or "").strip()

                    if not p_no or not stock_id or not f_loc_val or not t_loc_val:
                        continue

                    try:
                        qty = int(move_qtys[i])
                    except (ValueError, TypeError):
                        continue

                    if qty <= 0:
                        continue

                    if f_loc_val == t_loc_val:
                        continue

                    # 모델 조회
                    try:
                        part_obj = Part.objects.get(part_no=p_no)
                    except Part.DoesNotExist:
                        raise ValueError(f"존재하지 않는 품번입니다: [{p_no}]")

                    try:
                        from_wh = Warehouse.objects.get(code=f_loc_val)
                    except Warehouse.DoesNotExist:
                        raise ValueError(f"출발 창고를 찾을 수 없습니다: [{f_loc_val}]")
                    try:
                        to_wh = Warehouse.objects.get(code=t_loc_val)
                    except Warehouse.DoesNotExist:
                        raise ValueError(f"도착 창고를 찾을 수 없습니다: [{t_loc_val}]")

                    # 2-1. LOT별 재고(source_stock) 조회 및 차감
                    try:
                        source_stock = MaterialStock.objects.select_for_update().get(id=stock_id)
                    except MaterialStock.DoesNotExist:
                        raise ValueError(f"선택한 LOT의 재고를 찾을 수 없습니다.")

                    # 재고 검증
                    if source_stock.warehouse != from_wh:
                        raise ValueError(f"선택한 LOT가 출고 창고[{from_wh.name}]에 없습니다.")

                    if source_stock.part != part_obj:
                        raise ValueError(f"선택한 LOT가 품번[{p_no}]과 일치하지 않습니다.")

                    if int(source_stock.quantity) < qty:
                        raise ValueError(f"[{p_no}] LOT 재고 부족 (보유: {source_stock.quantity}, 요청: {qty})")

                    # LOT 정보 저장 (이동 시 동일 LOT 유지)
                    lot_no = source_stock.lot_no

                    # 출고 창고 재고 차감
                    MaterialStock.objects.filter(pk=source_stock.pk).update(
                        quantity=F('quantity') - qty
                    )

                    # 2-2. 받는 창고에 동일 LOT로 재고 증가
                    target_stock, _ = MaterialStock.objects.select_for_update().get_or_create(
                        warehouse=to_wh,
                        part=part_obj,
                        lot_no=lot_no,  # 동일 LOT 번호로 이동
                        defaults={'quantity': 0}
                    )

                    MaterialStock.objects.filter(pk=target_stock.pk).update(
                        quantity=F('quantity') + qty
                    )

                    # ✅ update(F()) 이후 실제 수량을 다시 읽어서 result_stock 숫자 저장
                    target_stock.refresh_from_db()

                    # FIFO 검증 및 사유 기록
                    fifo_remark = ""
                    if lot_no:
                        oldest_lot = MaterialStock.objects.filter(
                            warehouse=from_wh,
                            part=part_obj,
                            quantity__gt=0,
                            lot_no__isnull=False,
                        ).order_by('lot_no').first()

                        if oldest_lot and oldest_lot.lot_no < lot_no:
                            fifo_reason = fifo_reasons[i] if i < len(fifo_reasons) else ''
                            fifo_remark = f" [FIFO 위반: {fifo_reason or '사유 미입력'}]"

                    # 2-3. 이력(Transaction) 생성
                    trx_no = f"TRX-{timezone.now().strftime('%y%m%d%H%M%S%f')}-{request.user.id}-{i}"

                    lot_display = lot_no.strftime('%Y-%m-%d') if lot_no else 'NO LOT'

                    trx_obj = MaterialTransaction.objects.create(
                        transaction_no=trx_no,
                        transaction_type='TRANSFER',
                        date=transfer_datetime,
                        part=part_obj,
                        quantity=qty,
                        lot_no=lot_no,
                        warehouse_from=from_wh,
                        warehouse_to=to_wh,
                        result_stock=target_stock.quantity,
                        actor=request.user,
                        remark=f"재고이동 ({from_wh.name} -> {to_wh.name}) [LOT: {lot_display}]{fifo_remark}"
                    )

                    # 제조현장 이동 시 선택된 라벨 USED 처리 (RM + TAG)
                    if to_wh.is_production and i < len(label_ids_list):
                        import json
                        try:
                            raw_label_data = json.loads(label_ids_list[i]) if label_ids_list[i] else []
                        except (json.JSONDecodeError, TypeError):
                            raw_label_data = []

                        if raw_label_data:
                            from .models import ProcessTag
                            rm_ids = []
                            tag_ids = []
                            for item in raw_label_data:
                                if isinstance(item, dict):
                                    if item.get('type') == 'TAG':
                                        tag_ids.append(item['id'])
                                    else:
                                        rm_ids.append(item['id'])
                                else:
                                    rm_ids.append(item)  # 하위호환: 정수 = RM

                            if rm_ids:
                                RawMaterialLabel.objects.filter(
                                    id__in=rm_ids,
                                    part_no=p_no,
                                    lot_no=lot_no,
                                    status__in=['INSTOCK', 'PRINTED']
                                ).update(
                                    status='USED',
                                    used_at=transfer_datetime,
                                    used_by=request.user,
                                    used_transaction=trx_obj,
                                )

                            if tag_ids:
                                ProcessTag.objects.filter(
                                    id__in=tag_ids,
                                    part_no=p_no,
                                    lot_no=lot_no,
                                    status='PRINTED'
                                ).update(
                                    status='USED',
                                    used_at=transfer_datetime,
                                    used_by=request.user,
                                    used_warehouse=to_wh,
                                    used_transaction=trx_obj,
                                )

                    # ERP 연동 대기 목록에 추가
                    erp_pending.append({
                        'trx': trx_obj,
                        'qty': qty,
                        'from_code': from_wh.code,
                        'to_code': to_wh.code,
                    })

                    success_count += 1

            # ERP 재고이동 등록 (atomic 밖 - ERP 실패가 SCM 롤백하지 않도록)
            if erp_pending:
                from .erp_api import register_erp_stock_move
                erp_ok = 0
                erp_fail = 0
                erp_errors = []
                for item in erp_pending:
                    try:
                        ok, erp_no, err = register_erp_stock_move(
                            item['trx'], item['qty'], item['from_code'], item['to_code']
                        )
                        if ok:
                            erp_ok += 1
                        else:
                            erp_fail += 1
                            erp_errors.append(f"{item['trx'].part.part_no}: {err}")
                    except Exception as e:
                        logger.warning(f"ERP 재고이동 등록 예외: {e}")
                        erp_fail += 1
                        erp_errors.append(f"{item['trx'].part.part_no}: {e}")

                if erp_fail > 0:
                    err_detail = ' / '.join(erp_errors[:3])
                    messages.warning(request, f"ERP 연동: {erp_ok}건 성공, {erp_fail}건 실패 [{err_detail}] (SCM 재고이동은 정상 처리됨)")

            if success_count > 0:
                messages.success(request, f"총 {success_count}건 이동 완료되었습니다.")
            else:
                messages.warning(request, "처리된 내역이 없습니다.")

            return redirect('material:stock_move')

        except Exception as e:
            messages.error(request, f"오류 발생: {str(e)}")
            return redirect('material:stock_move')

@wms_permission_required('can_wms_stock_view')
def api_part_exists(request):
    part_no = (request.GET.get('part_no') or '').strip()
    exists = Part.objects.filter(part_no=part_no).exists()
    return JsonResponse({'exists': exists})


# =============================================================================
# 6. LOT 관리 - LOT별 재고 상세 조회 API
# =============================================================================
@wms_permission_required('can_wms_stock_view')
def get_lot_details(request, part_no):
    """
    특정 품목의 LOT별 재고 상세 정보를 JSON으로 반환 (WMS용)
    warehouse 파라미터가 있으면 해당 창고만 조회
    """
    try:
        part = Part.objects.filter(part_no=part_no).first()
        if not part:
            return JsonResponse({'error': '품목을 찾을 수 없습니다.'}, status=404)

        # MaterialStock에서 해당 품목의 전체 재고 합계 (NULL 포함)
        base_qs = MaterialStock.objects.filter(part=part).select_related('warehouse')

        # 창고 필터링 (warehouse 파라미터가 있으면)
        warehouse_code = request.GET.get('warehouse')
        if warehouse_code:
            base_qs = base_qs.filter(warehouse__code=warehouse_code)

        # 전체 재고 합계 (음수 NULL 포함)
        total_qty = base_qs.aggregate(total=Sum('quantity'))['total'] or 0

        # 표시용: quantity > 0인 LOT만 (NULL 제외, LOT만 보여줌)
        from django.db.models import F as _F
        lot_stocks = base_qs.filter(lot_no__isnull=False, quantity__gt=0).order_by(_F('lot_no').asc())

        lot_data = []
        lot_total = 0
        oldest_lot = None

        for stock in lot_stocks:
            days_old = (timezone.now().date() - stock.lot_no).days
            lot_info = {
                'warehouse': stock.warehouse.name,
                'warehouse_code': stock.warehouse.code,
                'lot_no': stock.lot_no.strftime('%Y-%m-%d'),
                'quantity': stock.quantity,
                'days_old': days_old,
                'is_null_lot': False,
            }
            lot_data.append(lot_info)
            lot_total += stock.quantity

            # 가장 오래된 LOT 추적 (FIFO 경고용)
            if oldest_lot is None or stock.lot_no < oldest_lot:
                oldest_lot = stock.lot_no

        # LOT에 배분 안 된 나머지 (ERP 관리분 - 마이너스 포함)
        unallocated = total_qty - lot_total
        if unallocated != 0:
            # NULL LOT 재고를 창고별로 표시
            null_stocks = base_qs.filter(lot_no__isnull=True).exclude(quantity=0)
            if null_stocks.exists():
                for ns in null_stocks:
                    lot_data.insert(0, {
                        'warehouse': ns.warehouse.name,
                        'warehouse_code': ns.warehouse.code,
                        'lot_no': 'ERP 재고',
                        'quantity': ns.quantity,
                        'days_old': 99999,
                        'is_null_lot': True,
                    })
            else:
                wh = base_qs.first().warehouse if base_qs.exists() else None
                lot_data.insert(0, {
                    'warehouse': wh.name if wh else '-',
                    'warehouse_code': wh.code if wh else '-',
                    'lot_no': 'ERP 재고',
                    'quantity': unallocated,
                    'days_old': 99999,
                    'is_null_lot': True,
                })

        # FIFO 경고 판정 (60일 이상 된 LOT가 있으면 경고)
        fifo_warning = False
        if oldest_lot:
            days_old = (timezone.now().date() - oldest_lot).days
            if days_old >= 60:
                fifo_warning = True

        return JsonResponse({
            'part_no': part.part_no,
            'part_name': part.part_name,
            'vendor_name': part.vendor.name if part.vendor else '-',
            'total_quantity': total_qty,
            'lot_details': lot_data,
            'fifo_warning': fifo_warning,
            'oldest_lot': oldest_lot.strftime('%Y-%m-%d') if oldest_lot else None,
            'oldest_days': (timezone.now().date() - oldest_lot).days if oldest_lot else 0,
            'weight_unit': part.weight_unit or 'EA',
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@wms_permission_required('can_wms_stock_view')
def api_get_available_lots(request):
    """
    [재고 이동용] 특정 품번 + 보내는 창고의 사용 가능한 LOT 목록 반환
    FIFO 순서(오래된 순)로 정렬하여 반환
    """
    try:
        part_no = request.GET.get('part_no', '').strip()
        warehouse_code = request.GET.get('warehouse_code', '').strip()

        if not part_no or not warehouse_code:
            return JsonResponse({'error': '품번과 창고 코드가 필요합니다.'}, status=400)

        # 품번 확인
        part = Part.objects.filter(part_no=part_no).first()
        if not part:
            return JsonResponse({'error': '존재하지 않는 품번입니다.'}, status=404)

        # 창고 확인
        warehouse = Warehouse.objects.filter(code=warehouse_code).first()
        if not warehouse:
            return JsonResponse({'error': '존재하지 않는 창고입니다.'}, status=404)

        # 해당 창고의 해당 품목 LOT별 재고 조회
        # 우선순위: lot_no=NULL(ERP 재고) 먼저 → 오래된 LOT 순 (FIFO)
        from django.db.models import F as _F
        lot_stocks = MaterialStock.objects.filter(
            warehouse=warehouse,
            part=part,
            quantity__gt=0
        ).order_by(_F('lot_no').asc(nulls_first=True))

        lots = []
        for stock in lot_stocks:
            if stock.lot_no:
                days_old = (timezone.now().date() - stock.lot_no).days
            else:
                days_old = 99999  # NULL = 가장 오래된 재고 (우선 소진)
            lot_info = {
                'stock_id': stock.id,
                'lot_no': stock.lot_no.strftime('%Y-%m-%d') if stock.lot_no else None,
                'quantity': stock.quantity,
                'days_old': days_old,
                'is_null_lot': stock.lot_no is None,
            }
            lots.append(lot_info)

        # 단위: weight_unit이 있으면 사용, 없으면 기본 EA
        unit = (part.weight_unit or 'EA').strip() or 'EA'

        return JsonResponse({
            'success': True,
            'part_name': part.part_name,
            'warehouse_name': warehouse.name,
            'unit': unit,
            'lots': lots
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# =============================================================================
# 6-1. LOT 배분 (NULL 재고 → LOT 배분) + LOT 수정
# =============================================================================

def _handle_lot_correct(request):
    """기존 LOT 수량 수정 처리 (LOT 간 수량 재배분)"""
    try:
        warehouse_id = request.POST.get('warehouse_id')
        if not warehouse_id:
            messages.error(request, "창고 정보가 없습니다.")
            return redirect('material:lot_allocation')

        warehouse = Warehouse.objects.get(id=warehouse_id)
        stock_ids = request.POST.getlist('stock_ids[]')
        new_qtys = request.POST.getlist('new_qtys[]')

        if not stock_ids:
            messages.error(request, "수정할 LOT가 없습니다.")
            return redirect('material:lot_allocation')

        total_corrected = 0

        with transaction.atomic():
            for stock_id, new_qty_str in zip(stock_ids, new_qtys):
                new_qty = int(new_qty_str)
                stock = MaterialStock.objects.select_for_update().get(pk=stock_id)
                old_qty = stock.quantity
                delta = new_qty - old_qty

                if delta == 0:
                    continue

                if new_qty < 0:
                    messages.error(request, f"{stock.part.part_no} LOT {stock.lot_no}: 음수 수량 불가")
                    return redirect('material:lot_allocation')

                # LOT 재고 업데이트
                stock.quantity = new_qty
                stock.save()

                # NULL 재고 역보정 (LOT +100 → NULL -100)
                null_stock, _ = MaterialStock.objects.select_for_update().get_or_create(
                    warehouse=warehouse, part=stock.part, lot_no=None,
                    defaults={'quantity': 0}
                )
                null_stock.quantity -= delta
                null_stock.save()

                # 트랜잭션 기록
                trx_no = f"LOTC-{timezone.now().strftime('%y%m%d%H%M%S%f')}-{request.user.id}"
                MaterialTransaction.objects.create(
                    transaction_no=trx_no,
                    transaction_type='LOT_CORRECT',
                    date=timezone.now(),
                    part=stock.part,
                    quantity=delta,
                    lot_no=stock.lot_no,
                    warehouse_to=warehouse,
                    result_stock=new_qty,
                    actor=request.user,
                    remark=f"LOT 수정: {stock.lot_no} {old_qty}→{new_qty} ({delta:+d})"
                )
                total_corrected += 1

        if total_corrected > 0:
            messages.success(request, f"LOT 수정 완료: {total_corrected}건")
        else:
            messages.info(request, "변경 사항이 없습니다.")

    except MaterialStock.DoesNotExist:
        messages.error(request, "존재하지 않는 재고 레코드입니다.")
    except (ValueError, TypeError) as e:
        messages.error(request, f"입력값 오류: {e}")
    except Exception as e:
        logger.error(f"LOT 수정 오류: {e}")
        messages.error(request, f"LOT 수정 처리 중 오류: {e}")

    return redirect('material:lot_allocation')


@wms_permission_required('can_wms_stock_edit')
def lot_allocation(request):
    """
    [WMS] LOT 배분 - NULL 재고를 LOT별로 배분 (재고 조사 결과 반영)
    """
    warehouses = Warehouse.objects.all().order_by('code')

    if request.method == 'POST':
        action = request.POST.get('action', 'allocate')

        # ── LOT 수정 (기존 LOT 수량 변경) ──
        if action == 'lot_correct':
            return _handle_lot_correct(request)

        try:
            warehouse_id = request.POST.get('warehouse_id')
            part_ids = request.POST.getlist('part_ids[]')

            if not warehouse_id or not part_ids:
                messages.error(request, "창고와 품목을 선택해주세요.")
                return redirect('material:lot_allocation')

            warehouse = Warehouse.objects.get(id=warehouse_id)
            from datetime import datetime

            total_parts = 0
            total_lots = 0
            total_ea = 0

            with transaction.atomic():
                for pid in part_ids:
                    part = Part.objects.get(id=pid)
                    lot_nos = request.POST.getlist(f'lot_nos_{pid}[]')
                    quantities = request.POST.getlist(f'qty_{pid}[]')

                    if not lot_nos or not quantities:
                        continue

                    # 배분 항목 파싱
                    alloc_items = []
                    part_alloc = 0
                    for i in range(len(lot_nos)):
                        lot_str = lot_nos[i].strip()
                        qty_str = quantities[i].strip()
                        if not lot_str or not qty_str:
                            continue
                        qty = int(qty_str)
                        if qty <= 0:
                            continue
                        try:
                            lot_date = datetime.strptime(lot_str, '%Y-%m-%d').date()
                        except (ValueError, TypeError):
                            messages.error(request, f"LOT 날짜 형식 오류: {part.part_no} - {lot_str}")
                            return redirect('material:lot_allocation')
                        alloc_items.append((lot_date, qty))
                        part_alloc += qty

                    if not alloc_items:
                        continue

                    # NULL 재고 검증
                    null_stock = MaterialStock.objects.filter(
                        warehouse=warehouse, part=part, lot_no__isnull=True
                    ).first()
                    null_qty = null_stock.quantity if null_stock else 0

                    if part_alloc > null_qty:
                        messages.error(request, f"{part.part_no}: 배분({part_alloc})이 NULL 재고({null_qty})를 초과합니다.")
                        return redirect('material:lot_allocation')

                    # 배분 실행
                    for lot_date, qty in alloc_items:
                        lot_stock, _ = MaterialStock.objects.get_or_create(
                            warehouse=warehouse, part=part, lot_no=lot_date,
                            defaults={'quantity': 0}
                        )
                        MaterialStock.objects.filter(pk=lot_stock.pk).update(
                            quantity=F('quantity') + qty
                        )
                        lot_stock.refresh_from_db()

                        trx_no = f"LOTA-{timezone.now().strftime('%y%m%d%H%M%S%f')}-{request.user.id}"
                        MaterialTransaction.objects.create(
                            transaction_no=trx_no,
                            transaction_type='LOT_ASSIGN',
                            date=timezone.now(),
                            part=part,
                            quantity=qty,
                            lot_no=lot_date,
                            warehouse_to=warehouse,
                            result_stock=lot_stock.quantity,
                            actor=request.user,
                            remark=f"LOT 배분: NULL→{lot_date} ({qty}EA)"
                        )

                    # NULL 재고 차감
                    MaterialStock.objects.filter(pk=null_stock.pk).update(
                        quantity=F('quantity') - part_alloc
                    )

                    total_parts += 1
                    total_lots += len(alloc_items)
                    total_ea += part_alloc

            if total_parts == 0:
                messages.warning(request, "배분할 LOT 정보가 없습니다.")
            else:
                messages.success(request, f"LOT 배분 완료: {total_parts}개 품목, {total_lots}건 LOT, 총 {total_ea}EA")

                # 배분 결과를 세션에 저장 (현품표 출력용)
                alloc_items_for_print = []
                for pid in part_ids:
                    p = Part.objects.get(id=pid)
                    lot_list = request.POST.getlist(f'lot_nos_{pid}[]')
                    qty_list = request.POST.getlist(f'qty_{pid}[]')
                    for ls, qs in zip(lot_list, qty_list):
                        ls, qs = ls.strip(), qs.strip()
                        if ls and qs and int(qs) > 0:
                            alloc_items_for_print.append({
                                'part_no': p.part_no,
                                'part_name': p.part_name,
                                'part_group': p.part_group or '',
                                'lot_no': ls,
                                'quantity': int(qs),
                                'weight_unit': p.weight_unit or 'EA',
                            })
                if alloc_items_for_print:
                    request.session['lot_alloc_result'] = alloc_items_for_print

            return redirect('material:lot_allocation')

        except Warehouse.DoesNotExist:
            messages.error(request, "존재하지 않는 창고입니다.")
        except Part.DoesNotExist:
            messages.error(request, "존재하지 않는 품목입니다.")
        except (ValueError, TypeError) as e:
            messages.error(request, f"입력값 오류: {e}")
        except Exception as e:
            logger.error(f"LOT 배분 오류: {e}")
            messages.error(request, f"LOT 배분 처리 중 오류: {e}")

        return redirect('material:lot_allocation')

    # GET: 페이지 렌더링
    # 최근 LOT_ASSIGN / LOT_CORRECT 이력 조회
    recent_assigns = MaterialTransaction.objects.filter(
        transaction_type__in=['LOT_ASSIGN', 'LOT_CORRECT']
    ).select_related('part', 'warehouse_to', 'actor').order_by('-date')[:50]

    # 배분 결과 (현품표 출력용)
    alloc_result = request.session.pop('lot_alloc_result', None)

    context = {
        'warehouses': warehouses,
        'recent_assigns': recent_assigns,
        'alloc_result': alloc_result,
    }
    return render(request, 'material/lot_allocation.html', context)


@wms_permission_required('can_wms_stock_view')
def api_null_stock_info(request):
    """
    [API] NULL 재고 정보 조회 (LOT 배분용)
    단일: GET ?warehouse_code=4200&part_no=SEAL-001
    복수: GET ?warehouse_code=4200&part_nos=ZR-001,ZR-002,ZR-003
    """
    from django.http import JsonResponse

    warehouse_code = request.GET.get('warehouse_code', '').strip()
    part_no = request.GET.get('part_no', '').strip()
    part_nos_str = request.GET.get('part_nos', '').strip()

    if not warehouse_code:
        return JsonResponse({'success': False, 'error': '창고 코드가 필요합니다.'}, status=400)

    try:
        warehouse = Warehouse.objects.filter(code=warehouse_code).first()
        if not warehouse:
            return JsonResponse({'success': False, 'error': f'창고 코드 {warehouse_code}를 찾을 수 없습니다.'})

        # 복수 품번 조회
        if part_nos_str:
            part_no_list = [p.strip() for p in part_nos_str.split(',') if p.strip()]
            parts = Part.objects.filter(part_no__in=part_no_list)
            items = []
            for part in parts:
                null_stock = MaterialStock.objects.filter(
                    warehouse=warehouse, part=part, lot_no__isnull=True
                ).first()
                null_qty = null_stock.quantity if null_stock else 0

                lot_stocks = MaterialStock.objects.filter(
                    warehouse=warehouse, part=part, lot_no__isnull=False
                ).order_by('lot_no')
                existing_lots = [
                    {'id': s.id, 'lot_no': s.lot_no.strftime('%Y-%m-%d'), 'quantity': s.quantity}
                    for s in lot_stocks
                ]
                total_qty = null_qty + sum(s.quantity for s in lot_stocks)

                items.append({
                    'part_no': part.part_no,
                    'part_name': part.part_name or '',
                    'part_id': part.id,
                    'null_qty': null_qty,
                    'existing_lots': existing_lots,
                    'total_qty': total_qty,
                    'weight_unit': part.weight_unit or 'EA',
                })

            return JsonResponse({'success': True, 'items': items})

        # 단일 품번 조회 (기존 호환)
        if not part_no:
            return JsonResponse({'success': False, 'error': '품번이 필요합니다.'}, status=400)

        part = Part.objects.filter(part_no=part_no).first()
        if not part:
            return JsonResponse({'success': False, 'error': f'품번 {part_no}를 찾을 수 없습니다.'})

        null_stock = MaterialStock.objects.filter(
            warehouse=warehouse, part=part, lot_no__isnull=True
        ).first()
        null_qty = null_stock.quantity if null_stock else 0

        lot_stocks = MaterialStock.objects.filter(
            warehouse=warehouse, part=part, lot_no__isnull=False
        ).order_by('lot_no')
        existing_lots = [
            {'id': s.id, 'lot_no': s.lot_no.strftime('%Y-%m-%d'), 'quantity': s.quantity}
            for s in lot_stocks
        ]
        total_qty = null_qty + sum(s.quantity for s in lot_stocks)

        return JsonResponse({
            'success': True,
            'null_qty': null_qty,
            'existing_lots': existing_lots,
            'total_qty': total_qty,
            'part_name': part.part_name or '',
            'part_id': part.id,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# 7. BOM 관리 (Bill of Materials)
# =============================================================================

from .models import Product, BOMItem
import csv
import io
from decimal import Decimal, InvalidOperation

@wms_permission_required('can_wms_bom_view')
def bom_list(request):
    """
    [WMS] BOM 관리 - 모품(제품) 목록 및 BOM 조회
    """
    q = request.GET.get('q', '').strip()
    account_type = request.GET.get('account_type', '')

    products = Product.objects.filter(is_active=True)

    if q:
        products = products.filter(
            Q(part_no__icontains=q) | Q(part_name__icontains=q)
        )

    if account_type:
        products = products.filter(account_type=account_type)

    products = products.order_by('part_no')

    # 페이징
    paginator = Paginator(products, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # 각 제품의 BOM 아이템 수 추가
    for product in page_obj:
        product.bom_count = product.bom_items.filter(is_active=True).count()

    # 동기화 상태 확인
    from django.core.cache import cache
    last_sync = cache.get('bom_last_sync', '')
    sync_running = cache.get('bom_sync_running', False)
    sync_result = cache.get('bom_sync_result')

    # 완료 결과가 있으면 메시지로 표시 후 삭제
    if sync_result and not sync_running:
        r = sync_result
        if r['synced'] > 0:
            messages.success(request, f"BOM 동기화 완료 ({r['finished_at']}): {r['synced']}건 동기화, {r['skipped']}건 건너뜀")
        if r['errors'] > 0:
            messages.warning(request, f"동기화 오류 {r['errors']}건: {', '.join(r['error_list'][:3])}")
        cache.delete('bom_sync_result')

    context = {
        'page_obj': page_obj,
        'q': q,
        'account_type': account_type,
        'last_sync': last_sync,
        'sync_running': sync_running,
    }
    return render(request, 'material/bom_list.html', context)


@login_required
@wms_permission_required('can_wms_bom_edit')
def bom_sync(request):
    """[WMS] ERP BOM 동기화 - 백그라운드 스레드로 실행"""
    if request.method != 'POST':
        return redirect('material:bom_list')

    from django.core.cache import cache

    # 이미 동기화 중이면 중복 실행 방지
    if cache.get('bom_sync_running'):
        messages.warning(request, '이미 BOM 동기화가 진행 중입니다. 완료될 때까지 기다려주세요.')
        return redirect('material:bom_list')

    import threading

    def _run_sync():
        try:
            from material.erp_api import sync_all_bom
            cache.set('bom_sync_running', True, timeout=7200)  # 2시간 타임아웃
            synced, skipped, errors, error_list = sync_all_bom()
            cache.set('bom_last_sync', timezone.now().strftime('%Y-%m-%d %H:%M'), timeout=None)
            cache.set('bom_sync_result', {
                'synced': synced, 'skipped': skipped,
                'errors': errors, 'error_list': error_list[:5],
                'finished_at': timezone.now().strftime('%Y-%m-%d %H:%M'),
            }, timeout=86400)
        except Exception as e:
            logger.error(f'BOM 동기화 예외: {e}')
            cache.set('bom_sync_result', {
                'synced': 0, 'skipped': 0, 'errors': 1,
                'error_list': [str(e)],
                'finished_at': timezone.now().strftime('%Y-%m-%d %H:%M'),
            }, timeout=86400)
        finally:
            cache.delete('bom_sync_running')

    thread = threading.Thread(target=_run_sync, daemon=True)
    thread.start()

    messages.info(request, 'BOM 동기화를 백그라운드에서 시작했습니다. 모품 8,000건+ 처리에 약 2~4시간 소요됩니다. 완료 시 이 페이지에서 결과를 확인할 수 있습니다.')
    return redirect('material:bom_list')


@wms_permission_required('can_wms_bom_view')
def bom_detail(request, part_no):
    """
    [WMS] 특정 제품의 BOM 상세 조회
    """
    product = get_object_or_404(Product, part_no=part_no)
    bom_items = product.bom_items.filter(is_active=True).order_by('seq')

    context = {
        'product': product,
        'bom_items': bom_items,
    }
    return render(request, 'material/bom_detail.html', context)


@login_required
@wms_permission_required('can_wms_bom_edit')
def bom_sync_single(request, part_no):
    """[WMS] 단일 모품 ERP BOM 동기화"""
    if request.method != 'POST':
        return redirect('material:bom_detail', part_no=part_no)

    try:
        from material.erp_api import sync_single_bom
        ok, count, err = sync_single_bom(part_no)
        if ok:
            messages.success(request, f'[{part_no}] ERP 동기화 완료: 자품 {count}개')
        else:
            messages.error(request, f'[{part_no}] 동기화 실패: {err}')
    except Exception as e:
        logger.error(f'단일 BOM 동기화 예외: {e}')
        messages.error(request, f'동기화 오류: {e}')

    return redirect('material:bom_detail', part_no=part_no)


@wms_permission_required('can_wms_bom_edit')
def bom_delete_all(request):
    """
    [WMS] BOM 전체 삭제
    """
    if request.method == 'POST':
        try:
            with transaction.atomic():
                bom_count = BOMItem.objects.count()
                product_count = Product.objects.count()
                BOMItem.objects.all().delete()
                Product.objects.all().delete()
                messages.success(request, f"BOM 전체 삭제 완료! (모품 {product_count}건, BOM 항목 {bom_count}건 삭제)")
        except Exception as e:
            messages.error(request, f"삭제 중 오류 발생: {str(e)}")
    return redirect('material:bom_upload')


@wms_permission_required('can_wms_bom_edit')
def bom_upload(request):
    """
    [WMS] BOM 데이터 엑셀(CSV) 업로드
    - 엑셀 헤더: No,모품번,모품명,규격,재고단위,계정구분,조달구분,BOM등록여부,순번,자품번,자품명,규격,재고단위,정미수량,LOSS(%),필요수량,사급구분,외주구분,시작일자,종료일자,도면번호,재질,주거래처,사용여부,BOM사용여부,비고
    - 주의: '규격'과 '재고단위' 컬럼이 모품/자품용으로 2번 나오므로 인덱스 기반으로 처리
    """
    # 현재 BOM 데이터 수량 (삭제 버튼 표시용)
    bom_count = BOMItem.objects.count()
    product_count = Product.objects.count()

    if request.method == 'POST':
        upload_file = request.FILES.get('bom_file')

        if not upload_file:
            messages.error(request, "파일을 선택해주세요.")
            return redirect('material:bom_upload')

        # 파일 확장자 확인
        file_name = upload_file.name.lower()
        if not (file_name.endswith('.csv') or file_name.endswith('.xlsx')):
            messages.error(request, "CSV 또는 XLSX 파일만 업로드 가능합니다.")
            return redirect('material:bom_upload')

        try:
            rows = []
            headers = []

            if file_name.endswith('.csv'):
                # CSV 파일 처리 - 인덱스 기반으로 읽기
                decoded_file = upload_file.read().decode('utf-8-sig')
                lines = decoded_file.strip().split('\n')
                if lines:
                    # 헤더 파싱 (중복 컬럼명 처리를 위해 인덱스 사용)
                    headers = lines[0].split(',')
                    for line in lines[1:]:
                        # CSV 파싱 (쉼표가 따옴표 안에 있을 수 있음)
                        import re
                        # 간단한 CSV 파싱 - 따옴표 안의 쉼표 처리
                        values = []
                        in_quotes = False
                        current = ''
                        for char in line:
                            if char == '"':
                                in_quotes = not in_quotes
                            elif char == ',' and not in_quotes:
                                values.append(current.strip().strip('"'))
                                current = ''
                            else:
                                current += char
                        values.append(current.strip().strip('"'))

                        if len(values) >= 10:  # 최소 필수 컬럼 수
                            rows.append(values)
            else:
                # XLSX 파일 처리
                wb = openpyxl.load_workbook(upload_file, read_only=True)
                ws = wb.active

                # 헤더 읽기
                headers = [cell.value or '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]

                for row in ws.iter_rows(min_row=2, values_only=True):
                    values = [v if v is not None else '' for v in row]
                    if len(values) >= 10:
                        rows.append(values)

            # 컬럼 인덱스 매핑 (헤더 기반)
            # No,모품번,모품명,규격,재고단위,계정구분,조달구분,BOM등록여부,순번,자품번,자품명,규격,재고단위,정미수량,LOSS(%),필요수량,...
            # 0   1     2     3    4       5       6       7          8    9     10    11   12      13       14      15
            col_map = {
                '모품번': 1, '모품명': 2, '모품규격': 3, '모품단위': 4,
                '계정구분': 5, '조달구분': 6, 'BOM등록여부': 7,
                '순번': 8, '자품번': 9, '자품명': 10, '자품규격': 11, '자품단위': 12,
                '정미수량': 13, 'LOSS': 14, '필요수량': 15,
                '사급구분': 16, '외주구분': 17, '시작일자': 18, '종료일자': 19,
                '도면번호': 20, '재질': 21, '주거래처': 22,
                '사용여부': 23, 'BOM사용여부': 24, '비고': 25
            }

            def get_val(row_data, idx, default=''):
                if idx < len(row_data):
                    return str(row_data[idx]).strip() if row_data[idx] else default
                return default

            def parse_decimal(val, default=0):
                if val is None or str(val).strip() == '':
                    return Decimal(default)
                try:
                    return Decimal(str(val).strip().replace(',', ''))
                except (InvalidOperation, ValueError):
                    return Decimal(default)

            def parse_int(val, default=1):
                if val is None or str(val).strip() == '':
                    return default
                try:
                    return int(float(str(val).strip()))
                except (ValueError, TypeError):
                    return default

            def parse_date(val):
                if not val or str(val).strip() == '':
                    return None
                try:
                    from datetime import datetime
                    val_str = str(val).strip()
                    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d'):
                        try:
                            return datetime.strptime(val_str, fmt).date()
                        except ValueError:
                            continue
                    return None
                except (TypeError, AttributeError):
                    return None

            product_count = 0
            bom_item_count = 0
            products_cache = {}  # 모품번 -> Product 객체 캐시

            with transaction.atomic():
                for row_data in rows:
                    # 모품번 가져오기
                    part_no = get_val(row_data, col_map['모품번'])
                    if not part_no:
                        continue

                    # Product 캐시에서 찾거나 생성
                    if part_no not in products_cache:
                        bom_registered_val = get_val(row_data, col_map['BOM등록여부'], 'Y')
                        is_bom = bom_registered_val.upper() in ('Y', 'YES', '1', 'TRUE', 'O', '등록')

                        product, created = Product.objects.update_or_create(
                            part_no=part_no,
                            defaults={
                                'part_name': get_val(row_data, col_map['모품명']),
                                'spec': get_val(row_data, col_map['모품규격']) or None,
                                'unit': get_val(row_data, col_map['모품단위'], 'EA'),
                                'account_type': get_val(row_data, col_map['계정구분'], '제품'),
                                'procurement_type': get_val(row_data, col_map['조달구분'], '생산'),
                                'is_bom_registered': is_bom,
                                'is_active': True,
                            }
                        )
                        products_cache[part_no] = product
                        if created:
                            product_count += 1

                    current_product = products_cache[part_no]

                    # 자품번이 있으면 BOMItem 생성
                    child_part_no = get_val(row_data, col_map['자품번'])
                    if not child_part_no:
                        continue

                    seq = parse_int(get_val(row_data, col_map['순번']), 1)

                    use_val = get_val(row_data, col_map['사용여부'], 'Y')
                    bom_use_val = get_val(row_data, col_map['BOM사용여부'], 'Y')

                    # BOMItem 생성 또는 업데이트
                    bom_item, created = BOMItem.objects.update_or_create(
                        product=current_product,
                        seq=seq,
                        child_part_no=child_part_no,
                        defaults={
                            'child_part_name': get_val(row_data, col_map['자품명']),
                            'child_spec': get_val(row_data, col_map['자품규격']) or None,
                            'child_unit': get_val(row_data, col_map['자품단위'], 'EA'),
                            'net_qty': parse_decimal(get_val(row_data, col_map['정미수량'])),
                            'loss_rate': parse_decimal(get_val(row_data, col_map['LOSS'])) if get_val(row_data, col_map['LOSS']) else None,
                            'required_qty': parse_decimal(get_val(row_data, col_map['필요수량'])),
                            'supply_type': get_val(row_data, col_map['사급구분'], '자재'),
                            'outsource_type': get_val(row_data, col_map['외주구분'], '무상'),
                            'start_date': parse_date(get_val(row_data, col_map['시작일자'])),
                            'end_date': parse_date(get_val(row_data, col_map['종료일자'])),
                            'drawing_no': get_val(row_data, col_map['도면번호']) or None,
                            'material': get_val(row_data, col_map['재질']) or None,
                            'vendor_name': get_val(row_data, col_map['주거래처']) or None,
                            'is_active': use_val.upper() in ('Y', 'YES', '1', 'TRUE', 'O', '사용', ''),
                            'is_bom_active': bom_use_val.upper() in ('Y', 'YES', '1', 'TRUE', 'O', '사용', ''),
                            'remark': get_val(row_data, col_map['비고']) or None,
                        }
                    )
                    if created:
                        bom_item_count += 1

            messages.success(request, f"BOM 업로드 완료! 모품 {product_count}건, BOM 항목 {bom_item_count}건 추가됨")
            return redirect('material:bom_list')

        except Exception as e:
            import traceback
            messages.error(request, f"파일 처리 중 오류 발생: {str(e)}")
            return redirect('material:bom_upload')

    return render(request, 'material/bom_upload.html', {
        'bom_count': bom_count,
        'product_count': product_count,
    })


def _calculate_bom_requirements(part_no, production_qty):
    """
    BOM 소요량 계산 공통 함수 (다단계 BOM 재귀 전개)
    - 자품번이 다른 Product의 모품번으로 등록되어 있으면 하위 BOM을 재귀 전개
    - 최하위 원자재만 결과에 포함 (반제품은 전개 후 제외)
    - structured_items: 반제품 구조 포함 계층형 결과 (화면/엑셀 표시용)
    """
    product = Product.objects.filter(part_no=part_no, is_active=True).first()
    if not product:
        return None, None, [], []

    # 재귀 전개 결과를 품번별로 합산 (flat)
    aggregated = {}
    # 계층형 결과 (반제품 구조 포함)
    structured_items = []
    _explode_bom(part_no, production_qty, aggregated, visited=set(),
                 structured_items=structured_items, level=1, parent_info=None)

    # 합산된 결과를 리스트로 변환 (flat - 기존 호환)
    result = []
    seq = 1
    for child_part_no, info in aggregated.items():
        # WMS 재고 조회
        stock_qty = 0
        part_obj = Part.objects.filter(part_no=child_part_no).first()
        if part_obj:
            stock_qty = MaterialStock.objects.filter(
                part=part_obj,
                quantity__gt=0
            ).aggregate(total=Sum('quantity'))['total'] or 0

        shortage = max(0, info['required_qty'] - stock_qty)

        result.append({
            'seq': seq,
            'child_part_no': child_part_no,
            'child_part_name': info['child_part_name'],
            'child_unit': info['child_unit'],
            'unit_qty': info['unit_qty'],
            'required_qty': info['required_qty'],
            'stock_qty': float(stock_qty),
            'shortage': float(shortage),
            'supply_type': info['supply_type'],
            'vendor_name': info['vendor_name'],
        })
        seq += 1

    # flat result에서 품번별 재고/부족 정보를 가져와서 structured_items에 반영
    flat_lookup = {item['child_part_no']: item for item in result}
    for sitem in structured_items:
        if sitem.get('is_semi'):
            continue  # 반제품 헤더는 재고 불필요
        cpno = sitem['child_part_no']
        flat_item = flat_lookup.get(cpno)
        if flat_item:
            sitem['stock_qty'] = flat_item['stock_qty']
            sitem['shortage'] = flat_item['shortage']
        else:
            sitem['stock_qty'] = 0
            sitem['shortage'] = float(sitem['required_qty'])

    return product, product.part_name if product else None, result, structured_items


def _explode_bom(part_no, production_qty, aggregated, visited,
                 structured_items=None, level=1, parent_info=None):
    """
    BOM 재귀 전개 내부 함수
    - 자품번이 Product로 등록되어 있고 하위 BOM이 있으면 재귀 전개
    - 없으면 최하위 원자재로 간주하여 aggregated에 합산
    - visited로 순환 참조 방지
    - structured_items: 계층형 결과 수집 (반제품 헤더 포함)
    """
    if part_no in visited:
        return
    visited.add(part_no)

    product = Product.objects.filter(part_no=part_no, is_active=True).first()
    if not product:
        return

    bom_items = product.bom_items.filter(is_active=True, is_bom_active=True).order_by('seq')

    for item in bom_items:
        required_qty = float(item.required_qty) * float(production_qty)

        # 자품번이 Product로 등록되어 있고 하위 BOM이 있는지 확인
        child_product = Product.objects.filter(part_no=item.child_part_no, is_active=True).first()
        has_child_bom = False
        if child_product:
            has_child_bom = child_product.bom_items.filter(is_active=True, is_bom_active=True).exists()

        if has_child_bom:
            # 반제품 헤더 추가 (structured_items)
            if structured_items is not None:
                structured_items.append({
                    'is_semi': True,
                    'level': level,
                    'child_part_no': item.child_part_no,
                    'child_part_name': item.child_part_name,
                    'child_unit': item.child_unit,
                    'unit_qty': float(item.required_qty),
                    'required_qty': required_qty,
                    'supply_type': item.supply_type,
                    'vendor_name': item.vendor_name,
                })
            # 반제품 → 하위 BOM 재귀 전개 (소요량을 곱해서 전달)
            _explode_bom(item.child_part_no, required_qty, aggregated, visited.copy(),
                         structured_items=structured_items, level=level + 1,
                         parent_info={'part_no': item.child_part_no, 'part_name': item.child_part_name})
        else:
            # 최하위 원자재 → 결과에 합산
            if item.child_part_no in aggregated:
                aggregated[item.child_part_no]['required_qty'] += required_qty
            else:
                aggregated[item.child_part_no] = {
                    'child_part_name': item.child_part_name,
                    'child_unit': item.child_unit,
                    'unit_qty': float(item.required_qty),
                    'required_qty': required_qty,
                    'supply_type': item.supply_type,
                    'vendor_name': item.vendor_name,
                }
            # structured_items에도 추가
            if structured_items is not None:
                structured_items.append({
                    'is_semi': False,
                    'level': level,
                    'child_part_no': item.child_part_no,
                    'child_part_name': item.child_part_name,
                    'child_unit': item.child_unit,
                    'unit_qty': float(item.required_qty),
                    'required_qty': required_qty,
                    'supply_type': item.supply_type,
                    'vendor_name': item.vendor_name,
                    'parent_part_no': parent_info['part_no'] if parent_info else None,
                })


@wms_permission_required('can_wms_bom_view')
def bom_calculate(request):
    """
    [WMS] 소요량 계산 - 제품 품번과 생산수량 입력 시 필요 자재량 계산
    - 단일 계산: 개별 품번 입력
    - 일괄 계산: 엑셀 업로드
    """
    result = None
    product = None
    production_qty = 0
    calc_type = None
    batch_results = None
    shortage_count = 0
    sufficient_count = 0
    total_material_count = 0
    total_shortage_count = 0
    total_sufficient_count = 0
    session_key = None
    structured = []
    single_session_key = None

    if request.method == 'POST':
        calc_type = request.POST.get('calc_type', 'single')

        if calc_type == 'single':
            # 단일 제품 계산
            part_no = request.POST.get('part_no', '').strip()
            production_qty = int(request.POST.get('production_qty', 0) or 0)

            if part_no and production_qty > 0:
                product, part_name, result, structured = _calculate_bom_requirements(part_no, production_qty)

                if product and result:
                    shortage_count = sum(1 for item in result if item['shortage'] > 0)
                    sufficient_count = len(result) - shortage_count
                    # structured_items를 세션에 저장 (엑셀 다운로드용)
                    import uuid as _uuid
                    single_session_key = str(_uuid.uuid4())
                    request.session[f'single_calc_{single_session_key}'] = {
                        'part_no': part_no,
                        'part_name': product.part_name,
                        'production_qty': production_qty,
                        'structured_items': structured,
                    }
                else:
                    structured = []
                    single_session_key = None
                    messages.warning(request, f"품번 '{part_no}'에 해당하는 BOM이 없습니다.")

        elif calc_type == 'batch':
            # 일괄 계산 (엑셀 업로드)
            upload_file = request.FILES.get('calc_file')

            if not upload_file:
                messages.error(request, "파일을 선택해주세요.")
            else:
                file_name = upload_file.name.lower()
                if not (file_name.endswith('.csv') or file_name.endswith('.xlsx')):
                    messages.error(request, "CSV 또는 XLSX 파일만 업로드 가능합니다.")
                else:
                    try:
                        rows = []

                        if file_name.endswith('.csv'):
                            decoded_file = upload_file.read().decode('utf-8-sig')
                            lines = decoded_file.strip().split('\n')
                            if len(lines) > 1:
                                headers = [h.strip() for h in lines[0].split(',')]
                                for line in lines[1:]:
                                    values = [v.strip().strip('"') for v in line.split(',')]
                                    if len(values) >= 2:
                                        rows.append(dict(zip(headers, values)))
                        else:
                            wb = openpyxl.load_workbook(upload_file, read_only=True)
                            ws = wb.active
                            headers = [cell.value or '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                            headers = [str(h).strip() for h in headers]

                            # 피벗 형태 감지: 헤더에 날짜 패턴(YYYY-MM-DD)이 있으면 피벗
                            import re
                            date_pattern = re.compile(r'^(\d{4}-\d{2}-\d{2})')
                            date_columns = []  # (col_index, date_str)
                            part_no_col = None

                            for idx, h in enumerate(headers):
                                m = date_pattern.match(h)
                                if m:
                                    date_columns.append((idx, m.group(1)))
                                elif h in ('\ud488\ubc88', '품번'):
                                    part_no_col = idx

                            if date_columns and part_no_col is not None:
                                # 피벗 형태 → unpivot (품번 × 날짜 → 행)
                                for row_data in ws.iter_rows(min_row=2, values_only=True):
                                    values = list(row_data)
                                    if len(values) <= part_no_col:
                                        continue
                                    pno = str(values[part_no_col] or '').strip()
                                    if not pno or not re.search(r'[A-Za-z0-9]', pno):
                                        continue  # 합계 행 등 제외
                                    for col_idx, date_str in date_columns:
                                        if col_idx < len(values):
                                            qty_val = values[col_idx]
                                            try:
                                                qty_num = int(float(str(qty_val).replace(',', '')))
                                            except (ValueError, TypeError):
                                                qty_num = 0
                                            if qty_num > 0:
                                                rows.append({
                                                    '\ud488\ubc88': pno,
                                                    '\uacc4\ud68d\uc218\ub7c9': qty_num,
                                                    '\ub0a0\uc9dc': date_str,
                                                })
                            else:
                                # 기존 행 단위 형태
                                for row in ws.iter_rows(min_row=2, values_only=True):
                                    values = [v if v is not None else '' for v in row]
                                    if len(values) >= 2:
                                        rows.append(dict(zip(headers, values)))

                        # 일괄 계산 수행
                        batch_results = []
                        for row in rows:
                            # 새 형식: 날짜/품번/계획수량  |  구 형식: 품번/수량/필요일자
                            part_no = str(row.get('품번', '')).strip()
                            qty = row.get('계획수량', None) or row.get('수량', 0)
                            need_date = row.get('날짜', None) or row.get('필요일자', '')

                            # need_date 정규화 → 'YYYY-MM-DD' 문자열로 통일
                            if need_date and hasattr(need_date, 'strftime'):
                                need_date = need_date.strftime('%Y-%m-%d')
                            elif need_date:
                                need_date = str(need_date).strip()
                                need_date = need_date.replace('.0', '')
                                if len(need_date) == 8 and need_date.isdigit():
                                    need_date = f'{need_date[:4]}-{need_date[4:6]}-{need_date[6:8]}'
                            else:
                                need_date = ''

                            if not part_no:
                                continue

                            try:
                                qty = int(float(str(qty).replace(',', '')))
                            except (ValueError, TypeError):
                                qty = 0

                            if qty <= 0:
                                continue

                            product_obj, part_name, items, structured_batch = _calculate_bom_requirements(part_no, qty)

                            batch_results.append({
                                'part_no': part_no,
                                'part_name': part_name or '-',
                                'qty': qty,
                                'need_date': need_date,
                                'items': items,
                                'structured_items': structured_batch,
                            })

                            # 통계
                            total_material_count += len(items)
                            total_shortage_count += sum(1 for item in items if item['shortage'] > 0)
                            total_sufficient_count += sum(1 for item in items if item['shortage'] == 0)

                        if batch_results:
                            # 세션에 결과 저장 (엑셀 다운로드용)
                            import uuid
                            session_key = str(uuid.uuid4())
                            request.session[f'batch_calc_{session_key}'] = batch_results
                            messages.success(request, f"{len(batch_results)}개 제품의 소요량 계산이 완료되었습니다.")
                        else:
                            messages.warning(request, "유효한 데이터가 없습니다. 파일 형식을 확인해주세요.")

                    except Exception as e:
                        import traceback
                        print(f"[BOM CALC ERROR] {traceback.format_exc()}", flush=True)
                        messages.error(request, f"파일 처리 중 오류: {str(e)}")

    # 제품 목록 (자동완성용)
    products = Product.objects.filter(is_active=True, is_bom_registered=True).order_by('part_no')

    context = {
        'products': products,
        'product': product,
        'production_qty': production_qty,
        'result': result,
        'structured_items': structured if calc_type == 'single' else None,
        'single_session_key': single_session_key if calc_type == 'single' else None,
        'calc_type': calc_type,
        'batch_results': batch_results,
        'shortage_count': shortage_count,
        'sufficient_count': sufficient_count,
        'total_material_count': total_material_count,
        'total_shortage_count': total_shortage_count,
        'total_sufficient_count': total_sufficient_count,
        'session_key': session_key,
    }
    return render(request, 'material/bom_calculate.html', context)


@wms_permission_required('can_wms_bom_view')
def bom_calc_template(request):
    """
    [WMS] 소요량 계산용 엑셀 양식 다운로드
    """
    from datetime import datetime, timedelta

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "소요량계산양식"

    # 헤더: 날짜, 품번, 계획수량
    headers = ['날짜', '품번', '계획수량']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

    # 예시 데이터 (오늘 기준으로 날짜 생성)
    today = datetime.now().date()
    date_cell1 = ws.cell(row=2, column=1, value=today + timedelta(days=4))
    date_cell1.number_format = 'YYYYMMDD'
    ws.cell(row=2, column=2, value="064133-0010")
    ws.cell(row=2, column=3, value=100)

    date_cell2 = ws.cell(row=3, column=1, value=today + timedelta(days=5))
    date_cell2.number_format = 'YYYYMMDD'
    ws.cell(row=3, column=2, value="064133-0020")
    ws.cell(row=3, column=3, value=50)

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12

    # A열(날짜) 전체를 YYYYMMDD 형식으로 지정
    for row in range(2, 1000):
        ws.cell(row=row, column=1).number_format = 'YYYYMMDD'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bom_calc_template.xlsx"'
    wb.save(response)
    return response


@wms_permission_required('can_wms_bom_view')
def bom_calc_export(request):
    """
    [WMS] 단일 제품 소요량 계산 결과 엑셀 다운로드
    """
    part_no = request.GET.get('part_no', '').strip()
    qty = int(request.GET.get('qty', 0) or 0)

    if not part_no or qty <= 0:
        messages.error(request, "품번과 수량이 필요합니다.")
        return redirect('material:bom_calculate')

    # 세션에서 structured_items 가져오기
    sk = request.GET.get('session_key', '')
    session_data = request.session.get(f'single_calc_{sk}') if sk else None

    if session_data and session_data.get('structured_items'):
        structured_items = session_data['structured_items']
    else:
        # 세션 없으면 다시 계산
        _, _, _, structured_items = _calculate_bom_requirements(part_no, qty)

    product, part_name, items, _ = _calculate_bom_requirements(part_no, qty)

    if not product or not items:
        messages.error(request, "BOM 데이터가 없습니다.")
        return redirect('material:bom_calculate')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "소요량계산결과"

    # 제목
    ws.merge_cells('A1:K1')
    ws['A1'] = f"소요량 계산 결과 - {part_no} ({part_name}) / 생산수량: {qty}개"
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)

    # 헤더 (ERP 스타일)
    headers = ['No', 'LEVEL', '순번', '자품번', '자품명', '단위', '정미수량', '필요수량', '현재고', '부족수량', '주거래처']
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # 반제품 헤더 스타일
    semi_fill = openpyxl.styles.PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    semi_font = openpyxl.styles.Font(bold=True)
    shortage_fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 데이터 (structured_items 기반)
    row_idx = 4
    seq_counter = {}  # level별 순번
    no = 1
    for sitem in structured_items:
        level = sitem.get('level', 1)

        if sitem.get('is_semi'):
            # 반제품 헤더 행
            ws.cell(row=row_idx, column=1, value=no)
            ws.cell(row=row_idx, column=2, value=level)
            ws.cell(row=row_idx, column=3, value='')
            ws.cell(row=row_idx, column=4, value=sitem['child_part_no'])
            ws.cell(row=row_idx, column=5, value=sitem['child_part_name'])
            ws.cell(row=row_idx, column=6, value=sitem['child_unit'])
            ws.cell(row=row_idx, column=7, value=sitem['unit_qty'])
            ws.cell(row=row_idx, column=8, value=sitem['required_qty'])
            ws.cell(row=row_idx, column=9, value='')
            ws.cell(row=row_idx, column=10, value='')
            ws.cell(row=row_idx, column=11, value=sitem.get('vendor_name') or '')
            # 반제품 스타일
            for col in range(1, 12):
                ws.cell(row=row_idx, column=col).fill = semi_fill
                ws.cell(row=row_idx, column=col).font = semi_font
            seq_counter[level + 1] = 0
        else:
            # 원자재 행
            seq_counter.setdefault(level, 0)
            seq_counter[level] += 1
            ws.cell(row=row_idx, column=1, value=no)
            ws.cell(row=row_idx, column=2, value=level)
            ws.cell(row=row_idx, column=3, value=seq_counter[level])
            ws.cell(row=row_idx, column=4, value=sitem['child_part_no'])
            ws.cell(row=row_idx, column=5, value=sitem['child_part_name'])
            ws.cell(row=row_idx, column=6, value=sitem['child_unit'])
            ws.cell(row=row_idx, column=7, value=sitem['unit_qty'])
            ws.cell(row=row_idx, column=8, value=sitem['required_qty'])
            ws.cell(row=row_idx, column=9, value=sitem.get('stock_qty', 0))
            ws.cell(row=row_idx, column=10, value=sitem.get('shortage', 0))
            ws.cell(row=row_idx, column=11, value=sitem.get('vendor_name') or '-')
            # 부족분 강조
            if sitem.get('shortage', 0) > 0:
                for col in range(1, 12):
                    ws.cell(row=row_idx, column=col).fill = shortage_fill

        no += 1
        row_idx += 1

    # 컬럼 너비 조정
    widths = [6, 8, 6, 18, 25, 8, 12, 12, 12, 12, 15]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="bom_calc_{part_no}.xlsx"'
    wb.save(response)
    return response


@wms_permission_required('can_wms_bom_view')
def bom_calc_batch_export(request):
    """
    [WMS] 일괄 소요량 계산 결과 엑셀 다운로드
    mode=structured 이면 반제품 LEVEL/순번 포함
    """
    session_key = request.GET.get('session_key', '')
    mode = request.GET.get('mode', 'flat')
    batch_results = request.session.get(f'batch_calc_{session_key}')

    if not batch_results:
        messages.error(request, "계산 결과가 없습니다. 다시 계산해주세요.")
        return redirect('material:bom_calculate')

    wb = openpyxl.Workbook()
    ws = wb.active
    header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    semi_fill = openpyxl.styles.PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    semi_font = openpyxl.styles.Font(bold=True)
    shortage_fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bom_missing_fill = openpyxl.styles.PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    if mode == 'structured':
        ws.title = "일괄소요량(BOM구조)"
        headers = ['모품번', '모품명', '생산수량', '필요일자', 'LEVEL', '자품번', '자품명', '단위', '정미수량', '필요수량', '현재고', '부족수량', '거래처']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row_idx = 2
        for batch in batch_results:
            structured = batch.get('structured_items', [])
            if structured:
                # 모품번 자체를 LEVEL 0 자품번 행으로 추가 (소요량 당겨오기용)
                ws.cell(row=row_idx, column=1, value=batch['part_no'])
                ws.cell(row=row_idx, column=2, value=batch['part_name'])
                ws.cell(row=row_idx, column=3, value=batch['qty'])
                ws.cell(row=row_idx, column=4, value=batch.get('need_date') or '')
                ws.cell(row=row_idx, column=5, value=0)
                ws.cell(row=row_idx, column=6, value=batch['part_no'])
                ws.cell(row=row_idx, column=7, value=batch['part_name'])
                ws.cell(row=row_idx, column=10, value=batch['qty'])
                for col in range(1, 14):
                    ws.cell(row=row_idx, column=col).fill = semi_fill
                    ws.cell(row=row_idx, column=col).font = semi_font
                row_idx += 1

                for sitem in structured:
                    # 필요일자는 모든 행에 채움
                    ws.cell(row=row_idx, column=4, value=batch.get('need_date') or '')

                    ws.cell(row=row_idx, column=5, value=sitem.get('level', 1))
                    ws.cell(row=row_idx, column=6, value=sitem['child_part_no'])
                    ws.cell(row=row_idx, column=7, value=sitem['child_part_name'])
                    ws.cell(row=row_idx, column=8, value=sitem.get('child_unit', ''))
                    ws.cell(row=row_idx, column=9, value=sitem.get('unit_qty', ''))
                    ws.cell(row=row_idx, column=10, value=sitem.get('required_qty', 0))
                    ws.cell(row=row_idx, column=13, value=sitem.get('vendor_name') or '-')

                    if sitem.get('is_semi'):
                        ws.cell(row=row_idx, column=11, value='')
                        ws.cell(row=row_idx, column=12, value='')
                        for col in range(1, 14):
                            ws.cell(row=row_idx, column=col).fill = semi_fill
                            ws.cell(row=row_idx, column=col).font = semi_font
                    else:
                        ws.cell(row=row_idx, column=11, value=sitem.get('stock_qty', 0))
                        ws.cell(row=row_idx, column=12, value=sitem.get('shortage', 0))
                        if sitem.get('shortage', 0) > 0:
                            for col in range(1, 14):
                                ws.cell(row=row_idx, column=col).fill = shortage_fill

                    row_idx += 1
            elif batch['items']:
                for item_idx, item in enumerate(batch['items']):
                    if item_idx == 0:
                        ws.cell(row=row_idx, column=1, value=batch['part_no'])
                        ws.cell(row=row_idx, column=2, value=batch['part_name'])
                        ws.cell(row=row_idx, column=3, value=batch['qty'])
                    # 필요일자는 모든 행에 채움
                    ws.cell(row=row_idx, column=4, value=batch.get('need_date') or '')
                    ws.cell(row=row_idx, column=5, value=1)
                    ws.cell(row=row_idx, column=6, value=item['child_part_no'])
                    ws.cell(row=row_idx, column=7, value=item['child_part_name'])
                    ws.cell(row=row_idx, column=8, value=item.get('child_unit', ''))
                    ws.cell(row=row_idx, column=9, value=item.get('unit_qty', ''))
                    ws.cell(row=row_idx, column=10, value=float(item['required_qty']))
                    ws.cell(row=row_idx, column=11, value=item['stock_qty'])
                    ws.cell(row=row_idx, column=12, value=item['shortage'])
                    ws.cell(row=row_idx, column=13, value=item.get('vendor_name') or '-')
                    if item['shortage'] > 0:
                        for col in range(1, 14):
                            ws.cell(row=row_idx, column=col).fill = shortage_fill
                    row_idx += 1
            else:
                ws.cell(row=row_idx, column=1, value=batch['part_no'])
                ws.cell(row=row_idx, column=2, value=batch['part_name'])
                ws.cell(row=row_idx, column=3, value=batch['qty'])
                ws.cell(row=row_idx, column=4, value=batch.get('need_date') or '')
                ws.cell(row=row_idx, column=6, value='BOM 없음')
                for col in range(1, 14):
                    ws.cell(row=row_idx, column=col).fill = bom_missing_fill
                row_idx += 1

        widths = [18, 20, 10, 12, 8, 18, 25, 8, 12, 12, 12, 12, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        filename = "bom_calc_batch_structured.xlsx"
    else:
        ws.title = "일괄소요량계산결과"
        headers = ['모품번', '모품명', '생산수량', '필요일자', '자품번', '자품명', '필요수량', '현재고', '부족수량', '거래처']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row_idx = 2
        for batch in batch_results:
            if batch['items']:
                for item_idx, item in enumerate(batch['items']):
                    if item_idx == 0:
                        ws.cell(row=row_idx, column=1, value=batch['part_no'])
                        ws.cell(row=row_idx, column=2, value=batch['part_name'])
                        ws.cell(row=row_idx, column=3, value=batch['qty'])
                    # 필요일자는 모든 행에 채움
                    ws.cell(row=row_idx, column=4, value=batch.get('need_date') or '')

                    ws.cell(row=row_idx, column=5, value=item['child_part_no'])
                    ws.cell(row=row_idx, column=6, value=item['child_part_name'])
                    ws.cell(row=row_idx, column=7, value=float(item['required_qty']))
                    ws.cell(row=row_idx, column=8, value=item['stock_qty'])
                    ws.cell(row=row_idx, column=9, value=item['shortage'])
                    ws.cell(row=row_idx, column=10, value=item.get('vendor_name') or '-')

                    if item['shortage'] > 0:
                        for col in range(1, 11):
                            ws.cell(row=row_idx, column=col).fill = shortage_fill

                    row_idx += 1
            else:
                ws.cell(row=row_idx, column=1, value=batch['part_no'])
                ws.cell(row=row_idx, column=2, value=batch['part_name'])
                ws.cell(row=row_idx, column=3, value=batch['qty'])
                ws.cell(row=row_idx, column=4, value=batch.get('need_date') or '')
                ws.cell(row=row_idx, column=5, value='BOM 없음')
                for col in range(1, 11):
                    ws.cell(row=row_idx, column=col).fill = bom_missing_fill
                row_idx += 1

        widths = [18, 20, 10, 12, 18, 20, 12, 12, 12, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        filename = "bom_calc_batch_result.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@wms_permission_required('can_wms_bom_view')
def bom_calc_demand_export(request):
    """
    [WMS] 소요량 변환 엑셀 다운로드 (필요일자/모품번/자품번/자품명/필요수량/거래처)
    - 동일 필요일자+자품번은 수량 합산
    """
    session_key = request.GET.get('session_key', '')
    batch_results = request.session.get(f'batch_calc_{session_key}')

    if not batch_results:
        messages.error(request, "계산 결과가 없습니다. 다시 계산해주세요.")
        return redirect('material:bom_calculate')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "소요량변환"

    headers = ['필요일자', '품번', '필요수량']
    header_fill = openpyxl.styles.PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # 자품번별 합산 (동일 필요일자+자품번 → 수량 합산)
    from collections import defaultdict
    agg = defaultdict(float)
    for batch in batch_results:
        for item in batch['items']:
            key = (batch.get('need_date', '') or '', item['child_part_no'])
            agg[key] += float(item['required_qty'])

    # 정렬: 필요일자 → 자품번
    row_idx = 2
    for (need_date, child_part_no) in sorted(agg.keys()):
        ws.cell(row=row_idx, column=1, value=need_date)
        ws.cell(row=row_idx, column=2, value=child_part_no)
        ws.cell(row=row_idx, column=3, value=round(agg[(need_date, child_part_no)], 2))
        row_idx += 1

    widths = [14, 20, 14]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bom_demand_export.xlsx"'
    wb.save(response)
    return response


@wms_permission_required('can_wms_bom_view')
def api_bom_calculate(request):
    """
    [API] 소요량 계산 AJAX 엔드포인트 (다단계 BOM 전개)
    """
    part_no = request.GET.get('part_no', '').strip()
    production_qty = int(request.GET.get('qty', 0) or 0)

    if not part_no or production_qty <= 0:
        return JsonResponse({'error': '품번과 생산수량을 입력하세요.'}, status=400)

    product, part_name, result, _structured = _calculate_bom_requirements(part_no, production_qty)

    if not product:
        return JsonResponse({'error': '해당 품번의 BOM을 찾을 수 없습니다.'}, status=404)

    # vendor_name이 None인 경우 '-'로 변환
    for item in result:
        if not item.get('vendor_name'):
            item['vendor_name'] = '-'

    return JsonResponse({
        'success': True,
        'product': {
            'part_no': product.part_no,
            'part_name': product.part_name,
            'account_type': product.account_type,
        },
        'production_qty': production_qty,
        'items': result,
    })


@wms_permission_required('can_wms_bom_edit')
def bom_sync_missing(request):
    """
    [WMS] BOM 없는 품번을 ERP에서 개별 동기화하는 AJAX API
    - POST: { part_no: '품번' }
    - 한 건씩 호출하여 프론트에서 진행률 관리
    """
    from django.http import JsonResponse
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': '잘못된 요청'}, status=400)

    part_no = data.get('part_no', '').strip()
    if not part_no:
        return JsonResponse({'success': False, 'error': '품번이 없습니다.'})

    from material.erp_api import sync_single_bom
    ok, count, err = sync_single_bom(part_no)

    if ok:
        return JsonResponse({
            'success': True,
            'part_no': part_no,
            'bom_count': count,
            'message': f'{part_no}: BOM {count}개 동기화 완료'
        })
    else:
        return JsonResponse({
            'success': False,
            'part_no': part_no,
            'error': err or 'ERP에 BOM 데이터 없음'
        })


@wms_permission_required('can_wms_bom_edit')
def bom_register_demand(request):
    """
    [WMS] BOM 일괄 소요량 계산 결과를 SCM 소요량(Demand)으로 등록
    - 품목 마스터(Part)에 존재하는 자재만 등록
    - part + due_date 기준으로 update_or_create (기존 데이터 업데이트)
    """
    from datetime import datetime
    import json

    if request.method != 'POST':
        messages.error(request, "잘못된 요청입니다.")
        return redirect('material:bom_calculate')

    session_key = request.POST.get('session_key', '')
    batch_results = request.session.get(f'batch_calc_{session_key}')

    if not batch_results:
        messages.error(request, "계산 결과가 없습니다. 다시 계산해주세요.")
        return redirect('material:bom_calculate')

    # 거래처 필터 파싱
    selected_vendors_json = request.POST.get('selected_vendors', '')
    selected_vendors = None
    if selected_vendors_json:
        try:
            selected_vendors = json.loads(selected_vendors_json)
            if not isinstance(selected_vendors, list):
                selected_vendors = None
        except (json.JSONDecodeError, TypeError):
            selected_vendors = None

    # 거래처 필터용: Part 품번 → vendor.name 매핑 캐시
    part_vendor_cache = {}
    if selected_vendors is not None:
        for p in Part.objects.select_related('vendor').filter(vendor__isnull=False):
            part_vendor_cache[p.part_no] = p.vendor.name

    # 동일 자품번+필요일자 기준으로 필요수량 합산
    demand_map = {}  # key: (child_part_no, need_date), value: required_qty 합계

    for batch in batch_results:
        need_date = batch.get('need_date', '')

        # 날짜 형식 파싱
        if need_date:
            if isinstance(need_date, str):
                try:
                    need_date = datetime.strptime(need_date, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    need_date = None
            elif hasattr(need_date, 'date'):
                need_date = need_date.date()
        else:
            need_date = None

        if not need_date:
            continue  # 필요일자 없는 항목은 스킵

        for item in batch.get('items', []):
            child_part_no = item.get('child_part_no', '')
            required_qty = item.get('required_qty', 0)

            if not child_part_no or required_qty <= 0:
                continue

            # 거래처 필터 적용: Part 마스터의 실제 vendor.name 기준으로 매칭
            if selected_vendors is not None:
                actual_vendor_name = part_vendor_cache.get(child_part_no, '')
                if not actual_vendor_name or actual_vendor_name not in selected_vendors:
                    continue

            key = (child_part_no, need_date)
            if key in demand_map:
                demand_map[key] += required_qty
            else:
                demand_map[key] = required_qty

    if not demand_map:
        messages.warning(request, "등록할 소요량 데이터가 없습니다. 필요일자가 입력된 항목이 있는지 확인해주세요.")
        return redirect('material:bom_calculate')

    # SCM 소요량 등록
    registered_count = 0
    updated_count = 0
    skipped_count = 0
    skipped_parts = []

    with transaction.atomic():
        for (child_part_no, need_date), total_qty in demand_map.items():
            # Part 마스터에서 조회
            part = Part.objects.filter(part_no=child_part_no).first()

            if not part:
                skipped_count += 1
                if child_part_no not in skipped_parts:
                    skipped_parts.append(child_part_no)
                continue

            # Demand update_or_create
            demand, created = Demand.objects.update_or_create(
                part=part,
                due_date=need_date,
                defaults={
                    'quantity': int(total_qty)
                }
            )

            if created:
                registered_count += 1
            else:
                updated_count += 1

    # 결과 메시지
    vendor_info = ""
    if selected_vendors:
        vendor_info = f" [거래처: {', '.join(selected_vendors)}]"
    result_msg = f"SCM 소요량 등록 완료{vendor_info}: 신규 {registered_count}건, 업데이트 {updated_count}건"
    if skipped_count > 0:
        result_msg += f", 스킵 {skipped_count}건 (품목마스터 미존재)"
        if len(skipped_parts) <= 5:
            result_msg += f" - {', '.join(skipped_parts)}"
        else:
            result_msg += f" - {', '.join(skipped_parts[:5])} 외 {len(skipped_parts) - 5}건"

    if registered_count > 0 or updated_count > 0:
        messages.success(request, result_msg)
    else:
        messages.warning(request, result_msg)

    return redirect('material:bom_calculate')


# =============================================================================
# 재고조사 (Inventory Check)
# =============================================================================

from .models import InventoryCheckSession, InventoryCheckSessionItem, ProcessTag


@wms_permission_required('can_wms_stock_view')
def inventory_check_list(request):
    """재고조사 목록"""
    qs = InventoryCheckSession.objects.select_related('warehouse', 'created_by').all()

    # 필터
    status = request.GET.get('status')
    warehouse_id = request.GET.get('warehouse')
    if status:
        qs = qs.filter(status=status)
    if warehouse_id:
        qs = qs.filter(warehouse_id=warehouse_id)

    paginator = Paginator(qs, 20)
    page = request.GET.get('page', 1)
    checks = paginator.get_page(page)

    warehouses = Warehouse.objects.filter(is_active=True)

    return render(request, 'material/inventory_check_list.html', {
        'checks': checks,
        'warehouses': warehouses,
        'status_choices': InventoryCheckSession.STATUS_CHOICES,
    })


@wms_permission_required('can_wms_stock_edit')
def inventory_check_create(request):
    """재고조사 세션 생성"""
    if request.method == 'POST':
        warehouse_id = request.POST.get('warehouse')
        check_date = request.POST.get('check_date') or timezone.localdate()
        remark = request.POST.get('remark', '')

        check = InventoryCheckSession.objects.create(
            check_no=InventoryCheckSession.generate_check_no(),
            warehouse_id=warehouse_id,
            check_date=check_date,
            status='IN_PROGRESS',
            created_by=request.user,
            remark=remark,
        )
        messages.success(request, f'재고조사 {check.check_no}가 시작되었습니다.')
        return redirect('material:inventory_check_scan', pk=check.pk)

    warehouses = Warehouse.objects.filter(is_active=True)
    return render(request, 'material/inventory_check_create.html', {
        'warehouses': warehouses,
    })


@wms_permission_required('can_wms_stock_edit')
def inventory_check_scan(request, pk):
    """재고조사 QR 스캔 페이지"""
    check = get_object_or_404(InventoryCheckSession.objects.select_related('warehouse'), pk=pk)

    if check.status not in ('DRAFT', 'IN_PROGRESS'):
        messages.warning(request, '이 조사는 이미 완료되었습니다.')
        return redirect('material:inventory_check_result', pk=check.pk)

    items = check.check_items.select_related('part').order_by('-scanned_at')[:50]

    return render(request, 'material/inventory_check_scan.html', {
        'check': check,
        'items': items,
    })


from django.views.decorators.http import require_POST

@wms_permission_required('can_wms_stock_edit')
@require_POST
def inventory_check_scan_api(request, pk):
    """재고조사 QR 스캔 API (AJAX)"""
    from django.http import JsonResponse
    import json
    import traceback

    # AJAX API는 로그인 안됐을 때 JSON 반환
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': '로그인이 필요합니다.'}, status=401)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    try:
        check = InventoryCheckSession.objects.get(pk=pk)
    except InventoryCheckSession.DoesNotExist:
        return JsonResponse({'success': False, 'error': '조사를 찾을 수 없습니다.'}, status=404)

    if check.status not in ('DRAFT', 'IN_PROGRESS'):
        return JsonResponse({'success': False, 'error': '조사가 이미 완료되었습니다.'})

    try:
        data = json.loads(request.body)
        raw_input = data.get('tag_id', '').strip()
    except json.JSONDecodeError:
        raw_input = request.POST.get('tag_id', '').strip()

    if not raw_input:
        return JsonResponse({'success': False, 'error': '태그ID가 없습니다.'})

    # QR 데이터 파싱: TAG_ID|품번|수량|LOT 형식에서 TAG_ID만 추출
    tag_id = raw_input.split('|')[0].strip() if '|' in raw_input else raw_input

    # 이미 스캔한 태그인지 확인
    if check.check_items.filter(tag_id=tag_id).exists():
        return JsonResponse({'success': False, 'error': f'이미 스캔된 태그입니다: {tag_id}', 'duplicate': True})

    # 현품표 조회
    try:
        tag = ProcessTag.objects.select_related('part').get(tag_id=tag_id)
    except ProcessTag.DoesNotExist:
        # RM 라벨인지 확인
        rm_label = RawMaterialLabel.objects.filter(label_id=tag_id).first()
        if rm_label and rm_label.status == 'CANCELLED':
            return JsonResponse({'success': False, 'error': f'취소된 라벨입니다. 사용할 수 없습니다: {tag_id}'})
        return JsonResponse({'success': False, 'error': f'현품표를 찾을 수 없습니다: {tag_id}'})

    if tag.status == 'CANCELLED':
        return JsonResponse({'success': False, 'error': f'취소된 현품표입니다. 사용할 수 없습니다: {tag_id}'})

    try:
        # 품목 마스터 조회
        part = tag.part
        part_no = tag.part_no
        part_name = tag.part_name

        if part:
            # 마스터에서 최신 품명 가져오기
            part_name = part.part_name

        # 스캔 수량만 기록 (품목별 합계 비교는 결과 화면에서)
        scanned_qty = tag.quantity

        # 항목 저장 (개별 비교 없이 스캔 기록만)
        item = InventoryCheckSessionItem.objects.create(
            check_session=check,
            process_tag=tag,
            tag_id=tag_id,
            part=part,
            part_no=part_no,
            part_name=part_name,
            lot_no=tag.lot_no,
            scanned_qty=scanned_qty,
            system_qty=0,  # 결과 화면에서 품목별로 계산
            discrepancy=0,
            is_matched=True,  # 임시
            scanned_by=request.user,
        )

        # 요약 업데이트
        check.update_summary()

        return JsonResponse({
            'success': True,
            'item': {
                'id': item.id,
                'tag_id': item.tag_id,
                'part_no': item.part_no,
                'part_name': item.part_name,
                'lot_no': item.lot_no.strftime('%Y-%m-%d') if item.lot_no else '-',
                'scanned_qty': item.scanned_qty,
                'system_qty': item.system_qty,
                'discrepancy': item.discrepancy,
                'is_matched': item.is_matched,
            },
            'summary': {
                'total_scanned': check.total_scanned,
                'total_matched': check.total_matched,
                'total_discrepancy': check.total_discrepancy,
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'서버 오류: {str(e)}',
            'traceback': traceback.format_exc()
        }, status=500)


@wms_permission_required('can_wms_stock_edit')
def inventory_check_complete(request, pk):
    """재고조사 완료"""
    check = get_object_or_404(InventoryCheckSession, pk=pk)

    if request.method == 'POST':
        check.status = 'COMPLETED'
        check.completed_by = request.user
        check.completed_at = timezone.now()
        check.update_summary()
        check.save()
        messages.success(request, f'재고조사 {check.check_no}가 완료되었습니다.')
        return redirect('material:inventory_check_result', pk=check.pk)

    return redirect('material:inventory_check_scan', pk=check.pk)


@wms_permission_required('can_wms_stock_view')
def inventory_check_result(request, pk):
    """재고조사 결과 - 품목별 집계"""
    from django.db.models import Sum, Count
    from collections import defaultdict

    check = get_object_or_404(InventoryCheckSession.objects.select_related('warehouse', 'created_by', 'completed_by'), pk=pk)
    items = check.check_items.select_related('part').order_by('-scanned_at')

    # 품목별 스캔 수량 집계
    part_scanned = check.check_items.values('part_no', 'part_name', 'part_id').annotate(
        total_scanned=Sum('scanned_qty'),
        tag_count=Count('id')
    ).order_by('part_no')

    # 품목별 결과 생성
    part_results = []
    total_matched = 0
    total_discrepancy = 0

    for row in part_scanned:
        part_no = row['part_no']
        part_id = row['part_id']
        scanned_total = row['total_scanned'] or 0
        tag_count = row['tag_count']

        # 시스템 재고 조회
        system_qty = MaterialStock.objects.filter(
            warehouse=check.warehouse,
            part_id=part_id,
        ).aggregate(total=Sum('quantity'))['total'] or 0

        discrepancy = scanned_total - system_qty
        is_matched = (discrepancy == 0)

        if is_matched:
            total_matched += 1
        else:
            total_discrepancy += 1

        part_results.append({
            'part_no': part_no,
            'part_name': row['part_name'],
            'tag_count': tag_count,
            'scanned_total': scanned_total,
            'system_qty': system_qty,
            'discrepancy': discrepancy,
            'is_matched': is_matched,
        })

    return render(request, 'material/inventory_check_result.html', {
        'check': check,
        'items': items,  # 개별 스캔 내역
        'part_results': part_results,  # 품목별 집계 결과
        'total_matched': total_matched,
        'total_discrepancy': total_discrepancy,
    })


# =============================================================================
# 재고 종합 조회 (피벗 테이블)
# =============================================================================

@wms_permission_required('can_wms_stock_view')
def inventory_summary(request):
    """재고조사 종합 집계표 - 창고별 피벗 테이블 (스캔 데이터 기반)"""
    from collections import defaultdict
    from orders.models import Part

    # 활성 창고 목록
    warehouses = Warehouse.objects.filter(is_active=True).order_by('code')

    # 필터
    part_no_q = request.GET.get('part_no', '').strip()
    part_name_q = request.GET.get('part_name', '').strip()
    category_q = request.GET.get('category', '').strip()

    # 스캔 데이터 조회 (InventoryCheckSessionItem에서)
    items = InventoryCheckSessionItem.objects.select_related(
        'check_session', 'check_session__warehouse', 'part'
    )

    if part_no_q:
        items = items.filter(part_no__icontains=part_no_q)
    if part_name_q:
        items = items.filter(part_name__icontains=part_name_q)
    if category_q:
        items = items.filter(part__part_group=category_q)

    # 품목별/창고별로 그룹화 (피벗 테이블 생성)
    part_data = defaultdict(lambda: {
        'part': None,
        'part_name': '',
        'total': 0,
        'warehouses': defaultdict(int)
    })

    for item in items:
        part_no = item.part_no
        wh_code = item.check_session.warehouse.code
        part_data[part_no]['part'] = item.part
        part_data[part_no]['part_name'] = item.part_name
        part_data[part_no]['total'] += item.scanned_qty
        part_data[part_no]['warehouses'][wh_code] += item.scanned_qty

    # 리스트로 변환하여 정렬
    summary_list = []
    for part_no, data in sorted(part_data.items()):
        row = {
            'part_no': part_no,
            'part': data['part'],
            'part_name': data.get('part_name', ''),
            'total': data['total'],
            'warehouse_qty': {wh.code: data['warehouses'].get(wh.code, 0) for wh in warehouses}
        }
        summary_list.append(row)

    # 창고별 소계
    warehouse_totals = {wh.code: 0 for wh in warehouses}
    grand_total = 0
    for row in summary_list:
        grand_total += row['total']
        for wh in warehouses:
            warehouse_totals[wh.code] += row['warehouse_qty'].get(wh.code, 0)

    # 품목군 목록 (필터용)
    categories = Part.objects.exclude(part_group__isnull=True).exclude(part_group='').values_list('part_group', flat=True).distinct()

    return render(request, 'material/inventory_summary.html', {
        'summary_list': summary_list,
        'warehouses': warehouses,
        'warehouse_totals': warehouse_totals,
        'grand_total': grand_total,
        'categories': categories,
        'part_no_q': part_no_q,
        'part_name_q': part_name_q,
        'category_q': category_q,
    })


@wms_permission_required('can_wms_stock_view')
def inventory_summary_excel(request):
    """재고조사 종합 집계표 - 엑셀 다운로드 (창고별 피벗)"""
    from collections import defaultdict
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # 활성 창고 목록
    warehouses = Warehouse.objects.filter(is_active=True).order_by('code')

    # 필터
    part_no_q = request.GET.get('part_no', '').strip()
    part_name_q = request.GET.get('part_name', '').strip()
    category_q = request.GET.get('category', '').strip()

    # 스캔 데이터 조회
    items = InventoryCheckSessionItem.objects.select_related(
        'check_session', 'check_session__warehouse', 'part'
    )

    if part_no_q:
        items = items.filter(part_no__icontains=part_no_q)
    if part_name_q:
        items = items.filter(part_name__icontains=part_name_q)
    if category_q:
        items = items.filter(part__part_group=category_q)

    # 품목별/창고별로 그룹화
    part_data = defaultdict(lambda: {
        'part': None,
        'part_name': '',
        'total': 0,
        'warehouses': defaultdict(int)
    })

    for item in items:
        part_no = item.part_no
        wh_code = item.check_session.warehouse.code
        part_data[part_no]['part'] = item.part
        part_data[part_no]['part_name'] = item.part_name
        part_data[part_no]['total'] += item.scanned_qty
        part_data[part_no]['warehouses'][wh_code] += item.scanned_qty

    # 엑셀 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "재고조사집계"

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    number_format = '#,##0'

    # 제목
    today = datetime.now().strftime('%y년 %m월')
    ws['A1'] = f"{today} 재고조사 통합 집계본"
    ws['A1'].font = Font(bold=True, size=14, color="2E7D32")
    ws.merge_cells('A1:F1')

    ws['A2'] = f"Ver. 2 [{datetime.now().strftime('%y.%m.%d')}]"
    ws['A2'].font = Font(size=9, color="666666")

    # 헤더 행 (4행부터)
    headers = ['품번', '품명', '규격', '계정', '단위', '품목군명', 'TOTAL']
    for wh in warehouses:
        headers.append(wh.name)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 데이터 행
    row_num = 5
    grand_total = 0
    warehouse_totals = {wh.code: 0 for wh in warehouses}

    for part_no, data in sorted(part_data.items()):
        part = data['part']
        ws.cell(row=row_num, column=1, value=part_no).border = thin_border
        ws.cell(row=row_num, column=2, value=data['part_name']).border = thin_border
        ws.cell(row=row_num, column=3, value=getattr(part, 'spec', '') or '' if part else '').border = thin_border
        ws.cell(row=row_num, column=4, value=getattr(part, 'account_type', '원재료') or '원재료' if part else '원재료').border = thin_border
        ws.cell(row=row_num, column=5, value=getattr(part, 'unit', 'EA') or 'EA' if part else 'EA').border = thin_border
        ws.cell(row=row_num, column=6, value=getattr(part, 'part_group', '') or '' if part else '').border = thin_border

        total_cell = ws.cell(row=row_num, column=7, value=data['total'])
        total_cell.border = thin_border
        total_cell.number_format = number_format
        total_cell.font = Font(bold=True)
        grand_total += data['total']

        for col_idx, wh in enumerate(warehouses, 8):
            qty = data['warehouses'].get(wh.code, 0)
            cell = ws.cell(row=row_num, column=col_idx, value=qty if qty > 0 else None)
            cell.border = thin_border
            if qty > 0:
                cell.number_format = number_format
                warehouse_totals[wh.code] += qty

        row_num += 1

    # 합계 행
    if part_data:
        ws.cell(row=row_num, column=1, value="합계").font = Font(bold=True)
        for col in range(1, 7):
            ws.cell(row=row_num, column=col).border = thin_border

        total_cell = ws.cell(row=row_num, column=7, value=grand_total)
        total_cell.border = thin_border
        total_cell.number_format = number_format
        total_cell.font = Font(bold=True)

        for col_idx, wh in enumerate(warehouses, 8):
            cell = ws.cell(row=row_num, column=col_idx, value=warehouse_totals[wh.code] or None)
            cell.border = thin_border
            if warehouse_totals[wh.code] > 0:
                cell.number_format = number_format
                cell.font = Font(bold=True)

    # 열 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    for i, wh in enumerate(warehouses, 8):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"inventory_check_summary_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# =============================================================================
# ERP 동기화 - 엑셀 변환 내보내기
# =============================================================================

@wms_permission_required('can_wms_stock_view')
def erp_sync(request):
    """
    ERP 동기화 페이지 - 입고/출고/이동 내역을 ERP 양식 엑셀로 변환
    """
    from datetime import datetime, timedelta

    # 기본 날짜 범위 (오늘)
    today = timezone.now().date()
    date_from = request.GET.get('date_from', today.strftime('%Y-%m-%d'))
    date_to = request.GET.get('date_to', today.strftime('%Y-%m-%d'))
    sync_type = request.GET.get('sync_type', 'IN')  # IN, OUT, TRANSFER

    # 날짜 파싱
    try:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    except ValueError:
        start_date = today
        end_date = today

    # 트랜잭션 유형별 필터링
    if sync_type == 'IN':
        type_filter = ['IN_SCM', 'IN_MANUAL']
        title = '입고 처리'
    elif sync_type == 'OUT':
        type_filter = ['OUT_PROD', 'OUT_RETURN']
        title = '출고 처리'
    else:  # TRANSFER
        type_filter = ['TRANSFER']
        title = '이동 처리'

    # 트랜잭션 조회
    transactions = MaterialTransaction.objects.filter(
        transaction_type__in=type_filter,
        date__date__gte=start_date,
        date__date__lte=end_date
    ).select_related('part', 'warehouse_from', 'warehouse_to', 'vendor', 'actor').order_by('-date')

    # 창고 목록 (이동처리용)
    warehouses = Warehouse.objects.filter(is_active=True).order_by('code')

    context = {
        'title': title,
        'sync_type': sync_type,
        'date_from': date_from,
        'date_to': date_to,
        'transactions': transactions,
        'warehouses': warehouses,
        'transaction_count': transactions.count(),
    }

    return render(request, 'material/erp_sync.html', context)


@wms_permission_required('can_wms_stock_view')
def erp_sync_export(request):
    """
    ERP 양식 엑셀 내보내기 - 더존 iCUBE 업로드 형식
    - 입고: 거래구분, 입고일자, 거래처코드, 환종, 환율, 과세구분, 단가구분, 창고코드, 품번, 입고수량, 재고단위수량, 장소코드
    """
    from datetime import datetime
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    sync_type = request.GET.get('sync_type', 'IN')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # 날짜 파싱
    today = timezone.now().date()
    try:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else today
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else today
    except ValueError:
        start_date = today
        end_date = today

    # 트랜잭션 유형별 필터링
    if sync_type == 'IN':
        type_filter = ['IN_SCM', 'IN_MANUAL']
        title = '입고등록'
    elif sync_type == 'OUT':
        type_filter = ['OUT_PROD', 'OUT_RETURN']
        title = '출고등록'
    else:  # TRANSFER
        type_filter = ['TRANSFER']
        title = '재고이동'

    # 트랜잭션 조회
    transactions = MaterialTransaction.objects.filter(
        transaction_type__in=type_filter,
        date__date__gte=start_date,
        date__date__lte=end_date
    ).select_related('part', 'warehouse_from', 'warehouse_to', 'vendor', 'actor').order_by('date')

    # 엑셀 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'sheet1'

    # 스타일
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
    code_fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')

    if sync_type == 'IN':
        # ========== 입고등록 ERP 양식 ==========
        # Row 1: 한글 헤더
        headers_kr = ['거래구분', '입고일자', '거래처코드', '환종', '환율', '과세구분', '단가구분', '창고코드', '품번', '입고수량', '재고단위수량', '장소코드']
        for col, header in enumerate(headers_kr, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Row 2: ERP 코드
        headers_code = ['PO_FG', 'RCV_DT', 'TR_CD', 'EXCH_CD', 'EXCH_RT', 'VAT_FG', 'UMVAT_FG', 'WH_CD', 'ITEM_CD', 'PO_QT', 'RCV_QT', 'LC_CD']
        for col, code in enumerate(headers_code, 1):
            cell = ws.cell(row=2, column=col, value=code)
            cell.font = Font(size=9)
            cell.fill = code_fill
            cell.alignment = Alignment(horizontal='center')

        # Row 3+: 데이터
        row_num = 3
        for trx in transactions:
            part = trx.part
            vendor = trx.vendor

            row_data = [
                '0',  # 거래구분: 0=DOMESTIC
                trx.date.strftime('%Y%m%d'),  # 입고일자: YYYYMMDD
                vendor.erp_code if vendor and vendor.erp_code else '',  # 거래처코드 (ERP코드)
                'KRW',  # 환종
                1,  # 환율
                '0',  # 과세구분: 0=매입과세
                '0',  # 단가구분: 0=부가세미포함
                trx.warehouse_to.code if trx.warehouse_to else '',  # 창고코드
                part.part_no if part else '',  # 품번
                abs(trx.quantity),  # 입고수량
                abs(trx.quantity),  # 재고단위수량 (동일)
                trx.warehouse_to.code if trx.warehouse_to else '',  # 장소코드 = 창고코드
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)

            row_num += 1

        # 열 너비 조정
        col_widths = {'A': 10, 'B': 12, 'C': 12, 'D': 8, 'E': 8, 'F': 10, 'G': 10, 'H': 10, 'I': 20, 'J': 12, 'K': 14, 'L': 10}

    elif sync_type == 'OUT':
        # ========== 출고등록 (임시 - 출고 양식 확인 후 수정 필요) ==========
        headers_kr = ['품번', '품명', '수량', '출고창고', '출고일자', '용도', '비고']
        for col, header in enumerate(headers_kr, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row_num = 2
        for trx in transactions:
            part = trx.part
            usage = '생산불출' if trx.transaction_type == 'OUT_PROD' else '반품출고'
            row_data = [
                part.part_no if part else '',
                part.part_name if part else '',
                abs(trx.quantity),
                trx.warehouse_from.code if trx.warehouse_from else '',
                trx.date.strftime('%Y%m%d'),
                usage,
                trx.remark or '',
            ]
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
            row_num += 1

        col_widths = {'A': 20, 'B': 25, 'C': 10, 'D': 12, 'E': 12, 'F': 12, 'G': 20}

    else:  # TRANSFER
        # ========== 재고이동 (임시 - 이동 양식 확인 후 수정 필요) ==========
        headers_kr = ['품번', '품명', '수량', '출발창고', '도착창고', '이동일자', '비고']
        for col, header in enumerate(headers_kr, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row_num = 2
        for trx in transactions:
            part = trx.part
            row_data = [
                part.part_no if part else '',
                part.part_name if part else '',
                abs(trx.quantity),
                trx.warehouse_from.code if trx.warehouse_from else '',
                trx.warehouse_to.code if trx.warehouse_to else '',
                trx.date.strftime('%Y%m%d'),
                trx.remark or '',
            ]
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
            row_num += 1

        col_widths = {'A': 20, 'B': 25, 'C': 10, 'D': 12, 'E': 12, 'F': 12, 'G': 20}

    # 열 너비 적용
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # 응답 생성
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    type_name = {'IN': 'incoming', 'OUT': 'outgoing', 'TRANSFER': 'transfer'}.get(sync_type, 'sync')
    filename = f"ERP_{title}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# =============================================================================
# 원재료 관리 - 창고 레이아웃, 입고, 랙 관리
# =============================================================================

from .models import RawMaterialRack, RawMaterialSetting, RawMaterialLabel


@wms_permission_required('can_wms_stock_view')
def raw_material_layout(request):
    """
    원재료 창고 레이아웃 - 실시간 재고 현황 표시
    실제 창고 구조 반영:
    - A열: 왼쪽 벽면 (2층/1층)
    - B열: 오른쪽 벽면 (2층/1층) - 마주보는 형태
    - 가운데: 통로
    """
    section = request.GET.get('section', '3F')

    # 감사모드 상태 조회
    from .models import WMSConfig
    config = WMSConfig.get_config()
    audit_mode_on = config.audit_mode

    # 해당 구역의 랙 목록 조회
    racks = RawMaterialRack.objects.filter(
        section=section,
        is_active=True
    ).select_related('part')

    def get_rack_info(rack):
        """랙별 재고 정보 조회 (3200 원재료창고 기준, 감사모드 지원)"""
        stock_qty = 0
        scanned_qty = 0
        stock_status = 'empty'
        base_qty = 0

        if rack.part:
            # 실재고 (3200 원재료창고)
            stock = MaterialStock.objects.filter(
                part=rack.part,
                warehouse__code='3200'
            ).aggregate(total=Sum('quantity'))
            actual_qty = stock['total'] or 0

            # 감사모드: 보정값(delta)이 있으면 실재고에 더함
            if audit_mode_on and rack.display_adjustment is not None:
                base_qty = actual_qty + rack.display_adjustment
            else:
                base_qty = actual_qty

            stock_qty = max(base_qty, 0)

            try:
                setting = rack.part.raw_material_setting
                safety = setting.safety_stock
                warning = setting.warning_stock
            except RawMaterialSetting.DoesNotExist:
                safety = 0
                warning = 0

            if stock_qty <= 0:
                stock_status = 'empty'
            elif stock_qty < warning:
                stock_status = 'danger'
            elif stock_qty < safety:
                stock_status = 'warning'
            else:
                stock_status = 'safe'

        return {
            'rack': rack,
            'stock_qty': stock_qty,
            'scanned_qty': scanned_qty,
            'stock_status': stock_status,
            'base_qty': base_qty,  # 감사모드 오버라이드 입력 시 표시용 (base = actual + adjustment)
        }

    # A열, B열 분리 후 col_num 오름차순 정렬
    wall_a = {'1': [], '2': []}
    wall_b = {'1': [], '2': []}

    for rack in racks:
        rack_info = get_rack_info(rack)
        floor = str(rack.row_num)

        if rack.row_label.upper() == 'A':
            wall_a[floor].append(rack_info)
        elif rack.row_label.upper() == 'B':
            wall_b[floor].append(rack_info)

    # col_num 오름차순 정렬
    for floor in wall_a:
        wall_a[floor] = sorted(wall_a[floor], key=lambda x: x['rack'].col_num)
    for floor in wall_b:
        wall_b[floor] = sorted(wall_b[floor], key=lambda x: x['rack'].col_num)

    # 유효기간 경고 데이터 (사이드바용)
    today = timezone.now().date()
    from datetime import timedelta
    expiry_expired = RawMaterialLabel.objects.filter(
        expiry_date__isnull=False,
        expiry_date__lt=today,
        status__in=['INSTOCK', 'PRINTED']
    ).count()
    expiry_imminent = RawMaterialLabel.objects.filter(
        expiry_date__isnull=False,
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30),
        status__in=['INSTOCK', 'PRINTED']
    ).count()
    expiry_warning = RawMaterialLabel.objects.filter(
        expiry_date__isnull=False,
        expiry_date__gt=today + timedelta(days=30),
        expiry_date__lte=today + timedelta(days=90),
        status__in=['INSTOCK', 'PRINTED']
    ).count()

    # 편집 권한 확인
    profile = getattr(request.user, 'userprofile', None)
    can_edit = request.user.is_superuser or (profile and getattr(profile, 'can_wms_stock_edit', False))

    # 도착 창고 목록 (3200 출발용 - 기본 3000)
    target_warehouses = Warehouse.objects.filter(
        is_active=True
    ).exclude(code='3200').order_by('code')

    context = {
        'section': section,
        'section_display': '3공장' if section == '3F' else '2공장',
        'wall_a': wall_a,
        'wall_b': wall_b,
        'sections': RawMaterialRack.SECTION_CHOICES,
        'expiry_expired': expiry_expired,
        'expiry_imminent': expiry_imminent,
        'expiry_warning': expiry_warning,
        'audit_mode': audit_mode_on,
        'can_edit': can_edit,
        'target_warehouses': target_warehouses,
    }

    return render(request, 'material/raw_material_layout.html', context)


# =============================================================================
# 감사모드 API
# =============================================================================
@wms_permission_required('can_wms_stock_edit')
def api_audit_mode_toggle(request):
    """감사모드 ON/OFF 토글"""
    from .models import WMSConfig
    config = WMSConfig.get_config()
    config.audit_mode = not config.audit_mode
    config.audit_mode_changed_at = timezone.now()
    config.audit_mode_changed_by = request.user
    config.save()
    return JsonResponse({
        'success': True,
        'audit_mode': config.audit_mode,
        'message': f"감사모드 {'활성화' if config.audit_mode else '비활성화'}"
    })


@wms_permission_required('can_wms_stock_edit')
def api_audit_mode_set_override(request):
    """특정 랙의 오버라이드 값 설정 (델타 방식: 입력값 - 실재고 = 보정값)"""
    data = json.loads(request.body)
    rack_id = data.get('rack_id')
    override_value = data.get('override_value')

    try:
        rack = RawMaterialRack.objects.select_related('part').get(id=rack_id)
    except RawMaterialRack.DoesNotExist:
        return JsonResponse({'success': False, 'error': '랙을 찾을 수 없습니다.'}, status=404)

    if override_value is not None and rack.part:
        desired = int(override_value)
        # 현재 실재고 조회
        actual = MaterialStock.objects.filter(
            part=rack.part,
            warehouse__code='3000'
        ).aggregate(total=Sum('quantity'))['total'] or 0
        rack.display_adjustment = desired - actual
    else:
        rack.display_adjustment = None
    rack.save(update_fields=['display_adjustment', 'updated_at'])

    return JsonResponse({
        'success': True,
        'rack_id': rack.id,
        'position_code': rack.position_code,
        'display_adjustment': rack.display_adjustment,
    })


@wms_permission_required('can_wms_stock_edit')
def api_audit_mode_clear_all(request):
    """모든 랙의 보정값 초기화"""
    count = RawMaterialRack.objects.filter(
        display_adjustment__isnull=False
    ).update(display_adjustment=None)
    return JsonResponse({
        'success': True,
        'cleared_count': count,
        'message': f'{count}개 랙의 오버라이드 값이 초기화되었습니다.'
    })


@wms_permission_required('can_wms_stock_view')
def raw_material_expiry(request):
    """
    유효기간 관리 - 품목설정에 등록된 품목 기준, LOT 입고일 + 보관기간으로 자동 계산
    """
    from datetime import timedelta
    from collections import defaultdict

    today = timezone.now().date()
    filter_status = request.GET.get('status', 'all')
    stock_search = request.GET.get('search', '').strip()

    # 품목설정에 등록된 품목만 대상
    settings_map = {}
    for s in RawMaterialSetting.objects.select_related('part').all():
        settings_map[s.part_id] = s

    all_items = []
    if settings_map:
        # 해당 품목들의 재고 (LOT 날짜 있고 수량 > 0)
        stock_qs = MaterialStock.objects.filter(
            part_id__in=settings_map.keys(),
            lot_no__isnull=False,
            quantity__gt=0,
        ).select_related('part', 'warehouse')

        if stock_search:
            from django.db.models import Q as _Q
            stock_qs = stock_qs.filter(
                _Q(part__part_no__icontains=stock_search) | _Q(part__part_name__icontains=stock_search)
            )

        # 품번 + LOT 기준 합산 (여러 창고에 걸칠 수 있으므로)
        grouped = defaultdict(lambda: {'total_qty': 0, 'shelf_life': 0})
        for stock in stock_qs:
            setting = settings_map.get(stock.part_id)
            if not setting:
                continue
            expiry_date = stock.lot_no + timedelta(days=setting.shelf_life_days)
            key = (stock.part.part_no, stock.part.part_name, stock.lot_no, expiry_date)
            grouped[key]['total_qty'] += stock.quantity
            grouped[key]['shelf_life'] = setting.shelf_life_days

        # 아이템 리스트 생성 + D-day/상태 계산
        for (part_no, part_name, lot_no, expiry_date), data in grouped.items():
            delta = (expiry_date - today).days
            if delta < 0:
                expiry_status = 'expired'
            elif delta <= 30:
                expiry_status = 'imminent'
            elif delta <= 90:
                expiry_status = 'warning'
            else:
                expiry_status = 'safe'
            all_items.append({
                'part_no': part_no,
                'part_name': part_name,
                'lot_no': lot_no,
                'expiry_date': expiry_date,
                'total_qty': data['total_qty'],
                'unit_display': 'kg',
                'shelf_life': data['shelf_life'],
                'd_day': delta,
                'expiry_status': expiry_status,
            })

    # 요약 카운트 (전체 기준)
    count_expired = sum(1 for i in all_items if i['expiry_status'] == 'expired')
    count_imminent = sum(1 for i in all_items if i['expiry_status'] == 'imminent')
    count_warning = sum(1 for i in all_items if i['expiry_status'] == 'warning')
    count_safe = sum(1 for i in all_items if i['expiry_status'] == 'safe')

    # 상태별 필터링
    if filter_status == 'expired':
        labels = [i for i in all_items if i['expiry_status'] == 'expired']
    elif filter_status == 'imminent':
        labels = [i for i in all_items if i['expiry_status'] == 'imminent']
    elif filter_status == 'warning':
        labels = [i for i in all_items if i['expiry_status'] == 'warning']
    elif filter_status == 'safe':
        labels = [i for i in all_items if i['expiry_status'] == 'safe']
    else:
        labels = all_items

    # 유효기간 순 정렬
    labels.sort(key=lambda x: x['expiry_date'])

    # 현장 투입 이력 (USED 라벨)
    from django.db.models import Q
    active_tab = request.GET.get('tab', 'stock')
    used_search = request.GET.get('used_search', '').strip()
    used_start = request.GET.get('used_start', '')
    used_end = request.GET.get('used_end', '')

    used_qs = RawMaterialLabel.objects.filter(
        status='USED',
        used_at__isnull=False,
    ).select_related('part', 'vendor', 'used_by').order_by('-used_at')

    if used_search:
        used_qs = used_qs.filter(
            Q(part_no__icontains=used_search) | Q(part_name__icontains=used_search)
        )
    if used_start:
        used_qs = used_qs.filter(used_at__date__gte=used_start)
    if used_end:
        used_qs = used_qs.filter(used_at__date__lte=used_end)

    # 투입 시점 잔여 D-DAY 계산 (고정값)
    for ul in used_qs:
        if ul.expiry_date and ul.used_at:
            ul.used_d_day = (ul.expiry_date - ul.used_at.date()).days
        else:
            ul.used_d_day = None

    context = {
        'labels': labels,
        'filter_status': filter_status,
        'today': today,
        'count_expired': count_expired,
        'count_imminent': count_imminent,
        'count_warning': count_warning,
        'count_safe': count_safe,
        'count_total': count_expired + count_imminent + count_warning + count_safe,
        'active_tab': active_tab,
        'used_labels': used_qs,
        'used_count': used_qs.count(),
        'used_search': used_search,
        'used_start': used_start,
        'used_end': used_end,
        'stock_search': stock_search,
    }

    return render(request, 'material/raw_material_expiry.html', context)


@wms_permission_required('can_wms_inout_view')
def raw_material_incoming(request):
    """
    원재료 입고 처리 - QR 라벨 발행 포함
    수입검사 OK 판정된 건에 대해서만 라벨 발행 가능
    각 라벨은 고유하며, 동일 입고 건에 대해 중복 발행 불가
    """
    from datetime import timedelta

    if request.method == 'POST':
        action = request.POST.get('action')

        # 라벨 취소 (소프트삭제 - CANCELLED 상태로 변경)
        if action == 'cancel_labels':
            inspection_id = request.POST.get('inspection_id')
            try:
                insp = ImportInspection.objects.get(id=inspection_id)
                trx = insp.inbound_transaction
                cancelled_count = RawMaterialLabel.objects.filter(
                    incoming_transaction=trx
                ).exclude(status='CANCELLED').update(
                    status='CANCELLED',
                )
                if cancelled_count > 0:
                    # 트랜잭션 remark에서 라벨 발행 기록 제거
                    remark = trx.remark or ''
                    import re
                    remark = re.sub(r'\s*\[라벨 \d+장 발행.*?\]', '', remark).strip()
                    trx.remark = remark
                    trx.save(update_fields=['remark'])
                    messages.success(request, f'{trx.part.part_no} - 라벨 {cancelled_count}장 취소 완료')
                else:
                    messages.info(request, '취소할 라벨이 없습니다.')
            except ImportInspection.DoesNotExist:
                messages.error(request, '해당 수입검사 건을 찾을 수 없습니다.')
            except Exception as e:
                messages.error(request, f'라벨 취소 실패: {str(e)}')
            return redirect('material:raw_material_incoming')

        # 라벨 발행 처리 (수입검사 OK 건)
        if action == 'print_labels':
            inspection_id = request.POST.get('inspection_id')

            try:
                inspection = ImportInspection.objects.get(id=inspection_id)

                # 검사 상태 확인
                if inspection.status != 'APPROVED':
                    messages.error(request, '수입검사 합격 판정이 필요합니다.')
                    return redirect('material:raw_material_incoming')

                trx = inspection.inbound_transaction
                part = trx.part
                qty = float(trx.quantity)
                lot = trx.lot_no or timezone.now().date()
                vendor = trx.vendor

                # 이미 라벨이 발행된 건인지 확인 (중복 방지)
                existing_labels = RawMaterialLabel.objects.filter(incoming_transaction=trx)
                if existing_labels.exists():
                    messages.warning(request, f'이미 라벨이 발행된 건입니다. (발행 라벨: {existing_labels.count()}장)')
                    label_ids = ','.join([str(l.id) for l in existing_labels])
                    return redirect(f'/wms/raw-material/label-print/?ids={label_ids}')

                # 모달에서 입력받은 값
                unit = request.POST.get('unit', 'KG')
                pkg_qty = float(request.POST.get('pkg_qty', 25))
                pkg_count = int(request.POST.get('pkg_count', 1))

                if pkg_qty <= 0 or pkg_count <= 0:
                    messages.error(request, '포장 단위수량과 포장 수는 0보다 커야 합니다.')
                    return redirect('material:raw_material_incoming')

                # 유효기간 - 모달에서 직접 선택한 날짜 사용
                expiry_date_str = request.POST.get('expiry_date', '').strip()
                if expiry_date_str:
                    from datetime import datetime as _dt
                    try:
                        expiry_date = _dt.strptime(expiry_date_str, '%Y-%m-%d').date()
                    except ValueError:
                        expiry_date = None
                else:
                    expiry_date = None

                # 라벨 발행만 수행 (재고는 입고처리 시 이미 추가됨)
                with transaction.atomic():
                    labels = []

                    for i in range(pkg_count):
                        label = RawMaterialLabel.objects.create(
                            label_id=RawMaterialLabel.generate_label_id(),
                            part=part,
                            part_no=part.part_no,
                            part_name=part.part_name,
                            lot_no=lot,
                            quantity=pkg_qty,
                            unit=unit,
                            expiry_date=expiry_date,
                            incoming_transaction=trx,
                            vendor=vendor,
                            status='INSTOCK',
                            printed_by=request.user,
                        )
                        labels.append(label)

                    # 입고 트랜잭션에 라벨 발행 기록
                    trx.remark = f'{trx.remark or ""} [라벨 {pkg_count}장 발행 ({pkg_qty}{dict(RawMaterialLabel.UNIT_CHOICES).get(unit, unit)}×{pkg_count})]'.strip()
                    trx.save()

                messages.success(request, f'{part.part_no} - {pkg_count}장 라벨 발행 완료 ({pkg_qty}{dict(RawMaterialLabel.UNIT_CHOICES).get(unit, unit)}×{pkg_count})')

                # 라벨 출력 페이지로 리다이렉트
                label_ids = ','.join([str(l.id) for l in labels])
                return redirect(f'/wms/raw-material/label-print/?ids={label_ids}')

            except ImportInspection.DoesNotExist:
                messages.error(request, '해당 수입검사 건을 찾을 수 없습니다.')
            except Exception as e:
                messages.error(request, f'라벨 발행 실패: {str(e)}')

            return redirect('material:raw_material_incoming')

    # 수입검사 합격(APPROVED) 건 목록 조회 - 전체 품목
    approved_inspections = []
    if ImportInspection:
        approved_inspections = ImportInspection.objects.filter(
            status='APPROVED'
        ).select_related(
            'inbound_transaction',
            'inbound_transaction__part',
            'inbound_transaction__vendor'
        ).order_by('-inspected_at')

        # 이미 라벨 발행된 건 표시 + SCM/WMS 구분 + 품목설정 로딩
        for insp in approved_inspections:
            trx = insp.inbound_transaction
            insp.is_scm = bool(trx.ref_delivery_order)
            insp.label_count = RawMaterialLabel.objects.filter(
                incoming_transaction=trx, label_type='PACKAGE'
            ).exclude(status='CANCELLED').count()
            insp.pallet_count = RawMaterialLabel.objects.filter(
                incoming_transaction=trx, label_type='PALLET'
            ).exclude(status='CANCELLED').count()
            # 품목설정 데이터 (모달 기본값용)
            try:
                setting = trx.part.raw_material_setting
                insp.setting_unit_weight = float(setting.unit_weight)
                insp.setting_shelf_life = setting.shelf_life_days
                insp.has_setting = True
            except RawMaterialSetting.DoesNotExist:
                insp.setting_unit_weight = 0
                insp.setting_shelf_life = 365
                insp.has_setting = False

    # 수입검사 대기중인 건 목록 - 전체 품목
    pending_inspections = []
    if ImportInspection:
        pending_inspections = ImportInspection.objects.filter(
            status='PENDING'
        ).select_related(
            'inbound_transaction',
            'inbound_transaction__part',
            'inbound_transaction__vendor'
        ).order_by('-inbound_transaction__date')[:10]

    # 최근 발행된 라벨 목록
    recent_labels = RawMaterialLabel.objects.select_related(
        'part', 'vendor'
    ).order_by('-printed_at')[:20]

    context = {
        'approved_inspections': approved_inspections,
        'pending_inspections': pending_inspections,
        'recent_labels': recent_labels,
        'today': timezone.now().date().strftime('%Y-%m-%d'),
    }

    return render(request, 'material/raw_material_incoming.html', context)


@wms_permission_required('can_wms_stock_view')
def raw_material_rack_manage(request):
    """
    랙 위치 관리 - 추가/수정/삭제, 품목 배치
    """
    section = request.GET.get('section', '3F')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_rack':
            position_code = request.POST.get('position_code')
            row_label = request.POST.get('row_label')
            row_num = request.POST.get('row_num')
            col_num = request.POST.get('col_num')

            if position_code and row_label:
                # 중복 체크 (같은 구역 내에서만)
                if RawMaterialRack.objects.filter(section=section, position_code=position_code).exists():
                    messages.error(request, f'[{section}] 위치코드 {position_code}는 이미 존재합니다.')
                else:
                    RawMaterialRack.objects.create(
                        section=section,
                        position_code=position_code,
                        row_label=row_label,
                        row_num=int(row_num or 1),
                        col_num=int(col_num or 1),
                    )
                    messages.success(request, f'랙 위치 {position_code} 추가 완료')

        elif action == 'assign_part':
            rack_id = request.POST.get('rack_id')
            part_id = request.POST.get('part_id')

            try:
                rack = RawMaterialRack.objects.get(id=rack_id)
            except RawMaterialRack.DoesNotExist:
                messages.error(request, '해당 랙을 찾을 수 없습니다.')
                return redirect(f'/wms/raw-material/rack-manage/?section={section}')
            rack.part_id = part_id if part_id else None
            rack.save()
            messages.success(request, f'{rack.position_code} 품목 배치 완료')

        elif action == 'swap_parts':
            # 랙 간 품목 교환
            source_rack_id = request.POST.get('source_rack_id')
            target_rack_id = request.POST.get('target_rack_id')

            try:
                source_rack = RawMaterialRack.objects.get(id=source_rack_id)
                target_rack = RawMaterialRack.objects.get(id=target_rack_id)
            except RawMaterialRack.DoesNotExist:
                messages.error(request, '해당 랙을 찾을 수 없습니다.')
                return redirect(f'/wms/raw-material/rack-manage/?section={section}')

            # 두 랙의 품목을 교환
            source_part = source_rack.part
            target_part = target_rack.part

            source_rack.part = target_part
            target_rack.part = source_part

            source_rack.save()
            target_rack.save()

            messages.success(request, f'{source_rack.position_code} ↔ {target_rack.position_code} 품목 교환 완료')

        elif action == 'clear_rack':
            rack_id = request.POST.get('rack_id')
            try:
                rack = RawMaterialRack.objects.get(id=rack_id)
            except RawMaterialRack.DoesNotExist:
                messages.error(request, '해당 랙을 찾을 수 없습니다.')
                return redirect(f'/wms/raw-material/rack-manage/?section={section}')
            position = rack.position_code
            rack.part = None
            rack.save(update_fields=['part'])
            messages.success(request, f'랙 위치 {position} 품목 제거 완료')

        elif action == 'delete_rack':
            rack_id = request.POST.get('rack_id')
            try:
                rack = RawMaterialRack.objects.get(id=rack_id)
            except RawMaterialRack.DoesNotExist:
                messages.error(request, '해당 랙을 찾을 수 없습니다.')
                return redirect(f'/wms/raw-material/rack-manage/?section={section}')
            position = rack.position_code
            rack.delete()
            messages.success(request, f'랙 위치 {position} 삭제 완료')

        return redirect(f'/wms/raw-material/rack-manage/?section={section}')

    racks = RawMaterialRack.objects.filter(section=section).select_related('part')
    parts = Part.objects.all().order_by('part_no')

    # 이미 랙에 배치된 품목 ID 목록
    assigned_parts = set(RawMaterialRack.objects.filter(part__isnull=False).values_list('part_id', flat=True))

    # A열, B열 분리 후 col_num 오름차순 정렬
    wall_a = {'1': [], '2': []}
    wall_b = {'1': [], '2': []}

    for rack in racks:
        floor = str(rack.row_num)
        if rack.row_label.upper() == 'A':
            wall_a[floor].append(rack)
        elif rack.row_label.upper() == 'B':
            wall_b[floor].append(rack)

    # col_num 오름차순 정렬
    for floor in wall_a:
        wall_a[floor] = sorted(wall_a[floor], key=lambda x: x.col_num)
    for floor in wall_b:
        wall_b[floor] = sorted(wall_b[floor], key=lambda x: x.col_num)

    # 각 층별 최대 col_num (다음 번호 계산용)
    max_col = {
        'A': {
            '1': max([r.col_num for r in wall_a['1']], default=0),
            '2': max([r.col_num for r in wall_a['2']], default=0),
        },
        'B': {
            '1': max([r.col_num for r in wall_b['1']], default=0),
            '2': max([r.col_num for r in wall_b['2']], default=0),
        }
    }

    context = {
        'section': section,
        'section_display': '3공장' if section == '3F' else '2공장',
        'wall_a': wall_a,
        'wall_b': wall_b,
        'max_col': max_col,
        'parts': parts,
        'assigned_parts': assigned_parts,
        'sections': RawMaterialRack.SECTION_CHOICES,
    }

    return render(request, 'material/raw_material_rack_manage.html', context)


@wms_permission_required('can_wms_stock_view')
def raw_material_setting(request):
    """
    원재료 품목 설정 - 안전재고, 보관기간 설정
    """
    if request.method == 'POST':
        part_id = request.POST.get('part_id')
        safety_stock = request.POST.get('safety_stock', 0)
        warning_stock = request.POST.get('warning_stock', 0)
        shelf_life_days = request.POST.get('shelf_life_days', 365)
        unit_weight = request.POST.get('unit_weight', 25)

        try:
            part = Part.objects.get(id=part_id)
        except Part.DoesNotExist:
            messages.error(request, '해당 품목을 찾을 수 없습니다.')
            return redirect('/wms/raw-material/setting/')
        setting, created = RawMaterialSetting.objects.update_or_create(
            part=part,
            defaults={
                'safety_stock': int(safety_stock),
                'warning_stock': int(warning_stock),
                'shelf_life_days': int(shelf_life_days),
                'unit_weight': float(unit_weight),
            }
        )
        messages.success(request, f'{part.part_no} 설정 저장 완료')
        return redirect('/wms/raw-material/setting/')

    # 랙에 배치된 품목 ID
    rack_part_ids = set(RawMaterialRack.objects.filter(
        part__isnull=False
    ).values_list('part_id', flat=True).distinct())

    # 이미 설정된 품목 (랙 배치 + 수동 추가 모두 포함)
    settings = RawMaterialSetting.objects.select_related('part').order_by('part__part_no')

    # 랙에 배치됐지만 아직 설정 안 된 품목은 드롭다운에 표시
    existing_part_ids = set(settings.values_list('part_id', flat=True))
    unsettled_rack_parts = Part.objects.filter(
        id__in=rack_part_ids - existing_part_ids
    ).order_by('part_no')

    context = {
        'settings': settings,
        'unsettled_rack_parts': unsettled_rack_parts,
    }

    return render(request, 'material/raw_material_setting.html', context)


@wms_permission_required('can_wms_stock_view')
def api_raw_material_labels(request):
    """
    수입검사 건의 발행된 라벨 목록 조회 API
    """
    from django.http import JsonResponse

    inspection_id = request.GET.get('inspection_id')
    if not inspection_id:
        return JsonResponse({'labels': []})

    try:
        inspection = ImportInspection.objects.get(id=inspection_id)
        trx = inspection.inbound_transaction
        labels = RawMaterialLabel.objects.filter(
            incoming_transaction=trx
        ).exclude(status='CANCELLED').order_by('id')

        result = []
        for label in labels:
            result.append({
                'id': label.id,
                'label_id': label.label_id,
                'part_no': label.part_no,
                'quantity': float(label.quantity),
                'unit': label.get_unit_display(),
                'lot_no': label.lot_no.strftime('%Y-%m-%d') if label.lot_no else '-',
                'expiry_date': label.expiry_date.strftime('%Y-%m-%d') if label.expiry_date else None,
                'printed_at': label.printed_at.strftime('%m/%d %H:%M') if label.printed_at else '-',
            })

        return JsonResponse({
            'labels': result,
            'part_no': trx.part.part_no,
            'part_name': trx.part.part_name,
            'total_qty': float(trx.quantity),
        })
    except ImportInspection.DoesNotExist:
        return JsonResponse({'labels': []})


@wms_permission_required('can_wms_stock_view')
def raw_material_label_print(request):
    """
    QR 라벨 출력 화면
    """
    ids = request.GET.get('ids', '')
    label_ids = [int(i) for i in ids.split(',') if i.isdigit()]

    labels = RawMaterialLabel.objects.filter(id__in=label_ids).select_related('part').order_by('id')

    context = {
        'labels': labels,
    }

    return render(request, 'material/raw_material_label_print.html', context)


@wms_permission_required('can_wms_stock_view')
def pallet_label_create(request):
    """
    파렛트 라벨 발행 API
    - 수입검사 합격 건에 대해 파렛트 라벨 발행
    - 사용자가 지정한 파렛트 단위 수량으로 분할 (예: 6,000kg에 1,000 지정 → 6장)
    - 미지정 시 전체 수량 1장 발행
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'})

    inspection_id = request.POST.get('inspection_id')
    # 사용자 지정 파렛트 단위 수량 (kg 또는 단위)
    pallet_unit_input = request.POST.get('pallet_unit_qty', '').strip()
    try:
        inspection = ImportInspection.objects.get(id=inspection_id)
        if inspection.status != 'APPROVED':
            return JsonResponse({'success': False, 'error': '수입검사 합격 판정이 필요합니다.'})

        trx = inspection.inbound_transaction
        part = trx.part
        qty = trx.quantity
        lot = trx.lot_no or timezone.now().date()
        vendor = trx.vendor

        # 이미 파렛트 라벨이 발행된 건인지 확인
        existing_pallets = RawMaterialLabel.objects.filter(
            incoming_transaction=trx, label_type='PALLET'
        ).exclude(status='CANCELLED')
        if existing_pallets.exists():
            label_ids_str = ','.join([str(l.id) for l in existing_pallets])
            return JsonResponse({
                'success': True,
                'already_exists': True,
                'label_ids': label_ids_str,
                'message': f'이미 파렛트 라벨이 발행되었습니다. ({existing_pallets.count()}장)'
            })

        # 단위 결정: 품목설정이 있으면 그 단위, 없으면 EA
        try:
            setting = part.raw_material_setting
            unit = setting.unit_weight_unit if hasattr(setting, 'unit_weight_unit') else 'KG'
        except Exception:
            unit = 'EA'

        # 파렛트 단위 수량 — 사용자가 지정하면 그 값으로 분할, 아니면 전체 1장
        from decimal import Decimal, InvalidOperation
        total_qty = Decimal(str(qty))
        pallet_unit_qty = None
        if pallet_unit_input:
            try:
                pallet_unit_qty = Decimal(pallet_unit_input)
                if pallet_unit_qty <= 0:
                    pallet_unit_qty = None
            except (InvalidOperation, ValueError):
                pallet_unit_qty = None

        # 분할 수량 계산
        split_qtys = []
        if pallet_unit_qty and total_qty > pallet_unit_qty:
            full_count = int(total_qty // pallet_unit_qty)
            remainder = total_qty - (pallet_unit_qty * full_count)
            for _ in range(full_count):
                split_qtys.append(pallet_unit_qty)
            if remainder > 0:
                split_qtys.append(remainder)
        else:
            split_qtys.append(total_qty)

        created_labels = []
        with transaction.atomic():
            for split_qty in split_qtys:
                label = RawMaterialLabel.objects.create(
                    label_id=RawMaterialLabel.generate_pallet_label_id(),
                    label_type='PALLET',
                    part=part,
                    part_no=part.part_no,
                    part_name=part.part_name,
                    lot_no=lot,
                    quantity=split_qty,
                    unit=unit,
                    incoming_transaction=trx,
                    vendor=vendor,
                    status='INSTOCK',
                    printed_by=request.user,
                )
                created_labels.append(label)

        label_ids_str = ','.join([str(l.id) for l in created_labels])
        return JsonResponse({
            'success': True,
            'label_ids': label_ids_str,
            'label_count': len(created_labels),
            'message': f'파렛트 라벨 {len(created_labels)}장 발행 완료 ({part.part_no}, 총 {qty}{unit})'
        })

    except ImportInspection.DoesNotExist:
        return JsonResponse({'success': False, 'error': '해당 수입검사 건을 찾을 수 없습니다.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@wms_permission_required('can_wms_stock_view')
def pallet_label_print(request):
    """
    파렛트 라벨 출력 화면 (A4 크기) - 복수 라벨 지원 (ids=1,2,3)
    """
    label_id = request.GET.get('id', '')
    ids_param = request.GET.get('ids', '')

    labels = []
    if ids_param:
        id_list = [int(x) for x in ids_param.split(',') if x.strip().isdigit()]
        labels = list(RawMaterialLabel.objects.filter(
            id__in=id_list, label_type='PALLET'
        ).select_related('part', 'vendor', 'incoming_transaction__warehouse_to'))
        # id_list 순서대로 정렬
        labels.sort(key=lambda l: id_list.index(l.id))
    elif label_id.isdigit():
        lbl = RawMaterialLabel.objects.filter(id=int(label_id), label_type='PALLET').select_related(
            'part', 'vendor', 'incoming_transaction__warehouse_to'
        ).first()
        if lbl:
            labels = [lbl]

    context = {
        'label': labels[0] if labels else None,  # 기존 단일 템플릿 호환
        'labels': labels,
    }
    return render(request, 'material/pallet_label_print.html', context)


@wms_permission_required('can_wms_stock_view')
def molding_loss_excel(request):
    """유실 사유별 상세 엑셀 다운로드"""
    from .models import MoldingDailyRecord, MoldingLossDetail
    from django.db.models import Sum
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    now = timezone.localtime()
    year = int(request.GET.get('year', now.year))
    month = int(request.GET.get('month', now.month))

    # 유실 상세 조회
    loss_details = MoldingLossDetail.objects.filter(
        record__date__year=year, record__date__month=month
    ).select_related('record__machine').order_by('record__date', 'record__machine__code', 'category')

    # 요약: 사유별 합계
    summary_qs = MoldingLossDetail.objects.filter(
        record__date__year=year, record__date__month=month
    ).values('category').annotate(total=Sum('minutes')).order_by('-total')
    summary = list(summary_qs)
    total_all = sum(s['total'] for s in summary) or 1

    wb = openpyxl.Workbook()

    # Sheet 1: 요약
    ws1 = wb.active
    ws1.title = '유실사유 요약'
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['순위', '사유', '유실시간(분)', '유실시간(시간)', '비율(%)']
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        c.border = thin

    for idx, s in enumerate(summary, 1):
        ws1.cell(row=idx+1, column=1, value=idx).border = thin
        ws1.cell(row=idx+1, column=2, value=s['category']).border = thin
        ws1.cell(row=idx+1, column=3, value=s['total']).border = thin
        ws1.cell(row=idx+1, column=4, value=round(s['total']/60, 2)).border = thin
        ws1.cell(row=idx+1, column=5, value=round(s['total']/total_all*100, 2)).border = thin

    # 합계 행
    total_row = len(summary) + 2
    ws1.cell(row=total_row, column=1, value='합계').font = Font(bold=True)
    ws1.cell(row=total_row, column=3, value=total_all).font = Font(bold=True)
    ws1.cell(row=total_row, column=4, value=round(total_all/60, 2)).font = Font(bold=True)

    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 10

    # Sheet 2: 상세 내역
    ws2 = wb.create_sheet('유실 상세 내역')
    headers2 = ['일자', '호기', '톤수', '근무조', '사유', '유실시간(분)']
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
        c.border = thin

    for idx, d in enumerate(loss_details, 2):
        r = d.record
        ws2.cell(row=idx, column=1, value=r.date.strftime('%Y-%m-%d')).border = thin
        ws2.cell(row=idx, column=2, value=r.machine.code if r.machine else '').border = thin
        ws2.cell(row=idx, column=3, value=r.machine.tonnage if r.machine else '').border = thin
        ws2.cell(row=idx, column=4, value=r.shift).border = thin
        ws2.cell(row=idx, column=5, value=d.category).border = thin
        ws2.cell(row=idx, column=6, value=d.minutes).border = thin

    for col_letter, w in [('A',12),('B',10),('C',8),('D',8),('E',16),('F',14)]:
        ws2.column_dimensions[col_letter].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="molding_loss_{year}{month:02d}.xlsx"'
    wb.save(response)
    return response


@wms_permission_required('can_wms_stock_view')
def api_molding_cavity(request):
    """성형 마스터에서 품번의 캐비티수 조회 (금형 등록 시 cv_count 자동입력용)"""
    from .models import MoldingMaster
    part_no = (request.GET.get('part_no') or '').strip()
    if not part_no:
        return JsonResponse({'success': False, 'error': '품번 필요'})
    mm = MoldingMaster.objects.filter(part_no=part_no).order_by('-id').first()
    if not mm or not mm.cavity:
        return JsonResponse({'success': False, 'cavity': 0})
    return JsonResponse({'success': True, 'cavity': mm.cavity})


@wms_permission_required('can_wms_stock_view')
def api_molding_master_lookup(request):
    """성형 마스터 등록용 품번 자동완성 정보 조회
    - 기존 MoldingMaster 동일 품번 있으면 그 정보 반환
    - 없으면 Part 마스터에서 품명/품목만 반환
    """
    from .models import MoldingMaster
    part_no = (request.GET.get('part_no') or '').strip()
    if not part_no:
        return JsonResponse({'success': False, 'error': '품번 필요'})

    # 1) 기존 성형 마스터 (가장 최근 것)
    existing = MoldingMaster.objects.filter(part_no=part_no).order_by('-id').first()
    if existing:
        return JsonResponse({
            'success': True,
            'source': 'molding_master',
            'part_no': existing.part_no,
            'part_name': existing.part_name or '',
            'item_group': existing.item_group or '',
            'material_type': existing.material_type or '',
            'material_part_no': existing.material_part_no or '',
            'material_name': existing.material_name or '',
        })

    # 2) Part 마스터
    part = Part.objects.filter(part_no=part_no).first()
    if part:
        return JsonResponse({
            'success': True,
            'source': 'part_master',
            'part_no': part.part_no,
            'part_name': part.part_name or '',
            'item_group': part.part_group or '',
            'material_type': '',
            'material_part_no': '',
            'material_name': '',
        })

    return JsonResponse({'success': False, 'error': '품번을 찾을 수 없습니다.'})


@wms_permission_required('can_wms_stock_view')
def api_part_search(request):
    """
    품목 검색 API - 품번/품명으로 검색
    ?exclude_setting=1 → 이미 품목설정에 등록된 품목 제외
    """
    from django.http import JsonResponse
    from django.db.models import Q

    query = request.GET.get('q', '').strip()
    exclude_setting = request.GET.get('exclude_setting', '')

    if len(query) < 2:
        return JsonResponse({'results': []})

    # 이미 랙에 배치된 품목 ID 목록
    assigned_parts = set(RawMaterialRack.objects.filter(
        part__isnull=False
    ).values_list('part_id', flat=True))

    # 품번 또는 품명으로 검색
    qs = Part.objects.filter(
        Q(part_no__icontains=query) | Q(part_name__icontains=query)
    )

    # 이미 설정된 품목 제외
    if exclude_setting == '1':
        existing_ids = set(RawMaterialSetting.objects.values_list('part_id', flat=True))
        qs = qs.exclude(id__in=existing_ids)

    parts = qs.order_by('part_no')[:30]

    results = []
    for part in parts:
        results.append({
            'id': part.id,
            'part_no': part.part_no,
            'part_name': part.part_name,
            'assigned': part.id in assigned_parts,
        })

    return JsonResponse({'results': results})


@wms_permission_required('can_wms_stock_view')
def api_labels_for_lot(request):
    """
    [API] 특정 품번+LOT의 사용 가능 라벨 목록 반환
    재고이동 시 라벨 선택용
    """
    part_no = (request.GET.get('part_no') or '').strip()
    lot_no_str = (request.GET.get('lot_no') or '').strip()

    if not part_no or not lot_no_str:
        return JsonResponse({'success': False, 'error': '품번과 LOT를 지정해주세요.'})

    try:
        from datetime import datetime as dt
        lot_date = dt.strptime(lot_no_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'success': False, 'error': 'LOT 형식이 올바르지 않습니다.'})

    from .models import ProcessTag

    # RM 라벨 조회
    rm_labels = RawMaterialLabel.objects.filter(
        part_no=part_no,
        lot_no=lot_date,
        status__in=['INSTOCK', 'PRINTED']
    ).order_by('printed_at')

    # 공정현품표(TAG) 조회
    tags = ProcessTag.objects.filter(
        part_no=part_no,
        lot_no=lot_date,
        status='PRINTED'
    ).order_by('printed_at')

    result = []
    today = timezone.now().date()

    for lb in rm_labels:
        d_day = (lb.expiry_date - today).days if lb.expiry_date else None
        result.append({
            'id': lb.id,
            'label_id': lb.label_id,
            'label_type': 'RM',
            'quantity': float(lb.quantity),
            'unit': lb.get_unit_display(),
            'expiry_date': lb.expiry_date.strftime('%Y-%m-%d') if lb.expiry_date else None,
            'd_day': d_day,
            'printed_at': timezone.localtime(lb.printed_at).strftime('%Y-%m-%d %H:%M') if lb.printed_at else None,
        })

    for tag in tags:
        result.append({
            'id': tag.id,
            'label_id': tag.tag_id,
            'label_type': 'TAG',
            'quantity': float(tag.quantity),
            'unit': 'EA',
            'expiry_date': None,
            'd_day': None,
            'printed_at': timezone.localtime(tag.printed_at).strftime('%Y-%m-%d %H:%M') if tag.printed_at else None,
        })

    # 발행일 기준 정렬 (오래된 것 먼저 = FIFO)
    result.sort(key=lambda x: x['printed_at'] or '')

    return JsonResponse({'success': True, 'labels': result, 'count': len(result)})


@login_required
@wms_permission_required('can_wms_stock_edit')
def cancel_stock_move(request, trx_id):
    """[WMS] 재고이동 취소 - 재고 원복 + ERP 삭제 + 라벨 원복 + 트랜잭션 삭제"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'})

    try:
        trx = MaterialTransaction.objects.get(pk=trx_id, transaction_type__in=['TRANSFER', 'TRF_ERP'])
    except MaterialTransaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': '해당 재고이동 건을 찾을 수 없습니다.'})

    if trx.transaction_type == 'TRF_ERP':
        return JsonResponse({'success': False, 'error': 'ERP에서 동기화된 재고이동 건은 SCM에서 취소할 수 없습니다. ERP(아마란스)에서 삭제해주세요.'})

    # 마감 체크
    is_closed, warning_msg, _ = check_closing_date(
        trx.date.date() if hasattr(trx.date, 'date') and callable(trx.date.date) else trx.date
    )
    if is_closed:
        return JsonResponse({'success': False, 'error': f'마감된 기간의 이동 건은 취소할 수 없습니다. ({warning_msg})'})

    trx_no = trx.transaction_no

    try:
        with transaction.atomic():
            from .models import RawMaterialLabel, ProcessTag

            # 1. 받는 창고 재고 차감
            target_stock = MaterialStock.objects.filter(
                warehouse=trx.warehouse_to,
                part=trx.part,
                lot_no=trx.lot_no
            ).first()

            if not target_stock or target_stock.quantity < trx.quantity:
                current = target_stock.quantity if target_stock else 0
                return JsonResponse({
                    'success': False,
                    'error': f'받는 창고 재고({current})가 이동 수량({int(trx.quantity)})보다 적어 취소할 수 없습니다.'
                })

            MaterialStock.objects.filter(pk=target_stock.pk).update(
                quantity=F('quantity') - trx.quantity
            )

            # 2. 보내는 창고 재고 복구
            source_stock, _ = MaterialStock.objects.get_or_create(
                warehouse=trx.warehouse_from,
                part=trx.part,
                lot_no=trx.lot_no,
                defaults={'quantity': 0}
            )
            MaterialStock.objects.filter(pk=source_stock.pk).update(
                quantity=F('quantity') + trx.quantity
            )

            # 3. 연결된 라벨 원복 (USED → INSTOCK / PRINTED)
            RawMaterialLabel.objects.filter(
                used_transaction=trx, status='USED'
            ).update(
                status='INSTOCK', used_at=None, used_by=None, used_transaction=None
            )
            ProcessTag.objects.filter(
                used_transaction=trx, status='USED'
            ).update(
                status='PRINTED', used_at=None, used_by=None, used_warehouse=None, used_transaction=None
            )

            # 4. ERP 재고이동 삭제
            erp_no = trx.erp_incoming_no
            erp_msg = ''
            if erp_no:
                try:
                    from .erp_api import delete_erp_stock_move
                    erp_ok, erp_err = delete_erp_stock_move(erp_no)
                    if erp_ok:
                        erp_msg = f' (ERP 삭제 완료: {erp_no})'
                    else:
                        erp_msg = f' (ERP 삭제 실패: {erp_err})'
                except Exception as e:
                    logger.error(f'ERP 재고이동 삭제 예외: {e}')
                    erp_msg = f' (ERP 삭제 오류: {e})'

            # 5. 트랜잭션 삭제
            trx.delete()

        return JsonResponse({
            'success': True,
            'message': f'재고이동 [{trx_no}] 취소 완료{erp_msg}'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'취소 처리 중 오류: {str(e)}'})


@wms_permission_required('can_wms_stock_view')
def api_transfer_detail(request, trx_id):
    """
    [API] 재고이동 트랜잭션 상세 + 연결된 라벨 목록 반환
    """
    try:
        trx = MaterialTransaction.objects.select_related(
            'part', 'warehouse_from', 'warehouse_to', 'actor'
        ).get(pk=trx_id, transaction_type__in=['TRANSFER', 'TRF_ERP'])
    except MaterialTransaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': '이동 내역을 찾을 수 없습니다.'})

    # 연결된 라벨 (RM + TAG)
    labels_data = []
    for lb in trx.used_labels.all().order_by('label_id'):
        used_d_day = None
        if lb.expiry_date and lb.used_at:
            used_d_day = (lb.expiry_date - lb.used_at.date()).days
        labels_data.append({
            'label_id': lb.label_id,
            'label_type': 'RM',
            'quantity': float(lb.quantity),
            'unit': lb.get_unit_display(),
            'expiry_date': lb.expiry_date.strftime('%Y-%m-%d') if lb.expiry_date else '-',
            'used_d_day': used_d_day,
        })

    for tag in trx.used_tags.all().order_by('tag_id'):
        labels_data.append({
            'label_id': tag.tag_id,
            'label_type': 'TAG',
            'quantity': float(tag.quantity),
            'unit': 'EA',
            'expiry_date': '-',
            'used_d_day': None,
        })

    return JsonResponse({
        'success': True,
        'trx': {
            'transaction_no': trx.transaction_no,
            'transaction_type': trx.transaction_type,
            'date': timezone.localtime(trx.date).strftime('%Y-%m-%d %H:%M'),
            'part_no': trx.part.part_no,
            'part_name': trx.part.part_name,
            'quantity': int(trx.quantity),
            'lot_no': trx.lot_no.strftime('%Y-%m-%d') if trx.lot_no else '-',
            'from_wh': f"({trx.warehouse_from.code}) {trx.warehouse_from.name}" if trx.warehouse_from else '-',
            'to_wh': f"({trx.warehouse_to.code}) {trx.warehouse_to.name}" if trx.warehouse_to else '-',
            'actor': trx.actor.username if trx.actor else '-',
            'remark': trx.remark or '',
            'erp_no': trx.erp_incoming_no or '',
            'erp_status': trx.erp_sync_status or 'NONE',
            'erp_message': trx.erp_sync_message or '',
        },
        'labels': labels_data,
        'label_count': len(labels_data),
    })


# =============================================================================
# ERP 재고 관리 (관리자)
# =============================================================================

@login_required
@wms_permission_required('can_wms_stock_edit')
def erp_stock_manage(request):
    """[WMS] ERP 재고 관리 - 재고 동기화 / ERP vs SCM 비교"""
    from django.conf import settings as django_settings
    from django.core.cache import cache
    from datetime import date

    context = {
        'erp_enabled': getattr(django_settings, 'ERP_ENABLED', False),
        'today': date.today().isoformat(),
    }

    action = request.POST.get('action', '') if request.method == 'POST' else ''

    # ── ERP 재고 동기화 (ERP 현재고 → SCM lot_no=NULL 조정) ──
    if action == 'sync_stock':
        from material.erp_api import sync_stock_from_erp
        result = sync_stock_from_erp()
        if result.get('error'):
            messages.error(request, f'재고 동기화 실패: {result["error"]}')
        else:
            messages.success(
                request,
                f'ERP 재고 동기화 완료: '
                f'조정 {result["adjusted"]}건 '
                f'(증가 {result["increased"]}, 감소 {result["decreased"]}), '
                f'생성 {result.get("created", 0)}건, '
                f'건너뜀(Part없음) {result.get("skipped_no_part", 0)}건, '
                f'건너뜀(창고없음) {result.get("skipped_no_wh", 0)}건'
            )
        return redirect('material:erp_stock_manage')

    # ── 수불 동기화 (입고/출고/생산출고/생산입고/재고이동) ──
    if action == 'sync_transactions':
        from material.erp_api import (
            sync_erp_incoming, sync_erp_issue, sync_erp_receipt,
            sync_erp_stock_transfer, sync_erp_adjustments, sync_erp_outgoing
        )

        sync_jobs = [
            ('구매입고', sync_erp_incoming),
            ('고객출고', sync_erp_outgoing),
            ('생산출고', sync_erp_issue),
            ('생산입고', sync_erp_receipt),
            ('재고이동', sync_erp_stock_transfer),
            ('재고조정', sync_erp_adjustments),
        ]

        total_synced = 0
        total_skipped = 0
        total_errors = 0
        details = []

        for idx, (label, func) in enumerate(sync_jobs):
            cache.set('erp_sync_progress', {
                'stage': f'{label} 동기화 중...',
                'percent': int((idx / len(sync_jobs)) * 100),
                'detail': f'{idx}/{len(sync_jobs)} 완료',
            }, timeout=300)

            try:
                synced, skipped, errs, err_list = func()
                total_synced += synced
                total_skipped += skipped
                total_errors += errs
                if synced > 0 or errs > 0:
                    details.append(f'{label}: 반영 {synced}건')
                for e in err_list[:2]:
                    messages.warning(request, f'[{label}] {e}')
            except Exception as e:
                total_errors += 1
                messages.warning(request, f'[{label}] 오류: {str(e)[:100]}')

        cache.set('erp_sync_progress', {
            'stage': '완료',
            'percent': 100,
            'detail': f'총 반영 {total_synced}건',
        }, timeout=300)

        if total_synced > 0:
            detail_str = ' / '.join(details) if details else ''
            messages.success(request, f'수불 동기화 완료: 반영 {total_synced}건, 건너뜀 {total_skipped}건 ({detail_str})')
        else:
            messages.info(request, f'수불 동기화: 신규 건 없음 (건너뜀 {total_skipped}건)')

        return redirect('material:erp_stock_manage')

    # ── ERP vs SCM 비교 ──
    if action == 'compare' or request.GET.get('compare'):
        from material.erp_api import compare_erp_stock
        diff_only = request.GET.get('diff_only', request.POST.get('diff_only', ''))
        ok, comparison, summary, err = compare_erp_stock()
        if ok:
            # 차이 있는 것만 필터
            if diff_only:
                comparison = [c for c in comparison if c['diff'] != 0]
            context['comparison'] = comparison
            context['summary'] = summary
            context['diff_only'] = diff_only
        else:
            messages.error(request, f'ERP 재고 비교 실패: {err}')
        if request.method == 'POST':
            return redirect(f'{request.path}?compare=1&diff_only={diff_only}')

    return render(request, 'material/erp_stock_manage.html', context)


@wms_permission_required('can_wms_stock_view')
def erp_stock_init_progress(request):
    """기초재고 셋팅 진행률 조회 API (AJAX 폴링용)"""
    from django.core.cache import cache
    from django.http import JsonResponse
    progress = cache.get('erp_stock_init_progress')
    if progress:
        return JsonResponse(progress)
    return JsonResponse({'stage': '', 'percent': 0})


@wms_permission_required('can_wms_stock_view')
def erp_sync_progress(request):
    """범용 ERP 동기화 진행률 조회 API (AJAX 폴링용)"""
    from django.core.cache import cache
    from django.http import JsonResponse
    key = request.GET.get('key', 'erp_sync_progress')
    # 허용된 키만 조회
    allowed_keys = {'erp_sync_progress', 'erp_link_vendor_progress'}
    if key not in allowed_keys:
        key = 'erp_sync_progress'
    progress = cache.get(key)
    if progress:
        return JsonResponse(progress)
    return JsonResponse({'stage': '', 'percent': 0})


# =============================================================================
# ERP 마스터 동기화 (거래처/품목)
# =============================================================================
@login_required
@wms_permission_required('can_wms_stock_edit')
def erp_master_sync(request):
    """ERP 마스터 데이터 동기화 (거래처, 품목)"""
    from django.conf import settings as django_settings

    context = {
        'erp_enabled': getattr(django_settings, 'ERP_ENABLED', False),
        'vendor_count': Vendor.objects.count(),
        'part_count': Part.objects.count(),
        'no_vendor_count': Part.objects.filter(vendor__isnull=True).count(),
    }

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'sync_vendors':
            from material.erp_api import sync_erp_vendors
            result = sync_erp_vendors()
            if result['errors']:
                for err in result['errors'][:5]:
                    messages.warning(request, err)
            messages.success(
                request,
                f"거래처 동기화 완료: 전체 {result['total']}건 "
                f"(신규 {result['created']}, 갱신 {result['updated']}, "
                f"건너뜀 {result['skipped']})"
            )
            context['vendor_result'] = result

        elif action == 'sync_items':
            from material.erp_api import sync_erp_items
            result = sync_erp_items()
            if result['errors']:
                for err in result['errors'][:5]:
                    messages.warning(request, err)
            messages.success(
                request,
                f"품목 동기화 완료: 전체 {result['total']}건 "
                f"(신규 {result['created']}, 갱신 {result['updated']}, "
                f"건너뜀 {result['skipped']})"
            )
            context['item_result'] = result

        elif action == 'link_vendors':
            from material.erp_api import link_vendor_by_incoming
            months = int(request.POST.get('months', 6))
            result = link_vendor_by_incoming(months=months)
            if result['errors']:
                for err in result['errors'][:5]:
                    messages.warning(request, err)
            messages.success(
                request,
                f"입고이력 기반 업체 연결 완료: "
                f"입고 {result['total_headers']}건 분석, "
                f"매핑 {result['matched']}건, "
                f"연결 {result['updated']}건"
            )
            context['link_result'] = result

        # 동기화 후 카운트 갱신
        context['vendor_count'] = Vendor.objects.count()
        context['part_count'] = Part.objects.count()
        context['no_vendor_count'] = Part.objects.filter(vendor__isnull=True).count()

    return render(request, 'material/erp_master_sync.html', context)


# =============================================================================
# 출고 관리 (Outgoing)
# =============================================================================

@wms_permission_required('can_wms_inout_edit')
def manual_outgoing(request):
    """[WMS] 자재 수기 출고 처리"""
    if request.method == 'POST':
        try:
            date_str = request.POST.get('date', timezone.now().date())
            warehouse_id = request.POST.get('warehouse_id')
            vendor_id = request.POST.get('vendor_id')

            part_ids = request.POST.getlist('part_ids[]')
            lot_nos = request.POST.getlist('lot_nos[]')
            quantities = request.POST.getlist('quantities[]')
            remarks = request.POST.getlist('remarks[]')

            if not part_ids:
                messages.error(request, "출고할 품목이 리스트에 없습니다.")
                return redirect('material:manual_outgoing')

            # 마감 기간 검증 (경고만 표시)
            from datetime import datetime
            if isinstance(date_str, str):
                check_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                check_date = date_str
            is_closed, warning_msg, _ = check_closing_date(check_date)
            if is_closed:
                messages.warning(request, warning_msg)

            # 출고 일자를 DateTimeField에 맞게 datetime 객체로 변환
            if isinstance(date_str, str):
                date_value = timezone.make_aware(datetime.strptime(date_str, '%Y-%m-%d'))
            else:
                date_value = timezone.now()

            success_count = 0

            with transaction.atomic():
                warehouse = Warehouse.objects.get(id=warehouse_id)
                vendor = Vendor.objects.get(id=vendor_id) if vendor_id else None

                for i in range(len(part_ids)):
                    p_id = part_ids[i]
                    qty = int(quantities[i])
                    rmk = remarks[i] if i < len(remarks) else ''
                    lot_no_str = lot_nos[i] if i < len(lot_nos) and lot_nos[i] else None

                    if qty <= 0:
                        continue

                    part = Part.objects.get(id=p_id)

                    # LOT 번호 처리
                    from datetime import datetime
                    lot_date = None
                    if lot_no_str:
                        try:
                            lot_date = datetime.strptime(lot_no_str, '%Y-%m-%d').date()
                        except (ValueError, TypeError):
                            pass

                    # (1) 재고 확인 및 차감
                    stock, _ = MaterialStock.objects.get_or_create(
                        warehouse=warehouse,
                        part=part,
                        lot_no=lot_date,
                        defaults={'quantity': 0}
                    )

                    if stock.quantity < qty:
                        messages.warning(
                            request,
                            f"재고 부족: {part.part_no} (현재고 {stock.quantity}, 출고요청 {qty}) - 가용수량만큼 출고됩니다."
                        )

                    MaterialStock.objects.filter(pk=stock.pk).update(
                        quantity=Greatest(F('quantity') - qty, Value(0))
                    )
                    stock.refresh_from_db()

                    # (2) 수불 이력 생성
                    trx_no = f"OUT-{timezone.now().strftime('%y%m%d%H%M%S')}-{request.user.id}-{i}"
                    MaterialTransaction.objects.create(
                        transaction_no=trx_no,
                        transaction_type='OUT_MANUAL',
                        date=date_value,
                        part=part,
                        quantity=-qty,
                        lot_no=lot_date,
                        warehouse_from=warehouse,
                        result_stock=stock.quantity,
                        vendor=vendor,
                        actor=request.user,
                        remark=rmk,
                    )

                    success_count += 1

            if success_count > 0:
                messages.success(request, f"총 {success_count}건 출고 처리가 완료되었습니다.")
            else:
                messages.warning(request, "저장된 항목이 없습니다.")

            return redirect('material:manual_outgoing')

        except Exception as e:
            messages.error(request, f"오류 발생: {str(e)}")
            return redirect('material:manual_outgoing')

    # === 출고처리 내역 (History) ===
    history_qs = MaterialTransaction.objects.filter(
        transaction_type__in=['OUT_MANUAL', 'OUT_ERP', 'OUT_PROD', 'OUT_RETURN']
    ).select_related(
        'part', 'warehouse_from', 'vendor', 'actor'
    ).order_by('-date', '-id')

    history_q = (request.GET.get('hq') or '').strip()
    if history_q:
        history_qs = history_qs.filter(
            Q(part__part_no__icontains=history_q) |
            Q(part__part_name__icontains=history_q) |
            Q(transaction_no__icontains=history_q)
        )

    history_wh = (request.GET.get('hwh') or '').strip()
    if history_wh:
        history_qs = history_qs.filter(warehouse_from_id=history_wh)

    history_start = (request.GET.get('hstart') or '').strip()
    history_end = (request.GET.get('hend') or '').strip()
    if history_start and history_start not in ('None', 'null'):
        history_qs = history_qs.filter(date__date__gte=history_start)
    if history_end and history_end not in ('None', 'null'):
        history_qs = history_qs.filter(date__date__lte=history_end)

    history_paginator = Paginator(history_qs, 15)
    history_page = history_paginator.get_page(request.GET.get('hpage'))

    # 삭제 가능 여부 + 수량 절대값
    for item in history_page:
        item.can_cancel = (item.transaction_type == 'OUT_MANUAL')
        item.display_quantity = abs(item.quantity)

    warehouses_qs = Warehouse.objects.filter(is_active=True).order_by('code')

    context = {
        'warehouses': warehouses_qs,
        'vendors': Vendor.objects.all().order_by('name'),
        'parts': Part.objects.select_related('vendor').all().order_by('part_no'),
        'today': timezone.now().date(),
        'history_page': history_page,
        'history_q': history_q,
        'history_wh': history_wh,
        'history_start': history_start,
        'history_end': history_end,
    }
    return render(request, 'material/manual_outgoing.html', context)


@wms_permission_required('can_wms_inout_edit')
def cancel_manual_outgoing(request, trx_id):
    """[WMS] 수기 출고 삭제 - 재고 복원 + 트랜잭션 삭제"""
    if request.method != 'POST':
        return redirect('material:manual_outgoing')

    trx = get_object_or_404(
        MaterialTransaction, pk=trx_id,
        transaction_type__in=['OUT_MANUAL', 'OUT_PROD', 'OUT_ERP', 'OUT_RETURN']
    )

    if trx.transaction_type == 'OUT_ERP':
        messages.error(request, "ERP에서 동기화된 출고 건은 WMS에서 삭제할 수 없습니다.")
        return redirect('material:manual_outgoing')

    is_closed, warning_msg, _ = check_closing_date(
        trx.date.date() if hasattr(trx.date, 'date') and callable(trx.date.date) else trx.date
    )
    if is_closed:
        messages.error(request, f"마감된 기간의 출고 건은 삭제할 수 없습니다. ({warning_msg})")
        return redirect('material:manual_outgoing')

    trx_no = trx.transaction_no
    trx_qty = abs(trx.quantity)

    try:
        with transaction.atomic():
            # 재고 복원
            stock, _ = MaterialStock.objects.get_or_create(
                warehouse=trx.warehouse_from,
                part=trx.part,
                lot_no=trx.lot_no,
                defaults={'quantity': 0}
            )
            MaterialStock.objects.filter(pk=stock.pk).update(
                quantity=F('quantity') + trx_qty
            )

            trx.delete()

        messages.success(request, f"출고 건 [{trx_no}] 삭제 완료 (재고 {trx_qty}개 복원)")

    except Exception as e:
        messages.error(request, f"취소 처리 중 오류 발생: {str(e)}")

    return redirect('material:manual_outgoing')


@wms_permission_required('can_wms_inout_view')
def outgoing_history(request):
    """[WMS] 출고 이력 조회"""
    import re

    qs = MaterialTransaction.objects.filter(
        transaction_type__in=['OUT_MANUAL', 'OUT_ERP', 'OUT_RETURN']
    ).select_related(
        'part', 'warehouse_from', 'actor', 'vendor'
    ).order_by('-date', '-id')

    # 검색 필터
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(part__part_no__icontains=q) |
            Q(part__part_name__icontains=q)
        )

    q_group = (request.GET.get('q_group') or '').strip()
    if q_group:
        qs = qs.filter(part__part_group__icontains=q_group)

    q_vendor = (request.GET.get('q_vendor') or '').strip()
    if q_vendor:
        qs = qs.filter(vendor__name__icontains=q_vendor)

    start_date = (request.GET.get('start_date') or '').strip()
    end_date = (request.GET.get('end_date') or '').strip()
    if start_date in ('None', 'null', 'NULL'):
        start_date = ''
    if end_date in ('None', 'null', 'NULL'):
        end_date = ''
    if start_date:
        qs = qs.filter(date__date__gte=start_date)
    if end_date:
        qs = qs.filter(date__date__lte=end_date)

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    # 표시용 구분/비고/수량(절대값)
    for item in page_obj:
        item.display_quantity = abs(item.quantity)
        if item.transaction_type == 'OUT_MANUAL':
            item.display_type = "수기출고"
            item.display_remark = item.remark or ""
        elif item.transaction_type == 'OUT_ERP':
            item.display_type = "ERP출고"
            remark = item.remark or ""
            item.display_remark = re.sub(r'^ERP출고\([^)]*\)\s*', '', remark)
        elif item.transaction_type == 'OUT_RETURN':
            item.display_type = "반품출고"
            item.display_remark = item.remark or ""
        else:
            item.display_type = "출고"
            item.display_remark = item.remark or ""

    # 품목군/거래처 목록
    part_groups = list(
        Part.objects.exclude(part_group__isnull=True).exclude(part_group='')
        .values_list('part_group', flat=True).distinct().order_by('part_group')
    )
    vendors = list(
        Vendor.objects.values('id', 'name', 'code').order_by('name')
    )

    context = {
        'page_obj': page_obj,
        'q': q,
        'q_group': q_group,
        'q_vendor': q_vendor,
        'start_date': start_date,
        'end_date': end_date,
        'part_groups': part_groups,
        'vendors': vendors,
    }
    return render(request, 'material/outgoing_history.html', context)


@login_required
@wms_permission_required('can_wms_inout_view')
def outgoing_history_excel(request):
    """[WMS] 출고 이력 엑셀 다운로드"""
    import re
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    qs = MaterialTransaction.objects.filter(
        transaction_type__in=['OUT_MANUAL', 'OUT_ERP', 'OUT_RETURN']
    ).select_related(
        'part', 'warehouse_from', 'actor', 'vendor'
    ).order_by('-date', '-id')

    # 검색 필터 (outgoing_history와 동일)
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(part__part_no__icontains=q) |
            Q(part__part_name__icontains=q)
        )
    q_group = (request.GET.get('q_group') or '').strip()
    if q_group:
        qs = qs.filter(part__part_group__icontains=q_group)
    q_vendor = (request.GET.get('q_vendor') or '').strip()
    if q_vendor:
        qs = qs.filter(vendor__name__icontains=q_vendor)
    start_date = (request.GET.get('start_date') or '').strip()
    end_date = (request.GET.get('end_date') or '').strip()
    if start_date in ('None', 'null', 'NULL'):
        start_date = ''
    if end_date in ('None', 'null', 'NULL'):
        end_date = ''
    if start_date:
        qs = qs.filter(date__date__gte=start_date)
    if end_date:
        qs = qs.filter(date__date__lte=end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "출고내역"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 제목
    period = ""
    if start_date and end_date:
        period = f" ({start_date} ~ {end_date})"
    elif start_date:
        period = f" ({start_date} ~)"
    elif end_date:
        period = f" (~ {end_date})"
    ws.append([f"출고 내역 조회{period}"])
    ws.merge_cells('A1:I1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])

    # 헤더
    headers = ['No', '일자', '구분', '거래처', '수불번호', '품번', '품명', '수량', 'LOT 번호', '출고 창고', '비고']
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    for idx, item in enumerate(qs.iterator(), 1):
        display_qty = abs(item.quantity)
        if item.transaction_type == 'OUT_MANUAL':
            display_type = "수기출고"
            display_remark = item.remark or ""
        elif item.transaction_type == 'OUT_ERP':
            display_type = "ERP출고"
            remark = item.remark or ""
            display_remark = re.sub(r'^ERP출고\([^)]*\)\s*', '', remark)
        elif item.transaction_type == 'OUT_RETURN':
            display_type = "반품출고"
            display_remark = item.remark or ""
        else:
            display_type = "출고"
            display_remark = item.remark or ""

        row = [
            idx,
            item.date.strftime("%Y-%m-%d %H:%M") if item.date else "",
            display_type,
            item.vendor.name if item.vendor else "-",
            item.transaction_no or "-",
            item.part.part_no if item.part else "",
            item.part.part_name if item.part else "",
            display_qty,
            item.lot_no.strftime("%Y-%m-%d") if item.lot_no else "-",
            f"{item.warehouse_from.name}" if item.warehouse_from else "-",
            display_remark,
        ]
        ws.append(row)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=idx + 3, column=col_num)
            cell.border = thin_border
            if col_num == 8:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

    # 열 너비
    widths = [6, 18, 10, 20, 22, 16, 30, 12, 14, 16, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"출고내역_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =============================================================================
# 입고 시 발주 매칭 API
# =============================================================================

@login_required
@wms_permission_required('can_wms_inout_edit')
def api_check_open_orders(request):
    """
    품번으로 ERP 미입고 발주 직접 조회
    - ERP API에서 최근 90일 발주 헤더 → 디테일 조회
    - 해당 품번의 잔량(발주수량 - 입고수량) > 0 인 건만 반환
    """
    from django.http import JsonResponse
    from material.erp_api import fetch_erp_po_headers, fetch_erp_po_details
    from django.conf import settings as conf_settings
    import datetime

    part_no = request.GET.get('part_no', '').strip()
    vendor_id = request.GET.get('vendor_id', '').strip()

    if not part_no:
        return JsonResponse({'orders': []})

    # ERP 연동 비활성화 시
    if not getattr(conf_settings, 'ERP_ENABLED', False):
        return JsonResponse({'orders': [], 'message': 'ERP 연동 비활성화'})

    # 업체 ERP 코드 조회: 1) Part.vendor 2) Header에서 선택한 vendor_id
    vendor_erp_code = ''
    part_obj = Part.objects.filter(part_no=part_no).select_related('vendor').first()
    if part_obj and part_obj.vendor and part_obj.vendor.erp_code:
        vendor_erp_code = part_obj.vendor.erp_code
    elif vendor_id:
        v = Vendor.objects.filter(id=vendor_id).first()
        if v and v.erp_code:
            vendor_erp_code = v.erp_code

    if not vendor_erp_code:
        return JsonResponse({'orders': [], 'message': '업체 ERP 코드 없음 (업체를 먼저 선택하세요)'})

    # 최근 6개월 발주 헤더 조회 (업체 코드 필수 → 해당 업체만 조회)
    today = timezone.localtime().date()
    date_from = (today - datetime.timedelta(days=180)).strftime('%Y%m%d')
    date_to = today.strftime('%Y%m%d')

    headers = fetch_erp_po_headers(date_from, date_to, tr_cd=vendor_erp_code)
    if not headers:
        return JsonResponse({'orders': []})

    results = []
    seen_keys = set()

    for header in headers:
        po_nb = header.get('poNb', '')
        po_dt = header.get('poDt', '')
        vendor_name = header.get('attrNm', '')

        details = fetch_erp_po_details(po_nb)
        for detail in details:
            item_cd = detail.get('itemCd', '')
            if item_cd != part_no:
                continue

            po_sq = str(detail.get('poSq', ''))
            po_qt = int(detail.get('poQt', 0) or 0)
            rcv_qt = int(detail.get('rcvQt', 0) or 0)
            due_dt = detail.get('dueDt', '')
            remain_qty = po_qt - rcv_qt

            if remain_qty <= 0:
                continue

            key = f"{po_nb}-{po_sq}"
            if key in seen_keys:
                continue
            seen_keys.add(key)

            # 날짜 포맷
            due_display = ''
            if due_dt and len(due_dt) == 8:
                due_display = f'{due_dt[:4]}-{due_dt[4:6]}-{due_dt[6:8]}'
            po_display = ''
            if po_dt and len(po_dt) == 8:
                po_display = f'{po_dt[:4]}-{po_dt[4:6]}-{po_dt[6:8]}'

            results.append({
                'id': 0,  # ERP 발주는 SCM Order ID 없음
                'erp_order_no': po_nb,
                'erp_order_seq': po_sq,
                'quantity': po_qt,
                'delivered_qty': rcv_qt,
                'remain_qty': remain_qty,
                'due_date': due_display,
                'po_date': po_display,
                'vendor_name': vendor_name,
            })

    # 납기일순 정렬
    results.sort(key=lambda x: x['due_date'])

    return JsonResponse({'orders': results})


# =============================================================================
# 성형 가동률 관리
# =============================================================================

@login_required
@wms_permission_required('can_wms_stock_view')
def molding_utilization(request):
    """성형 가동률 대시보드"""
    from .models import (MoldingMachine, MoldingDailyRecord, MoldingERPSyncLog,
                         MoldingWorkSetting, MOLDING_LOSS_CATEGORIES)
    import calendar

    year = int(request.GET.get('year', timezone.localtime().year))
    month = int(request.GET.get('month', timezone.localtime().month))
    _, days_in_month = calendar.monthrange(year, month)

    show_all = request.GET.get('show_all', '0') == '1'
    machines = MoldingMachine.objects.filter(is_active=True).order_by('tonnage', 'code')
    records = MoldingDailyRecord.objects.filter(
        date__year=year, date__month=month
    ).select_related('machine').prefetch_related('loss_details')

    setting = MoldingWorkSetting.get_setting(year, month)
    actual_work_days_util = len(set(r.date for r in records))

    # 실적 있는 호기 ID
    active_machine_ids = set(r.machine_id for r in records)

    # 호기별 월간 집계 (주간/야간 행 분리, 톤수별 그룹)
    from collections import defaultdict, OrderedDict
    tonnage_groups = OrderedDict()  # {톤수: {'rows': [...], 'summary': {...}}}

    for m in machines:
        m_records = [r for r in records if r.machine_id == m.id]

        # 전체보기가 아니면 실적 있는 호기만
        if not show_all and m.id not in active_machine_ids:
            continue

        tonnage = m.tonnage
        if tonnage not in tonnage_groups:
            tonnage_groups[tonnage] = {'rows': [], 'load_sum': 0, 'loss_sum': 0, 'active_count': 0, 'total_days': 0}

        has_shift_data = False
        for shift in ['주간', '야간']:
            s_records = [r for r in m_records if r.shift == shift]
            if not s_records and not show_all:
                continue
            if not s_records and show_all:
                # 전체보기 시 빈 행도 표시 (주간만)
                if shift == '야간':
                    continue
                daily_list = [None] * days_in_month
                tonnage_groups[tonnage]['rows'].append({
                    'machine': m, 'shift': shift,
                    'active_days': 0, 'avg_util': 0, 'avg_time': 0,
                    'daily': daily_list,
                })
                has_shift_data = True
                continue

            active_records = [r for r in s_records if r.status == '가동']
            active_days = len(active_records)
            # 설비가동률 = (부하시간 - 유실시간) / 부하시간 (가동일만)
            total_load = sum(r.base_minutes for r in active_records)
            total_loss = sum(r.loss_minutes for r in active_records)
            avg_util = ((total_load - total_loss) / total_load * 100) if total_load else 0
            # 시간가동률 = (부하시간 - 유실시간) / 근무시간 (주간+야간 전체)
            total_work = setting.work_days * (setting.day_shift_minutes + setting.night_shift_minutes)
            avg_time = ((total_load - total_loss) / total_work * 100) if total_work else 0
            total_work_actual = actual_work_days_util * (setting.day_shift_minutes + setting.night_shift_minutes)
            avg_time_actual = ((total_load - total_loss) / total_work_actual * 100) if total_work_actual else 0

            daily_map = {r.date.day: r for r in s_records}
            daily_list = []
            for d in range(1, days_in_month + 1):
                r = daily_map.get(d)
                if r:
                    daily_list.append({
                        'id': r.id,
                        'status': r.status,
                        'util': r.utilization_rate,
                        'erp_synced': r.erp_synced,
                        'input_completed': r.input_completed,
                        'operating': r.operating_minutes,
                        'loss': r.loss_minutes,
                    })
                else:
                    daily_list.append(None)

            tonnage_groups[tonnage]['rows'].append({
                'machine': m, 'shift': shift,
                'active_days': active_days, 'avg_util': avg_util, 'avg_time': avg_time,
                'avg_time_actual': avg_time_actual,
                'daily': daily_list,
            })
            if active_days > 0:
                tonnage_groups[tonnage]['load_sum'] += total_load
                tonnage_groups[tonnage]['loss_sum'] += total_loss
                tonnage_groups[tonnage]['active_count'] += 1
                tonnage_groups[tonnage]['total_days'] += active_days
            has_shift_data = True

    # 호기별 시간가동률 통합 (주간+야간 합산)
    # 같은 호기의 rows에 machine_time_rate (주+야 합산 %) + shift_span (rowspan 용) 세팅
    for tonnage_data in tonnage_groups.values():
        # 호기별 그룹핑
        machine_rows_map = defaultdict(list)
        for row in tonnage_data['rows']:
            machine_rows_map[row['machine'].id].append(row)

        for machine_id, rows_for_machine in machine_rows_map.items():
            # 주+야 합산 시간가동률
            combined_time = sum(r['avg_time'] or 0 for r in rows_for_machine)
            combined_time_actual = sum(r.get('avg_time_actual', 0) or 0 for r in rows_for_machine)
            shift_count = len(rows_for_machine)
            for idx, r in enumerate(rows_for_machine):
                r['machine_time_rate'] = combined_time
                r['machine_time_rate_actual'] = combined_time_actual
                # 첫 번째 행에만 표시 (rowspan)
                r['show_time_cell'] = (idx == 0)
                r['time_rowspan'] = shift_count if idx == 0 else 0

    # 톤수별 전체 활성 호기 수 (비가동 포함)
    tonnage_all_count = {}
    for m in machines:
        tonnage_all_count[m.tonnage] = tonnage_all_count.get(m.tonnage, 0) + 1

    # 소계 계산
    work_per_machine = setting.work_days * (setting.day_shift_minutes + setting.night_shift_minutes)
    for t, g in tonnage_groups.items():
        net = g['load_sum'] - g['loss_sum']
        g['avg_util'] = (net / g['load_sum'] * 100) if g['load_sum'] else 0
        # 시간가동률 분모 = 해당 톤수 전체 호기수 × 근무시간 (비가동 포함)
        g['machine_count'] = len(set(r['machine'].id for r in g['rows']))
        all_count = tonnage_all_count.get(t, g['machine_count'])
        total_work_all = all_count * work_per_machine
        g['avg_time'] = (net / total_work_all * 100) if total_work_all else 0
        total_work_all_actual = all_count * actual_work_days_util * (setting.day_shift_minutes + setting.night_shift_minutes)
        g['avg_time_actual'] = (net / total_work_all_actual * 100) if total_work_all_actual else 0
        g['all_machine_count'] = all_count

    # 기존 machine_summary 호환 (플랫 리스트)
    machine_summary = []
    for t, g in tonnage_groups.items():
        for row in g['rows']:
            machine_summary.append(row)

    # 전체 요약
    total_load_all = sum(g['load_sum'] for g in tonnage_groups.values())
    total_loss_all = sum(g['loss_sum'] for g in tonnage_groups.values())
    net_all = total_load_all - total_loss_all
    overall_util = (net_all / total_load_all * 100) if total_load_all else 0
    total_machines_count = machines.count()
    work_per_machine = setting.work_days * (setting.day_shift_minutes + setting.night_shift_minutes)
    overall_work = total_machines_count * work_per_machine
    overall_time = (net_all / overall_work * 100) if overall_work else 0
    # 실적 기준 시간가동률
    actual_work_days_util = len(set(r.date for r in records))
    work_per_machine_actual = actual_work_days_util * (setting.day_shift_minutes + setting.night_shift_minutes)
    overall_work_actual = total_machines_count * work_per_machine_actual
    overall_time_actual = (net_all / overall_work_actual * 100) if overall_work_actual else 0

    # 톤수별 집계
    tonnage_summary = {}
    for m in machines:
        t = m.tonnage
        if t not in tonnage_summary:
            tonnage_summary[t] = {'count': 0, 'util_sum': 0, 'active_count': 0}
        tonnage_summary[t]['count'] += 1
        m_active = [r for r in records if r.machine_id == m.id and r.status == '가동']
        if m_active:
            tonnage_summary[t]['util_sum'] += sum(r.utilization_rate for r in m_active) / len(m_active)
            tonnage_summary[t]['active_count'] += 1
    tonnage_list = sorted([
        {'tonnage': t, 'count': s['count'],
         'avg_util': s['util_sum'] / s['active_count'] if s['active_count'] else 0}
        for t, s in tonnage_summary.items()
    ], key=lambda x: x['tonnage'])

    sync_logs = MoldingERPSyncLog.objects.filter(year=year, month=month).order_by('-synced_at')[:5]
    loss_categories = [c[0] for c in MOLDING_LOSS_CATEGORIES]
    from .models import MOLDING_MGMT_LOSS
    mgmt_loss_categories = [c[0] for c in MOLDING_LOSS_CATEGORIES if c[0] in MOLDING_MGMT_LOSS]
    time_loss_categories = [c[0] for c in MOLDING_LOSS_CATEGORIES if c[0] not in MOLDING_MGMT_LOSS]

    context = {
        'year': year, 'month': month,
        'days_in_month': days_in_month,
        'day_range': range(1, days_in_month + 1),
        'machines': machines,
        'machine_summary': machine_summary,
        'overall_util': overall_util,
        'overall_time': overall_time,
        'overall_time_actual': overall_time_actual,
        'actual_work_days': actual_work_days_util,
        'tonnage_list': tonnage_list,
        'total_machines': machines.count(),
        'sync_logs': sync_logs,
        'setting': setting,
        'loss_categories': loss_categories,
        'mgmt_loss_categories': mgmt_loss_categories,
        'time_loss_categories': time_loss_categories,
        'tonnage_groups': tonnage_groups,
        'show_all': show_all,
        'no_tonnage_count': MoldingMachine.objects.filter(is_active=True, tonnage=0).count(),
    }
    return render(request, 'material/molding_utilization.html', context)


@login_required
@wms_permission_required('can_wms_stock_view')
def molding_erp_sync(request):
    """ERP 생산입고 데이터로 성형 가동 현황 동기화"""
    from .models import MoldingMachine, MoldingDailyRecord, MoldingERPSyncLog, MoldingWorkSetting
    from .erp_api import fetch_erp_receipt_list, call_erp_api
    from datetime import date as dt_date
    import calendar

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'})

    year = int(request.POST.get('year', timezone.localtime().year))
    month = int(request.POST.get('month', timezone.localtime().month))
    _, days_in_month = calendar.monthrange(year, month)

    date_from = f"{year}{month:02d}01"
    date_to = f"{year}{month:02d}{days_in_month:02d}"

    # ERP 생산입고 조회
    ok, data, err = fetch_erp_receipt_list(date_from, date_to)
    if not ok:
        return JsonResponse({'success': False, 'error': f'ERP 조회 실패: {err}'})

    if not data:
        return JsonResponse({'success': False, 'error': '해당 기간 생산입고 데이터가 없습니다.'})

    setting = MoldingWorkSetting.get_setting(year, month)

    # 생산실적 API 조회 (실가동시간 workTm)
    from django.conf import settings as django_settings
    wr_body = {
        'coCd': django_settings.ERP_COMPANY_CODE,
        'wrDtFrom': date_from,
        'wrDtTo': date_to,
    }
    wr_ok, wr_data, wr_err = call_erp_api('/apiproxy/api20A03S00901', wr_body)
    work_time_agg = {}  # {(machine_code, date_str, shift): work_minutes}
    bad_qty_agg = {}   # {(machine_code, date_str, shift): bad_qty}
    if wr_ok and wr_data:
        import re as _re2
        for r in wr_data.get('resultData', []) or []:
            en = (r.get('equipNm') or '').strip()
            if not _re2.match(r'^M\d', en):
                continue
            wr_dt = r.get('wrDt', '')
            if len(wr_dt) != 8:
                continue
            shift_nm = (r.get('wshftNm') or '').strip()
            shift = '야간' if shift_nm == '야간' else '주간'
            key = (en, wr_dt, shift)
            if key not in work_time_agg:
                work_time_agg[key] = 0
                bad_qty_agg[key] = 0
            work_time_agg[key] += int(float(r.get('workTm', 0) or 0))
            bad_qty_agg[key] += int(float(r.get('badQt', 0) or 0))

    try:
        with transaction.atomic():
            # equipNm이 M으로 시작하는 성형기 데이터만 필터
            import re as _re
            molding_data = [r for r in data if _re.match(r'^M\d', (r.get('equipNm') or ''))]

            # 호기+일자별 집계
            daily_agg = {}  # {(machine_code, date_str): {'parts': set, 'qty': 0}}
            for r in molding_data:
                machine_code = r['equipNm'].strip()
                rcv_dt = r.get('rcvDt', '')
                if len(rcv_dt) != 8:
                    continue

                shift_nm = (r.get('wshftNm') or '').strip()
                shift = '야간' if shift_nm == '야간' else '주간'

                key = (machine_code, rcv_dt, shift)
                if key not in daily_agg:
                    daily_agg[key] = {'part_qty': {}, 'tonnage': 0}
                item_cd = r.get('itemCd', '')
                qty = int(r.get('rcvQt', 0) or 0)
                daily_agg[key]['part_qty'][item_cd] = daily_agg[key]['part_qty'].get(item_cd, 0) + qty

            # 호기 마스터 갱신 및 레코드 생성
            record_count = 0
            machine_codes = set()
            for (mc, dt_str, shift), agg in daily_agg.items():
                # 호기 get_or_create
                machine, _ = MoldingMachine.objects.get_or_create(
                    code=mc, defaults={'tonnage': 0}
                )
                machine_codes.add(mc)

                # 날짜 변환
                rec_date = dt_date(int(dt_str[:4]), int(dt_str[4:6]), int(dt_str[6:8]))

                # 주간/야간에 따라 기준시간 결정
                base_min = setting.night_shift_minutes if shift == '야간' else setting.day_shift_minutes

                # 기존 수동 입력 보존: erp 필드만 업데이트
                record, created = MoldingDailyRecord.objects.get_or_create(
                    machine=machine, date=rec_date, shift=shift,
                    defaults={
                        'status': '가동',
                        'base_minutes': base_min,
                        'erp_synced': True,
                    }
                )
                record.status = '가동'
                # 품번별 수량 표시: "ZR700: 1000, ZR703: 2000"
                part_qty = agg['part_qty']
                record.product_part_no = ' | '.join(
                    f"{p}: {q:,}" for p, q in sorted(part_qty.items())
                )[:500]
                record.product_qty = sum(part_qty.values())
                record.defect_qty = bad_qty_agg.get((mc, dt_str, shift), 0)
                record.erp_synced = True
                if not record.input_completed:
                    record.base_minutes = base_min
                # 생산실적 API에서 실가동시간 반영
                wt_key = (mc, dt_str, shift)
                if wt_key in work_time_agg and not record.input_completed:
                    record.work_minutes = min(work_time_agg[wt_key], base_min)
                    record.operating_minutes = record.work_minutes
                    record.loss_minutes = max(base_min - record.work_minutes, 0)
                    # 가동률 계산
                    if base_min > 0:
                        rate = round(record.operating_minutes / base_min * 100, 1)
                        record.utilization_rate = rate
                        record.time_rate = rate
                record.save()
                record_count += 1

            # 동기화 이력
            MoldingERPSyncLog.objects.create(
                year=year, month=month,
                synced_by=request.user,
                record_count=record_count,
                machine_count=len(machine_codes),
                message=f'ERP 생산입고 {len(molding_data)}건 → {record_count}개 가동일 동기화 (호기 {len(machine_codes)}대)',
            )

        return JsonResponse({
            'success': True,
            'message': f'{year}년 {month}월: {record_count}개 가동일 동기화 완료 (호기 {len(machine_codes)}대)'
        })

    except Exception as e:
        logger.error(f'성형 ERP 동기화 오류: {e}', exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@wms_permission_required('can_wms_stock_view')
def molding_settings(request):
    """성형 가동률 근무시간 설정"""
    from .models import MoldingWorkSetting

    if request.method == 'POST':
        year = int(request.POST.get('year'))
        month = int(request.POST.get('month'))
        obj, _ = MoldingWorkSetting.objects.get_or_create(
            year=year, month=month,
            defaults={'work_days': MoldingWorkSetting.calc_weekdays(year, month)}
        )
        obj.day_shift_minutes = int(request.POST.get('day_shift_minutes', 670))
        obj.night_shift_minutes = int(request.POST.get('night_shift_minutes', 770))
        obj.work_days = int(request.POST.get('work_days', 20))
        obj.save()
        return JsonResponse({'success': True, 'message': f'{year}년 {month}월 설정 저장 완료'})

    return JsonResponse({'success': False, 'error': 'POST만 허용'})


@login_required
@wms_permission_required('can_wms_stock_view')
def api_molding_save_input(request):
    """성형 가동률 수동 입력 저장 (GET: 조회, POST: 저장)"""
    from .models import MoldingDailyRecord, MoldingLossDetail, MOLDING_LOSS_CATEGORIES
    import json

    # GET: 레코드 상세 조회 (모달에서 사용)
    if request.method == 'GET':
        record_id = request.GET.get('record_id')
        try:
            rec = MoldingDailyRecord.objects.select_related('machine').prefetch_related('loss_details').get(id=record_id)
        except MoldingDailyRecord.DoesNotExist:
            return JsonResponse({'error': '레코드 없음'})
        loss_details = {d.category: d.minutes for d in rec.loss_details.all()}
        return JsonResponse({'record': {
            'id': rec.id,
            'machine': rec.machine.code,
            'tonnage': rec.machine.tonnage,
            'date': rec.date.strftime('%Y-%m-%d'),
            'status': rec.status,
            'operating_minutes': rec.operating_minutes,
            'work_minutes': rec.work_minutes,
            'base_minutes': rec.base_minutes,
            'loss_minutes': rec.loss_minutes,
            'utilization_rate': rec.utilization_rate,
            'time_rate': rec.time_rate,
            'product_part_no': rec.product_part_no,
            'product_qty': rec.product_qty,
            'input_completed': rec.input_completed,
            'shift': rec.shift,
            'loss_details': loss_details,
        }})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'GET/POST만 허용'})

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'JSON 파싱 오류'})

    record_id = data.get('record_id')
    base_minutes = int(data.get('base_minutes', 0))
    loss_data = data.get('loss_details', {})  # {'계획정지': 30, '금형수리': 10, ...}

    try:
        record = MoldingDailyRecord.objects.get(id=record_id)
    except MoldingDailyRecord.DoesNotExist:
        return JsonResponse({'success': False, 'error': '레코드를 찾을 수 없습니다.'})

    try:
        with transaction.atomic():
            if base_minutes > 0:
                record.base_minutes = base_minutes

            # 유실 상세 갱신
            record.loss_details.all().delete()
            valid_cats = [c[0] for c in MOLDING_LOSS_CATEGORIES]
            for cat, mins in loss_data.items():
                mins = int(mins or 0)
                if mins > 0 and cat in valid_cats:
                    MoldingLossDetail.objects.create(
                        record=record, category=cat, minutes=mins
                    )

            # 가동률 자동 계산
            record.calculate_rates()
            record.input_completed = True
            record.input_by = request.user
            record.input_at = timezone.now()
            record.save()

        return JsonResponse({
            'success': True,
            'utilization_rate': record.utilization_rate,
            'time_rate': record.time_rate,
            'loss_minutes': record.loss_minutes,
        })

    except Exception as e:
        logger.error(f'성형 가동률 입력 오류: {e}', exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@wms_permission_required('can_wms_stock_view')
def api_molding_machines(request):
    """성형기 마스터 CRUD API"""
    from .models import MoldingMachine
    import json

    if request.method == 'GET':
        machines = MoldingMachine.objects.all().order_by('tonnage', 'code')
        data = [{
            'id': m.id, 'code': m.code, 'tonnage': m.tonnage,
            'line': m.line, 'is_active': m.is_active,
        } for m in machines]
        return JsonResponse({'machines': data})

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'JSON 파싱 오류'})

        action = data.get('action', '')

        if action == 'add':
            code = (data.get('code') or '').strip().upper()
            if not code:
                return JsonResponse({'success': False, 'error': '호기 코드를 입력하세요.'})
            if MoldingMachine.objects.filter(code=code).exists():
                return JsonResponse({'success': False, 'error': f'{code}는 이미 등록되어 있습니다.'})
            MoldingMachine.objects.create(
                code=code,
                tonnage=int(data.get('tonnage', 0)),
                line=(data.get('line') or '').strip(),
                is_active=True,
            )
            return JsonResponse({'success': True, 'message': f'{code} 등록 완료'})

        elif action == 'update':
            try:
                m = MoldingMachine.objects.get(id=data.get('id'))
            except MoldingMachine.DoesNotExist:
                return JsonResponse({'success': False, 'error': '호기를 찾을 수 없습니다.'})
            if 'tonnage' in data:
                m.tonnage = int(data['tonnage'])
            if 'line' in data:
                m.line = (data['line'] or '').strip()
            if 'is_active' in data:
                m.is_active = bool(data['is_active'])
            m.save()
            return JsonResponse({'success': True, 'message': f'{m.code} 수정 완료'})

        elif action == 'delete':
            try:
                m = MoldingMachine.objects.get(id=data.get('id'))
            except MoldingMachine.DoesNotExist:
                return JsonResponse({'success': False, 'error': '호기를 찾을 수 없습니다.'})
            code = m.code
            m.delete()
            return JsonResponse({'success': True, 'message': f'{code} 삭제 완료'})

        return JsonResponse({'success': False, 'error': '알 수 없는 액션'})

    return JsonResponse({'success': False, 'error': 'GET/POST만 허용'})


# =============================================================================
# 생산현장관리 메인 페이지
# =============================================================================
@login_required
@wms_permission_required('can_wms_stock_view')
def production_main(request):
    """생산현장관리 메인 페이지 - 바로가기 + 현황 요약"""
    from .models import MoldMaster, MoldingMachine, MoldingDailyRecord, MoldingWorkSetting
    from django.db.models import Sum

    now = timezone.localtime()
    today = now.date()
    year = now.year
    month = now.month

    # 성형 요약
    molding_records = MoldingDailyRecord.objects.filter(date__year=year, date__month=month)
    total_base = molding_records.aggregate(t=Sum('base_minutes'))['t'] or 0
    total_operating = molding_records.aggregate(t=Sum('operating_minutes'))['t'] or 0
    molding_util = round(total_operating / total_base * 100, 1) if total_base else 0
    molding_machine_count = MoldingMachine.objects.filter(is_active=True).count()
    molding_today_count = molding_records.filter(date=today, status='가동').values('machine').distinct().count()

    # 금형 요약
    mold_total = MoldMaster.objects.filter(is_active=True).count()
    # MT 필요 건수 (annotate 사용)
    from django.db.models.functions import Coalesce
    from django.db.models import Value, IntegerField
    molds_qs = MoldMaster.objects.filter(is_active=True).annotate(
        _ms=Coalesce(Sum('shot_records__shots'), Value(0), output_field=IntegerField())
    )
    mt_need = 0
    for m in molds_qs:
        total = m.total_shots_prev + m._ms
        since_mt = max(total - m.last_mt_shots, 0)
        if m.mt_interval > 0 and since_mt >= m.mt_interval:
            mt_need += 1

    context = {
        'year': year, 'month': month, 'today': today,
        'molding_util': molding_util,
        'molding_machine_count': molding_machine_count,
        'molding_today_count': molding_today_count,
        'mold_total': mold_total,
        'mt_need': mt_need,
    }
    return render(request, 'material/production_main.html', context)


# =============================================================================
# 성형 가동률 분석 대시보드 (상세)
# =============================================================================
@login_required
def molding_analytics(request):
    """성형 가동률 분석 대시보드 (6가지 차트 + KPI)"""
    from .models import (
        MoldingMachine, MoldingDailyRecord, MoldingLossDetail,
        MoldingWorkSetting, MOLDING_LOSS_CATEGORIES, MOLDING_MGMT_LOSS
    )
    from collections import defaultdict
    from django.db.models import Avg, Sum, Count, F
    import calendar
    import json
    from datetime import date
    from dateutil.relativedelta import relativedelta

    now = timezone.localtime()
    year = int(request.GET.get('year', now.year))
    month = int(request.GET.get('month', now.month))

    # ─── 해당 월 레코드 조회 ───
    records = MoldingDailyRecord.objects.filter(
        date__year=year, date__month=month
    ).select_related('machine').prefetch_related('loss_details')

    records_list = list(records)

    # ─── KPI 카드 ───
    total_base = sum(r.base_minutes for r in records_list) or 1
    total_operating = sum(r.operating_minutes for r in records_list)
    total_loss = sum(r.loss_minutes for r in records_list)
    total_work = sum(r.work_minutes for r in records_list)

    # 설비가동률 = operating / base (가동일 기준)
    kpi_utilization = round(total_operating / total_base * 100, 1) if total_base else 0
    # 시간가동률 = operating / 전체근무시간 (전체 호기 × 근무일 × 주야간)
    setting = MoldingWorkSetting.get_setting(year, month)
    all_machines = MoldingMachine.objects.filter(is_active=True).count()
    work_per_machine = setting.work_days * (setting.day_shift_minutes + setting.night_shift_minutes)
    total_work_capacity = all_machines * work_per_machine
    kpi_time_rate = round(total_operating / total_work_capacity * 100, 1) if total_work_capacity else 0
    # 실적 기준 시간가동률 (실제 데이터가 있는 일수 기준)
    actual_work_days = len(set(r.date for r in records_list))
    work_per_machine_actual = actual_work_days * (setting.day_shift_minutes + setting.night_shift_minutes)
    total_work_capacity_actual = all_machines * work_per_machine_actual
    kpi_time_rate_actual = round(total_operating / total_work_capacity_actual * 100, 1) if total_work_capacity_actual else 0
    kpi_total_loss_hours = round(total_loss / 60, 1)

    # 계획정지 = 유실사유 중 '계획정지'
    planned_loss = MoldingLossDetail.objects.filter(
        record__date__year=year, record__date__month=month,
        category='계획정지'
    ).aggregate(total=Sum('minutes'))['total'] or 0
    kpi_planned_hours = round(planned_loss / 60, 1)
    kpi_actual_loss_hours = round((total_loss - planned_loss) / 60, 1)

    # ─── 차트 1: 일별 설비가동률 추이 ───
    _, days_in_month = calendar.monthrange(year, month)
    daily_data = defaultdict(lambda: {'base': 0, 'operating': 0})
    for r in records_list:
        day = r.date.day
        daily_data[day]['base'] += r.base_minutes
        daily_data[day]['operating'] += r.operating_minutes

    daily_labels = []
    daily_rates = []
    for d in range(1, days_in_month + 1):
        if daily_data[d]['base'] > 0:
            daily_labels.append(f"{d}일")
            rate = round(daily_data[d]['operating'] / daily_data[d]['base'] * 100, 1)
            daily_rates.append(rate)

    # ─── 차트 2: 톤수별 가동률 비교 ───
    tonnage_data = defaultdict(lambda: {'base': 0, 'operating': 0})
    for r in records_list:
        t = r.machine.tonnage
        tonnage_data[t]['base'] += r.base_minutes
        tonnage_data[t]['operating'] += r.operating_minutes

    tonnage_labels = []
    tonnage_rates = []
    for t in sorted(tonnage_data.keys()):
        tonnage_labels.append(f"{t}t")
        b = tonnage_data[t]['base']
        rate = round(tonnage_data[t]['operating'] / b * 100, 1) if b else 0
        tonnage_rates.append(rate)

    # ─── 차트 3: 유실 사유별 비율 (donut) ───
    loss_details = MoldingLossDetail.objects.filter(
        record__date__year=year, record__date__month=month
    ).values('category').annotate(total=Sum('minutes')).order_by('-total')

    loss_labels = [d['category'] for d in loss_details]
    loss_values = [d['total'] for d in loss_details]

    # ─── 차트 4: 품목군별 생산수량 랭킹 TOP 10 ───
    from orders.models import Part
    import re as _re4
    part_group_qty = defaultdict(int)
    part_group_detail = defaultdict(lambda: defaultdict(int))  # {group: {pno: qty}}
    # Part 캐시 (N+1 방지)
    part_cache = {}
    for r in records_list:
        if not r.product_part_no:
            continue
        for m in _re4.finditer(r'([\w\-]+)\s*:\s*([\d,]+)', r.product_part_no):
            pno = m.group(1).strip()
            try:
                qty = int(m.group(2).replace(',', ''))
            except ValueError:
                qty = 0
            if pno:
                if pno not in part_cache:
                    p = Part.objects.filter(part_no=pno).first()
                    part_cache[pno] = (p.part_group if p and p.part_group else '미분류', p.part_name if p else pno)
                group, pname = part_cache[pno]
                part_group_qty[group] += qty
                part_group_detail[group][f"{pno} ({pname})"] += qty

    group_ranking = sorted(part_group_qty.items(), key=lambda x: x[1], reverse=True)[:10]
    ranking_top_labels = [g[0] for g in group_ranking]
    ranking_top_rates = [g[1] for g in group_ranking]
    # 상세 데이터 (모달용)
    group_detail_data = {}
    for grp, _ in group_ranking:
        items = sorted(part_group_detail[grp].items(), key=lambda x: -x[1])[:20]
        group_detail_data[grp] = [{'name': k, 'qty': v} for k, v in items]

    # ─── 차트 5: 호기별 가동일수 TOP 15 + 생산 상세 ───
    machine_days = defaultdict(set)
    machine_production = defaultdict(lambda: defaultdict(int))  # {code: {part_no: qty}}
    for r in records_list:
        if r.status == '가동':
            machine_days[r.machine.code].add(r.date)
        # 호기별 생산 상세 집계
        if r.product_part_no:
            for m in _re4.finditer(r'([\w\-]+)\s*:\s*([\d,]+)', r.product_part_no):
                pno = m.group(1).strip()
                qty = int(m.group(2).replace(',', ''))
                machine_production[r.machine.code][pno] += qty

    machine_days_list = sorted(
        [{'code': code, 'days': len(dates)} for code, dates in machine_days.items()],
        key=lambda x: -x['days']
    )[:15]
    machine_days_labels = [m['code'] for m in machine_days_list]
    machine_days_values = [m['days'] for m in machine_days_list]

    # 호기별 생산 상세 데이터 (JS용)
    machine_detail_data = {}
    for code, parts in machine_production.items():
        items = []
        for pno, qty in sorted(parts.items(), key=lambda x: -x[1]):
            name = pno
            if pno in part_cache:
                p = part_cache[pno]
                name = f"{pno} ({p.part_name})" if hasattr(p, 'part_name') and p.part_name else pno
            items.append({'name': name, 'qty': qty})
        machine_detail_data[code] = items

    # ─── 차트 6: 월별 트렌드 최근 6개월 ───
    current_date = date(year, month, 1)
    monthly_labels = []
    monthly_utilization = []
    monthly_time_rate = []

    for i in range(5, -1, -1):
        target = current_date - relativedelta(months=i)
        m_records = MoldingDailyRecord.objects.filter(
            date__year=target.year, date__month=target.month
        )
        m_base = m_records.aggregate(total=Sum('base_minutes'))['total'] or 0
        m_operating = m_records.aggregate(total=Sum('operating_minutes'))['total'] or 0

        # 시간가동률 분모 = 해당 월 전체 호기 × 근무시간
        m_setting = MoldingWorkSetting.get_setting(target.year, target.month)
        m_work_capacity = all_machines * m_setting.work_days * (m_setting.day_shift_minutes + m_setting.night_shift_minutes)

        monthly_labels.append(f"{target.year}.{target.month:02d}")
        monthly_utilization.append(round(m_operating / m_base * 100, 1) if m_base else 0)
        monthly_time_rate.append(round(m_operating / m_work_capacity * 100, 1) if m_work_capacity else 0)

    # ─── 톤수별 테이블 ───
    tonnage_all_count = defaultdict(int)
    for m in MoldingMachine.objects.filter(is_active=True):
        tonnage_all_count[m.tonnage] += 1

    # 톤수별 가동 호기 (한번이라도 실적 있는 호기)
    active_machine_by_tonnage = defaultdict(set)
    for r in records_list:
        active_machine_by_tonnage[r.machine.tonnage].add(r.machine_id)

    tonnage_table = []
    for t in sorted(tonnage_all_count.keys()):
        count = tonnage_all_count[t]
        active_count = len(active_machine_by_tonnage.get(t, set()))
        td = tonnage_data.get(t, {'base': 0, 'operating': 0})
        util = round(td['operating'] / td['base'] * 100, 1) if td['base'] else 0
        time_cap = count * setting.work_days * (setting.day_shift_minutes + setting.night_shift_minutes)
        time_r = round(td['operating'] / time_cap * 100, 1) if time_cap else 0
        time_cap_actual = count * actual_work_days * (setting.day_shift_minutes + setting.night_shift_minutes)
        time_r_actual = round(td['operating'] / time_cap_actual * 100, 1) if time_cap_actual else 0
        tonnage_table.append({
            'tonnage': t, 'count': count, 'active_count': active_count,
            'util': util, 'time_rate': time_r, 'time_rate_actual': time_r_actual,
        })

    # ─── 년도 목록 ───
    year_range = list(range(2024, now.year + 2))

    context = {
        'year': year,
        'month': month,
        'year_range': year_range,
        # KPI
        'kpi_utilization': kpi_utilization,
        'kpi_time_rate': kpi_time_rate,
        'kpi_time_rate_actual': kpi_time_rate_actual,
        'kpi_total_loss_hours': kpi_total_loss_hours,
        'actual_work_days': actual_work_days,
        'setting_work_days': setting.work_days,
        'kpi_planned_hours': kpi_planned_hours,
        'kpi_actual_loss_hours': kpi_actual_loss_hours,
        'total_records': len(records_list),
        'tonnage_table': tonnage_table,
        # Chart 1: Daily trend
        'daily_labels': json.dumps(daily_labels, ensure_ascii=False),
        'daily_rates': json.dumps(daily_rates),
        # Chart 2: Tonnage comparison
        'tonnage_labels': json.dumps(tonnage_labels, ensure_ascii=False),
        'tonnage_rates': json.dumps(tonnage_rates),
        # Chart 3: Loss breakdown
        'loss_labels': json.dumps(loss_labels, ensure_ascii=False),
        'loss_values': json.dumps(loss_values),
        # Chart 4: Machine ranking
        'ranking_top_labels': json.dumps(ranking_top_labels, ensure_ascii=False),
        'ranking_top_rates': json.dumps(ranking_top_rates),
        'group_detail_data': json.dumps(group_detail_data, ensure_ascii=False),
        # Chart 5: Machine days + detail
        'machine_days_labels': json.dumps(machine_days_labels, ensure_ascii=False),
        'machine_days_values': json.dumps(machine_days_values),
        'machine_detail_data': json.dumps(machine_detail_data, ensure_ascii=False),
        # Chart 6: Monthly trend
        'monthly_labels': json.dumps(monthly_labels, ensure_ascii=False),
        'monthly_utilization': json.dumps(monthly_utilization),
        'monthly_time_rate': json.dumps(monthly_time_rate),
    }
    return render(request, 'material/molding_analytics.html', context)


# =============================================================================
# 성형 마스터 (Molding Master)
# =============================================================================

@wms_permission_required('can_wms_stock_view')
def molding_master_list(request):
    """성형 마스터 목록 조회"""
    from .models import MoldingMaster, MoldingMachine

    q = request.GET.get('q', '').strip()
    item_group = request.GET.get('item_group', '').strip()
    material_part_no = request.GET.get('material_part_no', '').strip()
    machine_code = request.GET.get('machine_code', '').strip()

    qs = MoldingMaster.objects.select_related('machine').all()

    if q:
        qs = qs.filter(Q(part_no__icontains=q) | Q(part_name__icontains=q))
    if item_group:
        qs = qs.filter(item_group=item_group)
    if material_part_no:
        qs = qs.filter(material_part_no__icontains=material_part_no)
    if machine_code:
        qs = qs.filter(machine__code=machine_code)

    total_count = qs.count()
    paginator = Paginator(qs, 30)
    page = request.GET.get('page', 1)
    records = paginator.get_page(page)

    # 필터용 데이터
    item_groups = MoldingMaster.objects.values_list('item_group', flat=True).exclude(item_group='').distinct().order_by('item_group')
    machines = MoldingMachine.objects.filter(is_active=True).order_by('code')
    # 자동완성용
    part_nos = list(MoldingMaster.objects.values_list('part_no', flat=True).distinct().order_by('part_no'))
    material_nos = list(MoldingMaster.objects.values_list('material_part_no', flat=True).exclude(material_part_no='').distinct().order_by('material_part_no'))

    context = {
        'records': records,
        'total_count': total_count,
        'q': q,
        'item_group': item_group,
        'material_part_no': material_part_no,
        'machine_code': machine_code,
        'item_groups': item_groups,
        'machines': machines,
        'part_nos_json': json.dumps(part_nos, ensure_ascii=False),
        'material_nos_json': json.dumps(material_nos, ensure_ascii=False),
    }
    return render(request, 'material/molding_master_list.html', context)


@wms_permission_required('can_wms_stock_view')
def molding_master_save(request):
    """성형 마스터 저장 (생성/수정)"""
    from .models import MoldingMaster, MoldingMachine
    from decimal import Decimal, InvalidOperation

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST only'}, status=405)

    pk = request.POST.get('pk', '').strip()

    def to_decimal(val, default='0'):
        try:
            return Decimal(val) if val else Decimal(default)
        except (InvalidOperation, ValueError):
            return Decimal(default)

    def to_int(val, default=0):
        try:
            return int(val) if val else default
        except (ValueError, TypeError):
            return default

    part_no = request.POST.get('part_no', '').strip()
    material_type = request.POST.get('material_type', '').strip()

    if not part_no:
        messages.error(request, '품번은 필수 입력 항목입니다.')
        return redirect('material:molding_master_list')

    machine_code = request.POST.get('machine_code', '').strip()
    machine_obj = None
    if machine_code:
        machine_obj = MoldingMachine.objects.filter(code=machine_code).first()

    data = {
        'part_name': request.POST.get('part_name', '').strip(),
        'item_group': request.POST.get('item_group', '').strip(),
        'mold_type': request.POST.get('mold_type', '').strip(),
        'material_type': material_type,
        'material_part_no': request.POST.get('material_part_no', '').strip(),
        'material_name': request.POST.get('material_name', '').strip(),
        'machine': machine_obj,
        'machine_tonnage': to_int(request.POST.get('machine_tonnage')),
        'cycle_time': to_decimal(request.POST.get('cycle_time')),
        'cavity': to_int(request.POST.get('cavity')),
        'shot_time': to_decimal(request.POST.get('shot_time')),
        'gate_type': request.POST.get('gate_type', '').strip(),
        'hot_runner_time': to_decimal(request.POST.get('hot_runner_time')),
        'product_weight': to_decimal(request.POST.get('product_weight')),
        'tolerance': to_decimal(request.POST.get('tolerance')),
        'runner_weight': to_decimal(request.POST.get('runner_weight')),
        'total_material': to_decimal(request.POST.get('total_material')),
        'erp_qty': to_decimal(request.POST.get('erp_qty')),
        'remark': request.POST.get('remark', '').strip(),
        'updated_by': request.user,
    }

    try:
        if pk:
            obj = get_object_or_404(MoldingMaster, pk=pk)
            for k, v in data.items():
                setattr(obj, k, v)
            obj.part_no = part_no
            obj.save()
            messages.success(request, f'{part_no} ({material_type}) 수정 완료')
        else:
            obj = MoldingMaster(part_no=part_no, **data)
            obj.save()
            messages.success(request, f'{part_no} ({material_type}) 등록 완료')
    except Exception as e:
        messages.error(request, f'저장 실패: {e}')

    return redirect('material:molding_master_list')


@wms_permission_required('can_wms_stock_view')
def molding_master_delete(request):
    """성형 마스터 삭제"""
    from .models import MoldingMaster

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST only'}, status=405)

    pk = request.POST.get('pk')
    try:
        obj = get_object_or_404(MoldingMaster, pk=pk)
        label = str(obj)
        obj.delete()
        messages.success(request, f'{label} 삭제 완료')
    except Exception as e:
        messages.error(request, f'삭제 실패: {e}')

    return redirect('material:molding_master_list')


@wms_permission_required('can_wms_stock_view')
def molding_master_upload(request):
    """성형 마스터 엑셀 업로드"""
    from .models import MoldingMaster, MoldingMachine
    from decimal import Decimal, InvalidOperation

    if request.method != 'POST':
        return redirect('material:molding_master_list')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, '파일을 선택해주세요.')
        return redirect('material:molding_master_list')

    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active

        # 헤더 매핑 (첫 행)
        headers = [str(cell.value or '').strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        col_map = {}
        field_names = {
            '품번': 'part_no', '품명': 'part_name', '품목': 'item_group',
            '금형구분': 'mold_type', '재료구분': 'material_type',
            '재료품번': 'material_part_no', '재료품명': 'material_name',
            '설비호기': 'machine_code', '설비톤수': 'machine_tonnage',
            'CT(초)': 'cycle_time', 'CT': 'cycle_time',
            '캐비티수': 'cavity', '캐비티': 'cavity', 'CAVITY': 'cavity',
            'ST': 'shot_time',
            '게이트': 'gate_type', 'GATE': 'gate_type',
            'HT': 'hot_runner_time',
            '제품중량(g)': 'product_weight', '제품중량': 'product_weight',
            '공차': 'tolerance',
            '런너중량(g)': 'runner_weight', '런너중량': 'runner_weight',
            '총소요량(g)': 'total_material', '총소요량': 'total_material',
            'ERP수량(kg)': 'erp_qty', 'ERP수량': 'erp_qty',
            '비고': 'remark',
        }

        for idx, h in enumerate(headers):
            if h in field_names:
                col_map[field_names[h]] = idx

        if 'part_no' not in col_map:
            messages.error(request, '엑셀에 "품번" 열이 없습니다.')
            return redirect('material:molding_master_list')

        # 설비호기 맵 미리 로딩
        machine_map = {m.code: m for m in MoldingMachine.objects.all()}

        def safe_decimal(val, default='0'):
            if val is None or str(val).strip() == '':
                return Decimal(default)
            try:
                return Decimal(str(val).strip())
            except (InvalidOperation, ValueError):
                return Decimal(default)

        def safe_int(val, default=0):
            if val is None or str(val).strip() == '':
                return default
            try:
                return int(float(str(val).strip()))
            except (ValueError, TypeError):
                return default

        def get_val(row_data, field):
            idx = col_map.get(field)
            if idx is not None and idx < len(row_data):
                return row_data[idx]
            return None

        created = 0
        updated = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = list(row)
            part_no = str(get_val(row_data, 'part_no') or '').strip()
            if not part_no:
                continue

            material_type = str(get_val(row_data, 'material_type') or '').strip()
            machine_code_val = str(get_val(row_data, 'machine_code') or '').strip()
            machine_obj = machine_map.get(machine_code_val)

            defaults = {
                'part_name': str(get_val(row_data, 'part_name') or '').strip(),
                'item_group': str(get_val(row_data, 'item_group') or '').strip(),
                'mold_type': str(get_val(row_data, 'mold_type') or '').strip(),
                'material_part_no': str(get_val(row_data, 'material_part_no') or '').strip(),
                'material_name': str(get_val(row_data, 'material_name') or '').strip(),
                'machine': machine_obj,
                'machine_tonnage': safe_int(get_val(row_data, 'machine_tonnage')),
                'cycle_time': safe_decimal(get_val(row_data, 'cycle_time')),
                'cavity': safe_int(get_val(row_data, 'cavity')),
                'shot_time': safe_decimal(get_val(row_data, 'shot_time')),
                'gate_type': str(get_val(row_data, 'gate_type') or '').strip(),
                'hot_runner_time': safe_decimal(get_val(row_data, 'hot_runner_time')),
                'product_weight': safe_decimal(get_val(row_data, 'product_weight')),
                'tolerance': safe_decimal(get_val(row_data, 'tolerance')),
                'runner_weight': safe_decimal(get_val(row_data, 'runner_weight')),
                'total_material': safe_decimal(get_val(row_data, 'total_material')),
                'erp_qty': safe_decimal(get_val(row_data, 'erp_qty')),
                'remark': str(get_val(row_data, 'remark') or '').strip(),
                'updated_by': request.user,
            }

            try:
                obj, is_created = MoldingMaster.objects.update_or_create(
                    part_no=part_no,
                    material_type=material_type,
                    defaults=defaults,
                )
                if is_created:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append(f'행 {row_num}: {e}')

        wb.close()
        msg = f'업로드 완료 - 신규 {created}건, 수정 {updated}건'
        if errors:
            msg += f', 오류 {len(errors)}건'
        messages.success(request, msg)

    except Exception as e:
        messages.error(request, f'엑셀 처리 오류: {e}')

    return redirect('material:molding_master_list')


@wms_permission_required('can_wms_stock_view')
def molding_master_excel(request):
    """성형 마스터 엑셀 다운로드"""
    from .models import MoldingMaster
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '성형마스터'

    headers = [
        '품번', '품명', '품목', '금형구분', '재료구분',
        '재료품번', '재료품명', '설비호기', '설비톤수',
        'CT(초)', '캐비티수', 'ST', '게이트', 'HT',
        '제품중량(g)', '공차', '런너중량(g)', '총소요량(g)',
        'ERP수량(kg)', '비고',
    ]
    ws.append(headers)

    # 헤더 스타일
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    qs = MoldingMaster.objects.select_related('machine').all()
    for obj in qs:
        ws.append([
            obj.part_no, obj.part_name, obj.item_group, obj.mold_type, obj.material_type,
            obj.material_part_no, obj.material_name,
            obj.machine.code if obj.machine else '',
            obj.machine_tonnage,
            float(obj.cycle_time), obj.cavity, float(obj.shot_time),
            obj.gate_type, float(obj.hot_runner_time),
            float(obj.product_weight), float(obj.tolerance),
            float(obj.runner_weight), float(obj.total_material),
            float(obj.erp_qty), obj.remark,
        ])

    # 열 너비 자동 조절
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 3, 30)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="molding_master.xlsx"'
    wb.save(response)
    return response


@wms_permission_required('can_wms_stock_view')
def api_molding_master_detail(request, pk):
    """성형 마스터 상세 JSON API"""
    from .models import MoldingMaster

    obj = get_object_or_404(MoldingMaster, pk=pk)
    data = {
        'pk': obj.pk,
        'part_no': obj.part_no,
        'part_name': obj.part_name,
        'item_group': obj.item_group,
        'mold_type': obj.mold_type,
        'material_type': obj.material_type,
        'material_part_no': obj.material_part_no,
        'material_name': obj.material_name,
        'machine_code': obj.machine.code if obj.machine else '',
        'machine_tonnage': obj.machine_tonnage,
        'cycle_time': str(obj.cycle_time),
        'cavity': obj.cavity,
        'shot_time': str(obj.shot_time),
        'gate_type': obj.gate_type,
        'hot_runner_time': str(obj.hot_runner_time),
        'product_weight': str(obj.product_weight),
        'tolerance': str(obj.tolerance),
        'runner_weight': str(obj.runner_weight),
        'total_material': str(obj.total_material),
        'erp_qty': str(obj.erp_qty),
        'remark': obj.remark,
    }
    return JsonResponse(data)


# =============================================================================
# 금형 MT(정비) 관리
# =============================================================================

@login_required
@wms_permission_required('can_wms_stock_view')
def mold_mt_dashboard(request):
    """금형 MT 대시보드 - 금형 유지보수/MT 현황"""
    import json as _json
    from django.db.models import Sum, Value, IntegerField
    from django.db.models.functions import Coalesce
    from .models import MoldMaster as MoldMasterModel, MoldMTSetting

    # MT 기준을 DB에서 캐싱 (구분 → 숏트수)
    MT_INTERVAL_MAP = {}
    for s in MoldMTSetting.objects.all():
        MT_INTERVAL_MAP[s.material_type.upper()] = s.grade_a
    if not MT_INTERVAL_MAP:
        MT_INTERVAL_MAP = {'A': 50000, 'B': 30000}

    # annotate로 월별 숏트 합계를 DB에서 한번에 계산
    qs = MoldMasterModel.objects.filter(is_active=True).annotate(
        _monthly_shots=Coalesce(Sum('shot_records__shots'), Value(0), output_field=IntegerField())
    )

    # 필터 (다중 선택 지원: getlist)
    q = request.GET.get('q', '').strip()
    grades = request.GET.getlist('grade')
    item_groups_filter = request.GET.getlist('item_group')
    materials_filter = request.GET.getlist('material_type')
    status_filters = request.GET.getlist('status')

    if q:
        qs = qs.filter(Q(part_no__icontains=q) | Q(mold_name__icontains=q))
    if grades:
        qs = qs.filter(grade__in=grades)
    if item_groups_filter:
        qs = qs.filter(item_group__in=item_groups_filter)
    if materials_filter:
        qs = qs.filter(material_type__in=materials_filter)

    molds_list = list(qs)

    # Python에서 한번에 계산 (DB 추가 조회 없음)
    for m in molds_list:
        m.c_total_shots = m.total_shots_prev + m._monthly_shots
        m.c_remaining = m.guarantee_shots - m.c_total_shots
        m.c_over_guarantee = m.c_total_shots > m.guarantee_shots

        # MT interval: A=50,000 / B=30,000
        interval = MT_INTERVAL_MAP.get(m.grade.upper(), 30000) if m.grade else 30000
        m.c_mt_interval = interval

        shots_since = max(m.c_total_shots - m.last_mt_shots, 0)
        m.c_shots_since_mt = shots_since
        m.c_mt_pct = min(round(shots_since / interval * 100, 1), 100) if interval > 0 else 0
        m.c_is_mt_due = shots_since >= interval

    # 상태별 분류 계산
    total_count = len(molds_list)
    mt_due_count = 0
    over_guarantee_count = 0
    normal_count = 0

    for m in molds_list:
        if m.c_over_guarantee:
            over_guarantee_count += 1
        if m.c_is_mt_due or m.c_mt_pct >= 80:
            mt_due_count += 1
        if not m.c_over_guarantee and not m.c_is_mt_due and m.c_mt_pct < 80:
            normal_count += 1

    # 상태 필터 (다중)
    if status_filters:
        filtered = []
        for m in molds_list:
            if m.c_over_guarantee and 'over' in status_filters:
                filtered.append(m)
            elif (m.c_is_mt_due or m.c_mt_pct >= 80) and not m.c_over_guarantee and 'mt_due' in status_filters:
                filtered.append(m)
            elif not m.c_is_mt_due and m.c_mt_pct < 80 and not m.c_over_guarantee and 'normal' in status_filters:
                filtered.append(m)
        molds_list = filtered

    # 기본 정렬: MT 진행률 높은순, 진행률 0이면 pk순
    molds_list.sort(key=lambda m: (-m.c_mt_pct if m.c_mt_pct > 0 else 0, m.c_mt_pct == 0, m.pk))

    # 페이지네이션
    paginator = Paginator(molds_list, 60)
    page = request.GET.get('page', 1)
    molds = paginator.get_page(page)

    # 자동완성/필터용 데이터 (annotate 없는 순수 쿼리)
    _base = MoldMasterModel.objects.filter(is_active=True)
    part_nos = sorted(set(
        f"{pn} / {mn}" for pn, mn in _base.values_list('part_no', 'mold_name') if pn
    ))
    grade_list = sorted(set(_base.exclude(grade__isnull=True).exclude(grade='').values_list('grade', flat=True)))
    material_list = sorted(set(_base.exclude(material_type__isnull=True).exclude(material_type='').values_list('material_type', flat=True)))
    item_group_list = sorted(set(_base.exclude(item_group__isnull=True).exclude(item_group='').values_list('item_group', flat=True)))

    # MT 기준 설정
    mt_settings = [{'grade': k, 'interval': v} for k, v in MT_INTERVAL_MAP.items()]

    # 페이지네이션용 쿼리스트링 (page 제외)
    from urllib.parse import urlencode
    qs_params = []
    if q:
        qs_params.append(('q', q))
    for g in grades:
        qs_params.append(('grade', g))
    for ig in item_groups_filter:
        qs_params.append(('item_group', ig))
    for mt in materials_filter:
        qs_params.append(('material_type', mt))
    for sf in status_filters:
        qs_params.append(('status', sf))
    pagination_qs = urlencode(qs_params)

    context = {
        'molds': molds,
        'total_count': total_count,
        'mt_due_count': mt_due_count,
        'over_guarantee_count': over_guarantee_count,
        'normal_count': normal_count,
        'q': q,
        'grades': grades,
        'item_groups_filter': item_groups_filter,
        'materials_filter': materials_filter,
        'status_filters': status_filters,
        'pagination_qs': pagination_qs,
        'part_nos_json': _json.dumps(part_nos, ensure_ascii=False),
        'grade_list': grade_list,
        'material_list': material_list,
        'item_group_list': item_group_list,
        'mt_settings': mt_settings,
    }
    return render(request, 'material/mold_mt_dashboard.html', context)


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_mt_upload(request):
    """금형 마스터 엑셀 업로드 (신규 양식: 기종/부품번호/금형명/등급/보증수량/누적숏트/이전MT/MT기준)"""
    from .models import MoldMaster as MoldMasterModel
    import openpyxl

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'success': False, 'error': '파일을 선택해주세요.'})

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        created = 0
        updated = 0

        # 1행=헤더, 2행부터 데이터 (열 순서: 기종/부품번호/금형명/등급/보증수량/누적숏트/이전MT/MT기준)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            part_no = str(row[1] or '').strip()
            if not part_no:
                continue

            def safe_int(val, default=0):
                try:
                    return int(val) if val else default
                except (ValueError, TypeError):
                    return default

            defaults = {
                'item_group': str(row[0] or '').strip(),
                'mold_name': str(row[2] or '').strip() if len(row) > 2 else '',
                'grade': str(row[3] or '').strip().upper() if len(row) > 3 else '',
                'guarantee_shots': safe_int(row[4] if len(row) > 4 else None, 500000),
                'total_shots_prev': safe_int(row[5] if len(row) > 5 else None, 0),
                'last_mt_shots': safe_int(row[6] if len(row) > 6 else None, 0),
            }

            obj, is_created = MoldMasterModel.objects.update_or_create(
                part_no=part_no, defaults=defaults
            )

            if is_created:
                created += 1
            else:
                updated += 1

        return JsonResponse({
            'success': True,
            'message': f'업로드 완료: 신규 {created}건, 수정 {updated}건'
        })

    except Exception as e:
        logger.exception('금형 MT 엑셀 업로드 오류')
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_mt_add(request):
    """금형 수기 등록"""
    from .models import MoldMaster as MoldMasterModel

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    part_no = request.POST.get('part_no', '').strip()
    if not part_no:
        return JsonResponse({'success': False, 'error': '부품번호는 필수입니다.'})

    if MoldMasterModel.objects.filter(part_no=part_no).exists():
        return JsonResponse({'success': False, 'error': f'부품번호 {part_no}가 이미 존재합니다.'})

    def safe_int(val, default=0):
        try:
            return int(val) if val else default
        except (ValueError, TypeError):
            return default

    MoldMasterModel.objects.create(
        item_group=request.POST.get('item_group', '').strip(),
        part_no=part_no,
        mold_name=request.POST.get('mold_name', '').strip(),
        grade=request.POST.get('grade', '').strip().upper(),
        material_type=request.POST.get('material_type', '').strip(),
        cv_count=safe_int(request.POST.get('cv_count'), 1),
        guarantee_shots=safe_int(request.POST.get('guarantee_shots'), 500000),
        total_shots_prev=safe_int(request.POST.get('total_shots_prev'), 0),
        last_mt_shots=safe_int(request.POST.get('last_mt_shots'), 0),
    )

    return JsonResponse({'success': True, 'message': f'{part_no} 등록 완료'})


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_mt_excel(request):
    """금형 마스터 엑셀 다운로드"""
    from .models import MoldMaster as MoldMasterModel, MoldShotRecord
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '금형 유지보수 관리'

    headers = [
        'No', '기종(전체)', '기종(세부)', '금형명', '부품번호',
        '이관일자', '이관처', '보증수명', 'C/V수', '재료구분',
        '총사용숏트수(이전)', '등급',
        '1월', '2월', '3월', '4월', '5월', '6월',
        '7월', '8월', '9월', '10월', '11월', '12월',
        '누적숏트', '잔량', 'MT진행률(%)', '상태'
    ]

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    ids_param = request.GET.get('ids', '').strip()
    if ids_param:
        id_list = [int(x) for x in ids_param.split(',') if x.strip().isdigit()]
        molds = MoldMasterModel.objects.filter(pk__in=id_list).order_by('item_group', 'part_no')
    else:
        molds = MoldMasterModel.objects.filter(is_active=True).order_by('item_group', 'part_no')

    for idx, m in enumerate(molds, 1):
        # 월별 숏트
        monthly = {}
        for rec in m.shot_records.filter(year=2026):
            monthly[rec.month] = rec.shots

        status = '보증초과' if m.is_over_guarantee else ('MT필요' if m.is_mt_due else '정상')

        row_data = [
            idx, m.item_group, m.item_group_detail, m.mold_name, m.part_no,
            m.transfer_date, m.transfer_from, m.guarantee_shots, m.cv_count,
            m.material_type, m.total_shots_prev, m.grade,
        ]
        for mon in range(1, 13):
            row_data.append(monthly.get(mon, 0))
        row_data += [m.total_shots, m.remaining_shots, m.mt_progress_pct, status]

        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col_num, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # 컬럼 너비 자동 조정
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        max_length = len(str(headers[col_num - 1]))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 25)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="mold_mt_management.xlsx"'
    wb.save(response)
    return response


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_mt_complete(request):
    """MT 완료 처리"""
    from .models import MoldMaster as MoldMasterModel, MoldMTLog

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    mold_id = request.POST.get('mold_id')
    mt_date = request.POST.get('mt_date')
    description = request.POST.get('description', '')
    performed_by = request.POST.get('performed_by', '')

    if not mold_id or not mt_date:
        return JsonResponse({'success': False, 'error': '필수 항목이 누락되었습니다.'})

    try:
        from datetime import datetime as dt
        mold = MoldMasterModel.objects.get(pk=mold_id)
        current_shots = mold.total_shots

        MoldMTLog.objects.create(
            mold=mold,
            mt_date=dt.strptime(mt_date, '%Y-%m-%d').date(),
            accumulated_shots=current_shots,
            description=description,
            performed_by=performed_by,
        )

        mold.last_mt_shots = current_shots
        mold.save(update_fields=['last_mt_shots', 'updated_at'])

        return JsonResponse({
            'success': True,
            'message': f'{mold.part_no} MT 완료 처리되었습니다. (누적: {current_shots:,}숏)'
        })
    except MoldMasterModel.DoesNotExist:
        return JsonResponse({'success': False, 'error': '금형을 찾을 수 없습니다.'})
    except Exception as e:
        logger.exception('MT 완료 처리 오류')
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_mt_settings(request):
    """MT 기준 설정 CRUD (구분별 숏트수) - MoldMTSetting을 구분별로 재활용"""
    from .models import MoldMTSetting
    import json

    if request.method == 'GET':
        # MoldMTSetting의 material_type을 구분으로, grade_a를 interval로 사용
        settings_list = []
        for s in MoldMTSetting.objects.all():
            settings_list.append({
                'id': s.pk,
                'grade': s.material_type,
                'interval': s.grade_a,
            })
        return JsonResponse({'success': True, 'data': settings_list})

    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action', 'save')

            if action == 'delete':
                pk = data.get('id')
                MoldMTSetting.objects.filter(pk=pk).delete()
                return JsonResponse({'success': True, 'message': '삭제 완료'})

            grade = data.get('grade', '').strip().upper()
            if not grade:
                return JsonResponse({'success': False, 'error': '구분을 입력해주세요.'})

            interval = int(data.get('interval', 30000))
            pk = data.get('id')

            if pk:
                MoldMTSetting.objects.filter(pk=pk).update(
                    material_type=grade, grade_a=interval
                )
                created = False
            else:
                MoldMTSetting.objects.create(
                    material_type=grade, grade_a=interval,
                    grade_b=interval, grade_c=interval,
                    grade_d=interval, grade_e=interval
                )
                created = True

            return JsonResponse({
                'success': True,
                'message': f'구분 {grade} {"등록" if created else "수정"} 완료 (MT기준: {interval:,})'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': '허용되지 않는 메서드'}, status=405)


@login_required
@wms_permission_required('can_wms_stock_view')
def api_mold_mt_detail(request, pk):
    """금형 MT 상세 JSON API"""
    from .models import MoldMaster as MoldMasterModel

    mold = get_object_or_404(MoldMasterModel, pk=pk)

    # 월별 숏트 이력
    shot_records = list(mold.shot_records.order_by('year', 'month').values(
        'year', 'month', 'shots', 'source'
    ))

    # MT 이력
    mt_logs = list(mold.mt_logs.order_by('-mt_date').values(
        'mt_date', 'accumulated_shots', 'description', 'performed_by', 'created_at'
    ))

    data = {
        'pk': mold.pk,
        'part_no': mold.part_no,
        'mold_name': mold.mold_name,
        'item_group': mold.item_group,
        'item_group_detail': mold.item_group_detail,
        'transfer_date': mold.transfer_date,
        'transfer_from': mold.transfer_from,
        'guarantee_shots': mold.guarantee_shots,
        'cv_count': mold.cv_count,
        'material_type': mold.material_type,
        'total_shots_prev': mold.total_shots_prev,
        'grade': mold.grade,
        'total_shots': mold.total_shots,
        'remaining_shots': mold.remaining_shots,
        'shots_since_last_mt': mold.shots_since_last_mt,
        'mt_interval': mold.mt_interval,
        'mt_progress_pct': mold.mt_progress_pct,
        'is_mt_due': mold.is_mt_due,
        'is_over_guarantee': mold.is_over_guarantee,
        'last_mt_shots': mold.last_mt_shots,
        'remark': mold.remark,
        'shot_records': shot_records,
        'mt_logs': mt_logs,
    }
    return JsonResponse(data)


@login_required
@wms_permission_required('can_wms_stock_view')
def api_mold_mt_edit(request, pk):
    """금형 마스터 수정 API"""
    from .models import MoldMaster as MoldMasterModel

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    mold = get_object_or_404(MoldMasterModel, pk=pk)

    def safe_int(val, default=0):
        try:
            return int(val) if val else default
        except (ValueError, TypeError):
            return default

    mold.part_no = request.POST.get('part_no', mold.part_no).strip()
    mold.mold_name = request.POST.get('mold_name', mold.mold_name).strip()
    mold.item_group = request.POST.get('item_group', mold.item_group).strip()
    mold.item_group_detail = request.POST.get('item_group_detail', mold.item_group_detail).strip()
    mold.material_type = request.POST.get('material_type', mold.material_type).strip()
    mold.grade = request.POST.get('grade', mold.grade).strip().upper()
    mold.cv_count = safe_int(request.POST.get('cv_count'), mold.cv_count)
    mold.guarantee_shots = safe_int(request.POST.get('guarantee_shots'), mold.guarantee_shots)
    mold.total_shots_prev = safe_int(request.POST.get('total_shots_prev'), mold.total_shots_prev)
    mold.last_mt_shots = safe_int(request.POST.get('last_mt_shots'), mold.last_mt_shots)
    mold.transfer_date = request.POST.get('transfer_date', mold.transfer_date).strip()
    mold.transfer_from = request.POST.get('transfer_from', mold.transfer_from).strip()
    mold.remark = request.POST.get('remark', mold.remark).strip()
    mold.save()

    return JsonResponse({'success': True, 'message': f'{mold.part_no} 수정 완료'})


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_mt_erp_sync(request):
    """ERP 생산실적 데이터로 금형 숏트수 동기화 (양품+불량 포함)"""
    from .models import MoldMaster as MoldMasterModel, MoldShotRecord
    from .erp_api import call_erp_api
    from django.conf import settings as django_settings
    import calendar

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    year = int(request.POST.get('year', timezone.localtime().year))
    month = int(request.POST.get('month', timezone.localtime().month))
    _, days_in_month = calendar.monthrange(year, month)

    date_from = f"{year}{month:02d}01"
    date_to = f"{year}{month:02d}{days_in_month:02d}"

    # 생산실적 API 사용 (badQt 포함)
    body = {
        'coCd': django_settings.ERP_COMPANY_CODE,
        'wrDtFrom': date_from,
        'wrDtTo': date_to,
    }
    ok, raw, err = call_erp_api('/apiproxy/api20A03S00901', body)
    if not ok:
        return JsonResponse({'success': False, 'error': f'ERP 조회 실패: {err}'})

    data = (raw.get('resultData', []) or []) if raw else []
    if not data:
        return JsonResponse({'success': False, 'error': '해당 기간 생산실적 데이터가 없습니다.'})

    # 금형 마스터 매핑 (part_no -> MoldMaster)
    mold_map = {}
    for m in MoldMasterModel.objects.filter(is_active=True):
        mold_map[m.part_no] = m

    # ERP 데이터 집계: part_no별 총 생산수량 (workQt = 양품+불량 포함)
    part_qty_agg = {}
    for r in data:
        item_cd = (r.get('itemCd') or '').strip()
        work_qt = int(float(r.get('workQt', 0) or 0))
        if item_cd and work_qt > 0:
            part_qty_agg[item_cd] = part_qty_agg.get(item_cd, 0) + work_qt

    synced = 0
    skipped = 0

    try:
        with transaction.atomic():
            for part_no, total_qty in part_qty_agg.items():
                mold = mold_map.get(part_no)
                if not mold:
                    skipped += 1
                    continue

                # 숏트수 = 생산수량 / C/V수
                cv = mold.cv_count if mold.cv_count > 0 else 1
                shots = total_qty // cv

                if shots > 0:
                    MoldShotRecord.objects.update_or_create(
                        mold=mold,
                        year=year,
                        month=month,
                        defaults={'shots': shots, 'source': 'ERP'}
                    )
                    synced += 1

        return JsonResponse({
            'success': True,
            'message': f'ERP 동기화 완료: {year}년 {month}월 / 동기화 {synced}건, 미매칭 {skipped}건'
        })
    except Exception as e:
        logger.exception('금형 MT ERP 동기화 오류')
        return JsonResponse({'success': False, 'error': str(e)})


def _send_repair_notification(event_type, obj):
    """금형 수리 알림 발송 헬퍼"""
    from admin_app.notifications import send_notification
    priority_labels = {'A': 'A(긴급)', 'B': 'B(보통)', 'C': 'C(낮음)'}
    requester_email = obj.requested_by.email if obj.requested_by and obj.requested_by.email else ''

    context_vars = {
        'part_no': obj.part_no,
        'mold_name': obj.mold_name,
        'item_group': obj.item_group,
        'priority': priority_labels.get(obj.priority, obj.priority),
        'status': obj.get_status_display(),
        'repair_types': ', '.join(obj.repair_type_list) or '-',
        'request_content': obj.request_content,
        'requester': (obj.requested_by.get_full_name() or obj.requested_by.username) if obj.requested_by else '-',
        'repair_by': obj.repair_by or '-',
        'repair_content': obj.repair_content or '-',
        'received_date': str(obj.received_date or '-'),
        'expected_date': str(obj.expected_date or '-'),
        'completed_date': str(obj.completed_date or '-'),
    }
    send_notification(event_type, context_vars=context_vars,
                      reference_id=obj.pk, requester_email=requester_email)


# =============================================================================
# 금형 수리 의뢰 / 이력
# =============================================================================
@login_required
@wms_permission_required('can_wms_stock_view')
def mold_repair_list(request):
    """금형 수리 의뢰 목록"""
    from .models import MoldRepairRequest, MOLD_REPAIR_STATUS
    import json

    status_filter = request.GET.get('status', '')
    q = request.GET.get('q', '').strip()

    qs = MoldRepairRequest.objects.all().select_related('mold', 'requested_by')
    if status_filter:
        qs = qs.filter(status=status_filter)
    if q:
        qs = qs.filter(Q(part_no__icontains=q) | Q(mold_name__icontains=q) | Q(request_content__icontains=q))

    # 상태별 건수
    from django.db.models import Count
    _counts = dict(MoldRepairRequest.objects.values_list('status').annotate(c=Count('id')).values_list('status', 'c'))
    total_count = sum(_counts.values())
    status_tabs = [(code, label, _counts.get(code, 0)) for code, label in MOLD_REPAIR_STATUS]

    paginator = Paginator(list(qs), 30)
    page = request.GET.get('page', 1)
    repairs = paginator.get_page(page)

    # 금형 마스터 데이터 (자동완성용)
    from .models import MoldMaster
    mold_autocomplete = json.dumps([
        {'part_no': m['part_no'], 'mold_name': m['mold_name'], 'item_group': m['item_group']}
        for m in MoldMaster.objects.filter(is_active=True).values('part_no', 'mold_name', 'item_group')
    ], ensure_ascii=False)

    context = {
        'repairs': repairs,
        'status_filter': status_filter,
        'q': q,
        'status_tabs': status_tabs,
        'total_count': total_count,
        'mold_autocomplete': mold_autocomplete,
    }
    return render(request, 'material/mold_repair_list.html', context)


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_repair_create(request):
    """금형 수리 의뢰 등록"""
    from .models import MoldRepairRequest, MoldMaster
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    part_no = request.POST.get('part_no', '').strip()
    if not part_no:
        return JsonResponse({'success': False, 'error': '품번을 입력해주세요.'})

    # 금형 마스터 연결
    mold = MoldMaster.objects.filter(part_no=part_no).first()

    repair_types = request.POST.getlist('repair_types')

    mold_name = request.POST.get('mold_name', mold.mold_name if mold else '').strip()
    item_group = request.POST.get('item_group', mold.item_group if mold else '').strip()
    priority = request.POST.get('priority', 'B')
    request_content = request.POST.get('request_content', '').strip()

    obj = MoldRepairRequest.objects.create(
        mold=mold,
        part_no=part_no,
        mold_name=mold_name,
        item_group=item_group,
        priority=priority,
        request_content=request_content,
        repair_types=','.join(repair_types),
        requested_by=request.user,
    )

    # 알림 발송
    try:
        _send_repair_notification('MOLD_REPAIR_REQUESTED', obj)
    except Exception as e:
        logger.warning(f'수리의뢰 알림 발송 실패: {e}')

    return JsonResponse({'success': True, 'message': f'수리의뢰 등록 완료 (#{obj.pk})', 'id': obj.pk})


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_repair_history_create(request):
    """금형 수리 이력 직접 등록 (의뢰 단계 건너뛰고 완료 상태로)"""
    from .models import MoldRepairRequest, MoldMaster

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    part_no = request.POST.get('part_no', '').strip()
    if not part_no:
        return JsonResponse({'success': False, 'error': '품번을 입력해주세요.'})

    mold = MoldMaster.objects.filter(part_no=part_no).first()
    repair_types = request.POST.getlist('repair_types')

    def safe_decimal(val):
        try:
            return float(val) if val else 0
        except (ValueError, TypeError):
            return 0

    def safe_int(val):
        try:
            return int(val) if val else 0
        except (ValueError, TypeError):
            return 0

    def safe_date(val):
        val = (val or '').strip()
        if not val:
            return None
        try:
            from datetime import datetime
            return datetime.strptime(val, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return None

    obj = MoldRepairRequest.objects.create(
        mold=mold,
        part_no=part_no,
        mold_name=request.POST.get('mold_name', mold.mold_name if mold else '').strip(),
        item_group=request.POST.get('item_group', mold.item_group if mold else '').strip(),
        priority=request.POST.get('priority', 'B'),
        status='COMPLETED',  # 바로 완료 상태
        request_content=request.POST.get('request_content', '').strip(),
        repair_types=','.join(repair_types),
        requested_by=request.user,
        received_date=safe_date(request.POST.get('received_date')),
        repair_request_date=safe_date(request.POST.get('repair_request_date')),
        completed_date=safe_date(request.POST.get('completed_date')) or timezone.now().date(),
        repair_content=request.POST.get('repair_content', '').strip(),
        repair_by=request.POST.get('repair_by', '').strip(),
        hr_milling=safe_decimal(request.POST.get('hr_milling')),
        hr_lathe=safe_decimal(request.POST.get('hr_lathe')),
        hr_grinding=safe_decimal(request.POST.get('hr_grinding')),
        hr_welding=safe_decimal(request.POST.get('hr_welding')),
        hr_high_speed=safe_decimal(request.POST.get('hr_high_speed')),
        hr_edm=safe_decimal(request.POST.get('hr_edm')),
        hr_wire=safe_decimal(request.POST.get('hr_wire')),
        hr_mt=safe_decimal(request.POST.get('hr_mt')),
        hr_polishing=safe_decimal(request.POST.get('hr_polishing')),
        hr_assembly=safe_decimal(request.POST.get('hr_assembly')),
        hr_other=safe_decimal(request.POST.get('hr_other')),
        cost_welding=safe_int(request.POST.get('cost_welding')),
        cost_tapping=safe_int(request.POST.get('cost_tapping')),
        cost_milpin=safe_int(request.POST.get('cost_milpin')),
        cost_purchase=safe_int(request.POST.get('cost_purchase')),
        cost_outsource=safe_int(request.POST.get('cost_outsource')),
        shot_count=safe_int(request.POST.get('shot_count')),
        first_article=request.POST.get('first_article', '').strip(),
        ng_content=request.POST.get('ng_content', '').strip(),
    )

    return JsonResponse({
        'success': True,
        'message': f'{part_no} 수리이력 등록 완료 (완료 상태)',
        'id': obj.pk,
    })


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_repair_update(request, pk):
    """금형 수리 의뢰 수정 (접수/수리/완료 처리 포함)"""
    from .models import MoldRepairRequest

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    obj = get_object_or_404(MoldRepairRequest, pk=pk)
    old_status = obj.status

    # 상태 변경
    new_status = request.POST.get('status', '')
    if new_status:
        obj.status = new_status

    # 일정
    for field in ['received_date', 'repair_request_date', 'expected_date', 'completed_date']:
        val = request.POST.get(field, '').strip()
        if val:
            setattr(obj, field, val)

    # 수리 내용
    repair_content = request.POST.get('repair_content', '').strip()
    if repair_content:
        obj.repair_content = repair_content
    repair_by = request.POST.get('repair_by', '').strip()
    if repair_by:
        obj.repair_by = repair_by

    # 수리유형
    repair_types = request.POST.getlist('repair_types')
    if repair_types:
        obj.repair_types = ','.join(repair_types)

    # 중요도
    priority = request.POST.get('priority', '')
    if priority:
        obj.priority = priority

    # 사내수리 HR
    def safe_decimal(val):
        try:
            return float(val) if val else 0
        except (ValueError, TypeError):
            return 0

    for hr_field in ['hr_milling', 'hr_lathe', 'hr_grinding', 'hr_welding', 'hr_high_speed',
                     'hr_edm', 'hr_wire', 'hr_mt', 'hr_polishing', 'hr_assembly', 'hr_other']:
        val = request.POST.get(hr_field, '')
        if val != '':
            setattr(obj, hr_field, safe_decimal(val))

    # 외주금액
    def safe_int(val):
        try:
            return int(val) if val else 0
        except (ValueError, TypeError):
            return 0

    for cost_field in ['cost_welding', 'cost_tapping', 'cost_milpin', 'cost_purchase', 'cost_outsource']:
        val = request.POST.get(cost_field, '')
        if val != '':
            setattr(obj, cost_field, safe_int(val))

    # 기타
    shot_count = request.POST.get('shot_count', '')
    if shot_count != '':
        obj.shot_count = safe_int(shot_count)
    first_article = request.POST.get('first_article', '').strip()
    if first_article:
        obj.first_article = first_article
    ng_content = request.POST.get('ng_content', '').strip()
    if ng_content:
        obj.ng_content = ng_content

    obj.save()

    # 상태 변경 시 알림 발송
    if new_status and new_status != old_status:
        status_event_map = {
            'REQUESTED': 'MOLD_REPAIR_REQUESTED',
            'RECEIVED': 'MOLD_REPAIR_RECEIVED',
            'IN_PROGRESS': 'MOLD_REPAIR_IN_PROGRESS',
            'COMPLETED': 'MOLD_REPAIR_COMPLETED',
        }
        event_type = status_event_map.get(new_status)
        if event_type:
            try:
                _send_repair_notification(event_type, obj)
            except Exception as e:
                logger.warning(f'수리 알림 발송 실패: {e}')

    status_display = obj.get_status_display()
    return JsonResponse({'success': True, 'message': f'{obj.part_no} {status_display} 처리 완료'})


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_repair_delete(request, pk):
    """금형 수리 의뢰 삭제"""
    from .models import MoldRepairRequest

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST만 허용'}, status=405)

    obj = get_object_or_404(MoldRepairRequest, pk=pk)
    part_no = obj.part_no
    obj.delete()
    return JsonResponse({'success': True, 'message': f'{part_no} 수리의뢰 삭제 완료'})


@login_required
@wms_permission_required('can_wms_stock_view')
def mold_repair_detail(request, pk):
    """금형 수리 의뢰 상세 JSON"""
    from .models import MoldRepairRequest

    obj = get_object_or_404(MoldRepairRequest, pk=pk)

    data = {
        'pk': obj.pk,
        'part_no': obj.part_no,
        'mold_name': obj.mold_name,
        'item_group': obj.item_group,
        'priority': obj.priority,
        'priority_display': obj.get_priority_display(),
        'status': obj.status,
        'status_display': obj.get_status_display(),
        'request_content': obj.request_content,
        'repair_types': obj.repair_type_list,
        'requested_by': obj.requested_by.get_full_name() or obj.requested_by.username if obj.requested_by else '-',
        'requested_at': timezone.localtime(obj.requested_at).strftime('%Y-%m-%d %H:%M') if obj.requested_at else '',
        'received_date': str(obj.received_date or ''),
        'repair_request_date': str(obj.repair_request_date or ''),
        'expected_date': str(obj.expected_date or ''),
        'completed_date': str(obj.completed_date or ''),
        'repair_content': obj.repair_content,
        'repair_by': obj.repair_by,
        'hr_milling': float(obj.hr_milling), 'hr_lathe': float(obj.hr_lathe),
        'hr_grinding': float(obj.hr_grinding), 'hr_welding': float(obj.hr_welding),
        'hr_high_speed': float(obj.hr_high_speed), 'hr_edm': float(obj.hr_edm),
        'hr_wire': float(obj.hr_wire), 'hr_mt': float(obj.hr_mt),
        'hr_polishing': float(obj.hr_polishing), 'hr_assembly': float(obj.hr_assembly),
        'hr_other': float(obj.hr_other),
        'total_hr': obj.total_hr,
        'cost_welding': obj.cost_welding, 'cost_tapping': obj.cost_tapping,
        'cost_milpin': obj.cost_milpin, 'cost_purchase': obj.cost_purchase,
        'cost_outsource': obj.cost_outsource,
        'total_outsource_cost': obj.total_outsource_cost,
        'shot_count': obj.shot_count,
        'first_article': obj.first_article,
        'ng_content': obj.ng_content,
    }
    return JsonResponse(data)